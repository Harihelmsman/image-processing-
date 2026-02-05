#!/usr/bin/env python3
"""
Batch Advanced Labeled Editor - Process entire folders of images
Navigate with A (previous) and D (next) keys
Saves output to separate folder
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime


class EditMode(Enum):
    """Available editing modes"""
    HIGHLIGHT = "highlight"
    BLUR = "blur"
    PIXELATE = "pixelate"
    DARKEN = "darken"
    GRAYSCALE = "grayscale"
    INVERT = "invert"
    OUTLINE = "outline"


class BatchLabeledEditor:
    """Batch editor for processing multiple images in a folder"""
    
    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)
        
        # Setup output folder
        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{timestamp}"
        
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # Load all image files
        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No images found in {input_folder}")
        
        self.current_index = 0
        self.total_images = len(self.image_files)
        
        # Editor state
        self.original_image = None
        self.scaled_image = None
        self.display_image = None
        self.output_image = None
        self.circles = []
        self.drawing = False
        self.center = None
        self.current_radius = 0
        self.current_label = ""
        self.label_input_mode = False
        
        # Current editing mode
        self.current_mode = EditMode.HIGHLIGHT
        
        # Effect parameters
        self.blur_kernel = 25
        self.pixelate_size = 10
        self.highlight_alpha = 0.4
        
        # Label settings
        self.label_font = cv2.FONT_HERSHEY_SIMPLEX
        self.label_scale = 0.5
        self.label_thickness = 1
        self.show_labels = True
        
        # Mode colors
        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR: (255, 0, 0),
            EditMode.PIXELATE: (0, 0, 255),
            EditMode.DARKEN: (128, 128, 128),
            EditMode.GRAYSCALE: (200, 200, 200),
            EditMode.INVERT: (255, 255, 0),
            EditMode.OUTLINE: (0, 255, 255),
        }
        
        # Track saved images and their states
        self.saved_status = {}
        self.image_states = {}  # Store circle data for each image
        
        # Window setup
        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)
        
        # Load first image
        self._load_current_image()
        self._print_instructions()
    
    def _load_image_files(self):
        """Load all image files from folder"""
        extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}
        files = []
        
        for ext in extensions:
            files.extend(self.input_folder.glob(f'*{ext}'))
            files.extend(self.input_folder.glob(f'*{ext.upper()}'))
        
        return sorted(files)
    
    def _load_current_image(self):
        """Load current image and restore its state if previously edited"""
        current_file = self.image_files[self.current_index]
        
        self.original_image = cv2.imread(str(current_file))
        if self.original_image is None:
            print(f"Error loading: {current_file}")
            return False
        
        self._scale_image()
        
        # Check if this image was previously edited (has saved state)
        if current_file.name in self.image_states:
            # Restore previous editing state
            self.circles = self.image_states[current_file.name].copy()
            print(f"\n✓ Loaded: {current_file.name} ({self.current_index + 1}/{self.total_images}) - RESTORED PREVIOUS EDITS")
        else:
            # Fresh image, no previous edits
            self.circles = []
            print(f"\nLoaded: {current_file.name} ({self.current_index + 1}/{self.total_images})")
        
        # Reset input state
        self.drawing = False
        self.label_input_mode = False
        self.current_label = ""
        
        # Apply effects if there are circles
        self.output_image = self.scaled_image.copy()
        if self.circles:
            self._apply_all_effects()
        
        self.display_image = self.output_image.copy()
        
        return True
    
    def _print_instructions(self):
        """Print usage instructions"""
        print("\n" + "="*70)
        print("BATCH LABELED EDITOR - FOLDER PROCESSING")
        print("="*70)
        print(f"\nInput Folder:  {self.input_folder}")
        print(f"Output Folder: {self.output_folder}")
        print(f"Total Images:  {self.total_images}")
        print("\n🔄 Auto-Save Feature:")
        print("  - Your work is automatically saved in memory when you navigate")
        print("  - Press 'A' or 'D' to move between images - edits are preserved!")
        print("  - Press 'S' to permanently save to disk")
        print("\n🎨 Editing Modes:")
        for idx, mode in enumerate(EditMode, 1):
            print(f"  {idx}: {mode.value.upper()}")
        print("\n⌨️  Navigation Controls:")
        print("  A         - Previous image (work auto-saved)")
        print("  D         - Next image (work auto-saved)")
        print("  S         - Save current image to disk")
        print("  SHIFT+S   - Save to disk and go to next image")
        print("\n✏️  Editing Controls:")
        print("  1-7       - Switch editing mode")
        print("  C         - Clear all circles on current image")
        print("  U         - Undo last circle")
        print("  L         - List all labeled objects")
        print("  E         - Edit label of last circle")
        print("  T         - Toggle label visibility")
        print("\n💾 Label Input:")
        print("  Type      - Enter label text (appears above circle)")
        print("  ENTER     - Confirm label")
        print("  ESC       - Skip label")
        print("\n🚪 Other:")
        print("  Q         - Quit and save progress")
        print("  SHIFT+Q   - Quit without saving current")
        print("\n💡 Status Indicators:")
        print("  ✓ SAVED          - Image saved to disk")
        print("  ✎ EDITED         - Has edits but not saved to disk")
        print("  NO EDITS         - Fresh/untouched image")
        print("="*70 + "\n")
    
    def _scale_image(self):
        """Scale image for display if needed"""
        max_height = 900
        max_width = 1400
        
        height, width = self.original_image.shape[:2]
        self.scale_factor = 1.0
        
        if height > max_height or width > max_width:
            scale_h = max_height / height
            scale_w = max_width / width
            self.scale_factor = min(scale_h, scale_w)
            
            new_width = int(width * self.scale_factor)
            new_height = int(height * self.scale_factor)
            
            self.scaled_image = cv2.resize(
                self.original_image,
                (new_width, new_height),
                interpolation=cv2.INTER_AREA
            )
        else:
            self.scaled_image = self.original_image.copy()
    
    def _mouse_callback(self, event, x, y, flags, param):
        """Handle mouse events"""
        if self.label_input_mode:
            return
        
        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center = (x, y)
            self.current_radius = 0
            
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            self.current_radius = int(np.sqrt((x - self.center[0])**2 + 
                                             (y - self.center[1])**2))
            self._update_display()
            
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                self._enter_label_input_mode()
    
    def _enter_label_input_mode(self):
        """Enter label input mode"""
        self.label_input_mode = True
        self.current_label = ""
        print(f"\n→ Circle drawn in {self.current_mode.value.upper()} mode")
        print("  Enter label (or ESC to skip):")
        self._update_display_with_input()
    
    def _exit_label_input_mode(self, save=True):
        """Exit label input mode"""
        self.label_input_mode = False
        
        if save:
            label = self.current_label.strip()
            self.circles.append({
                'center': self.center,
                'radius': self.current_radius,
                'mode': self.current_mode,
                'label': label
            })
            
            if label:
                print(f"  ✓ Added: '{label}' [{self.current_mode.value}]")
            else:
                print(f"  ✓ Added unlabeled [{self.current_mode.value}]")
            
            self._apply_all_effects()
        else:
            print("  ✗ Cancelled")
        
        self.current_label = ""
        self._update_display()
    
    def _apply_effect(self, image, circle):
        """Apply specific effect to circular region"""
        mask = np.zeros(image.shape[:2], dtype=np.uint8)
        cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
        
        mode = circle['mode']
        
        if mode == EditMode.HIGHLIGHT:
            highlighted = cv2.addWeighted(
                image, 1 - self.highlight_alpha,
                np.full_like(image, 255), self.highlight_alpha, 0
            )
            image = np.where(mask[:, :, np.newaxis] == 255, highlighted, image)
        
        elif mode == EditMode.BLUR:
            blurred = cv2.GaussianBlur(image, (self.blur_kernel, self.blur_kernel), 0)
            image = np.where(mask[:, :, np.newaxis] == 255, blurred, image)
        
        elif mode == EditMode.PIXELATE:
            h, w = image.shape[:2]
            temp = cv2.resize(
                image,
                (w // self.pixelate_size, h // self.pixelate_size),
                interpolation=cv2.INTER_LINEAR
            )
            pixelated = cv2.resize(temp, (w, h), interpolation=cv2.INTER_NEAREST)
            image = np.where(mask[:, :, np.newaxis] == 255, pixelated, image)
        
        elif mode == EditMode.DARKEN:
            darkened = cv2.addWeighted(image, 0.5, np.zeros_like(image), 0.5, 0)
            image = np.where(mask[:, :, np.newaxis] == 255, darkened, image)
        
        elif mode == EditMode.GRAYSCALE:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            gray_bgr = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
            image = np.where(mask[:, :, np.newaxis] == 255, gray_bgr, image)
        
        elif mode == EditMode.INVERT:
            inverted = cv2.bitwise_not(image)
            image = np.where(mask[:, :, np.newaxis] == 255, inverted, image)
        
        return image
    
    def _apply_all_effects(self):
        """Apply all effects"""
        self.output_image = self.scaled_image.copy()
        
        for circle in self.circles:
            self.output_image = self._apply_effect(self.output_image, circle)
            
            # Draw circle border
            color = self.mode_colors[circle['mode']]
            cv2.circle(self.output_image, circle['center'],
                      circle['radius'], color, 2)
    
    def _draw_label(self, image, circle, number):
        """Draw label for a circle"""
        if not self.show_labels or not circle['label']:
            return
        
        center = circle['center']
        radius = circle['radius']
        label = circle['label']
        
        # Position label
        label_x = center[0] - radius
        label_y = center[1] - radius - 15
        
        # Adjust if out of bounds
        if label_y < 25:
            label_y = center[1] + radius + 30
        if label_x < 10:
            label_x = 10
        
        # Format label
        mode_short = circle['mode'].value[:3].upper()
        full_label = f"#{number} [{mode_short}] {label}"
        
        # Get text size
        (text_w, text_h), baseline = cv2.getTextSize(
            full_label, self.label_font, self.label_scale, self.label_thickness
        )
        
        # Draw background
        padding = 4
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     (0, 0, 0), -1)
        
        # Draw border
        color = self.mode_colors[circle['mode']]
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     color, 1)
        
        # Draw text
        cv2.putText(image, full_label, (label_x, label_y),
                   self.label_font, self.label_scale,
                   (255, 255, 255), self.label_thickness)
        
        # Draw connector line
        line_start = (label_x + text_w // 2, label_y + baseline + padding)
        cv2.line(image, line_start, center, color, 1)
    
    def _draw_typing_label(self, image, center, radius, label, mode):
        """Draw label text above circle in real-time while typing"""
        # Position label
        label_x = center[0] - radius
        label_y = center[1] - radius - 15
        
        # Adjust if out of bounds
        if label_y < 25:
            label_y = center[1] + radius + 30
        if label_x < 10:
            label_x = 10
        
        # Show current text with cursor
        mode_short = mode.value[:3].upper()
        display_label = f"[{mode_short}] {label}_"
        
        # Get text size with larger font
        (text_w, text_h), baseline = cv2.getTextSize(
            display_label, self.label_font, self.label_scale + 0.2, self.label_thickness + 1
        )
        
        # Draw background with mode color
        padding = 6
        color = self.mode_colors[mode]
        
        # Dark background
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     (0, 0, 0), -1)
        
        # Bright colored border
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     color, 2)
        
        # Draw text
        cv2.putText(image, display_label, (label_x, label_y),
                   self.label_font, self.label_scale + 0.2,
                   (255, 255, 255), self.label_thickness + 1)
        
        # Draw connector line
        line_start = (label_x + text_w // 2, label_y + baseline + padding)
        cv2.line(image, line_start, center, color, 2)
    
    def _update_display(self):
        """Update display"""
        self.display_image = self.output_image.copy()
        
        # Draw labels
        for idx, circle in enumerate(self.circles):
            self._draw_label(self.display_image, circle, idx + 1)
        
        # Draw current circle
        if self.drawing and self.current_radius > 0:
            color = self.mode_colors[self.current_mode]
            cv2.circle(self.display_image, self.center,
                      self.current_radius, color, 2)
        
        # Draw UI
        self._draw_ui()
        cv2.imshow(self.window_name, self.display_image)
    
    def _update_display_with_input(self):
        """Update display during label input"""
        self.display_image = self.output_image.copy()
        
        # Draw existing labels
        for idx, circle in enumerate(self.circles):
            self._draw_label(self.display_image, circle, idx + 1)
        
        # Draw current circle
        color = self.mode_colors[self.current_mode]
        cv2.circle(self.display_image, self.center,
                  self.current_radius, color, 3)
        
        # Draw the label text ABOVE the circle in real-time
        if self.current_label:
            self._draw_typing_label(self.display_image, self.center,
                                   self.current_radius, self.current_label,
                                   self.current_mode)
        
        # Draw input box
        self._draw_input_box()
        cv2.imshow(self.window_name, self.display_image)
    
    def _draw_input_box(self):
        """Draw label input box"""
        h, w = self.display_image.shape[:2]
        box_height = 70
        box_y = h - box_height
        
        # Background
        overlay = self.display_image.copy()
        cv2.rectangle(overlay, (0, box_y), (w, h), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.8, self.display_image, 0.2, 0, self.display_image)
        
        # Border
        color = self.mode_colors[self.current_mode]
        cv2.rectangle(self.display_image, (0, box_y), (w, h), color, 2)
        
        # Text
        mode_text = f"[{self.current_mode.value.upper()}]"
        cv2.putText(self.display_image, mode_text, (10, box_y + 25),
                   self.label_font, 0.6, color, 2)
        
        prompt = "Label:"
        cv2.putText(self.display_image, prompt, (150, box_y + 25),
                   self.label_font, 0.6, (255, 255, 255), 1)
        
        input_text = self.current_label + "_"
        cv2.putText(self.display_image, input_text, (10, box_y + 55),
                   self.label_font, 0.7, (0, 255, 255), 2)
    
    def _draw_ui(self):
        """Draw UI overlay"""
        overlay = self.display_image.copy()
        h, w = self.display_image.shape[:2]
        
        # Top bar - Mode and image info
        bar_height = 80
        cv2.rectangle(overlay, (0, 0), (w, bar_height), (0, 0, 0), -1)
        
        # Mode indicator
        mode_text = f"Mode: {self.current_mode.value.upper()}"
        cv2.putText(overlay, mode_text, (15, 30),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, self.mode_colors[self.current_mode], 2)
        
        # Image navigation info
        current_file = self.image_files[self.current_index]
        nav_text = f"Image: {self.current_index + 1}/{self.total_images}"
        cv2.putText(overlay, nav_text, (15, 60),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)
        
        # Filename
        filename = current_file.name
        if len(filename) > 50:
            filename = filename[:47] + "..."
        cv2.putText(overlay, filename, (250, 30),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Status indicator - show both saved and edited status
        has_edits = current_file.name in self.image_states or len(self.circles) > 0
        is_saved = self.saved_status.get(current_file.name, False)
        
        if is_saved:
            status_text = "✓ SAVED"
            status_color = (0, 255, 0)
        elif has_edits:
            status_text = "✎ EDITED (NOT SAVED)"
            status_color = (0, 165, 255)  # Orange
        else:
            status_text = "NO EDITS"
            status_color = (100, 100, 100)  # Gray
        
        cv2.putText(overlay, status_text, (w - 250, 30),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, status_color, 2)
        
        # Objects count
        obj_text = f"Objects: {len(self.circles)}"
        cv2.putText(overlay, obj_text, (250, 60),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Labels status
        labels_text = f"Labels: {'ON' if self.show_labels else 'OFF'}"
        cv2.putText(overlay, labels_text, (400, 60),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Bottom bar - Navigation hints
        nav_hint = "◄ A: Previous | D: Next ► | S: Save | Shift+S: Save+Next | Q: Quit"
        cv2.putText(overlay, nav_hint, (15, h - 15),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        cv2.addWeighted(overlay, 0.7, self.display_image, 0.3, 0, self.display_image)
    
    def _previous_image(self):
        """Go to previous image"""
        if self.current_index > 0:
            # Save current state in memory before leaving
            current_file = self.image_files[self.current_index]
            if self.circles:
                self.image_states[current_file.name] = [circle.copy() for circle in self.circles]
                print(f"  → Auto-saved current work state")
            
            self.current_index -= 1
            self._load_current_image()
            self._update_display()
        else:
            print("Already at first image")
    
    def _next_image(self):
        """Go to next image"""
        if self.current_index < self.total_images - 1:
            # Save current state in memory before leaving
            current_file = self.image_files[self.current_index]
            if self.circles:
                self.image_states[current_file.name] = [circle.copy() for circle in self.circles]
                print(f"  → Auto-saved current work state")
            
            self.current_index += 1
            self._load_current_image()
            self._update_display()
        else:
            print("Already at last image")
    
    def _list_labels(self):
        """List all labels"""
        print("\n" + "="*70)
        print(f"LABELED OBJECTS - {self.image_files[self.current_index].name}")
        print("="*70)
        if not self.circles:
            print("No objects marked yet.")
        else:
            for idx, circle in enumerate(self.circles, 1):
                label = circle['label'] if circle['label'] else "(no label)"
                mode = circle['mode'].value
                print(f"#{idx}: {label}")
                print(f"     Mode: {mode}, Position: {circle['center']}, "
                      f"Radius: {circle['radius']}px")
        print("="*70 + "\n")
    
    def _edit_last_label(self):
        """Edit last circle's label"""
        if not self.circles:
            print("No objects to edit!")
            return
        
        last_circle = self.circles[-1]
        current_label = last_circle['label']
        
        print(f"\nCurrent label: '{current_label}'")
        print("Enter new label (ESC to cancel):")
        
        self.current_label = current_label
        self.label_input_mode = True
        self._update_display_with_input()
        
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            
            if key == 27:  # ESC
                print("  ✗ Edit cancelled")
                self.label_input_mode = False
                self.current_label = ""
            elif key == 13:  # ENTER
                last_circle['label'] = self.current_label.strip()
                print(f"  ✓ Updated to: '{self.current_label.strip()}'")
                self.label_input_mode = False
                self.current_label = ""
                self._apply_all_effects()
            elif key == 8:  # BACKSPACE
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key)
                self._update_display_with_input()
        
        self._update_display()
    
    def save_current(self):
        """Save current image and labels"""
        current_file = self.image_files[self.current_index]
        
        # Store the current state in memory for this image
        self.image_states[current_file.name] = [circle.copy() for circle in self.circles]
        
        # Prepare output filename
        output_image_path = self.output_folder / current_file.name
        output_json_path = output_image_path.with_suffix('.json')
        output_txt_path = output_image_path.with_suffix('.txt')
        
        # Save image
        if self.scale_factor != 1.0:
            h, w = self.original_image.shape[:2]
            final = cv2.resize(self.output_image, (w, h),
                              interpolation=cv2.INTER_LANCZOS4)
        else:
            final = self.output_image
        
        cv2.imwrite(str(output_image_path), final)
        
        # Save labels as JSON
        labels_data = {
            'source_image': current_file.name,
            'timestamp': datetime.now().isoformat(),
            'objects': []
        }
        
        for idx, circle in enumerate(self.circles, 1):
            labels_data['objects'].append({
                'id': idx,
                'label': circle['label'],
                'mode': circle['mode'].value,
                'center': circle['center'],
                'radius': circle['radius']
            })
        
        with open(output_json_path, 'w') as f:
            json.dump(labels_data, f, indent=2)
        
        # Save labels as text
        with open(output_txt_path, 'w') as f:
            f.write(f"Labeled Objects - {current_file.name}\n")
            f.write("="*70 + "\n\n")
            for idx, circle in enumerate(self.circles, 1):
                label = circle['label'] if circle['label'] else "(no label)"
                f.write(f"#{idx}: {label}\n")
                f.write(f"  Mode: {circle['mode'].value}\n")
                f.write(f"  Position: {circle['center']}\n")
                f.write(f"  Radius: {circle['radius']}px\n\n")
        
        # Mark as saved
        self.saved_status[current_file.name] = True
        
        print(f"\n✓ Saved: {output_image_path.name}")
        print(f"  - Image: {output_image_path}")
        print(f"  - Labels: {output_json_path}")
        print(f"  - Text: {output_txt_path}")
        print(f"  - State stored in memory (will be restored when you return)")
    
    def generate_summary(self):
        """Generate summary report for all processed images"""
        summary_path = self.output_folder / "processing_summary.txt"
        excel_path = self.output_folder / "processing_summary.xlsx"
        
        # Text summary
        with open(summary_path, 'w') as f:
            f.write("BATCH PROCESSING SUMMARY\n")
            f.write("="*70 + "\n\n")
            f.write(f"Input Folder:  {self.input_folder}\n")
            f.write(f"Output Folder: {self.output_folder}\n")
            f.write(f"Total Images:  {self.total_images}\n")
            f.write(f"Processed:     {len(self.saved_status)}\n\n")
            
            f.write("Saved Images:\n")
            f.write("-"*70 + "\n")
            for filename in sorted(self.saved_status.keys()):
                f.write(f"  ✓ {filename}\n")
            
            if len(self.saved_status) < self.total_images:
                f.write("\nNot Processed:\n")
                f.write("-"*70 + "\n")
                for img_file in self.image_files:
                    if img_file.name not in self.saved_status:
                        f.write(f"  ✗ {img_file.name}\n")
        
        print(f"\n✓ Summary saved: {summary_path}")
        
        # Excel summary
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Processing Summary"
            
            # Header row styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Set headers
            headers = ["Image Name", "Number of Labels", "Label Names"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Collect data from JSON files
            row = 2
            for img_file in sorted(self.image_files):
                if img_file.name not in self.saved_status:
                    continue
                
                json_path = self.output_folder / img_file.with_suffix('.json').name
                
                if json_path.exists():
                    with open(json_path, 'r') as f:
                        data = json.load(f)
                    
                    # Image name
                    ws.cell(row=row, column=1).value = img_file.name
                    ws.cell(row=row, column=1).border = border
                    
                    # Number of labels
                    num_labels = len(data['objects'])
                    ws.cell(row=row, column=2).value = num_labels
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2).border = border
                    
                    # Label names (comma-separated)
                    labels = [obj['label'] for obj in data['objects'] if obj['label']]
                    label_names = ", ".join(labels) if labels else "(no labels)"
                    ws.cell(row=row, column=3).value = label_names
                    ws.cell(row=row, column=3).border = border
                    
                    row += 1
            
            # Add summary statistics at the bottom
            row += 1
            summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            summary_font = Font(bold=True, size=11)
            
            ws.cell(row=row, column=1).value = "SUMMARY"
            ws.cell(row=row, column=1).font = summary_font
            ws.cell(row=row, column=1).fill = summary_fill
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws.cell(row=row, column=1).value = "Total Images Processed"
            ws.cell(row=row, column=2).value = len(self.saved_status)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
            ws.cell(row=row, column=1).value = "Total Objects Labeled"
            # Calculate total objects
            total_objects = 0
            for img_file in self.image_files:
                if img_file.name in self.saved_status:
                    json_path = self.output_folder / img_file.with_suffix('.json').name
                    if json_path.exists():
                        with open(json_path, 'r') as f:
                            data = json.load(f)
                        total_objects += len(data['objects'])
            ws.cell(row=row, column=2).value = total_objects
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 60
            
            # Save Excel file
            wb.save(str(excel_path))
            print(f"✓ Excel summary saved: {excel_path}")
            
        except ImportError:
            print("⚠️  openpyxl not installed. Excel summary not created.")
            print("   Install with: pip install openpyxl")
        except Exception as e:
            print(f"⚠️  Could not create Excel summary: {e}")
    
    def run(self):
        """Main loop"""
        self._update_display()
        
        while True:
            key = cv2.waitKey(1) & 0xFF
            
            # Handle label input mode
            if self.label_input_mode:
                if key == 27:  # ESC
                    self._exit_label_input_mode(save=True)
                elif key == 13:  # ENTER
                    self._exit_label_input_mode(save=True)
                elif key == 8:  # BACKSPACE
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key)
                    self._update_display_with_input()
                continue
            
            # Navigation
            if key == ord('a'):  # Previous
                self._previous_image()
            elif key == ord('d'):  # Next
                self._next_image()
            
            # Save
            elif key == ord('s'):
                if self.circles:
                    self.save_current()
                    self._update_display()
                else:
                    print("No objects to save")
            elif key == ord('S'):  # Save and next
                if self.circles:
                    self.save_current()
                    self._next_image()
                else:
                    print("No objects to save")
            
            # Editing
            elif key == ord('c'):
                self.circles.clear()
                self.output_image = self.scaled_image.copy()
                self._update_display()
                print("✓ Cleared all circles")
            elif key == ord('u'):
                if self.circles:
                    removed = self.circles.pop()
                    label = removed['label'] if removed['label'] else "(unlabeled)"
                    print(f"✓ Removed: {label}")
                    self._apply_all_effects()
                    self._update_display()
            elif key == ord('l'):
                self._list_labels()
            elif key == ord('e'):
                self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels: {'ON' if self.show_labels else 'OFF'}")
                self._update_display()
            
            # Mode switching
            elif ord('1') <= key <= ord('7'):
                modes = list(EditMode)
                self.current_mode = modes[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._update_display()
            
            # Quit
            elif key == ord('q'):
                break
            elif key == ord('Q'):  # Quit without saving current
                break
        
        cv2.destroyAllWindows()
        
        # Generate summary
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Processed {len(self.saved_status)}/{self.total_images} images")
        else:
            print("\n⚠️  No images were saved")


def main():
    parser = argparse.ArgumentParser(
        description="Batch Advanced Labeled Editor - Process entire folders"
    )
    parser.add_argument("input_folder", type=str, 
                       help="Input folder containing images")
    parser.add_argument("--output", "-o", type=str, default=None,
                       help="Output folder (default: labeled_output_TIMESTAMP)")
    
    args = parser.parse_args()
    
    try:
        editor = BatchLabeledEditor(args.input_folder, args.output)
        editor.run()
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())