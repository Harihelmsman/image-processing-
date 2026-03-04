"""
Batch Labeled Editor - Refactored & Production Ready
Changes from original:
  [NEW v2 additions]
    1. Centered image on dark canvas
       - Image is centred in a fixed MAX_DISPLAY_WIDTH × MAX_DISPLAY_HEIGHT frame
       - Surroundings are CANVAS_BG_COLOR (near-black) — clearly empty/blank
       - A subtle 1-px border marks the image boundary
       - Mouse hit-testing updated to account for the centring offset

    2. Anti-aliased text (no more broken pixel fonts)
       - Every cv2.putText call now passes lineType=cv2.LINE_AA
       - MIN_FONT_SCALE constant prevents text from shrinking below readability
       - Applies to labels, UI bar, input box — everywhere

    3. Interpolation toggle  ( I  key )
       - INTER_LINEAR  (LN) — smooth bilinear zoom, default for normal editing
       - INTER_NEAREST (NR) — hard pixel edges, ideal for pixel-level inspection
       - Display status shown next to zoom level: "Zoom: 3.50× [NR]"
       - Toggle only affects the viewport; saved images are unchanged
"""

import cv2
import csv
import json
import sys
import time
import argparse
from collections import OrderedDict
from contextlib import contextmanager
from datetime import datetime
from enum import Enum
from pathlib import Path

import numpy as np


class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


@contextmanager
def _temporary_circles(editor, replacement):
    original = editor.circles
    try:
        editor.circles = replacement
        yield
    finally:
        editor.circles = original


class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50

    ZOOM_DEBOUNCE_MS = 50
    MIN_ZOOM         = 0.5
    MAX_ZOOM         = 10.0

    MEMORY_EFFICIENT_MODE = True
    MAX_CACHED_STATES     = 5

    LABEL_BG_ALPHA        = 0.5
    LABEL_BORDER_THICKNESS = 1

    DEFAULT_BLUR_KERNEL    = 25
    DEFAULT_PIXELATE_SIZE  = 10
    DEFAULT_HIGHLIGHT_ALPHA = 0.4

    MAX_DISPLAY_HEIGHT = 900
    MAX_DISPLAY_WIDTH  = 1400
    # Near-black background so the image area is always clearly delimited
    CANVAS_BG_COLOR = (18, 18, 18)

    IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}

    MODE_COLORS = {
        EditMode.HIGHLIGHT: (0, 255, 0),
        EditMode.BLUR:      (255, 0, 0),
        EditMode.PIXELATE:  (0, 0, 255),
        EditMode.DARKEN:    (128, 128, 128),
        EditMode.GRAYSCALE: (200, 200, 200),
        EditMode.INVERT:    (255, 255, 0),
        EditMode.OUTLINE:   (0, 255, 255),
    }

    # Anti-aliased line type — eliminates jagged/broken pixel artefacts on text
    TEXT_LINE_TYPE = cv2.LINE_AA
    MIN_FONT_SCALE = 0.35

    def __init__(self, input_folder: str, output_folder: str | None = None):
        self.input_folder = Path(input_folder)

        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.current_index = 0
        self.total_images  = len(self.image_files)

        self.original_image: np.ndarray | None = None
        self.scaled_image:   np.ndarray | None = None
        self.output_image:   np.ndarray | None = None
        self.display_image:  np.ndarray | None = None
        self.scale_factor = 1.0

        self.circles: list[dict] = []
        self._undo_stack: list[list[dict]] = []
        self._redo_stack: list[list[dict]] = []
        self.drawing         = False
        self.center: tuple[int, int] | None = None
        self.current_radius  = 0
        self.current_label   = ""
        self.label_input_mode = False

        self.zoom_level  = 1.0
        self.pan_x       = 0
        self.pan_y       = 0
        self.is_panning  = False
        self.pan_start_x = 0
        self.pan_start_y = 0
        self.last_zoom_time = 0.0

        # Interpolation toggle: False=LINEAR (smooth), True=NEAREST (pixel-crisp)
        self._use_nearest_interp: bool = False

        self.current_mode   = EditMode.HIGHLIGHT
        self._blur_kernel   = self.DEFAULT_BLUR_KERNEL
        self._pixelate_size = self.DEFAULT_PIXELATE_SIZE
        self.highlight_alpha = self.DEFAULT_HIGHLIGHT_ALPHA

        self.label_font  = cv2.FONT_HERSHEY_SIMPLEX
        self.show_labels = True

        self.saved_status: dict[str, bool]        = {}
        self.image_states: OrderedDict[str, dict] = OrderedDict()

        # Canvas centring offsets (pixels); recomputed per image
        self._canvas_w     = self.MAX_DISPLAY_WIDTH
        self._canvas_h     = self.MAX_DISPLAY_HEIGHT
        self._img_offset_x = 0
        self._img_offset_y = 0

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")

        self._print_instructions()

    # ─── validated properties ────────────────────────────────────────────────

    @property
    def blur_kernel(self) -> int:
        return self._blur_kernel

    @blur_kernel.setter
    def blur_kernel(self, value: int):
        if not isinstance(value, int): raise TypeError("Blur kernel must be int")
        if value <= 0:                 raise ValueError("Blur kernel must be positive")
        if value % 2 == 0:            raise ValueError("Blur kernel must be odd")
        self._blur_kernel = value

    @property
    def pixelate_size(self) -> int:
        return self._pixelate_size

    @pixelate_size.setter
    def pixelate_size(self, value: int):
        if not isinstance(value, int): raise TypeError("Pixelate size must be int")
        if value <= 0:                 raise ValueError("Pixelate size must be positive")
        self._pixelate_size = value

    # ─── undo / redo ────────────────────────────────────────────────────────

    def _push_undo(self):
        self._undo_stack.append([c.copy() for c in self.circles])
        self._redo_stack.clear()

    def _undo(self):
        if not self._undo_stack:
            print("Nothing to undo"); return
        self._redo_stack.append([c.copy() for c in self.circles])
        self.circles = self._undo_stack.pop()
        self._apply_all_effects(); self._update_display()
        print(f"↩ Undo — {len(self.circles)} object(s) remain")

    def _redo(self):
        if not self._redo_stack:
            print("Nothing to redo"); return
        self._undo_stack.append([c.copy() for c in self.circles])
        self.circles = self._redo_stack.pop()
        self._apply_all_effects(); self._update_display()
        print(f"↪ Redo — {len(self.circles)} object(s)")

    # ─── LRU cache ──────────────────────────────────────────────────────────

    def _touch_state(self, filename: str):
        if filename in self.image_states:
            self.image_states.move_to_end(filename)

    def _evict_old_states(self):
        if not self.MEMORY_EFFICIENT_MODE: return
        while len(self.image_states) > self.MAX_CACHED_STATES:
            self.image_states.popitem(last=False)

    # ─── file discovery ─────────────────────────────────────────────────────

    def _load_image_files(self) -> list[Path]:
        found: set[Path] = set()
        for ext in self.IMAGE_EXTENSIONS:
            found.update(self.input_folder.glob(f'*{ext}'))
            found.update(self.input_folder.glob(f'*{ext.upper()}'))
        files = sorted(found)
        if not files: return []

        valid: list[Path] = []
        for f in files:
            img = cv2.imread(str(f))
            if img is not None and img.shape[0] >= self.MIN_IMAGE_SIZE and img.shape[1] >= self.MIN_IMAGE_SIZE:
                valid.append(f)
            else:
                print(f"⚠️  Skipping {f.name}: too small or unreadable")

        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n{'='*70}\n⚠️  WARNING: Large batch — {len(valid)} images found\n{'='*70}\n")
            if input("   Continue anyway? (y/N): ").strip().lower() != 'y':
                print("Operation cancelled."); sys.exit(0)

        return valid

    # ─── image loading ───────────────────────────────────────────────────────

    def _load_current_image(self) -> bool:
        if self.current_index >= len(self.image_files): return False
        cf = self.image_files[self.current_index]
        try:
            img = cv2.imread(str(cf))
            if img is None: raise IOError("cv2.imread returned None")
            h, w = img.shape[:2]
            if h < self.MIN_IMAGE_SIZE or w < self.MIN_IMAGE_SIZE:
                raise ValueError(f"Image too small: {w}x{h}")

            self.original_image = img
            self._scale_image()
            self._recalc_canvas_offsets()
            self.zoom_level = 1.0
            self.pan_x = self.pan_y = 0

            self._undo_stack.clear(); self._redo_stack.clear()
            self.circles = self._restore_circles(cf)
            self.drawing = self.label_input_mode = False
            self.current_label = ""
            self.output_image = self.scaled_image.copy()
            if self.circles: self._apply_all_effects()
            self.display_image = self.output_image.copy()
            return True
        except Exception as e:
            print(f"\n❌ Error loading {cf.name}: {e}")
            if self.current_index < self.total_images - 1:
                print("  Skipping to next image…")
                self.current_index += 1
                return self._load_current_image()
            return False

    def _recalc_canvas_offsets(self):
        """Centre the scaled image inside the fixed canvas."""
        if self.scaled_image is None: return
        ih, iw = self.scaled_image.shape[:2]
        self._canvas_w = self.MAX_DISPLAY_WIDTH
        self._canvas_h = self.MAX_DISPLAY_HEIGHT
        self._img_offset_x = (self._canvas_w - iw) // 2
        self._img_offset_y = (self._canvas_h - ih) // 2

    def _restore_circles(self, cf: Path) -> list[dict]:
        jp = self.output_folder / cf.with_suffix('.json').name
        if jp.exists():
            try:
                with open(jp) as f: data = json.load(f)
                circles = []
                for obj in data.get('objects', []):
                    try:    mode = EditMode(obj.get('mode', 'highlight'))
                    except: mode = EditMode.HIGHLIGHT
                    circles.append({'center': tuple(obj['center']), 'radius': obj['radius'],
                                    'mode': mode, 'label': obj.get('label', '')})
                print(f"✓ {cf.name} ({self.current_index+1}/{self.total_images}) — restored {len(circles)} objects")
                self._touch_state(cf.name); return circles
            except Exception as e:
                print(f"⚠️  Error reading JSON: {e}")
        if cf.name in self.image_states:
            print(f"✓ {cf.name} ({self.current_index+1}/{self.total_images}) — from memory")
            self._touch_state(cf.name)
            return [c.copy() for c in self.image_states[cf.name]['circles']]
        print(f"  {cf.name} ({self.current_index+1}/{self.total_images})")
        return []

    def _scale_image(self):
        h, w = self.original_image.shape[:2]
        self.scale_factor = 1.0
        if h > self.MAX_DISPLAY_HEIGHT or w > self.MAX_DISPLAY_WIDTH:
            self.scale_factor = min(self.MAX_DISPLAY_HEIGHT / h, self.MAX_DISPLAY_WIDTH / w)
            nw, nh = int(w * self.scale_factor), int(h * self.scale_factor)
            self.scaled_image = cv2.resize(self.original_image, (nw, nh), interpolation=cv2.INTER_NEAREST)
        else:
            self.scaled_image = self.original_image.copy()

    # ─── effects ────────────────────────────────────────────────────────────

    def _apply_effect(self, image: np.ndarray, circle: dict) -> np.ndarray:
        mask = np.zeros(image.shape[:2], dtype=np.uint8)
        cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
        sel  = mask[:, :, np.newaxis] == 255
        mode = circle['mode']
        try:
            if mode == EditMode.HIGHLIGHT:
                layer = cv2.addWeighted(image, 1 - self.highlight_alpha,
                                        np.full_like(image, 255), self.highlight_alpha, 0)
                image = np.where(sel, layer, image)
            elif mode == EditMode.BLUR:
                k = self.blur_kernel
                image = np.where(sel, cv2.GaussianBlur(image, (k, k), 0), image)
            elif mode == EditMode.PIXELATE:
                h, w = image.shape[:2]; p = self.pixelate_size
                small = cv2.resize(image, (max(1,w//p), max(1,h//p)), interpolation=cv2.INTER_NEAREST)
                image = np.where(sel, cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST), image)
            elif mode == EditMode.DARKEN:
                image = np.where(sel, cv2.addWeighted(image, .5, np.zeros_like(image), .5, 0), image)
            elif mode == EditMode.GRAYSCALE:
                gray = cv2.cvtColor(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                image = np.where(sel, gray, image)
            elif mode == EditMode.INVERT:
                image = np.where(sel, cv2.bitwise_not(image), image)
        except Exception as e:
            print(f"⚠️  Effect error ({mode.value}): {e}")
        return image

    def _apply_all_effects(self):
        self.output_image = self.scaled_image.copy()
        for c in self.circles:
            self.output_image = self._apply_effect(self.output_image, c)
            cv2.circle(self.output_image, c['center'], c['radius'], self.MODE_COLORS[c['mode']], 2)

    # ─── label drawing (all text uses LINE_AA) ───────────────────────────────

    def _get_dynamic_label_params(self, image_size):
        d = min(image_size)
        if d < 200: return max(self.MIN_FONT_SCALE, 0.3), 1
        if d < 400: return max(self.MIN_FONT_SCALE, 0.4), 1
        if d < 800: return max(self.MIN_FONT_SCALE, 0.5), 2
        return 0.6, 2

    def _draw_transparent_bg(self, image, pt1, pt2, color=(0,0,0), alpha=None):
        if alpha is None: alpha = self.LABEL_BG_ALPHA
        ov = image.copy(); cv2.rectangle(ov, pt1, pt2, color, -1)
        cv2.addWeighted(ov, alpha, image, 1 - alpha, 0, image)

    def _check_label_collision(self, new_rect, existing):
        x1,y1,x2,y2 = new_rect
        for ex1,ey1,ex2,ey2 in existing:
            b=5
            if not (x2+b < ex1 or x1-b > ex2 or y2+b < ey1 or y1-b > ey2): return True
        return False

    def _find_label_position(self, center, radius, text_w, text_h, baseline,
                              padding, image_size, existing_rects):
        img_h, img_w = image_size
        tw = text_w + 2*padding; th = text_h + baseline + 2*padding
        cx, cy = center
        attempts = [
            (cx-radius-tw//2, cy-radius-th-10), (cx-radius-tw//2, cy+radius+20),
            (cx-radius-tw-10, cy-th//2),         (cx+radius+10,    cy-th//2),
            (cx+radius+10,    cy-radius-th-10),  (cx-radius-tw-10, cy-radius-th-10),
            (cx+radius+10,    cy+radius+20),      (cx-radius-tw-10, cy+radius+20),
        ]
        for offset in (0, 30, 60, 90, 120):
            for bx, by in attempts:
                for dx in (0, offset, -offset):
                    for dy in (0, offset, -offset):
                        px, py = bx+dx, by+dy; ly = py+text_h
                        rect = (px-padding, ly-text_h-padding, px+text_w+padding, ly+baseline+padding)
                        if (rect[0]>=padding and rect[1]>=padding and
                                rect[2]<=img_w-padding and rect[3]<=img_h-padding and
                                not self._check_label_collision(rect, existing_rects)):
                            return px, ly, rect
        px = max(padding, min(cx, img_w-tw-padding)); py = padding+text_h
        return px, py, (px-padding, py-text_h-padding, px+text_w+padding, py+baseline+padding)

    def _draw_all_labels_smart(self, image: np.ndarray):
        if not self.show_labels: return
        ih, iw = image.shape[:2]
        scale, thick = self._get_dynamic_label_params((ih, iw))
        pad = max(2, int(4*scale/0.5)); existing = []

        for idx, c in enumerate(self.circles, 1):
            if not c['label']: continue
            ms   = c['mode'].value[:3].upper()
            text = f"#{idx} [{ms}] {c['label']}"
            (tw,th), bl = cv2.getTextSize(text, self.label_font, scale, thick)
            lx, ly, rect = self._find_label_position(c['center'], c['radius'],
                                                      tw, th, bl, pad, (ih,iw), existing)
            existing.append(rect)
            color = self.MODE_COLORS[c['mode']]
            self._draw_transparent_bg(image, (int(rect[0]),int(rect[1])), (int(rect[2]),int(rect[3])))
            cv2.rectangle(image, (lx-pad,ly-th-pad), (lx+tw+pad,ly+bl+pad), color, 2)
            cv2.putText(image, text, (int(lx),int(ly)), self.label_font, scale,
                        (255,255,255), thick, lineType=self.TEXT_LINE_TYPE)
            cv2.line(image, (int(lx+tw//2), int(ly+bl+pad)), c['center'], color, 1,
                     lineType=self.TEXT_LINE_TYPE)

    def _draw_typing_label(self, image, center, radius, label, mode):
        ih, iw = image.shape[:2]
        scale, thick = self._get_dynamic_label_params((ih, iw))
        pad = max(4, int(6*scale/0.5))
        ms  = mode.value[:3].upper(); disp = f"[{ms}] {label}_"
        (tw,th), bl = cv2.getTextSize(disp, self.label_font, scale+0.1, thick+1)
        tot_w = tw+2*pad; cx,cy = center
        lx = ly = None
        for px, py in [(cx-radius-tw//2, cy-radius-tot_w-10), (cx-radius-tw//2, cy+radius+20)]:
            if px>=pad and py>=th+pad and px+tot_w<=iw-pad and py+bl+pad<=ih-pad:
                lx, ly = px, py+th; break
        if lx is None:
            lx = max(pad, min(cx-radius, iw-tot_w-pad)); ly = max(th+pad, cy-radius-10)
        lx = int(max(pad, min(lx, iw-tot_w-pad))); ly = int(max(th+pad, min(ly, ih-bl-pad)))
        color = self.MODE_COLORS[mode]
        cv2.rectangle(image, (lx-pad,ly-th-pad), (lx+tw+pad,ly+bl+pad), (0,0,0), -1)
        cv2.rectangle(image, (lx-pad,ly-th-pad), (lx+tw+pad,ly+bl+pad), color, 2)
        cv2.putText(image, disp, (lx,ly), self.label_font, scale+0.1,
                    (255,255,255), thick+1, lineType=self.TEXT_LINE_TYPE)
        cv2.line(image, (lx+tw//2, ly+bl+pad), center, color, 2, lineType=self.TEXT_LINE_TYPE)

    # ─── interpolation toggle ────────────────────────────────────────────────

    def _toggle_interpolation(self):
        """I key: flip between smooth (LINEAR) and pixel-crisp (NEAREST) zoom."""
        self._use_nearest_interp = not self._use_nearest_interp
        label = "NEAREST — pixel inspect" if self._use_nearest_interp else "LINEAR  — smooth"
        print(f"✓ Zoom interpolation → {label}")
        self._update_display()

    def _zoom_interp(self) -> int:
        return cv2.INTER_NEAREST if self._use_nearest_interp else cv2.INTER_LINEAR

    # ─── centred zoom / pan canvas ───────────────────────────────────────────

    def _screen_to_image(self, sx: int, sy: int) -> tuple[int, int]:
        ix = int((sx - self._img_offset_x - self.pan_x) / self.zoom_level)
        iy = int((sy - self._img_offset_y - self.pan_y) / self.zoom_level)
        return ix, iy

    def _get_zoomed_view(self, image: np.ndarray) -> np.ndarray:
        """
        Render the image onto a fixed dark canvas:
          - Image always centred at zoom=1 / pan=0
          - Surrounding area is CANVAS_BG_COLOR (clearly empty, never stretched)
          - Zoom + pan animate the image within that canvas
          - Interpolation mode follows self._use_nearest_interp
        """
        ih, iw = image.shape[:2]
        cw, ch = self._canvas_w, self._canvas_h
        nw = max(1, int(iw * self.zoom_level))
        nh = max(1, int(ih * self.zoom_level))

        zoomed = cv2.resize(image, (nw, nh), interpolation=self._zoom_interp())
        canvas = np.full((ch, cw, 3), self.CANVAS_BG_COLOR, dtype=np.uint8)

        tx = self._img_offset_x + int(self.pan_x)
        ty = self._img_offset_y + int(self.pan_y)

        src_x0 = max(0, -tx);  dst_x0 = max(0, tx)
        src_y0 = max(0, -ty);  dst_y0 = max(0, ty)
        src_x1 = min(nw, cw - tx)
        src_y1 = min(nh, ch - ty)
        cw_ = src_x1 - src_x0
        ch_ = src_y1 - src_y0

        if cw_ > 0 and ch_ > 0:
            canvas[dst_y0:dst_y0+ch_, dst_x0:dst_x0+cw_] = \
                zoomed[src_y0:src_y0+ch_, src_x0:src_x0+cw_]

        # Subtle border around image area
        cv2.rectangle(canvas,
                      (max(0, dst_x0-1), max(0, dst_y0-1)),
                      (min(cw-1, dst_x0+cw_), min(ch-1, dst_y0+ch_)),
                      (55, 55, 55), 1)
        return canvas

    # ─── display update ──────────────────────────────────────────────────────

    def _get_dynamic_ui_height(self, image_height: int) -> int:
        return max(1, max(30, min(100, int(image_height * 0.15))))

    def _update_display(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        if self.drawing and self.current_radius > 0 and self.center is not None:
            cv2.circle(tmp, self.center, self.current_radius, self.MODE_COLORS[self.current_mode], 2)
        self.display_image = self._get_zoomed_view(tmp)
        self._draw_ui()
        cv2.imshow(self.window_name, self.display_image)

    def _update_display_with_input(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        if self.center is not None and self.current_radius > 0:
            cv2.circle(tmp, self.center, self.current_radius, self.MODE_COLORS[self.current_mode], 3)
            self._draw_typing_label(tmp, self.center, self.current_radius, self.current_label, self.current_mode)
        self.display_image = self._get_zoomed_view(tmp)
        self._draw_input_box()
        cv2.imshow(self.window_name, self.display_image)

    def _draw_input_box(self):
        h, w = self.display_image.shape[:2]
        bh = min(70, int(h * 0.15)); by = h - bh
        ov = self.display_image.copy()
        cv2.rectangle(ov, (0,by), (w,h), (0,0,0), -1)
        cv2.addWeighted(ov, 0.8, self.display_image, 0.2, 0, self.display_image)
        color = self.MODE_COLORS[self.current_mode]
        cv2.rectangle(self.display_image, (0,by), (w,h), color, 2)
        fs = max(self.MIN_FONT_SCALE, min(0.6, bh/120))
        cv2.putText(self.display_image, f"[{self.current_mode.value.upper()}]",
                    (10, by+int(bh*.35)), self.label_font, fs, color, 2,
                    lineType=self.TEXT_LINE_TYPE)
        cv2.putText(self.display_image, "Label:",
                    (150, by+int(bh*.35)), self.label_font, fs, (255,255,255), 1,
                    lineType=self.TEXT_LINE_TYPE)
        cv2.putText(self.display_image, self.current_label+"_",
                    (10, by+int(bh*.78)), self.label_font, fs*1.2, (0,255,255), 2,
                    lineType=self.TEXT_LINE_TYPE)

    def _draw_ui(self):
        ov = self.display_image.copy()
        h, w = self.display_image.shape[:2]
        bh = self._get_dynamic_ui_height(h)
        fs = max(self.MIN_FONT_SCALE, min(0.6, bh/180))
        cv2.rectangle(ov, (0,0), (w,bh), (0,0,0), -1)

        color = self.MODE_COLORS[self.current_mode]
        ly = int(bh*.25)
        cv2.putText(ov, f"Mode: {self.current_mode.value.upper()}", (15,ly),
                    cv2.FONT_HERSHEY_SIMPLEX, fs, color, 2, lineType=self.TEXT_LINE_TYPE)
        ly += int(bh*.25)
        tag = "NR" if self._use_nearest_interp else "LN"
        cv2.putText(ov, f"Zoom: {self.zoom_level:.2f}x [{tag}]", (15,ly),
                    cv2.FONT_HERSHEY_SIMPLEX, fs*.8, (100,200,255), 1, lineType=self.TEXT_LINE_TYPE)
        ly += int(bh*.25)
        cv2.putText(ov, f"Image: {self.current_index+1}/{self.total_images}", (15,ly),
                    cv2.FONT_HERSHEY_SIMPLEX, fs*.8, (200,200,200), 1, lineType=self.TEXT_LINE_TYPE)

        fname = self.image_files[self.current_index].name
        mc = max(20, w//20)
        if len(fname) > mc: fname = fname[:mc-3]+"…"
        cv2.putText(ov, fname, (250,int(bh*.25)),
                    cv2.FONT_HERSHEY_SIMPLEX, fs*.8, (200,200,200), 1, lineType=self.TEXT_LINE_TYPE)

        cf = self.image_files[self.current_index]
        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = len(self.circles) > 0 or cf.name in self.image_states
        if is_saved:      status, sc = "SAVED",    (0,255,0)
        elif has_edits:   status, sc = "EDITED",   (0,165,255)
        else:             status, sc = "NO EDITS", (100,100,100)
        cv2.putText(ov, status, (w-180, int(bh*.25)),
                    cv2.FONT_HERSHEY_SIMPLEX, fs, sc, 2, lineType=self.TEXT_LINE_TYPE)

        oc = (0,165,255) if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES else (200,200,200)
        cv2.putText(ov, f"Objects: {len(self.circles)}  Blur:{self.blur_kernel}  Px:{self.pixelate_size}",
                    (250, int(bh*.5)), cv2.FONT_HERSHEY_SIMPLEX, fs*.8, oc, 1,
                    lineType=self.TEXT_LINE_TYPE)

        hint = "Wheel:Zoom | R-Click:Pan | A/D:Nav | I:Interp | S:Save | R:Reset | U:Undo | H:Help | Q:Quit"
        cv2.putText(ov, hint, (15,h-15),
                    cv2.FONT_HERSHEY_SIMPLEX, min(0.45, w/1600), (200,200,200), 1,
                    lineType=self.TEXT_LINE_TYPE)

        cv2.addWeighted(ov, 0.7, self.display_image, 0.3, 0, self.display_image)

    # ─── mouse ──────────────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time()*1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS: return
            self.last_zoom_time = now
            factor = 1.1 if flags > 0 else 0.9
            new_zoom = max(self.MIN_ZOOM, min(self.MAX_ZOOM, self.zoom_level * factor))
            ix, iy = self._screen_to_image(x, y)
            self.zoom_level = new_zoom
            self.pan_x = int(x - self._img_offset_x - ix * self.zoom_level)
            self.pan_y = int(y - self._img_offset_y - iy * self.zoom_level)
            self._update_display(); return

        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning = True
            self.pan_start_x = x - self.pan_x; self.pan_start_y = y - self.pan_y; return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False; return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.pan_x = int(x - self.pan_start_x); self.pan_y = int(y - self.pan_start_y)
            self._update_display(); return

        if self.label_input_mode or self.is_panning: return

        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True; self.center = self._screen_to_image(x,y); self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing and self.center is not None:
            ix, iy = self._screen_to_image(x, y)
            self.current_radius = int(np.hypot(ix-self.center[0], iy-self.center[1]))
            self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠️  Many circles ({len(self.circles)}); recommended ≤{self.MAX_RECOMMENDED_CIRCLES}")
                self._enter_label_input_mode()

    # ─── label input ────────────────────────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True; self.current_label = ""
        print(f"\n→ Circle drawn [{self.current_mode.value.upper()}]  Enter label (ENTER=confirm, ESC=skip):")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save: bool = True):
        self.label_input_mode = False
        if save:
            self._push_undo(); label = self.current_label.strip()
            self.circles.append({'center': self.center, 'radius': self.current_radius,
                                  'mode': self.current_mode, 'label': label})
            print(f"  ✓ Added {'\"'+label+'\"' if label else '(unlabeled)'} [{self.current_mode.value}]")
            self._apply_all_effects()
        else:
            print("  ✗ Cancelled")
        self.current_label = ""; self._update_display()

    # ─── navigation ─────────────────────────────────────────────────────────

    def _guard_navigation(self) -> bool:
        if self.drawing or self.label_input_mode:
            print("⚠️  Complete current action before navigating"); return True
        return False

    def _auto_save_if_needed(self):
        if self.circles: self.save_current(auto_save=True)

    def _previous_image(self):
        if self._guard_navigation(): return
        if self.current_index > 0:
            self._auto_save_if_needed(); self.current_index -= 1
            self._load_current_image(); self._update_display()
        else: print("Already at first image")

    def _next_image(self):
        if self._guard_navigation(): return
        if self.current_index < self.total_images - 1:
            self._auto_save_if_needed(); self.current_index += 1
            self._load_current_image(); self._update_display()
        else: print("Already at last image")

    # ─── copy from previous ─────────────────────────────────────────────────

    def _copy_from_previous(self):
        if self.current_index == 0: print("No previous image"); return
        pf = self.image_files[self.current_index - 1]
        jp = self.output_folder / pf.with_suffix('.json').name
        src = None
        if jp.exists():
            try:
                with open(jp) as f: data = json.load(f)
                src = []
                for obj in data.get('objects', []):
                    try: mode = EditMode(obj.get('mode','highlight'))
                    except: mode = EditMode.HIGHLIGHT
                    src.append({'center': tuple(obj['center']), 'radius': obj['radius'],
                                'mode': mode, 'label': obj.get('label','')})
            except Exception as e: print(f"⚠️  Could not read previous JSON: {e}")
        elif pf.name in self.image_states:
            src = [c.copy() for c in self.image_states[pf.name]['circles']]
        if not src: print("No annotations found on previous image"); return
        self._push_undo(); self.circles.extend(src)
        self._apply_all_effects(); self._update_display()
        print(f"✓ Copied {len(src)} circle(s) from {pf.name}")

    # ─── utility commands ────────────────────────────────────────────────────

    def _list_labels(self):
        fname = self.image_files[self.current_index].name
        print(f"\n{'='*60}\nLABELS — {fname}\n{'='*60}")
        if not self.circles: print("(none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'} | {c['mode'].value} | pos={c['center']} | r={c['radius']}px")
        print('='*60)

    def _show_memory_status(self):
        print(f"\n{'='*60}\nMEMORY STATUS\n{'='*60}")
        print(f"  Cached states : {len(self.image_states)} / {self.MAX_CACHED_STATES}")
        print(f"  Saved to disk : {len(self.saved_status)}")
        print(f"  Total images  : {self.total_images}")
        print(f"  Undo depth    : {len(self._undo_stack)}")
        print(f"  Redo depth    : {len(self._redo_stack)}")
        print(f"  Interpolation : {'NEAREST' if self._use_nearest_interp else 'LINEAR'}")
        print('='*60)

    def _edit_last_label(self):
        if not self.circles: print("No objects to edit"); return
        last = self.circles[-1]
        print(f"\n  Current label: '{last['label']}'")
        print("  Enter new label (ENTER=confirm, ESC=cancel):")
        self.current_label = last['label']; self.label_input_mode = True
        sc, sr = self.center, self.current_radius
        self.center = None; self.current_radius = 0
        self._update_display_with_input()
        while True:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:   self.label_input_mode = False; print("  ✗ Edit cancelled"); break
            elif key == 13: self._push_undo(); last['label'] = self.current_label.strip(); self.label_input_mode = False; print("  ✓ Label updated"); break
            elif key == 8:  self.current_label = self.current_label[:-1]; self._update_display_with_input()
            elif 32 <= key <= 126: self.current_label += chr(key); self._update_display_with_input()
        self.center = sc; self.current_radius = sr; self.current_label = ""
        self._apply_all_effects(); self._update_display()



    # ─── save ────────────────────────────────────────────────────────────────

    def save_current(self, auto_save: bool = False):
        if not self.circles:
            if not auto_save: print("No objects to save"); return
        cf = self.image_files[self.current_index]
        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._touch_state(cf.name); self._evict_old_states()
        try:
            self._save_image_to_disk(cf); self._save_json_to_disk(cf)
            self.saved_status[cf.name] = True
            if not auto_save:
                print(f"\n✓ Saved {cf.name}\n  Image : {self.output_folder/cf.name}\n  Objects: {len(self.circles)}")
            else:
                print(f"  ✓ Auto-saved {len(self.circles)} object(s)")
        except Exception as e:
            print(f"❌ Save error: {e}")

    def _build_scaled_circles(self):
        if self.scale_factor == 1.0: return [c.copy() for c in self.circles]
        sf = self.scale_factor
        return [{'center':(int(c['center'][0]/sf),int(c['center'][1]/sf)),
                 'radius':int(c['radius']/sf),'mode':c['mode'],'label':c['label']}
                for c in self.circles]

    def _save_image_to_disk(self, cf: Path):
        out = self.output_folder / cf.name
        sc  = self._build_scaled_circles()
        if self.scale_factor != 1.0:
            final = self.original_image.copy()
            for s in sc:
                final = self._apply_effect(final, s)
                cv2.circle(final, s['center'], s['radius'], self.MODE_COLORS[s['mode']], 3)
            with _temporary_circles(self, sc): self._draw_all_labels_smart(final)
        else:
            final = self.output_image.copy(); self._draw_all_labels_smart(final)
        if not cv2.imwrite(str(out), final): raise IOError(f"cv2.imwrite failed for {out}")

    def _save_json_to_disk(self, cf: Path):
        jp = (self.output_folder/cf.name).with_suffix('.json')
        sc = self._build_scaled_circles()
        data = {'source_image': cf.name, 'timestamp': datetime.now().isoformat(),
                'objects': [{'id':i,'label':c['label'],'mode':c['mode'].value,
                              'center':list(c['center']),'radius':c['radius']}
                             for i,c in enumerate(sc,1)]}
        with open(jp,'w') as f: json.dump(data, f, indent=2)

    # ─── summary ─────────────────────────────────────────────────────────────

    def generate_summary(self):
        rows = []; total = 0
        for img in sorted(self.image_files):
            if img.name not in self.saved_status: continue
            jp = self.output_folder / img.with_suffix('.json').name
            if not jp.exists(): continue
            with open(jp) as f: data = json.load(f)
            objs = data.get('objects',[])
            labels = [o['label'] for o in objs if o.get('label')]
            total += len(objs)
            rows.append({'name':img.name,'num_objects':len(objs),'labels':','.join(labels) or '(no labels)'})

        cp = self.output_folder/"processing_summary.csv"
        with open(cp,'w',newline='') as f:
            w = csv.DictWriter(f, fieldnames=['Image Name','Objects','Labels'])
            w.writeheader()
            for r in rows: w.writerow({'Image Name':r['name'],'Objects':r['num_objects'],'Labels':r['labels']})
            w.writerow({}); w.writerow({'Image Name':'TOTAL IMAGES','Objects':len(rows),'Labels':''})
            w.writerow({'Image Name':'TOTAL OBJECTS','Objects':total,'Labels':''})
        print(f"✓ CSV summary: {cp}")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Processing Summary"
            hf = PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
            hn = Font(bold=True,color="FFFFFF",size=12)
            for col,hdr in enumerate(['Image Name','Objects','Labels'],1):
                c = ws.cell(row=1,column=col,value=hdr); c.fill=hf; c.font=hn
                c.alignment=Alignment(horizontal='center',vertical='center')
            for ri,r in enumerate(rows,2):
                ws.cell(row=ri,column=1,value=r['name'])
                ws.cell(row=ri,column=2,value=r['num_objects']).alignment=Alignment(horizontal='center')
                ws.cell(row=ri,column=3,value=r['labels'])
            sr = len(rows)+3
            ws.merge_cells(f'A{sr}:C{sr}')
            ws.cell(row=sr,column=1,value='SUMMARY').font=Font(bold=True,size=11)
            ws.cell(row=sr+1,column=1,value='Total Images').font=Font(bold=True); ws.cell(row=sr+1,column=2,value=len(rows))
            ws.cell(row=sr+2,column=1,value='Total Objects').font=Font(bold=True); ws.cell(row=sr+2,column=2,value=total)
            ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=10; ws.column_dimensions['C'].width=50
            xl = self.output_folder/"processing_summary.xlsx"; wb.save(str(xl)); print(f"✓ Excel summary: {xl}")
        except ImportError: print("  (openpyxl not installed — Excel skipped)")
        except Exception as e: print(f"⚠️  Excel error: {e}")

    # ─── help ────────────────────────────────────────────────────────────────

    def _print_instructions(self):
        print("\n" + "="*70)
        print("BATCH LABELED EDITOR  (v2)")
        print("="*70)
        print(f"  Input  : {self.input_folder}")
        print(f"  Output : {self.output_folder}")
        print(f"  Images : {self.total_images}")
        print("\n🔍 Zoom & Pan")
        print("  Wheel            Zoom (0.5×–10×)")
        print("  Right-drag       Pan")
        print("  R                Reset zoom/pan")
        print("  I                Toggle interpolation:")
        print("                     LN = INTER_LINEAR  — smooth (default)")
        print("                     NR = INTER_NEAREST — crisp pixels (inspect mode)")
        print("\n⌨️  Navigation")
        print("  A / D            Previous / Next image")
        print("  S                Save   |   Shift+S  Save+Next")
        print("\n✏️  Editing")
        print("  Left-drag        Draw circle")
        print("  1–7              Switch mode (Highlight/Blur/Pixelate/Darken/Gray/Invert/Outline)")
        print("  [ / ]            Blur kernel −/+")
        print("  { / }            Pixelate size −/+")
        print("  U / Ctrl+Z       Undo   |   Ctrl+Y  Redo")
        print("  C                Clear all")
        print("  E                Edit last label")
        print("  P                Copy circles from previous image")
        print("  T                Toggle label visibility")
        print("  L                List all objects")
        print("  M                Memory status")
        print("\n🚪 Other  |  Q Quit  |  H Help")
        print("="*70 + "\n")

    # ─── main loop ───────────────────────────────────────────────────────────

    def run(self):
        self._update_display()
        dispatch = {
            ord('a'): self._previous_image,  ord('A'): self._previous_image,
            ord('d'): self._next_image,       ord('D'): self._next_image,
            ord('r'): self._reset_zoom,       ord('R'): self._reset_zoom,
            ord('i'): self._toggle_interpolation,
            ord('I'): self._toggle_interpolation,
            ord('l'): self._list_labels,      ord('L'): self._list_labels,
            ord('e'): self._edit_last_label,  ord('E'): self._edit_last_label,
            ord('t'): self._toggle_labels,    ord('T'): self._toggle_labels,
            ord('m'): self._show_memory_status, ord('M'): self._show_memory_status,
            ord('h'): self._print_instructions, ord('H'): self._print_instructions,
            ord('p'): self._copy_from_previous, ord('P'): self._copy_from_previous,
          
        }

        while True:
            key = cv2.waitKey(1) & 0xFF

            if self.label_input_mode:
                if   key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13:  self._exit_label_input_mode(save=True)
                elif key == 8:   self.current_label = self.current_label[:-1]; self._update_display_with_input()
                elif 32<=key<=126: self.current_label += chr(key); self._update_display_with_input()
                continue

            if key == 255: continue

            if key in (ord('q'), ord('Q')):
                cf = self.image_files[self.current_index]
                if self.circles and not self.saved_status.get(cf.name):
                    print("Saving before exit…"); self.save_current(auto_save=True)
                break

            if key == ord('s'):  self.save_current(); self._update_display(); continue
            if key == ord('S'):  self.save_current(); self._next_image(); continue

            if key in (26, ord('u'), ord('U')): self._undo(); continue
            if key == 25: self._redo(); continue

            if key in (ord('c'), ord('C')):
                self._push_undo(); self.circles.clear()
                self.output_image = self.scaled_image.copy()
                self._update_display(); print("✓ Cleared all objects"); continue

            if ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._update_display(); continue

            h = dispatch.get(key)
            if h: h()

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done! {len(self.saved_status)}/{self.total_images} saved → {self.output_folder}")
        else:
            print("\n⚠️  No images were saved")

    def _reset_zoom(self):
        self.zoom_level = 1.0; self.pan_x = self.pan_y = 0
        self._update_display(); print("✓ Zoom reset to 100%")

    def _toggle_labels(self):
        self.show_labels = not self.show_labels
        print(f"✓ Labels: {'ON' if self.show_labels else 'OFF'}")
        self._update_display()


def main():
    parser = argparse.ArgumentParser(description="Batch Labeled Editor")
    parser.add_argument("input_folder")
    parser.add_argument("--output", "-o", default=None)
    args = parser.parse_args()
    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted"); return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}"); import traceback; traceback.print_exc(); return 1
    return 0

if __name__ == "__main__":
    sys.exit(main())
