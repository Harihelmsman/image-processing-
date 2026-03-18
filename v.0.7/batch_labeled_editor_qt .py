#!/usr/bin/env python3
"""
Batch Labeled Editor — Qt Edition  (Full-Resolution · Performance-Optimised)
PySide6 + OpenCV  ·  Full-Res Display + Save  ·  Async Pixmap  ·  Qt Label Overlay

Navigate: A / D    Zoom: Mouse Wheel    Pan: Right-click drag    Draw: Left-click drag

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ARCHITECTURE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ┌─ numpy pipeline (full-resolution, never downscaled) ───────────┐
  │  _display_img  full-res array loaded once per image            │
  │  _effects_img  display_img + ROI-clipped circle effects        │
  │                Add circle  = ONE ROI op  O(r²)                 │
  │                Undo        = ROI snapshot restore  O(r²)       │
  │                Clear       = display_img.copy()   O(1)         │
  └────────────────────────────────────────────────────────────────┘
  ┌─ PixmapConverter (worker thread) ─────────────────────────────┐
  │  BGR numpy → QPixmap runs OFF the main thread.                 │
  │  UI stays responsive even when converting a 180 MB array.     │
  │  Previous converter is cancelled before each new one starts.   │
  └────────────────────────────────────────────────────────────────┘
  ┌─ Qt scene (GPU-composited, zero numpy cost per frame) ─────────┐
  │  QGraphicsPixmapItem  ← full-res QPixmap (static between ops)  │
  │  GhostCircleItem      ← live preview during drag (60fps cap)   │
  │  AnnotationItem × n   ← badge + label + connector via QPainter │
  │    add    = O(1)  append to scene                              │
  │    undo   = O(1)  removeItem                                   │
  │    toggle = O(n)  setVisible                                   │
  │    pan/zoom = 0   pure Qt transform, zero numpy                │
  └────────────────────────────────────────────────────────────────┘
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import gc
import sys
import json
import math
import argparse
from datetime import datetime
from enum import Enum
from pathlib import Path

import cv2
import numpy as np

from PySide6.QtCore import Qt, QThread, Signal, QPointF, QRectF, QTimer
from PySide6.QtGui import (
    QImage, QPixmap, QPainter, QPen, QColor,
    QWheelEvent, QMouseEvent, QBrush, QAction, QPalette, QFont, QFontMetrics,
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QGraphicsItem, QDialog, QLineEdit,
    QDialogButtonBox, QFormLayout, QVBoxLayout, QMessageBox, QToolBar,
    QLabel, QHBoxLayout, QListWidget, QListWidgetItem, QSizePolicy,
    QPushButton,
)

# ── Constants ──────────────────────────────────────────────────────────────────

MAX_DISPLAY_PX          = 4000   # longest-edge cap; looks sharp on any monitor
MAX_BATCH_SIZE          = 200
MIN_IMAGE_SIZE          = 50
MAX_RECOMMENDED_CIRCLES = 50

_FONT_CV = cv2.FONT_HERSHEY_DUPLEX   # OpenCV font — effects layer only
_AA      = cv2.LINE_AA


# ── Edit modes ─────────────────────────────────────────────────────────────────

class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


_MODE_BGR: dict[EditMode, tuple] = {
    EditMode.HIGHLIGHT: (0,   255, 0),
    EditMode.BLUR:      (255, 80,  80),
    EditMode.PIXELATE:  (80,  80,  255),
    EditMode.DARKEN:    (160, 160, 160),
    EditMode.GRAYSCALE: (210, 210, 210),
    EditMode.INVERT:    (255, 255, 0),
    EditMode.OUTLINE:   (0,   220, 255),
}

_MODE_QT: dict[EditMode, QColor] = {
    EditMode.HIGHLIGHT: QColor(0,   255, 0),
    EditMode.BLUR:      QColor(255, 80,  80),
    EditMode.PIXELATE:  QColor(80,  80,  255),
    EditMode.DARKEN:    QColor(160, 160, 160),
    EditMode.GRAYSCALE: QColor(210, 210, 210),
    EditMode.INVERT:    QColor(255, 255, 0),
    EditMode.OUTLINE:   QColor(0,   220, 255),
}


# ── Utility ────────────────────────────────────────────────────────────────────

def bgr_to_pixmap(img: np.ndarray) -> QPixmap:
    """BGR numpy uint8 → QPixmap (zero-copy QImage then detach)."""
    rgb  = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    h, w, ch = rgb.shape
    qi   = QImage(rgb.data, w, h, w * ch, QImage.Format.Format_RGB888)
    return QPixmap.fromImage(qi.copy())


def _rects_overlap(r1: tuple, r2: tuple, buf: int = 4) -> bool:
    return not (r1[2] + buf < r2[0] or r1[0] - buf > r2[2] or
                r1[3] + buf < r2[1] or r1[1] - buf > r2[3])


# ── Background image loader ────────────────────────────────────────────────────

class ImageLoader(QThread):
    """
    Worker thread: disk read + smart downscale.
    Emits loaded(display_img, scale, orig_w, orig_h) once done.
    Full-res array is freed inside the thread — no peak-RAM overlap.
    """
    loaded = Signal(object, float, int, int)
    error  = Signal(str)

    def __init__(self, path: Path, max_px: int = MAX_DISPLAY_PX):
        super().__init__()
        self._path  = path
        self._max_px = max_px
        self._abort = False

    def abort(self): self._abort = True

    def run(self):
        try:
            full = cv2.imread(str(self._path))
            if full is None:
                self.error.emit(f"Cannot read: {self._path.name}"); return
            if self._abort: return

            orig_h, orig_w = full.shape[:2]
            if min(orig_h, orig_w) < MIN_IMAGE_SIZE:
                self.error.emit(f"{self._path.name}: too small"); return

            # ── Full-resolution display — NO downscaling ───────────────────
            # QGraphicsView zooms/pans natively on the QPixmap.  We never
            # need a pre-scaled working copy for display.  scale=1.0 means
            # annotation coords are always identical to original pixel coords,
            # eliminating the entire coordinate-conversion layer.
            if not self._abort:
                self.loaded.emit(full, 1.0, orig_w, orig_h)

        except Exception as exc:
            if not self._abort:
                self.error.emit(str(exc))



# ── Async numpy → QPixmap converter ──────────────────────────────────────────

class PixmapConverter(QThread):
    """
    Converts a BGR numpy array to QPixmap in a worker thread.

    For full-resolution images (60 MB JPEG → 180 MB numpy) the
    cv2.cvtColor + QImage + QPixmap.fromImage pipeline takes 200–500 ms.
    Running it here keeps the main thread — and therefore the Qt event
    loop, AnnotationItem overlays, and all UI interactions — fully
    responsive while the conversion runs.

    Usage pattern (see MainWindow._schedule_pixmap_update):
      • Caller makes a .copy() of the numpy array, passes it to __init__.
      • Old converter (if still running) is cancelled before starting new one.
      • ready signal delivers QPixmap to the main thread via queued connection.
    """
    ready = Signal(object)   # QPixmap

    def __init__(self, img: np.ndarray):
        super().__init__()
        self._img       = img          # already a private copy from caller
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        if self._cancelled:
            return
        try:
            px = bgr_to_pixmap(self._img)
            if not self._cancelled:
                self.ready.emit(px)
        except Exception as e:
            print(f"  PixmapConverter error: {e}")
        finally:
            self._img = None           # release memory ASAP




class GhostCircleItem(QGraphicsItem):
    """Dashed circle shown while the user drags to draw an annotation."""

    def __init__(self):
        super().__init__()
        self._cx = self._cy = self._r = 0.0
        self._color = QColor(0, 255, 0)
        self.setVisible(False)
        self.setZValue(30)

    def update_circle(self, cx: float, cy: float, r: float, color: QColor):
        self.prepareGeometryChange()
        self._cx, self._cy, self._r, self._color = cx, cy, r, color
        self.update()

    def boundingRect(self) -> QRectF:
        d = self._r + 8
        return QRectF(self._cx - d, self._cy - d, 2 * d, 2 * d)

    def paint(self, painter: QPainter, option, widget=None):
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(self._color, 2.0, Qt.PenStyle.DashLine)
        pen.setCosmetic(True)
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawEllipse(QPointF(self._cx, self._cy), self._r, self._r)
        s, pen2 = 7.0, QPen(self._color, 1.0)
        pen2.setCosmetic(True)
        painter.setPen(pen2)
        painter.drawLine(QPointF(self._cx-s, self._cy), QPointF(self._cx+s, self._cy))
        painter.drawLine(QPointF(self._cx, self._cy-s), QPointF(self._cx, self._cy+s))


# ── Annotation overlay item ────────────────────────────────────────────────────

class AnnotationItem(QGraphicsItem):
    """
    One Qt scene item per annotation circle.

    Coordinate system: SCENE PIXELS (= image pixels when display_scale=1.0).

    Why we dropped ItemIgnoresTransformations
    ──────────────────────────────────────────
    With ItemIgnoresTransformations the label rect lives in screen space,
    not scene space.  That makes it impossible to do collision-avoidance
    between labels at construction time — we would need the view transform
    which isn't available (and changes with every zoom/pan).

    Instead, all geometry — circle centre, dot, label box, connector — is
    stored in scene coordinates.  Font and dot sizes are proportional to the
    image dimensions so that at the typical fit-view zoom level the labels
    render at a comfortable fixed screen size (~14-20 px text height).

    Label placement algorithm
    ─────────────────────────
    For every new circle we try candidate positions in a clockwise sweep
    (0°…355° in 5° steps) at four ring distances from the circle edge.
    The first candidate whose bounding box:
      (a) fits within the image bounds, and
      (b) does not overlap any previously placed label rect
    wins.  placed_rects is updated in-place by the caller, so subsequent
    items automatically avoid earlier ones.
    """

    Z_VALUE = 20

    def __init__(
        self,
        index:         int,
        circle:        dict,
        display_scale: float,
        img_w:         int,
        img_h:         int,
        placed_rects:  list,   # label rects already placed — mutated here
        blocked_rects: list = (),  # other circles' bounding rects (avoidance only)
    ):
        super().__init__()
        self.setZValue(self.Z_VALUE)
        # NO ItemIgnoresTransformations — pure scene coordinates throughout

        ds  = display_scale          # always 1.0 in current code
        cx  = circle["center"][0] * ds
        cy  = circle["center"][1] * ds
        r   = max(1.0, circle["radius"] * ds)
        col = _MODE_QT[circle["mode"]]
        lbl = circle["label"] or f"Object #{index}"
        tag = f"#{index}  {lbl}"

        longest = max(img_w, img_h) * ds

        # Font size: proportional to image so text is ~16 px on screen at
        # fit-view zoom regardless of source resolution.
        # fit-view zoom ≈ viewport_px / longest  →  fsize * zoom ≈ 16 px
        # target: fsize ≈ longest / 60  (gives ~16 px on a 1440 px viewport)
        fsize = max(12, int(longest / 60))
        pad   = max(6,  fsize // 3)

        font = QFont()
        font.setPixelSize(fsize)
        font.setBold(True)
        fm = QFontMetrics(font)

        tw = fm.horizontalAdvance(tag)
        th = fm.height()
        W  = tw + 2 * pad + 8    # +8 for left accent bar
        H  = th + 2 * pad

        # Dot radius: 10 % of circle radius, clamped to a readable range
        dot_r = max(int(longest / 180), min(int(r * 0.15), int(longest / 60)))
        dot_r = max(4, dot_r)

        # ── Collision-aware label placement ────────────────────────────────
        iw = img_w * ds;  ih = img_h * ds
        lx, ly = self._place_label(cx, cy, r, W, H, pad, iw, ih,
                                   placed_rects, blocked_rects)
        rect   = (lx, ly, lx + W, ly + H)
        placed_rects.append(rect)

        # Connector: circle-edge → nearest point on label box mid-edge
        label_cx = lx + W / 2
        # Connect to top edge if label is above circle, else bottom edge
        label_ey = ly if (ly + H / 2) < cy else ly + H
        ang  = math.atan2(label_ey - cy, label_cx - cx)
        ex   = cx + r * math.cos(ang)
        ey   = cy + r * math.sin(ang)

        # ── Store all geometry (immutable after __init__) ──────────────────
        self._cx      = cx;   self._cy      = cy
        self._ex      = ex;   self._ey      = ey
        self._lbl_ex  = label_cx;  self._lbl_ey = label_ey
        self._col     = col
        self._tag     = tag
        self._lrect   = QRectF(lx, ly, W, H)
        self._dot_r   = float(dot_r)
        self._accent_w = max(4, dot_r // 2)
        self._font    = font

    # ── Placement search ────────────────────────────────────────────────────

    @staticmethod
    def _place_label(
        cx: float, cy: float, r: float,
        W: float,  H: float,  pad: float,
        iw: float, ih: float,
        placed: list,
        blocked: list = (),
    ) -> tuple[float, float]:
        """
        Return (lx, ly) — top-left of label box — guaranteed non-overlapping.

        placed  — already-placed LABEL rects (x1,y1,x2,y2).
                  Checked with buf=6 for a small breathing gap.
        blocked — OTHER CIRCLES' bounding rects (cx-r, cy-r, cx+r, cy+r).
                  Checked with buf=pad so labels never land inside a
                  neighbour circle's area.  Kept separate from `placed`
                  so undo's _placed_rects.pop() is unaffected.

        Three-phase search:

        Phase 1 — Radial sweep (24 rings × 120 angles, 3° step)
          Rings start at the circle edge and grow up to the image diagonal.
          Compass directions tried first at each ring before the fine grid.

        Phase 2 — Full-image sweep-line scan (fallback)
          For each candidate row, computes blocked x-intervals from both
          labels AND circles, then jumps to the first free x slot.
          Never misses a gap however narrow.

        Phase 3 — Last resort
          Clamp to bottom-right. Overlaps are possible but the label
          stays visible and never crashes.
        """
        gap  = max(16, r * 0.12)
        LBUF = 6    # breathing gap between labels
        CBUF = pad  # gap between label and neighbouring circle edge

        def _free(rect) -> bool:
            if any(_rects_overlap(rect, p, buf=LBUF) for p in placed):
                return False
            if any(_rects_overlap(rect, b, buf=CBUF) for b in blocked):
                return False
            return True

        # ── Phase 1: radial sweep ──────────────────────────────────────────
        diag = math.hypot(iw, ih)
        rings: list[float] = []
        d    = r + gap
        step = max(H * 0.6, 16.0)
        while d < diag * 1.5 and len(rings) < 24:
            rings.append(d)
            d    += step
            step  = min(step * 1.18, H * 3)

        pref   = [0, 180, 270, 90, 315, 225, 45, 135]
        rest   = [a for a in range(0, 360, 3) if a not in pref]
        angles = pref + rest     # 120 unique angles at 3° resolution

        for dist in rings:
            for deg in angles:
                rad = math.radians(deg)
                lx  = cx + dist * math.cos(rad) - W / 2
                ly  = cy + dist * math.sin(rad) - H / 2
                if lx < pad or ly < pad or lx + W > iw - pad or ly + H > ih - pad:
                    continue
                if _free((lx, ly, lx + W, ly + H)):
                    return lx, ly

        # ── Phase 2: sweep-line scan ───────────────────────────────────────
        # For each row y, collect blocked x-intervals from both labels and
        # circle rects using interval arithmetic, then find the first free
        # x slot.  Never misses a valid gap.
        placed_ids   = {id(p) for p in placed}   # O(1) membership; avoids
        all_blockers = list(placed) + list(blocked)  # float-equality bugs
        y = pad
        while y + H <= ih - pad:
            segs: list[tuple] = []
            for b in all_blockers:
                buf = LBUF if id(b) in placed_ids else CBUF
                if b[3] + buf >= y and b[1] - buf <= y + H:
                    segs.append((b[0] - buf, b[2] + buf))
            segs.sort()
            merged: list[tuple] = []
            for lo, hi in segs:
                if merged and lo <= merged[-1][1]:
                    merged[-1] = (merged[-1][0], max(merged[-1][1], hi))
                else:
                    merged.append((lo, hi))
            x = pad
            for lo, hi in merged:
                if x + W <= lo:
                    break
                if hi > x:
                    x = hi
            if x + W <= iw - pad:
                if _free((x, y, x + W, y + H)):
                    return x, y
            y += 1.0

        # ── Phase 3a: relax circle constraints — avoid labels only ─────────
        # Phase 2 failed because every row is blocked by massive circles that
        # fill the image.  Ignore circle bboxes; only keep labels separated.
        def _label_free(rect) -> bool:
            return not any(_rects_overlap(rect, p, buf=LBUF) for p in placed)

        y = pad
        while y + H <= ih - pad:
            x = pad
            while x + W <= iw - pad:
                if _label_free((x, y, x + W, y + H)):
                    return x, y
                x += max(1.0, W * 0.25)
            y += max(1.0, H * 0.5)

        # ── Phase 3b: absolute last resort — stacked grid ─────────────────
        # Divide the image width into label-wide columns and assign this
        # label to its cell by index.  Guarantees labels never stack at the
        # exact same position regardless of how many circles are packed in.
        cols = max(1, int((iw - 2 * pad) / (W + 4)))
        n    = len(placed)
        lx   = pad + (n % cols) * (W + 4)
        ly   = pad + (n // cols) * (H + 4)
        return max(pad, min(lx, iw - W - pad)), max(pad, min(ly, ih - H - pad))

    # ── save_current still calls this for OpenCV label layout ──────────────

    @staticmethod
    def _find_label_pos(
        cx: float, cy: float, r: float,
        W: float, H: float, pad: float,
        iw: float, ih: float,
        placed: list,
        blocked: list = (),
    ) -> tuple:
        """Wrapper used by save_current. Returns (text_x, text_y, rect).
        Appends rect to placed so successive calls avoid each other."""
        lx, ly = AnnotationItem._place_label(cx, cy, r, W, H, pad, iw, ih,
                                             placed, blocked)
        rect   = (lx, ly, lx + W, ly + H)
        placed.append(rect)
        return lx + pad, ly + H - pad, rect

    # ── QGraphicsItem interface ─────────────────────────────────────────────

    def boundingRect(self) -> QRectF:
        dr  = self._dot_r + 3
        dot = QRectF(self._cx - dr, self._cy - dr, 2 * dr, 2 * dr)
        return dot.united(self._lrect).adjusted(-4, -4, 4, 4)

    def paint(self, painter: QPainter, option, widget=None):
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        col = self._col
        dr  = self._dot_r

        # ── Connector line ─────────────────────────────────────────────────
        pen_c = QPen(col, max(1.0, dr * 0.35))
        pen_c.setCosmetic(False)
        painter.setOpacity(0.75)
        painter.setPen(pen_c)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawLine(QPointF(self._ex,     self._ey),
                         QPointF(self._lbl_ex, self._lbl_ey))

        # ── Anchor dot at circle centre ────────────────────────────────────
        painter.setOpacity(1.0)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(255, 255, 255)))
        painter.drawEllipse(QPointF(self._cx, self._cy), dr + 2, dr + 2)
        painter.setBrush(QBrush(col))
        painter.drawEllipse(QPointF(self._cx, self._cy), dr, dr)

        # ── Label box ──────────────────────────────────────────────────────
        x1 = self._lrect.x();    y1 = self._lrect.y()
        x2 = self._lrect.right(); y2 = self._lrect.bottom()

        painter.setOpacity(0.92)
        painter.setBrush(QBrush(QColor(12, 12, 20)))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(self._lrect, 4, 4)

        # Left accent bar
        painter.setOpacity(1.0)
        painter.setBrush(QBrush(col))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(
            QRectF(x1, y1, float(self._accent_w), y2 - y1), 2, 2)

        # Border
        painter.setOpacity(0.85)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(QPen(col, max(1.0, dr * 0.2)))
        painter.drawRoundedRect(self._lrect, 4, 4)

        # Label text
        painter.setOpacity(1.0)
        painter.setFont(self._font)
        painter.setPen(QPen(QColor(245, 245, 245)))
        inner = QRectF(x1 + self._accent_w + 4, y1, x2 - x1 - self._accent_w - 8, y2 - y1)
        painter.drawText(
            inner,
            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
            self._tag)


# ── Image viewer ───────────────────────────────────────────────────────────────

class ImageViewer(QGraphicsView):
    """
    QGraphicsView: zoom / pan / draw circle.

    The scene contains:
      [0] QGraphicsPixmapItem  — base image + circle effects (numpy-baked)
      [1] GhostCircleItem      — live preview during drag
      [2..n] AnnotationItem    — one per circle (Qt-rendered, zero numpy cost)

    Pan and zoom are pure Qt transform ops — no numpy work at all.
    """

    circle_drawn = Signal(tuple, int)   # (cx_orig, cy_orig), r_orig
    zoom_changed = Signal(float)

    ZOOM_FACTOR = 1.20
    MIN_SCALE   = 0.02
    MAX_SCALE   = 20.0

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)

        self._pixmap_item = QGraphicsPixmapItem()
        self._pixmap_item.setTransformationMode(
            Qt.TransformationMode.SmoothTransformation)
        self._pixmap_item.setZValue(0)
        self._scene.addItem(self._pixmap_item)

        self._ghost = GhostCircleItem()
        self._scene.addItem(self._ghost)

        self._ann_items: list[AnnotationItem] = []

        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.setOptimizationFlag(
            QGraphicsView.OptimizationFlag.DontAdjustForAntialiasing, True)
        self.setViewportUpdateMode(
            QGraphicsView.ViewportUpdateMode.SmartViewportUpdate)
        self.setDragMode(QGraphicsView.DragMode.NoDrag)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.NoAnchor)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.NoAnchor)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setBackgroundBrush(QBrush(QColor(18, 18, 18)))

        self._draw_enabled     = False   # enabled only after image loads
        self._draw_mode        = EditMode.HIGHLIGHT
        self._display_scale    = 1.0
        self._img_w            = 0    # image pixel width  (set by set_image_size)
        self._img_h            = 0    # image pixel height
        self._panning          = False
        self._pan_origin       = None
        self._drawing          = False
        self._draw_start_scene = None

        # Ghost throttle: cap circle-preview redraws at ~60 fps
        # (mouseMoveEvent can fire at >500 Hz on high-DPI screens)
        self._last_ghost_ns    = 0
        self._GHOST_INTERVAL   = 16_000_000   # 16 ms in nanoseconds

    # ── Public API ─────────────────────────────────────────────────────────────

    def set_pixmap(self, pixmap: QPixmap):
        self._pixmap_item.setPixmap(pixmap)
        self._scene.setSceneRect(QRectF(pixmap.rect()))

    def fit_view(self):
        r = self._scene.sceneRect()
        if not r.isEmpty():
            self.fitInView(r.adjusted(-20, -20, 20, 20),
                           Qt.AspectRatioMode.KeepAspectRatio)

    def set_draw_enabled(self, v: bool):
        self._draw_enabled = v
        if not v:
            self._ghost.setVisible(False)
            self._drawing = False

    def set_draw_mode(self,  mode: EditMode): self._draw_mode     = mode
    def set_display_scale(self, ds: float):   self._display_scale = ds
    def set_image_size(self, w: int, h: int): self._img_w = w; self._img_h = h

    @property
    def current_zoom(self) -> float: return self.transform().m11()

    # ── Annotation items ───────────────────────────────────────────────────────

    def add_annotation_item(self, item: AnnotationItem):
        self._scene.addItem(item)
        self._ann_items.append(item)

    def remove_last_annotation_item(self):
        if self._ann_items:
            self._scene.removeItem(self._ann_items.pop())

    def clear_annotation_items(self):
        for it in self._ann_items:
            self._scene.removeItem(it)
        self._ann_items.clear()

    def set_annotations_visible(self, v: bool):
        for it in self._ann_items:
            it.setVisible(v)

    # ── Key forwarding ─────────────────────────────────────────────────────────

    def keyPressEvent(self, event):
        if self.parent() is not None:
            QApplication.sendEvent(self.parent(), event)
        else:
            super().keyPressEvent(event)

    # ── Mouse events ───────────────────────────────────────────────────────────

    def wheelEvent(self, event: QWheelEvent):
        delta  = event.angleDelta().y()
        factor = self.ZOOM_FACTOR if delta > 0 else (1.0 / self.ZOOM_FACTOR)
        new_s  = self.current_zoom * factor
        if not (self.MIN_SCALE <= new_s <= self.MAX_SCALE):
            return
        anchor    = event.position().toPoint()
        old_scene = self.mapToScene(anchor)
        self.scale(factor, factor)
        diff = self.mapToScene(anchor) - old_scene
        self.translate(diff.x(), diff.y())
        self.zoom_changed.emit(self.current_zoom)

    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.RightButton:
            self._panning    = True
            self._pan_origin = event.position()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
        elif event.button() == Qt.MouseButton.LeftButton and self._draw_enabled:
            self._drawing = True
            raw = self.mapToScene(event.position().toPoint())
            # Clamp centre to image bounds so circles can never start outside
            if self._img_w > 0 and self._img_h > 0:
                cx = max(0.0, min(raw.x(), float(self._img_w - 1)))
                cy = max(0.0, min(raw.y(), float(self._img_h - 1)))
                self._draw_start_scene = QPointF(cx, cy)
            else:
                self._draw_start_scene = raw
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent):
        if self._panning and self._pan_origin is not None:
            d = event.position() - self._pan_origin
            self._pan_origin = event.position()
            s = self.current_zoom
            self.translate(d.x() / s, d.y() / s)
        elif self._drawing and self._draw_start_scene is not None:
            import time
            now = time.monotonic_ns()
            if now - self._last_ghost_ns < self._GHOST_INTERVAL:
                return
            self._last_ghost_ns = now

            cur = self.mapToScene(event.position().toPoint())
            r   = math.hypot(cur.x() - self._draw_start_scene.x(),
                             cur.y() - self._draw_start_scene.y())
            # Clamp radius so ghost stays inside image
            if self._img_w > 0 and self._img_h > 0:
                cx = self._draw_start_scene.x()
                cy = self._draw_start_scene.y()
                max_r = min(cx, cy,
                            self._img_w - 1 - cx,
                            self._img_h - 1 - cy)
                r = min(r, max(1.0, max_r))
            if r > 3:
                self._ghost.update_circle(
                    self._draw_start_scene.x(),
                    self._draw_start_scene.y(),
                    r, _MODE_QT[self._draw_mode])
                self._ghost.setVisible(True)
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.RightButton and self._panning:
            self._panning = False
            self.setCursor(Qt.CursorShape.ArrowCursor)
        elif event.button() == Qt.MouseButton.LeftButton and self._drawing:
            self._drawing = False
            self._ghost.setVisible(False)
            if self._draw_start_scene is not None:
                cur    = self.mapToScene(event.position().toPoint())
                r_disp = math.hypot(cur.x() - self._draw_start_scene.x(),
                                    cur.y() - self._draw_start_scene.y())
                # Clamp radius: circle must not cross any image edge
                if self._img_w > 0 and self._img_h > 0:
                    cx = self._draw_start_scene.x()
                    cy = self._draw_start_scene.y()
                    max_r = min(cx, cy,
                                self._img_w - 1 - cx,
                                self._img_h - 1 - cy)
                    r_disp = min(r_disp, max(1.0, max_r))
                if r_disp > 5:
                    ds      = self._display_scale
                    cx_orig = int(self._draw_start_scene.x() / ds)
                    cy_orig = int(self._draw_start_scene.y() / ds)
                    r_orig  = max(1, int(r_disp / ds))
                    self.circle_drawn.emit((cx_orig, cy_orig), r_orig)
            self._draw_start_scene = None
        else:
            super().mouseReleaseEvent(event)


# ── Annotation dialog ──────────────────────────────────────────────────────────

class AnnotationDialog(QDialog):
    """
    Compact semi-transparent frameless dialog, positioned near the circle.

    Label field has an inline quick-pick list of 4 common error types:
      • Overlapping   • Label Mismatch   • Label Missing   • Inaccuracy

    All 4 are shown immediately when the dialog opens.
    Typing filters the list live (case-insensitive substring match).
    Clicking a suggestion fills the field and moves focus to Detail.
    The user can also ignore the list and type freely.
    """

    # The 4 predefined error suggestions (ordered for display)
    SUGGESTIONS = [
        "Overlapping",
        "Label Mismatch",
        "Label Missing",
        "Inaccuracy",
    ]

    def __init__(self, mode: EditMode, parent=None, screen_pos=None):
        super().__init__(parent)
        self._screen_pos = screen_pos

        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setWindowOpacity(0.95)
        self.setMinimumWidth(320)
        self.setMaximumWidth(420)

        qc     = _MODE_QT[mode]
        accent = qc.name()
        h, s, v, _ = qc.getHsvF()
        soft = QColor.fromHsvF(h, s * 0.55, min(1.0, v * 1.1)).name()

        self.setStyleSheet(f"""
            QDialog {{
                background: rgba(16, 16, 22, 235);
                border: 1px solid {accent};
                border-radius: 6px; color: #d0d0d0;
            }}
            QLabel  {{ color: #888; font-size: 11px; }}
            QLineEdit {{
                background: rgba(34, 34, 46, 200); color: #e8e8e8;
                border: none; border-bottom: 1px solid {accent};
                border-radius: 3px; padding: 5px 7px; font-size: 13px;
            }}
            QLineEdit:focus {{ border-bottom: 2px solid {accent}; }}
            QListWidget {{
                background: rgba(22, 22, 34, 240);
                border: 1px solid rgba(255,255,255,30);
                border-radius: 4px;
                outline: none;
                font-size: 12px;
                color: #ccc;
            }}
            QListWidget::item {{
                padding: 6px 10px;
                border-bottom: 1px solid rgba(255,255,255,10);
            }}
            QListWidget::item:hover {{
                background: rgba(255,255,255,12);
                color: #fff;
            }}
            QListWidget::item:selected {{
                background: {accent};
                color: #111;
                font-weight: bold;
            }}
            QDialogButtonBox QPushButton {{
                background: transparent; color: {accent};
                border: 1px solid {accent}; border-radius: 3px;
                padding: 4px 16px; font-size: 11px; font-weight: 600;
            }}
            QDialogButtonBox QPushButton:default {{
                background: {accent}; color: #111;
            }}
            QDialogButtonBox QPushButton:hover {{
                background: rgba(255,255,255,20);
            }}
        """)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 10)
        outer.setSpacing(0)

        # Mode pill header
        pill = QLabel(f"  ● {mode.value.upper()}")
        pill.setStyleSheet(f"""
            QLabel {{
                background: {soft}; color: rgba(0,0,0,200);
                font-size: 10px; font-weight: 700; letter-spacing: 1px;
                padding: 4px 10px;
                border-top-left-radius: 5px; border-top-right-radius: 5px;
            }}
        """)
        outer.addWidget(pill)

        form = QFormLayout()
        form.setContentsMargins(14, 10, 14, 4)
        form.setSpacing(6)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        outer.addLayout(form)

        # ── Label field ────────────────────────────────────────────────────
        self._lbl = QLineEdit()
        self._lbl.setPlaceholderText("Type or pick a common error below…")
        form.addRow("Label", self._lbl)

        # ── Suggestion list (inline, always visible under the label field) ─
        # A thin section header
        hint = QLabel("  Quick pick")
        hint.setStyleSheet(
            "QLabel { color:#555; font-size:10px; padding: 6px 14px 2px 14px; }")
        outer.addWidget(hint)

        self._sugg_list = QListWidget()
        self._sugg_list.setFocusPolicy(Qt.FocusPolicy.NoFocus)   # keep focus on _lbl
        self._sugg_list.setSelectionMode(
            QListWidget.SelectionMode.SingleSelection)
        self._sugg_list.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._sugg_list.setFixedHeight(len(self.SUGGESTIONS) * 34)
        self._sugg_list.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        outer.addWidget(self._sugg_list)
        outer.setContentsMargins(0, 0, 0, 0)

        self._populate_suggestions(self.SUGGESTIONS)

        # Clicking a suggestion → fill label, jump to detail
        self._sugg_list.itemClicked.connect(self._on_suggestion_clicked)

        # Typing in label field → live filter
        self._lbl.textChanged.connect(self._filter_suggestions)

        # ── Detail field ───────────────────────────────────────────────────
        form2 = QFormLayout()
        form2.setContentsMargins(14, 8, 14, 4)
        form2.setSpacing(6)
        form2.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        outer.addLayout(form2)

        self._desc = QLineEdit()
        self._desc.setPlaceholderText("optional detail")
        form2.addRow("Detail", self._desc)

        # ── Buttons ────────────────────────────────────────────────────────
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        ok = btns.button(QDialogButtonBox.StandardButton.Ok)
        if ok: ok.setDefault(True)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(14, 4, 14, 0)
        btn_row.addStretch()
        btn_row.addWidget(btns)
        outer.addLayout(btn_row)

        self._lbl.setFocus()

    # ── Suggestion helpers ─────────────────────────────────────────────────

    def _populate_suggestions(self, items: list[str]):
        """Rebuild list widget from `items`, preserving row height."""
        self._sugg_list.clear()
        for text in items:
            it = QListWidgetItem(f"  {text}")
            self._sugg_list.addItem(it)
        # Adjust height to number of visible rows (max 4)
        row_h = 34
        self._sugg_list.setFixedHeight(max(1, len(items)) * row_h)

    def _filter_suggestions(self, text: str):
        """Show only suggestions that contain the typed text (substring)."""
        q = text.strip().lower()
        if not q:
            # Empty field → show all 4
            self._populate_suggestions(self.SUGGESTIONS)
            self._sugg_list.setVisible(True)
            return
        matched = [s for s in self.SUGGESTIONS if q in s.lower()]
        if matched:
            self._populate_suggestions(matched)
            self._sugg_list.setVisible(True)
        else:
            # No match → hide list (user is typing something custom)
            self._sugg_list.setVisible(False)

    def _on_suggestion_clicked(self, item: QListWidgetItem):
        """Fill the label field with the chosen suggestion, focus detail."""
        self._lbl.setText(item.text().strip())
        self._lbl.setStyleSheet(
            self._lbl.styleSheet() +
            " QLineEdit { border-bottom: 2px solid #5af; }")
        self._sugg_list.setVisible(False)
        self._desc.setFocus()

    # ── Positioning ────────────────────────────────────────────────────────

    def showEvent(self, event):
        super().showEvent(event)
        if self._screen_pos is not None:
            self.adjustSize()
            screen = QApplication.primaryScreen().availableGeometry()
            x = int(self._screen_pos.x()) + 30
            y = int(self._screen_pos.y()) - self.height() // 2
            x = max(screen.left(), min(x, screen.right()  - self.width()))
            y = max(screen.top(),  min(y, screen.bottom() - self.height()))
            self.move(x, y)

    @property
    def label(self)       -> str: return self._lbl.text().strip()
    @property
    def description(self) -> str: return self._desc.text().strip()


# ── Main window ────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    """
    Annotation workflow controller.

    Cache model
    -----------
    _display_img   numpy   Raw downscaled image (never modified).
    _effects_img   numpy   display_img + all circle effects + circle outlines.
                           Rebuilt incrementally (one ROI per new circle).
                           Full rebuild only on undo/clear (still fast: ROI ops).
    _base_pixmap   QPixmap Derived from _effects_img. Updated after each numpy op.

    Labels are NOT stored in numpy.  They live as AnnotationItem Qt scene items,
    managed directly by ImageViewer.  This decouples all label operations from
    the numpy pipeline entirely.

    _placed_rects  list    Collision-avoidance state for label layout.
                           Rebuilt from scratch only when item list changes.
    """

    def __init__(self, input_folder: str, output_folder: str | None = None,
                 max_display_px: int = MAX_DISPLAY_PX):
        super().__init__()
        self.setWindowTitle("Batch Labeled Editor  ·  Qt")
        self.resize(1440, 900)

        self._input_folder  = Path(input_folder)
        self._max_display   = max_display_px
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._output_folder = (Path(output_folder) if output_folder
                               else self._input_folder.parent / f"labeled_output_{ts}")
        self._output_folder.mkdir(parents=True, exist_ok=True)

        self._image_files: list[Path] = self._discover_images()
        self._total   = len(self._image_files)
        self._idx     = 0
        self._loader: ImageLoader | None = None

        # ── numpy state ───────────────────────────────────────────────────────
        self._display_img:   np.ndarray | None = None
        self._effects_img:   np.ndarray | None = None
        self._display_scale: float             = 1.0
        self._orig_w:        int               = 0
        self._orig_h:        int               = 0

        # ── annotation state ──────────────────────────────────────────────────
        self._circles:       list[dict]  = []
        self._placed_rects:  list[tuple] = []    # for incremental label layout
        self._saved_status:  dict[str, bool] = {}
        self._img_circles:   dict[str, list] = {}

        # ── UI state ──────────────────────────────────────────────────────────
        self._mode        = EditMode.HIGHLIGHT
        self._blur_kernel = 25
        self._pix_size    = 10
        self._hl_alpha    = 0.4
        self._show_labels = True

        self._mode_actions: dict[EditMode, QAction] = {}
        self._build_ui()
        self._build_toolbar()
        self._build_statusbar()
        self._setup_shortcuts()

        self._zoom_timer = QTimer(self)
        self._zoom_timer.setSingleShot(True)
        self._zoom_timer.setInterval(60)
        self._zoom_timer.timeout.connect(self._update_status_bar)

        # Async numpy → QPixmap worker (replaces the coalescing timer).
        # Keeps the UI interactive during full-res pixmap conversion.
        self._px_converter: PixmapConverter | None = None
        # Threads that have been cancelled but not yet finished.
        # Keeping Python refs here prevents "QThread destroyed while running".
        self._zombie_converters: list = []

        if not self._image_files:
            QMessageBox.critical(self, "No images",
                                 f"No images found in:\n{self._input_folder}")
            sys.exit(1)

        self._load_image(0)
        self._print_instructions()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        self._viewer = ImageViewer(self)
        self._viewer.circle_drawn.connect(self._on_circle_drawn)
        self._viewer.zoom_changed.connect(lambda _: self._zoom_timer.start())
        self.setCentralWidget(self._viewer)

        # ── Loading overlay (shown while ImageLoader + PixmapConverter run) ──
        self._loading_overlay = QLabel(self._viewer)
        self._loading_overlay.setText("  Loading…  ")
        self._loading_overlay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._loading_overlay.setStyleSheet("""
            QLabel {
                background: rgba(14, 14, 22, 210);
                color: #7ab4ff;
                font-size: 15px;
                font-weight: bold;
                border-radius: 10px;
                padding: 10px 24px;
            }
        """)
        self._loading_overlay.adjustSize()
        self._loading_overlay.hide()

    def _show_loading(self):
        """Centre the overlay in the viewer and make it visible."""
        ow = self._loading_overlay.width()
        oh = self._loading_overlay.height()
        vw = self._viewer.width()
        vh = self._viewer.height()
        self._loading_overlay.move((vw - ow) // 2, (vh - oh) // 2)
        self._loading_overlay.raise_()
        self._loading_overlay.show()
        QApplication.processEvents()   # paint immediately, before disk I/O

    def _hide_loading(self):
        self._loading_overlay.hide()

    def _build_toolbar(self):
        tb = QToolBar("Controls", self)
        tb.setMovable(False)
        tb.setStyleSheet("""
            QToolBar { background:#232323; border:none; spacing:4px; padding:2px 6px; }
            QToolButton {
                color:#ccc; background:#2e2e2e; border:1px solid #444;
                border-radius:3px; padding:3px 8px; font-size:11px;
            }
            QToolButton:checked { border-color:#5af; color:#5af; }
            QToolButton:hover   { background:#3a3a3a; }
        """)
        self.addToolBar(tb)

        modes = list(EditMode)
        abbr  = ["HLT","BLR","PIX","DRK","GRY","INV","OUT"]
        for key, ab, mode in zip("1234567", abbr, modes):
            act = QAction(f"{key}:{ab}", self)
            act.setCheckable(True); act.setChecked(mode == EditMode.HIGHLIGHT)
            act.setToolTip(f"Mode: {mode.value}  ({key})")
            act.triggered.connect(lambda _, m=mode: self._set_mode(m))
            tb.addAction(act)
            self._mode_actions[mode] = act

        tb.addSeparator()
        for label, tip, slot in [
            ("Save",   "Save (S)",              self.save_current),
            ("Undo",   "Undo last (U)",         self._undo),
            ("Edit",   "Edit last label (E)",   self._edit_last_label),
            ("Clear",  "Clear all (C)",         self._clear),
            ("Labels", "Toggle labels (T)",     self._toggle_labels),
            ("Fit",    "Fit view (R)",           self._reset_zoom),
            ("List",   "List annotations (L)",  self._list_objects),
            ("Help",   "Help (H)",              self._show_help),
        ]:
            act = QAction(label, self)
            act.setToolTip(tip)
            act.triggered.connect(slot)
            tb.addAction(act)

        tb.addSeparator()
        excel_act = QAction("📊 Save Excel", self)
        excel_act.setToolTip("Export processing summary to Excel (.xlsx)")
        excel_act.triggered.connect(self._save_excel_with_feedback)
        tb.addAction(excel_act)

    def _build_statusbar(self):
        sb = self.statusBar()
        sb.setStyleSheet("QStatusBar { background:#1a1a1a; color:#999; }")

        def _lbl(bold=False):
            w = QLabel("–")
            w.setStyleSheet(f"color:{'#eee' if bold else '#999'}; padding:0 8px;"
                            + (" font-weight:bold;" if bold else ""))
            return w

        self._lbl_file  = _lbl()
        self._lbl_nav   = _lbl(bold=True)
        self._lbl_mode  = _lbl()
        self._lbl_obj   = _lbl()
        self._lbl_res   = _lbl()
        self._lbl_zoom  = _lbl()
        self._lbl_saved = _lbl(bold=True)

        sb.addWidget(self._lbl_file)
        for w in (self._lbl_res, self._lbl_zoom, self._lbl_obj,
                  self._lbl_mode, self._lbl_nav, self._lbl_saved):
            sb.addPermanentWidget(w)

    def _setup_shortcuts(self):
        modes = list(EditMode)
        self._key_dispatch = {
            Qt.Key.Key_A: self._prev_image,
            Qt.Key.Key_D: self._next_image,
            Qt.Key.Key_S: self.save_current,
            Qt.Key.Key_C: self._clear,
            Qt.Key.Key_U: self._undo,
            Qt.Key.Key_E: self._edit_last_label,
            Qt.Key.Key_T: self._toggle_labels,
            Qt.Key.Key_R: self._reset_zoom,
            Qt.Key.Key_L: self._list_objects,
            Qt.Key.Key_H: self._show_help,
            Qt.Key.Key_Q: self.close,
            # NOTE: lambda must take NO args — fn() called with no args below
            **{getattr(Qt.Key, f"Key_{i+1}"): (lambda m=m: self._set_mode(m))
               for i, m in enumerate(modes)},
        }

    def keyPressEvent(self, event):
        key  = event.key()
        mods = event.modifiers()
        if key == Qt.Key.Key_S and (mods & Qt.KeyboardModifier.ShiftModifier):
            self._save_and_next(); return
        fn = self._key_dispatch.get(key)
        if fn:
            fn(); return
        super().keyPressEvent(event)

    # ── File discovery ─────────────────────────────────────────────────────────

    def _discover_images(self) -> list[Path]:
        exts  = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}
        files = sorted(f for e in exts for f in self._input_folder.glob(f"*{e}"))
        valid = [f for f in files if f.stat().st_size >= 512]
        print(f"  Found {len(valid)} image(s) in {self._input_folder}")
        return valid

    # ── Image loading ──────────────────────────────────────────────────────────

    def _free_current_image(self):
        """Release ALL pixel buffers before loading the next image."""
        # Cancel active converter — move to zombie list (not None) so the
        # Python reference stays alive until the OS thread actually exits.
        if self._px_converter is not None:
            self._px_converter.cancel()
            try:
                self._px_converter.ready.disconnect()
            except RuntimeError:
                pass
            self._zombie_converters.append(self._px_converter)
            self._px_converter = None
        # Cancel any still-running zombies from previous rapid navigation
        for z in list(self._zombie_converters):
            z.cancel()
        self._display_img = self._effects_img = None
        self._viewer.clear_annotation_items()
        gc.collect()

    def _load_image(self, idx: int):
        if not (0 <= idx < self._total): return
        self._idx = idx
        f = self._image_files[idx]

        if self._loader and self._loader.isRunning():
            self._loader.abort(); self._loader.wait(300)

        # Restore circles from JSON or memory cache
        json_path = self._output_folder / f.with_suffix(".json").name
        if json_path.exists():
            try:
                data = json.loads(json_path.read_text())
                self._circles = [{
                    "center":      tuple(o["center"]),
                    "radius":      o["radius"],
                    "mode":        EditMode(o.get("mode", "highlight")),
                    "label":       o.get("label", ""),
                    "description": o.get("description", ""),
                } for o in data.get("objects", [])]
                print(f"  {f.name} — restored {len(self._circles)} annotation(s) from JSON")
            except Exception as e:
                print(f"  JSON load error: {e}"); self._circles = []
        elif f.name in self._img_circles:
            self._circles = [c.copy() for c in self._img_circles[f.name]]
        else:
            self._circles = []

        self._placed_rects = []
        self._free_current_image()
        self._viewer.set_draw_enabled(False)

        self._loader = ImageLoader(f, self._max_display)
        self._loader.loaded.connect(self._on_image_loaded,
                                    Qt.ConnectionType.QueuedConnection)
        self._loader.error.connect(self._on_image_error,
                                   Qt.ConnectionType.QueuedConnection)
        self._loader.start()
        self.setWindowTitle(f"Batch Labeled Editor  —  {f.name}  (loading…)")
        self._show_loading()
        self._update_status_bar()

    def _on_image_loaded(self, disp: np.ndarray, scale: float,
                         orig_w: int, orig_h: int):
        self._display_img   = disp
        self._display_scale = scale
        self._orig_w        = orig_w
        self._orig_h        = orig_h

        self._viewer.set_display_scale(scale)
        self._viewer.set_image_size(orig_w, orig_h)
        dh, dw = disp.shape[:2]

        # Build effects numpy from scratch (all restored circles)
        self._rebuild_effects_full()
        # Create Qt annotation items for all restored circles
        self._rebuild_all_annotation_items()

        self._viewer.fit_view()
        self._viewer.set_draw_enabled(True)
        self._hide_loading()
        self._update_status_bar()

        f = self._image_files[self._idx]
        self.setWindowTitle(f"Batch Labeled Editor  —  {f.name}")
        print(f"  {f.name}  [{orig_w}×{orig_h}  disp:{dw}×{dh}  "
              f"scale:{scale:.2f}  circles:{len(self._circles)}]")

    def _on_image_error(self, msg: str):
        print(f"  Load error: {msg}")
        self._hide_loading()
        self.statusBar().showMessage(f"Load failed: {msg}", 4000)
        self._viewer.set_draw_enabled(True)
        if self._idx < self._total - 1:
            self._load_image(self._idx + 1)

    # ── Effects (numpy) ────────────────────────────────────────────────────────

    def _circle_disp(self, c: dict) -> dict:
        """Original-pixel circle → display-pixel circle."""
        ds = self._display_scale
        return {**c,
                "center": (int(c["center"][0] * ds), int(c["center"][1] * ds)),
                "radius": max(1, int(c["radius"] * ds))}

    def _apply_effect(self, image: np.ndarray, dc: dict) -> np.ndarray:
        """
        Apply one circle's visual effect to `image` IN PLACE.

        ROI-clipped approach:
          All work is confined to the tight bounding box of the circle.
          For a 4000×3000 px image with r=200, blur/pixelate/grayscale
          process 160 000 px instead of 12 000 000 px — a 75× speedup.
          The boolean mask is also only the size of the bounding box,
          not the full image, saving both memory and allocation time.
        """
        try:
            cx, cy = dc["center"]
            r      = dc["radius"]
            ih, iw = image.shape[:2]

            # Tight bounding box clamped to image
            rx1 = max(0,  cx - r);  rx2 = min(iw, cx + r + 1)
            ry1 = max(0,  cy - r);  ry2 = min(ih, cy + r + 1)
            if rx2 <= rx1 or ry2 <= ry1:
                return image

            roi  = image[ry1:ry2, rx1:rx2]   # view — no copy

            # Circle mask for the ROI only (tiny allocation)
            mask = np.zeros(roi.shape[:2], dtype=np.uint8)
            cv2.circle(mask, (cx - rx1, cy - ry1), r, 255, -1)
            mb   = mask == 255

            mode = dc["mode"]

            if mode == EditMode.HIGHLIGHT:
                roi[mb] = cv2.addWeighted(
                    roi, 1 - self._hl_alpha,
                    np.full_like(roi, 255), self._hl_alpha, 0)[mb]

            elif mode == EditMode.BLUR:
                k   = self._blur_kernel
                # Blur only the ROI; pad=r ensures edge pixels are correct
                pad = r
                px1 = max(0, rx1-pad); py1 = max(0, ry1-pad)
                px2 = min(iw, rx2+pad); py2 = min(ih, ry2+pad)
                padded   = image[py1:py2, px1:px2]
                blurred  = cv2.GaussianBlur(padded, (k|1, k|1), 0)
                # Extract the inner region that matches roi
                oy = ry1 - py1;  ox = rx1 - px1
                roi[mb] = blurred[oy:oy+roi.shape[0], ox:ox+roi.shape[1]][mb]

            elif mode == EditMode.PIXELATE:
                sz    = max(1, self._pix_size)
                rh, rw = roi.shape[:2]
                small = cv2.resize(roi, (max(1, rw//sz), max(1, rh//sz)),
                                   interpolation=cv2.INTER_NEAREST)
                pix   = cv2.resize(small, (rw, rh), interpolation=cv2.INTER_NEAREST)
                roi[mb] = pix[mb]

            elif mode == EditMode.DARKEN:
                roi[mb] = (roi[mb] * 0.5).astype(np.uint8)

            elif mode == EditMode.GRAYSCALE:
                gray = cv2.cvtColor(
                    cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                roi[mb] = gray[mb]

            elif mode == EditMode.INVERT:
                roi[mb] = 255 - roi[mb]

            # OUTLINE: no fill; circle outline drawn separately after this call

        except Exception as e:
            print(f"  Effect error ({dc.get('mode','?')}): {e}")
        return image

    def _apply_circle_to_effects(self, c: dict):
        """
        INCREMENTAL add: apply one circle's effect to _effects_img in-place.

        SNAPSHOT BBOX includes a 3-pixel margin beyond r on every side to
        fully capture the cv2.circle outline (thickness=2 + LINE_AA fringe
        extends up to r+2 from centre).  Without this, undo leaves a ghost
        arc of ~160 pixels that the smaller bbox never restores.
        """
        dc   = self._circle_disp(c)
        cx, cy = dc["center"]
        r    = dc["radius"]
        ih, iw = self._effects_img.shape[:2]

        # Extend by STROKE_PAD so the outline is fully inside the snapshot
        STROKE_PAD = 3   # thickness=2 outline + 1 px AA safety margin
        rx1 = max(0, cx - r - STROKE_PAD);  rx2 = min(iw, cx + r + STROKE_PAD + 1)
        ry1 = max(0, cy - r - STROKE_PAD);  ry2 = min(ih, cy + r + STROKE_PAD + 1)

        # ── Save pre-effect ROI snapshot (for O(r²) undo) ─────────────────
        c["_undo_roi"]  = self._effects_img[ry1:ry2, rx1:rx2].copy()
        c["_undo_bbox"] = (rx1, ry1, rx2, ry2)

        self._apply_effect(self._effects_img, dc)
        cv2.circle(self._effects_img, dc["center"], dc["radius"],
                   _MODE_BGR[c["mode"]], 2, _AA)

    def _rebuild_effects_full(self):
        """
        Full rebuild: all circles applied from scratch.
        Used at image load and fallback undo (JSON-restored circles).
        Each circle only processes its ROI; even 100 circles is fast.
        """
        if self._display_img is None:
            return
        self._effects_img = self._display_img.copy()
        for c in self._circles:
            self._apply_circle_to_effects(c)
        self._schedule_pixmap_update()

    def _update_pixmap(self, pixmap: "QPixmap | None" = None):
        """
        Receive a converted QPixmap from PixmapConverter and push it
        to the viewer.  Called via Qt queued connection from the worker
        thread — guaranteed to run on the main thread.
        """
        if pixmap is None:
            if self._effects_img is not None:
                self._viewer.set_pixmap(bgr_to_pixmap(self._effects_img))
        else:
            self._viewer.set_pixmap(pixmap)
        self._hide_loading()

    def _schedule_pixmap_update(self):
        """
        Start an async numpy → QPixmap conversion on a worker thread.

        Cancelled threads are moved to _zombie_converters so Python keeps
        them alive until Qt finishes the thread — preventing the
        'QThread: Destroyed while thread is still running' warning/crash.
        Each zombie self-removes via its finished signal.
        """
        if self._effects_img is None:
            return

        # Retire the previous converter safely
        if self._px_converter is not None:
            old = self._px_converter
            old.cancel()
            try:
                old.ready.disconnect()
            except RuntimeError:
                pass
            # Keep alive until the OS thread actually exits
            self._zombie_converters.append(old)
            old.finished.connect(
                lambda c=old: self._zombie_converters.remove(c)
                if c in self._zombie_converters else None)
            self._px_converter = None

        img_copy = self._effects_img.copy()   # private copy — no race
        conv = PixmapConverter(img_copy)
        conv.ready.connect(self._update_pixmap, Qt.ConnectionType.QueuedConnection)
        self._px_converter = conv
        conv.start()

    # ── Annotation items (Qt) ──────────────────────────────────────────────────

    def _rebuild_all_annotation_items(self):
        """
        Recreate ALL Qt annotation items from _circles.
        Runs collision-avoidance placement for all at once.
        Called only at image load, undo, clear — NOT on add.

        Each label avoids both other labels (via placed_rects) AND the
        bounding boxes of all other circles (via blocked_rects), so labels
        never land inside a neighbouring circle's annotation region.
        """
        self._viewer.clear_annotation_items()
        self._placed_rects = []
        if self._display_img is None:
            return
        dh, dw = self._display_img.shape[:2]
        ds = self._display_scale

        # Precompute bounding rect for every circle (in scene / display coords)
        all_crects = [
            (c["center"][0]*ds - c["radius"]*ds,
             c["center"][1]*ds - c["radius"]*ds,
             c["center"][0]*ds + c["radius"]*ds,
             c["center"][1]*ds + c["radius"]*ds)
            for c in self._circles
        ]

        for i, c in enumerate(self._circles, 1):
            # Block every OTHER circle's region
            blocked = all_crects[:i-1] + all_crects[i:]
            item = AnnotationItem(i, c, ds, dw, dh,
                                  self._placed_rects, blocked)
            item.setVisible(self._show_labels)
            self._viewer.add_annotation_item(item)

    def _push_new_annotation_item(self, c: dict):
        """
        INCREMENTAL add: create one AnnotationItem for the most recently
        appended circle, reusing the existing _placed_rects for collision.
        O(1) Qt scene op — zero numpy work.
        """
        if self._display_img is None:
            return
        dh, dw = self._display_img.shape[:2]
        ds     = self._display_scale
        index  = len(self._circles)   # already appended

        # All OTHER circles are blocked zones (the new circle is last, index-1)
        blocked = [
            (oc["center"][0]*ds - oc["radius"]*ds,
             oc["center"][1]*ds - oc["radius"]*ds,
             oc["center"][0]*ds + oc["radius"]*ds,
             oc["center"][1]*ds + oc["radius"]*ds)
            for j, oc in enumerate(self._circles)
            if j != index - 1     # skip the circle being added
        ]
        item = AnnotationItem(index, c, ds, dw, dh,
                              self._placed_rects, blocked)
        item.setVisible(self._show_labels)
        self._viewer.add_annotation_item(item)

    # ── Circle annotation flow ─────────────────────────────────────────────────

    def _on_circle_drawn(self, center_orig: tuple, radius_orig: int):
        if len(self._circles) >= MAX_RECOMMENDED_CIRCLES:
            print(f"  Note: {len(self._circles)} annotations")

        # Position dialog near the circle on screen
        sp = QPointF(center_orig[0] * self._display_scale,
                     center_orig[1] * self._display_scale)
        screen_pt = self._viewer.mapToGlobal(self._viewer.mapFromScene(sp))

        dlg = AnnotationDialog(self._mode, self, screen_pos=screen_pt)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        c = {
            "center":      center_orig,
            "radius":      radius_orig,
            "mode":        self._mode,
            "label":       dlg.label,
            "description": dlg.description,
        }
        self._circles.append(c)
        print(f"  Added '{c['label'] or '(unlabeled)'}'  [{self._mode.value}]  "
              f"r={radius_orig}  pos={center_orig}")

        # FAST PATH: incremental effect + one new Qt item
        self._apply_circle_to_effects(c)
        self._schedule_pixmap_update()
        self._push_new_annotation_item(c)
        self._update_status_bar()

    # ── Commands ───────────────────────────────────────────────────────────────

    def _set_mode(self, mode: EditMode):
        self._mode = mode
        self._viewer.set_draw_mode(mode)
        for m, act in self._mode_actions.items():
            act.setChecked(m == mode)
        self._update_status_bar()
        print(f"  Mode → {mode.value.upper()}")

    def _undo(self):
        if not self._circles:
            self.statusBar().showMessage("Nothing to undo", 1500); return
        rem = self._circles.pop()

        # ── O(r²) ROI restore ──────────────────────────────────────────────
        # If the removed circle carried a pre-effect snapshot, paste it back.
        # This avoids a full effects rebuild (O(n × W×H)).
        # Only falls back to full rebuild if snapshot is missing (e.g. circles
        # restored from JSON on load, where no snapshot was captured).
        roi_data  = rem.get("_undo_roi")
        roi_bbox  = rem.get("_undo_bbox")
        if roi_data is not None and roi_bbox is not None and \
                self._effects_img is not None:
            rx1, ry1, rx2, ry2 = roi_bbox
            self._effects_img[ry1:ry2, rx1:rx2] = roi_data
            # Circle outline was drawn on top of the effect; it's now erased
            # by the ROI restore.  Re-draw outlines for circles that overlap
            # this bbox (uncommon, fast).
            for c in self._circles:
                dc = self._circle_disp(c)
                cx2, cy2 = dc["center"]; r2 = dc["radius"]
                if (cx2-r2 < rx2 and cx2+r2 > rx1 and
                        cy2-r2 < ry2 and cy2+r2 > ry1):
                    cv2.circle(self._effects_img, (cx2, cy2), r2,
                               _MODE_BGR[c["mode"]], 2, _AA)
        else:
            # Fallback: full rebuild (circles loaded from JSON have no snapshots)
            self._rebuild_effects_full()

        if self._placed_rects:
            self._placed_rects.pop()

        print(f"  Removed: {rem['label'] or '(unlabeled)'}")
        self._viewer.remove_last_annotation_item()
        self._schedule_pixmap_update()
        self._update_status_bar()

    def _clear(self):
        if not self._circles: return
        self._circles.clear()
        self._placed_rects.clear()
        self._viewer.clear_annotation_items()
        if self._display_img is not None:
            self._effects_img = self._display_img.copy()
            self._schedule_pixmap_update()
        self._update_status_bar()
        print("  Cleared all annotations")

    def _toggle_labels(self):
        self._show_labels = not self._show_labels
        self._viewer.set_annotations_visible(self._show_labels)
        print(f"  Labels {'ON' if self._show_labels else 'OFF'}")

    def _reset_zoom(self): self._viewer.fit_view(); self._update_status_bar()

    def _list_objects(self):
        f = self._image_files[self._idx]
        print(f"\n{'='*64}\nAnnotations — {f.name}  [{self._orig_w}×{self._orig_h}]")
        if not self._circles: print("  (none)")
        for i, c in enumerate(self._circles, 1):
            print(f"  #{i}  {c['label'] or '(no label)':30s}  [{c['mode'].value}]"
                  f"  r={c['radius']}  pos={c['center']}")
            if c.get("description"):
                print(f"       {c['description']}")
        print("="*64 + "\n")

    def _edit_last_label(self):
        if not self._circles:
            self.statusBar().showMessage("No annotations to edit", 1500); return
        last = self._circles[-1]

        sp = QPointF(last["center"][0] * self._display_scale,
                     last["center"][1] * self._display_scale)
        screen_pt = self._viewer.mapToGlobal(self._viewer.mapFromScene(sp))

        dlg = AnnotationDialog(last["mode"], self, screen_pos=screen_pt)
        dlg._lbl.setText(last["label"])
        dlg._desc.setText(last.get("description", ""))
        dlg._lbl.selectAll(); dlg._lbl.setFocus()

        if dlg.exec() != QDialog.DialogCode.Accepted: return

        old = last["label"]
        last["label"]       = dlg.label
        last["description"] = dlg.description
        print(f"  Edited: '{old}' → '{last['label']}'")

        # Only the label text changed — rebuild Qt items (fast, no numpy)
        self._rebuild_all_annotation_items()
        self._update_status_bar()

    def _show_help(self):
        QMessageBox.information(self, "Keyboard Shortcuts", """\
Navigation
  A / D       Previous / Next image
  S           Save   |   Shift+S  Save + Next
  Q           Quit (auto-saves)

Annotations
  Left drag   Draw circle
  E           Edit last annotation
  U           Undo last   |   C  Clear all
  T           Toggle labels
  1 – 7       Mode: Highlight Blur Pixelate Darken Grayscale Invert Outline

View
  Mouse wheel  Zoom (anchored at cursor)
  Right drag   Pan
  R            Fit to window
  L            List annotations in console
  H            This dialog
""")

    # ── Navigation ─────────────────────────────────────────────────────────────

    def _cache_circles(self):
        f = self._image_files[self._idx]
        self._img_circles[f.name] = [c.copy() for c in self._circles]

    def _navigate(self, delta: int):
        if self._loader and self._loader.isRunning(): return
        new_idx = self._idx + delta
        if not (0 <= new_idx < self._total):
            edge = "first" if delta < 0 else "last"
            self.statusBar().showMessage(f"Already at the {edge} image", 1500); return
        # Always cache and save/clean when leaving — even if circles is empty.
        # Without this, undo-all then navigate leaves a stale JSON on disk,
        # causing the undone circles to reappear on the next load.
        self._cache_circles()
        self.save_current(auto_save=True)
        self._load_image(new_idx)

    def _prev_image(self):    self._navigate(-1)
    def _next_image(self):    self._navigate(+1)
    def _save_and_next(self): self.save_current(); self._next_image()

    # ── Save (full resolution) ─────────────────────────────────────────────────

    def save_current(self, auto_save: bool = False):
        """
        Reload source at FULL resolution, apply effects + labels, write output.

        Empty-circles + auto_save path:
          Deletes any existing output image and JSON so a stale save from a
          previous session cannot restore circles that have since been undone.
        """
        f   = self._image_files[self._idx]
        out = self._output_folder / f.name
        js  = out.with_suffix(".json")

        # ── No circles: clean up stale files and return ────────────────────
        if not self._circles:
            if auto_save:
                deleted = []
                for p in (out, js):
                    if p.exists():
                        p.unlink()
                        deleted.append(p.name)
                if deleted:
                    print(f"  Cleaned stale output: {', '.join(deleted)}")
                self._saved_status.pop(f.name, None)
            else:
                self.statusBar().showMessage("No annotations to save", 2000)
            return

        f   = self._image_files[self._idx]
        out = self._output_folder / f.name
        js  = out.with_suffix(".json")
        self._cache_circles()

        try:
            full = cv2.imread(str(f))
            if full is None: raise IOError(f"Cannot reload {f.name}")

            fh, fw = full.shape[:2]
            fsize  = max(14, max(fw, fh) // 220)

            for c in self._circles:
                self._apply_effect(full, c)      # orig coords = display coords at scale=1
                cv2.circle(full, c["center"], c["radius"],
                           _MODE_BGR[c["mode"]], max(2, fsize//8), _AA)

            # Draw labels directly via OpenCV at full res
            placed: list[tuple] = []
            # Precompute all circle bounding rects for label avoidance
            all_crects_sv = [
                (c["center"][0] - c["radius"], c["center"][1] - c["radius"],
                 c["center"][0] + c["radius"], c["center"][1] + c["radius"])
                for c in self._circles
            ]
            for i, c in enumerate(self._circles, 1):
                cx, cy = c["center"];  r = c["radius"]
                col    = _MODE_BGR[c["mode"]]
                lbl    = c["label"] or f"Object #{i}"
                tag    = f"#{i}  {lbl}"
                sc     = max(0.5, fsize / 28)
                th     = 1

                (tw, fh_cv), bl = cv2.getTextSize(tag, _FONT_CV, sc, th)
                pad = max(6, fsize // 3)
                W, H = tw + 2*pad + 6, fh_cv + bl + 2*pad

                # Block every OTHER circle so labels can't land inside them
                sv_blocked = all_crects_sv[:i-1] + all_crects_sv[i:]
                lx, ly, rect = AnnotationItem._find_label_pos(
                    cx, cy, r, W, H, pad, fw, fh, placed, sv_blocked)
                # Note: _find_label_pos already appends rect to placed
                x1, y1, x2, y2 = [int(v) for v in rect]

                # ── Anchor dot at circle centre (no number) ───────────────────
                dot_r = max(8, fsize // 2)
                # White outer ring for visibility
                cv2.circle(full, (cx, cy), dot_r + 3, (255, 255, 255), -1, _AA)
                # Coloured dot
                cv2.circle(full, (cx, cy), dot_r, col, -1, _AA)

                # ── Label box (solid opaque background for clear visibility) ──
                if y2 > y1 and x2 > x1:
                    roi = full[y1:y2, x1:x2].copy()
                    bg  = np.full_like(roi, (12, 12, 18))
                    full[y1:y2, x1:x2] = cv2.addWeighted(bg, 0.82, roi, 0.18, 0)
                cv2.rectangle(full, (x1, y1), (x1+4, y2),  col,           -1, _AA)
                cv2.rectangle(full, (x1, y1), (x2,   y2),  col,            1, _AA)
                # Shadow pass then bright text for maximum contrast
                cv2.putText(full, tag, (int(lx)+8, int(ly)+1),
                            _FONT_CV, sc, (0, 0, 0),       th+1, _AA)
                cv2.putText(full, tag, (int(lx)+7, int(ly)),
                            _FONT_CV, sc, (240, 240, 240), th,   _AA)

                # connector
                mid_x = (x1+x2)//2
                mid_y = y2 if ly > cy else y1
                ang = math.atan2(mid_y-cy, mid_x-cx)
                ex, ey = int(cx+r*math.cos(ang)), int(cy+r*math.sin(ang))
                cv2.line(full, (mid_x, mid_y), (ex, ey), col, 1, _AA)
                cv2.circle(full, (ex, ey), 4, col, -1, _AA)
                cv2.circle(full, (ex, ey), 4, (200,200,200), 1, _AA)

            if not cv2.imwrite(str(out), full):
                raise IOError("cv2.imwrite failed")
            del full; gc.collect()

            data = {
                "source_image":  f.name,
                "original_size": [self._orig_w, self._orig_h],
                "timestamp":     datetime.now().isoformat(),
                "objects": [{
                    "id":          i,
                    "label":       c["label"],
                    "description": c.get("description", ""),
                    "mode":        c["mode"].value,
                    "center":      list(c["center"]),
                    "radius":      c["radius"],
                } for i, c in enumerate(self._circles, 1)],
            }
            js.write_text(json.dumps(data, indent=2))
            self._saved_status[f.name] = True
            msg = (f"Auto-saved {len(self._circles)} annotation(s)  [{f.name}]"
                   if auto_save else
                   f"Saved  {f.name}  ({len(self._circles)} annotation(s))")
            print(f"  {msg}\n  → {out}\n  → {js}")
            self.statusBar().showMessage(msg, 3000)
            self._update_status_bar()

        except Exception as e:
            print(f"  Save error: {e}")
            self.statusBar().showMessage(f"Save failed: {e}", 5000)

    # ── Excel summary ──────────────────────────────────────────────────────────

    def _save_excel_with_feedback(self):
        """
        Toolbar button handler — save Excel and show a brief status message.
        Warns if there are no saved annotations yet.
        """
        if not self._saved_status:
            QMessageBox.information(
                self, "No Data Yet",
                "No images have been saved yet.\n"
                "Annotate and save at least one image first.")
            return
        self.generate_summary()
        xp = self._output_folder / "processing_summary.xlsx"
        if xp.exists():
            self.statusBar().showMessage(
                f"Excel saved \u2192 {xp.name}", 4000)
        else:
            self.statusBar().showMessage("Excel export failed — see console", 4000)


    def generate_summary(self):
        xp = self._output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
            hfill = PatternFill(start_color="2E4DA7", end_color="2E4DA7", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF", size=11)
            border = Border(**{s: Side(style="thin")
                               for s in ("left","right","top","bottom")})
            ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
            tl = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

            def styled(cell, fill=None, font=None, align=None):
                if fill:  cell.fill      = fill
                if font:  cell.font      = font
                if align: cell.alignment = align
                cell.border = border

            for col, hdr in enumerate(["#","Image","Errors","Details"], 1):
                styled(ws.cell(row=1, column=col, value=hdr),
                       fill=hfill, font=hfont, align=ca)
            ws.row_dimensions[1].height = 26

            row = 2; n = 0
            for img in sorted(self._image_files):
                if img.name not in self._saved_status: continue
                jp = self._output_folder / img.with_suffix(".json").name
                if not jp.exists(): continue
                data = json.loads(jp.read_text())
                objs = data.get("objects", []); n += 1
                styled(ws.cell(row=row, column=1, value=n), align=ca)
                ws.cell(row=row, column=1).font = Font(bold=True, size=10)
                styled(ws.cell(row=row, column=2, value=img.name), align=tl)
                c3 = ws.cell(row=row, column=3, value=len(objs))
                styled(c3, align=ca)
                color = ("C6EFCE","276221") if len(objs)==0 else \
                        ("FFEB9C","9C5700") if len(objs)<=3 else ("FFC7CE","9C0006")
                c3.fill = PatternFill(start_color=color[0], end_color=color[0], fill_type="solid")
                c3.font = Font(color=color[1], bold=True)
                lines = [f"{i}. {o.get('label','').strip() or '(unlabeled)'}"
                         + (f" — {o['description']}" if o.get("description","").strip() else "")
                         for i, o in enumerate(objs, 1)] or ["(none)"]
                styled(ws.cell(row=row, column=4, value="\n".join(lines)), align=tl)
                ws.row_dimensions[row].height = max(20, len(objs)*15+6)
                row += 1

            row += 1
            sf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            sb = Font(bold=True, size=11)
            total = sum(len(json.loads((self._output_folder/f_.with_suffix(".json").name).read_text())["objects"])
                        for f_ in self._image_files
                        if f_.name in self._saved_status
                        and (self._output_folder/f_.with_suffix(".json").name).exists())
            for label, val in [("Total Images", len(self._saved_status)),
                                ("Total Errors", total),
                                ("Avg / Image", round(total/max(1,len(self._saved_status)),1))]:
                for col, v in ((2, label), (3, val)):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = sb; c.fill = sf; c.border = border
                    if col == 3: c.alignment = ca
                ws.row_dimensions[row].height = 20; row += 1

            for col, w in zip("ABCD", (5, 34, 12, 62)):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"
            wb.save(str(xp))
            print(f"  Excel summary → {xp}")
        except ImportError:
            print("  pip install openpyxl")
        except Exception as e:
            print(f"  Excel error: {e}")

    # ── Status bar ─────────────────────────────────────────────────────────────

    def _update_status_bar(self):
        if not self._image_files: return
        f         = self._image_files[self._idx]
        is_saved  = self._saved_status.get(f.name, False)
        has_edits = len(self._circles) > 0 or f.name in self._img_circles
        state     = "SAVED" if is_saved else ("EDITED" if has_edits else "NO EDITS")
        colors    = {"SAVED": "#5f5", "EDITED": "#5af", "NO EDITS": "#666"}

        self._lbl_file.setText(f.name[:55])
        self._lbl_nav.setText(f"{self._idx+1} / {self._total}")
        self._lbl_mode.setText(f"Mode: {self._mode.value.upper()}")
        self._lbl_obj.setText(f"Objects: {len(self._circles)}")
        self._lbl_res.setText(f"{self._orig_w}×{self._orig_h}")
        self._lbl_zoom.setText(f"Zoom: {self._viewer.current_zoom:.2f}×")
        self._lbl_saved.setText(state)
        self._lbl_saved.setStyleSheet(
            f"color:{colors[state]}; font-weight:bold; padding:0 8px;")

    # ── Close ──────────────────────────────────────────────────────────────────

    def closeEvent(self, event):
        # ── Auto-save unsaved changes on the current image ─────────────────
        if self._circles:
            f = self._image_files[self._idx]
            if not self._saved_status.get(f.name):
                self.save_current(auto_save=True)
        if self._loader and self._loader.isRunning():
            self._loader.abort(); self._loader.wait(500)

        # ── Excel export popup ──────────────────────────────────────────────
        if self._saved_status:
            dlg = QDialog(self)
            dlg.setWindowTitle("Save Excel Summary?")
            dlg.setWindowFlags(
                Qt.WindowType.Dialog |
                Qt.WindowType.FramelessWindowHint)
            dlg.setStyleSheet("""
                QDialog {
                    background: #1e1e2e;
                    border: 1px solid #444;
                    border-radius: 10px;
                }
                QLabel#title {
                    color: #eee;
                    font-size: 14px;
                    font-weight: bold;
                }
                QLabel#sub {
                    color: #999;
                    font-size: 11px;
                }
                QPushButton {
                    padding: 7px 22px;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: bold;
                }
                QPushButton#btn_save {
                    background: #2e6fdb;
                    color: #fff;
                    border: none;
                }
                QPushButton#btn_save:hover  { background: #3a83f5; }
                QPushButton#btn_nosave {
                    background: #333;
                    color: #bbb;
                    border: 1px solid #555;
                }
                QPushButton#btn_nosave:hover { background: #444; color: #fff; }
            """)

            vbox = QVBoxLayout(dlg)
            vbox.setContentsMargins(28, 22, 28, 18)
            vbox.setSpacing(8)

            ico = QLabel("📊")
            ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ico.setStyleSheet("font-size: 32px; background: transparent;")
            vbox.addWidget(ico)

            title = QLabel("Save Excel Summary?")
            title.setObjectName("title")
            title.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.addWidget(title)

            n_saved = len(self._saved_status)
            sub = QLabel(f"{n_saved} image{'s' if n_saved != 1 else ''} annotated  ·  "
                         f"output: {self._output_folder.name}")
            sub.setObjectName("sub")
            sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.addWidget(sub)

            vbox.addSpacing(10)

            hbox = QHBoxLayout()
            hbox.setSpacing(10)

            btn_save   = QPushButton("💾  Save")
            btn_nosave = QPushButton("Don't Save")
            btn_save.setObjectName("btn_save")
            btn_nosave.setObjectName("btn_nosave")
            btn_save.setFixedHeight(34)
            btn_nosave.setFixedHeight(34)

            btn_save.clicked.connect(lambda: dlg.done(1))
            btn_nosave.clicked.connect(lambda: dlg.done(0))

            hbox.addWidget(btn_save)
            hbox.addWidget(btn_nosave)
            vbox.addLayout(hbox)

            # Centre popup over main window
            dlg.adjustSize()
            cx = self.x() + (self.width()  - dlg.width())  // 2
            cy = self.y() + (self.height() - dlg.height()) // 2
            dlg.move(cx, cy)

            if dlg.exec() == 1:
                self.generate_summary()
                xp = self._output_folder / "processing_summary.xlsx"
                print(f"\n  Excel summary \u2192 {xp}")

            print(f"\n  Done \u2014 {len(self._saved_status)}/{self._total} saved"
                  f"\n  Output: {self._output_folder}")

        event.accept()

    @staticmethod
    def _print_instructions():
        print("\n" + "="*66)
        print("  BATCH LABELED EDITOR  ·  Qt  (Optimised)")
        print("="*66)
        print("  Left drag    Draw circle    Right drag  Pan")
        print("  Wheel        Zoom           R           Fit to window")
        print("  A / D        Prev / Next    S           Save")
        print("  Shift+S      Save + Next    U           Undo")
        print("  E            Edit last      C           Clear all")
        print("  T            Toggle labels  L           List")
        print("  1–7          Edit mode      H           Help   Q  Quit")
        print("="*66 + "\n")


# ── Dark palette ───────────────────────────────────────────────────────────────

def _dark_palette() -> QPalette:
    p = QPalette()
    p.setColor(QPalette.ColorRole.Window,          QColor(30, 30, 30))
    p.setColor(QPalette.ColorRole.WindowText,      QColor(220,220,220))
    p.setColor(QPalette.ColorRole.Base,            QColor(18, 18, 18))
    p.setColor(QPalette.ColorRole.AlternateBase,   QColor(25, 25, 25))
    p.setColor(QPalette.ColorRole.Text,            QColor(220,220,220))
    p.setColor(QPalette.ColorRole.Button,          QColor(45, 45, 45))
    p.setColor(QPalette.ColorRole.ButtonText,      QColor(220,220,220))
    p.setColor(QPalette.ColorRole.Highlight,       QColor(60,120,200))
    p.setColor(QPalette.ColorRole.HighlightedText, QColor(255,255,255))
    return p


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="Batch Labeled Editor — Qt")
    ap.add_argument("input_folder")
    ap.add_argument("--output",     "-o",   default=None)
    ap.add_argument("--display-px", type=int, default=MAX_DISPLAY_PX,
                    help=f"Longest-edge cap for display (default {MAX_DISPLAY_PX}; "
                         f"0 = full resolution)")
    args = ap.parse_args()
    dpx  = args.display_px or 999_999   # 0 → effectively unlimited

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setPalette(_dark_palette())
    app.setStyleSheet("""
        QToolTip { background:#2a2a2a; color:#ddd; border:1px solid #555; padding:4px; }
        QScrollBar { width:0; height:0; }
    """)

    try:
        win = MainWindow(args.input_folder, args.output, dpx)
        win.show()
        return app.exec()
    except KeyboardInterrupt:
        print("\n  Interrupted"); return 130
    except Exception as e:
        import traceback
        print(f"\n  Fatal: {e}"); traceback.print_exc(); return 1


if __name__ == "__main__":
    sys.exit(main())
