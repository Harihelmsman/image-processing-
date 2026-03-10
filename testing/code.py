#!/usr/bin/env python3
"""
Batch Labeled Editor - Production Version (Performance Optimised)
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag

Performance architecture
────────────────────────
• Composite cache  — labels baked into image-space ONCE; reused across all
                     zoom/pan frames. Invalidated only when circles change.
• Interaction mode — INTER_LINEAR + no unsharp-mask during zoom/pan/draw;
                     HQ Lanczos fires HQ_SETTLE_MS after last event.
• Dirty tiers      — _circles_dirty (rebuild cache) vs _dirty (re-render).
                     Idle ticks with !_dirty are skipped entirely.
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime
import time
import sys

FONT        = cv2.FONT_HERSHEY_DUPLEX
AA          = cv2.LINE_AA
CANVAS_W    = 1440
CANVAS_H    = 900
CANVAS_BG   = (18, 18, 18)
CANVAS_PAD  = 40
HQ_SETTLE_MS = 180   # ms after last interaction before HQ render fires


class Viewport:
    ZOOM_STEPS     = [0.125,0.167,0.25,0.333,0.5,0.667,1.0,1.5,2.0,3.0,4.0,6.0,8.0,12.0,16.0]
    SNAP_THRESHOLD = 0.04

    def __init__(self, cw, ch):
        self.cw=cw; self.ch=ch; self._scale=1.0; self._tx=0.0; self._ty=0.0

    def fit(self, iw, ih, pad=CANVAS_PAD):
        self._scale = min((self.cw-2*pad)/iw, (self.ch-2*pad)/ih)
        self._snap(); self._recenter(iw, ih)

    def zoom_at(self, sx, sy, factor, iw, ih):
        ix=(sx-self._tx)/self._scale; iy=(sy-self._ty)/self._scale
        min_s=min(0.05,(self.cw-2*CANVAS_PAD)/max(iw,1))
        self._scale=max(min_s,min(16.0,self._scale*factor)); self._snap()
        self._tx=sx-ix*self._scale; self._ty=sy-iy*self._scale

    def pan(self, dx, dy): self._tx+=dx; self._ty+=dy
    def reset(self, iw, ih): self.fit(iw, ih)
    def img_to_screen(self, ix, iy): return int(ix*self._scale+self._tx), int(iy*self._scale+self._ty)
    def screen_to_img(self, sx, sy): return int((sx-self._tx)/self._scale), int((sy-self._ty)/self._scale)

    @property
    def scale(self): return self._scale

    def render(self, image, fast=False):
        ih,iw=image.shape[:2]; tx=self._tx; ty=self._ty; sc=self._scale
        canvas=np.full((self.ch,self.cw,3),CANVAS_BG,dtype=np.uint8)
        sx0=max(0.0,(-tx)/sc); sy0=max(0.0,(-ty)/sc)
        sx1=min(float(iw),(self.cw-tx)/sc); sy1=min(float(ih),(self.ch-ty)/sc)
        if sx1<=sx0 or sy1<=sy0: return canvas
        cx0=max(0,int(np.floor(sx0))); cy0=max(0,int(np.floor(sy0)))
        cx1=min(iw,int(np.ceil(sx1))); cy1=min(ih,int(np.ceil(sy1)))
        crop=image[cy0:cy1,cx0:cx1]
        if crop.size==0: return canvas
        dw=max(1,int(round((cx1-cx0)*sc))); dh=max(1,int(round((cy1-cy0)*sc)))
        if fast:
            scaled=cv2.resize(crop,(dw,dh),interpolation=cv2.INTER_LINEAR)
        else:
            interp=cv2.INTER_LANCZOS4 if sc>=1.0 else cv2.INTER_AREA
            scaled=cv2.resize(crop,(dw,dh),interpolation=interp)
            if sc>=1.0:
                blur=cv2.GaussianBlur(scaled,(0,0),sigmaX=1.0)
                scaled=cv2.addWeighted(scaled,1.4,blur,-0.4,0)
        dx=max(0,int(tx+cx0*sc)); dy=max(0,int(ty+cy0*sc))
        pw=min(dw,self.cw-dx); ph=min(dh,self.ch-dy)
        if pw>0 and ph>0: canvas[dy:dy+ph,dx:dx+pw]=scaled[:ph,:pw]
        zw=int(iw*sc); zh=int(ih*sc); itx=int(tx); ity=int(ty)
        cv2.rectangle(canvas,(max(0,itx-1),max(0,ity-1)),
                      (min(self.cw-1,itx+zw),min(self.ch-1,ity+zh)),(45,45,45),1,AA)
        return canvas

    def _snap(self):
        for s in self.ZOOM_STEPS:
            if abs(self._scale-s)/s<self.SNAP_THRESHOLD: self._scale=s; return

    def _recenter(self, iw, ih):
        self._tx=(self.cw-iw*self._scale)/2; self._ty=(self.ch-ih*self._scale)/2


class EditMode(Enum):
    HIGHLIGHT="highlight"; BLUR="blur"; PIXELATE="pixelate"
    DARKEN="darken"; GRAYSCALE="grayscale"; INVERT="invert"; OUTLINE="outline"


class BatchLabeledEditor:
    MAX_BATCH_SIZE=200; MAX_RECOMMENDED_CIRCLES=30; MIN_IMAGE_SIZE=50
    ZOOM_DEBOUNCE_MS=30; MAX_CACHED_STATES=5

    def __init__(self, input_folder, output_folder=None):
        self.input_folder=Path(input_folder)
        if output_folder:
            self.output_folder=Path(output_folder)
        else:
            ts=datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder=self.input_folder.parent/f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True,exist_ok=True)
        self.image_files=self._load_image_files()
        if not self.image_files: raise ValueError(f"No valid images in {input_folder}")
        self.total_images=len(self.image_files); self.current_index=0
        self.original_image=self.scaled_image=self.display_image=self.output_image=None
        self.circles=[]; self.drawing=False; self.center=None; self.current_radius=0
        self.current_label=""; self.current_description=""
        self.label_input_mode=False; self.description_input_mode=False
        self.vp=Viewport(CANVAS_W,CANVAS_H)
        self.is_panning=False; self._pan_last_x=self._pan_last_y=0; self.last_zoom_time=0
        # ── composite cache ──────────────────────────────────────────────────
        self._composite_cache=None   # image-space baked composite
        self._circles_dirty=True     # True → rebuild cache before next render
        # ── interaction state ─────────────────────────────────────────────────
        self._interacting=False       # True → use fast INTER_LINEAR render
        self._last_interact_ms=0.0    # epoch-ms of last zoom/pan/draw event
        # ── general ──────────────────────────────────────────────────────────
        self._dirty=True
        self.current_mode=EditMode.HIGHLIGHT
        self._blur_kernel=25; self._pixelate_size=10; self.highlight_alpha=0.4
        self.show_labels=True
        self.mode_colors={
            EditMode.HIGHLIGHT:(0,255,0),   EditMode.BLUR:(255,80,80),
            EditMode.PIXELATE:(80,80,255),  EditMode.DARKEN:(160,160,160),
            EditMode.GRAYSCALE:(210,210,210),EditMode.INVERT:(255,255,0),
            EditMode.OUTLINE:(0,220,255),
        }
        self.saved_status={}; self.image_states={}; self.state_access_order=[]
        self.window_name="Batch Labeled Editor"
        cv2.namedWindow(self.window_name,cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name,CANVAS_W,CANVAS_H)
        cv2.setMouseCallback(self.window_name,self._mouse_callback)
        if not self._load_current_image(): raise RuntimeError("Failed to load first image")
        self._print_instructions()

    @property
    def blur_kernel(self): return self._blur_kernel
    @blur_kernel.setter
    def blur_kernel(self,v):
        if not isinstance(v,int) or v<=0 or v%2==0: raise ValueError("must be positive odd int")
        self._blur_kernel=v
    @property
    def pixelate_size(self): return self._pixelate_size
    @pixelate_size.setter
    def pixelate_size(self,v):
        if not isinstance(v,int) or v<=0: raise ValueError("must be positive int")
        self._pixelate_size=v

    # ── file loading ──────────────────────────────────────────────────────────
    def _load_image_files(self):
        exts={'.jpg','.jpeg','.png','.bmp','.tiff','.tif'}
        files=sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid=[]
        for f in files:
            try:
                if f.stat().st_size<512: print(f"⚠  Skipping {f.name}: too small"); continue
                valid.append(f)
            except Exception as e: print(f"⚠  Skipping {f.name}: {e}")
        if len(valid)>self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images")
            if input("   Continue? (y/N): ").strip().lower()!='y': sys.exit(0)
        return valid

    def _scale_image(self):
        self.scale_factor=1.0; self.scaled_image=self.original_image.copy()

    def _load_current_image(self):
        if self.current_index>=len(self.image_files): return False
        f=self.image_files[self.current_index]
        try:
            self.original_image=cv2.imread(str(f))
            if self.original_image is None: raise IOError("cv2.imread returned None")
            if min(self.original_image.shape[:2])<self.MIN_IMAGE_SIZE: raise ValueError("too small")
            self._scale_image()
            ih,iw=self.original_image.shape[:2]; self.vp.fit(iw,ih)
            json_path=self.output_folder/f.with_suffix('.json').name
            if json_path.exists():
                try:
                    with open(json_path) as jf: data=json.load(jf)
                    self.circles=[{'center':tuple(o['center']),'radius':o['radius'],
                        'mode':EditMode(o.get('mode','highlight')),
                        'label':o.get('label',''),'description':o.get('description','')}
                        for o in data.get('objects',[])]
                    print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images}) — restored {len(self.circles)} objects")
                except Exception as e: print(f"⚠  JSON load: {e}"); self.circles=[]
            elif f.name in self.image_states:
                self.circles=self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images}) — from memory")
            else:
                self.circles=[]; print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")
            self._update_state_access(f.name)
            self.drawing=self.label_input_mode=self.description_input_mode=False
            self.current_label=self.current_description=""
            self.output_image=self.scaled_image.copy()
            self._circles_dirty=True; self._composite_cache=None
            if self.circles: self._apply_all_effects()
            self.display_image=self.output_image.copy(); self._dirty=True
            return True
        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index<len(self.image_files)-1:
                self.current_index+=1; return self._load_current_image()
            return False

    def _update_state_access(self,name):
        if name in self.state_access_order: self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order)>self.MAX_CACHED_STATES:
            for n in self.state_access_order[:-self.MAX_CACHED_STATES]: self.image_states.pop(n,None)
            self.state_access_order=self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── mouse ──────────────────────────────────────────────────────────────────
    def _mouse_callback(self, event, x, y, flags, param):
        ih,iw=self.scaled_image.shape[:2]
        if event==cv2.EVENT_MOUSEWHEEL:
            now=time.time()*1000
            if now-self.last_zoom_time<self.ZOOM_DEBOUNCE_MS: return
            self.last_zoom_time=now
            self.vp.zoom_at(x,y,1.2 if flags>0 else 0.8,iw,ih)
            self._mark_interacting(); self._dirty=True; self._update_display(); return
        if event==cv2.EVENT_RBUTTONDOWN:
            self.is_panning=True; self._pan_last_x=x; self._pan_last_y=y; return
        if event==cv2.EVENT_RBUTTONUP:
            self.is_panning=False; self._dirty=True; return
        if event==cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.vp.pan(x-self._pan_last_x,y-self._pan_last_y)
            self._pan_last_x=x; self._pan_last_y=y
            self._mark_interacting(); self._dirty=True; self._update_display(); return
        if self.label_input_mode or self.description_input_mode or self.is_panning: return
        if event==cv2.EVENT_LBUTTONDOWN:
            self.drawing=True; self.center=self.vp.screen_to_img(x,y); self.current_radius=0
        elif event==cv2.EVENT_MOUSEMOVE and self.drawing:
            ix,iy=self.vp.screen_to_img(x,y)
            self.current_radius=int(np.hypot(ix-self.center[0],iy-self.center[1]))
            self._mark_interacting(); self._dirty=True; self._update_display()
        elif event==cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing=False
            if self.current_radius>5:
                if len(self.circles)>=self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles — performance may degrade")
                self._enter_label_input_mode()

    def _mark_interacting(self):
        self._interacting=True; self._last_interact_ms=time.time()*1000

    def _check_settle(self):
        if not self._interacting: return False
        if (time.time()*1000-self._last_interact_ms)>=HQ_SETTLE_MS:
            self._interacting=False; return True
        return False

    # ── label/desc flow ───────────────────────────────────────────────────────
    def _enter_label_input_mode(self):
        self.label_input_mode=True; self.current_label=""
        print(f"\n→ [{self.current_mode.value.upper()}] type label, ENTER confirm, ESC cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self,save):
        self.label_input_mode=False
        if save: self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled"); self.current_label=""; self.drawing=False
            self._dirty=True; self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode=True; self.current_description=""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode=False
        lbl=self.current_label.strip(); desc=self.current_description.strip()
        self.circles.append({'center':self.center,'radius':self.current_radius,
            'mode':self.current_mode,'label':lbl,'description':desc})
        tag=f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}"+(" + description" if desc else "")+f" [{self.current_mode.value}]")
        self.current_label=self.current_description=""
        self._apply_all_effects(); self._dirty=True; self._update_display()

    # ── effects ───────────────────────────────────────────────────────────────
    def _apply_effect(self, image, circle):
        try:
            mask=np.zeros(image.shape[:2],dtype=np.uint8)
            cv2.circle(mask,circle['center'],circle['radius'],255,-1)
            m=circle['mode']
            if m==EditMode.HIGHLIGHT:
                lit=cv2.addWeighted(image,1-self.highlight_alpha,np.full_like(image,255),self.highlight_alpha,0)
                image=np.where(mask[:,:,np.newaxis]==255,lit,image)
            elif m==EditMode.BLUR:
                image=np.where(mask[:,:,np.newaxis]==255,cv2.GaussianBlur(image,(self._blur_kernel,)*2,0),image)
            elif m==EditMode.PIXELATE:
                h,w=image.shape[:2]
                small=cv2.resize(image,(max(1,w//self._pixelate_size),max(1,h//self._pixelate_size)),interpolation=cv2.INTER_NEAREST)
                image=np.where(mask[:,:,np.newaxis]==255,cv2.resize(small,(w,h),interpolation=cv2.INTER_NEAREST),image)
            elif m==EditMode.DARKEN:
                image=np.where(mask[:,:,np.newaxis]==255,cv2.addWeighted(image,0.5,np.zeros_like(image),0.5,0),image)
            elif m==EditMode.GRAYSCALE:
                gray=cv2.cvtColor(cv2.cvtColor(image,cv2.COLOR_BGR2GRAY),cv2.COLOR_GRAY2BGR)
                image=np.where(mask[:,:,np.newaxis]==255,gray,image)
            elif m==EditMode.INVERT:
                image=np.where(mask[:,:,np.newaxis]==255,cv2.bitwise_not(image),image)
        except Exception as e: print(f"⚠  Effect error: {e}")
        return image

    def _apply_all_effects(self):
        """Recompute output_image and invalidate composite cache."""
        self.output_image=self.scaled_image.copy()
        for c in self.circles:
            self.output_image=self._apply_effect(self.output_image,c)
            cv2.circle(self.output_image,c['center'],c['radius'],self.mode_colors[c['mode']],2,AA)
        self._circles_dirty=True; self._composite_cache=None

    # ── composite cache ───────────────────────────────────────────────────────
    def _rebuild_composite_cache(self):
        """Bake labels into image-space once — zoom/pan then only calls vp.render()."""
        tmp=self.output_image.copy()
        if self.show_labels: self._draw_all_labels_smart(tmp)
        self._composite_cache=tmp; self._circles_dirty=False

    def _get_composite(self):
        if self._circles_dirty or self._composite_cache is None: self._rebuild_composite_cache()
        return self._composite_cache

    # ── label drawing (called once into cache, not every frame) ───────────────
    def _label_font_params(self,img_shape):
        md=min(img_shape[:2])
        if md<200: return 0.30,1
        elif md<400: return 0.40,1
        elif md<800: return 0.50,1
        else: return 0.55,1

    def _rects_overlap(self,r1,r2,buf=4):
        return not(r1[2]+buf<r2[0] or r1[0]-buf>r2[2] or r1[3]+buf<r2[1] or r1[1]-buf>r2[3])

    def _find_label_pos(self,center,radius,tw,fh,bl,pad,img_shape,placed):
        ih,iw=img_shape[:2]; W=tw+2*pad; H=fh+bl+2*pad; cx,cy=center
        cands=[(cx-radius,cy-radius-H-10),(cx-radius,cy+radius+10),
               (cx-radius-W-8,cy-H//2),(cx+radius+8,cy-H//2),
               (cx+radius+8,cy-radius-H-10),(cx-radius-W-8,cy-radius-H-10),
               (cx+radius+8,cy+radius+10),(cx-radius-W-8,cy+radius+10)]
        for extra in (40,80,120):
            cands+=[(px,py+extra) for px,py in cands[:4]]
            cands+=[(px,py-extra) for px,py in cands[:4]]
        for px,py in cands:
            rect=(int(px),int(py),int(px+W),int(py+H))
            if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad: continue
            if not any(self._rects_overlap(rect,r) for r in placed): return int(px)+pad,int(py)+fh+pad,rect
        bottom=max((r[3] for r in placed),default=0)+8
        lx=max(pad,min(cx-W//2,iw-W-pad)); ly=min(bottom+fh+pad,ih-bl-pad)
        return lx,ly,(lx-pad,ly-fh-pad,lx+tw+pad,ly+bl+pad)

    def _draw_all_labels_smart(self,image):
        if not self.show_labels: return
        placed=[]; ih,iw=image.shape[:2]; md=min(ih,iw)
        sc,th=(0.45,1) if md<400 else (0.58,1) if md<900 else (0.72,2)
        for i,c in enumerate(self.circles,1):
            label=c['label'] or f"Error #{i}"; color=self.mode_colors[c['mode']]
            cx,cy=c['center']; r=c['radius']
            bt=str(i); (bw,bh),_=cv2.getTextSize(bt,FONT,sc*0.9,th)
            br=max(bw,bh)//2+6
            cv2.circle(image,(cx+2,cy+2),br,(0,0,0),-1,AA)
            cv2.circle(image,(cx,cy),br,color,-1,AA)
            cv2.circle(image,(cx,cy),br,(255,255,255),2,AA)
            cv2.putText(image,bt,(cx-bw//2,cy+bh//2),FONT,sc*0.9,(0,0,0),th+1,AA)
            text=f"#{i}  [{c['mode'].value[:3].upper()}]  {label}"
            (tw,fh),bl=cv2.getTextSize(text,FONT,sc,th); pad=max(5,int(5*sc/0.5))
            lx,ly,rect=self._find_label_pos((cx,cy),r,tw,fh,bl,pad,image.shape,placed)
            placed.append(rect); x1,y1,x2,y2=rect
            for ge,ga in ((6,0.25),(3,0.45)):
                glow=image.copy()
                cv2.rectangle(glow,(x1-ge,y1-ge),(x2+ge,y2+ge),color,-1,AA)
                cv2.addWeighted(glow,ga,image,1-ga,0,image)
            cv2.rectangle(image,(x1,y1),(x2,y2),(15,15,20),-1)
            cv2.rectangle(image,(x1,y1),(x1+3,y2),color,-1,AA)
            cv2.rectangle(image,(x1,y1),(x2,y2),color,2,AA)
            tx=lx+4
            cv2.putText(image,text,(tx+1,ly+1),FONT,sc,(0,0,0),th+1,AA)
            cv2.putText(image,text,(tx,ly),FONT,sc,(255,255,255),th,AA)
            mx=(x1+x2)//2; my=y2 if ly>cy else y1
            ang=np.arctan2(my-cy,mx-cx)
            ex=int(cx+r*np.cos(ang)); ey=int(cy+r*np.sin(ang))
            cv2.line(image,(mx,my),(ex,ey),color,2,AA)
            cv2.circle(image,(ex,ey),4,color,-1,AA)
            cv2.circle(image,(ex,ey),4,(255,255,255),1,AA)

    def _draw_typing_label(self,image,center,radius,label,mode):
        sc,th=self._label_font_params(image.shape); sc+=0.1; th+=1
        text=f"[{mode.value[:3].upper()}] {label}_"
        (tw,fh),bl=cv2.getTextSize(text,FONT,sc,th); pad=max(4,int(6*sc/0.5))
        ih,iw=image.shape[:2]; color=self.mode_colors[mode]
        px,py=center[0]-radius,center[1]-radius-fh-2*pad-10
        if not(px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px,py=center[0]-radius,center[1]+radius+10
        if not(px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px=max(pad,min(center[0]-radius,iw-tw-2*pad-pad))
            py=max(fh+pad,center[1]-radius-fh-2*pad-10)
        lx=int(px); ly=int(py+fh+pad)
        cv2.rectangle(image,(lx-pad,ly-fh-pad),(lx+tw+pad,ly+bl+pad),(10,10,10),-1)
        cv2.rectangle(image,(lx-pad,ly-fh-pad),(lx+tw+pad,ly+bl+pad),color,2,AA)
        cv2.putText(image,text,(lx,ly),FONT,sc,(255,255,255),th,AA)
        cv2.line(image,(lx+tw//2,ly+bl+pad),center,color,2,AA)

    # ── display ───────────────────────────────────────────────────────────────
    def _present(self): self._dirty=False; cv2.imshow(self.window_name,self.display_image)

    def _update_display(self):
        """
        O(viewport-resize) per frame during zoom/pan — label work only when cache stale.
        """
        if not self._dirty: return
        composite=self._get_composite()
        if self.drawing and self.current_radius>0:
            src=composite.copy()
            cv2.circle(src,self.center,self.current_radius,self.mode_colors[self.current_mode],2,AA)
        else:
            src=composite
        self.display_image=self.vp.render(src,fast=self._interacting)
        self._draw_ui(); self._present()

    def _update_display_with_input(self):
        tmp=self._get_composite().copy()
        cv2.circle(tmp,self.center,self.current_radius,self.mode_colors[self.current_mode],3,AA)
        self._draw_typing_label(tmp,self.center,self.current_radius,self.current_label,self.current_mode)
        self.display_image=self.vp.render(tmp,fast=False)
        self._draw_input_box("Label"); self._present()

    def _update_display_with_description_input(self):
        tmp=self._get_composite().copy()
        cv2.circle(tmp,self.center,self.current_radius,self.mode_colors[self.current_mode],3,AA)
        if self.current_label: self._draw_typing_label(tmp,self.center,self.current_radius,self.current_label,self.current_mode)
        self.display_image=self.vp.render(tmp,fast=False)
        self._draw_input_box("Description (optional)"); self._present()

    def _draw_input_box(self,prompt):
        h,w=self.display_image.shape[:2]; bh=64; by=h-bh; color=self.mode_colors[self.current_mode]
        ov=self.display_image.copy(); cv2.rectangle(ov,(0,by),(w,h),(12,12,12),-1)
        cv2.addWeighted(ov,0.88,self.display_image,0.12,0,self.display_image)
        cv2.rectangle(self.display_image,(0,by),(w,h),color,2,AA)
        cv2.putText(self.display_image,f"[{self.current_mode.value.upper()}]",(12,by+22),FONT,0.52,color,1,AA)
        cv2.putText(self.display_image,f"{prompt}:",(160,by+22),FONT,0.52,(200,200,200),1,AA)
        text=(self.current_description if self.description_input_mode else self.current_label)+"_"
        cv2.putText(self.display_image,text,(12,by+50),FONT,0.65,(0,240,240),1,AA)

    def _draw_ui(self):
        ov=self.display_image.copy(); h,w=self.display_image.shape[:2]; bh=84
        cv2.rectangle(ov,(0,0),(w,bh),(12,12,12),-1)
        color=self.mode_colors[self.current_mode]; cf=self.image_files[self.current_index]
        cv2.putText(ov,f"Mode: {self.current_mode.value.upper()}",(14,24),FONT,0.55,color,1,AA)
        cv2.putText(ov,f"Zoom: {self.vp.scale:.2f}x",(14,48),FONT,0.45,(100,190,255),1,AA)
        cv2.putText(ov,f"{self.current_index+1} / {self.total_images}",(14,70),FONT,0.43,(160,160,160),1,AA)
        fname=cf.name if len(cf.name)<=46 else cf.name[:43]+"..."
        cv2.putText(ov,fname,(200,24),FONT,0.46,(185,185,185),1,AA)
        n=len(self.circles); oc=(40,160,255) if n>=self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(ov,f"Objects: {n}",(200,48),FONT,0.43,oc,1,AA)
        qi="FAST" if self._interacting else "HQ"
        qc=(60,180,255) if self._interacting else (60,220,100)
        cv2.putText(ov,qi,(200,70),FONT,0.38,qc,1,AA)
        is_saved=self.saved_status.get(cf.name,False)
        has_edits=cf.name in self.image_states or n>0
        st,sc=("SAVED",(60,200,60)) if is_saved else ("EDITED",(40,160,255)) if has_edits else ("NO EDITS",(80,80,80))
        cv2.putText(ov,st,(w-155,24),FONT,0.54,sc,1,AA)
        hint="Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  R:Reset  C:Clear  U:Undo  H:Help  Q:Quit"
        cv2.putText(ov,hint,(12,h-10),FONT,0.36,(110,110,110),1,AA)
        cv2.addWeighted(ov,0.82,self.display_image,0.18,0,self.display_image)

    # ── navigation ────────────────────────────────────────────────────────────
    def _guard_state(self):
        if self.drawing: print("⚠  Finish drawing"); return True
        if self.label_input_mode: print("⚠  Finish label input (ESC)"); return True
        if self.description_input_mode: print("⚠  Finish description (ENTER)"); return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index>0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index-=1; self._load_current_image(); self._update_display()
        else: print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index<self.total_images-1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index+=1; self._load_current_image(); self._update_display()
        else: print("Already at last image")

    # ── misc ──────────────────────────────────────────────────────────────────
    def _list_labels(self):
        print("\n"+"="*60)
        print(f"Objects — {self.image_files[self.current_index].name}"); print("="*60)
        if not self.circles: print("  (none)")
        for i,c in enumerate(self.circles,1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  r={c['radius']}  pos={c['center']}")
            if c.get('description'): print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        print(f"\nMemory — {len(self.image_states)} states cached (max {self.MAX_CACHED_STATES})  |  {len(self.saved_status)} saved\n")

    def _edit_last_label(self):
        if not self.circles: print("No objects to edit"); return
        last=self.circles[-1]
        print(f"\nEditing: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label=last['label']; self.label_input_mode=True
        self._update_display_with_input()
        while self.label_input_mode:
            key=cv2.waitKey(0)&0xFF
            if key==27: self.label_input_mode=False; self.current_label=""; self._dirty=True; self._update_display(); return
            elif key==13: last['label']=self.current_label.strip(); self.label_input_mode=False; break
            elif key==8: self.current_label=self.current_label[:-1]; self._update_display_with_input()
            elif 32<=key<=126: self.current_label+=chr(key); self._update_display_with_input()
        print(f"\nEditing description: '{last.get('description','')}'  (ENTER=save, ESC=skip)")
        self.current_description=last.get('description',''); self.description_input_mode=True
        self._update_display_with_description_input()
        while self.description_input_mode:
            key=cv2.waitKey(0)&0xFF
            if key==13: last['description']=self.current_description.strip(); self.description_input_mode=False; break
            elif key==27: self.description_input_mode=False; break
            elif key==8: self.current_description=self.current_description[:-1]; self._update_display_with_description_input()
            elif 32<=key<=126: self.current_description+=chr(key); self._update_display_with_description_input()
        self.current_label=self.current_description=""; self._circles_dirty=True
        print("  ✓ Updated"); self._apply_all_effects(); self._update_display()

    # ── save ──────────────────────────────────────────────────────────────────
    def save_current(self,auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save"); return
        cf=self.image_files[self.current_index]; out=self.output_folder/cf.name; js=out.with_suffix('.json')
        self.image_states[cf.name]={'circles':[c.copy() for c in self.circles]}
        self._update_state_access(cf.name)
        try:
            if self.scale_factor!=1.0:
                final=self.original_image.copy()
                sc_circles=[{'center':(int(c['center'][0]/self.scale_factor),int(c['center'][1]/self.scale_factor)),
                    'radius':int(c['radius']/self.scale_factor),'mode':c['mode'],'label':c['label']} for c in self.circles]
                for sc in sc_circles:
                    final=self._apply_effect(final,sc)
                    cv2.circle(final,sc['center'],sc['radius'],self.mode_colors[sc['mode']],3,AA)
                old=self.circles; self.circles=sc_circles; self._draw_all_labels_smart(final); self.circles=old
            else:
                final=self.output_image.copy(); self._draw_all_labels_smart(final)
            if not cv2.imwrite(str(out),final): raise IOError(f"imwrite failed: {out}")
            data={'source_image':cf.name,'timestamp':datetime.now().isoformat(),
                'objects':[{'id':i,'label':c['label'],'description':c.get('description',''),
                    'mode':c['mode'].value,'center':list(c['center']),'radius':c['radius']}
                    for i,c in enumerate(self.circles,1)]}
            with open(js,'w') as f: json.dump(data,f,indent=2)
            self.saved_status[cf.name]=True
            if auto_save: print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else: print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)\n  → {out}\n  → {js}")
            self._cleanup_old_states()
        except Exception as e: print(f"❌ Save error: {e}")

    # ── excel summary ─────────────────────────────────────────────────────────
    def generate_summary(self):
        xp=self.output_folder/"processing_summary.xlsx"
        try:
            import openpyxl; from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
            wb=openpyxl.Workbook(); ws=wb.active; ws.title="Summary"
            hfill=PatternFill(start_color="2E4DA7",end_color="2E4DA7",fill_type="solid")
            hfont=Font(bold=True,color="FFFFFF",size=11)
            border=Border(**{s:Side(style='thin') for s in ('left','right','top','bottom')})
            center=Alignment(horizontal='center',vertical='center',wrap_text=True)
            top_l=Alignment(horizontal='left',vertical='top',wrap_text=True)
            def styled(cell,fill=None,font=None,align=None,bord=True):
                if fill: cell.fill=fill
                if font: cell.font=font
                if align: cell.alignment=align
                if bord: cell.border=border
            for col,hdr in enumerate(["#","Image Name","Error Count","Errors (numbered)"],1):
                c=ws.cell(row=1,column=col,value=hdr); styled(c,fill=hfill,font=hfont,align=center)
            ws.row_dimensions[1].height=28
            data_row=2; img_num=0
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                jp=self.output_folder/img.with_suffix('.json').name
                if not jp.exists(): continue
                with open(jp) as f: data=json.load(f)
                objects=data.get('objects',[]); img_num+=1
                c1=ws.cell(row=data_row,column=1,value=img_num); styled(c1,align=center); c1.font=Font(bold=True,size=10)
                c2=ws.cell(row=data_row,column=2,value=img.name); styled(c2,align=top_l)
                c3=ws.cell(row=data_row,column=3,value=len(objects)); styled(c3,align=center)
                if len(objects)==0: c3.fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid"); c3.font=Font(color="276221",bold=True)
                elif len(objects)<=3: c3.fill=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid"); c3.font=Font(color="9C5700",bold=True)
                else: c3.fill=PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid"); c3.font=Font(color="9C0006",bold=True)
                lines=[f"{i}. {o.get('label','').strip() or '(unlabeled)'}"+(" — "+o['description'].strip() if o.get('description','').strip() else "") for i,o in enumerate(objects,1)]
                c4=ws.cell(row=data_row,column=4,value="\n".join(lines) if lines else "(no errors)"); styled(c4,align=top_l)
                ws.row_dimensions[data_row].height=max(20,len(objects)*15+6)
                if img_num%2==0:
                    alt=PatternFill(start_color="F2F5FB",end_color="F2F5FB",fill_type="solid")
                    for col in (1,2,4): ws.cell(row=data_row,column=col).fill=alt
                data_row+=1
            data_row+=1
            sfill=PatternFill(start_color="D9E1F2",end_color="D9E1F2",fill_type="solid"); sfont=Font(bold=True,size=11)
            def summary_row(label,value):
                nonlocal data_row
                lc=ws.cell(row=data_row,column=2,value=label); lc.font=sfont; lc.fill=sfill; lc.border=border
                vc=ws.cell(row=data_row,column=3,value=value); vc.font=sfont; vc.fill=sfill; vc.border=border; vc.alignment=center
                ws.row_dimensions[data_row].height=20; data_row+=1
            summary_row("Total Images Processed",len(self.saved_status))
            total_obj=sum(len(json.load(open(self.output_folder/f.with_suffix('.json').name))['objects']) for f in self.image_files if f.name in self.saved_status and (self.output_folder/f.with_suffix('.json').name).exists())
            summary_row("Total Errors Found",total_obj)
            if self.saved_status: summary_row("Avg Errors / Image",round(total_obj/len(self.saved_status),1))
            ws.column_dimensions['A'].width=5; ws.column_dimensions['B'].width=34
            ws.column_dimensions['C'].width=13; ws.column_dimensions['D'].width=62
            ws.freeze_panes="A2"; wb.save(str(xp)); print(f"✓ Excel summary: {xp}")
        except ImportError: print("⚠  pip install openpyxl")
        except Exception as e: print(f"⚠  Excel error: {e}")

    def _print_instructions(self):
        print("\n"+"="*58+"\nBATCH LABELED EDITOR\n"+"="*58)
        print(f"  Input : {self.input_folder}\n  Output: {self.output_folder}\n  Images: {self.total_images}\n")
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*58+"\n")

    # ── main loop ─────────────────────────────────────────────────────────────
    def run(self):
        self._dirty=True; self._update_display()
        while True:
            key=cv2.waitKey(1)&0xFF
            # idle tick
            if key==0xFF:
                if self._check_settle():   # interaction ended — fire one HQ render
                    self._dirty=True; self._update_display()
                elif not self._dirty: continue
            # description input
            if self.description_input_mode:
                if key==27:
                    self.description_input_mode=False; self.current_label=self.current_description=""
                    print("  ✗ Circle discarded"); self._dirty=True; self._update_display()
                elif key==13: self._exit_description_input_mode()
                elif key==8: self.current_description=self.current_description[:-1]; self._update_display_with_description_input()
                elif 32<=key<=126: self.current_description+=chr(key); self._update_display_with_description_input()
                continue
            # label input
            if self.label_input_mode:
                if key==27: self._exit_label_input_mode(save=False)
                elif key==13: self._exit_label_input_mode(save=True)
                elif key==8: self.current_label=self.current_label[:-1]; self._update_display_with_input()
                elif 32<=key<=126: self.current_label+=chr(key); self._update_display_with_input()
                continue
            # commands
            if key==ord('a'): self._previous_image()
            elif key==ord('d'): self._next_image()
            elif key==ord('r'):
                ih,iw=self.scaled_image.shape[:2]; self.vp.reset(iw,ih)
                self._interacting=False; self._dirty=True; self._update_display(); print("✓ Zoom reset")
            elif key==ord('s'): self.save_current(); self._dirty=True; self._update_display()
            elif key==ord('S'): self.save_current(); self._next_image()
            elif key==ord('c'):
                self.circles.clear(); self.output_image=self.scaled_image.copy()
                self._circles_dirty=True; self._composite_cache=None
                self._dirty=True; self._update_display(); print("✓ Cleared")
            elif key==ord('u'):
                if self.circles:
                    rem=self.circles.pop(); print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._apply_all_effects(); self._dirty=True; self._update_display()
            elif key==ord('l'): self._list_labels()
            elif key==ord('e'): self._edit_last_label()
            elif key==ord('t'):
                self.show_labels=not self.show_labels; self._circles_dirty=True; self._composite_cache=None
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}"); self._dirty=True; self._update_display()
            elif key==ord('m'): self._show_memory_status()
            elif key in (ord('h'),ord('H')): self._print_instructions()
            elif ord('1')<=key<=ord('7'):
                self.current_mode=list(EditMode)[key-ord('1')]; print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._dirty=True; self._update_display()
            elif key==ord('q'):
                if self.circles and not self.saved_status.get(self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break
        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved\n   Output: {self.output_folder}")
        else: print("\n⚠  No images were saved")


def main():
    ap=argparse.ArgumentParser(description="Batch Labeled Editor")
    ap.add_argument("input_folder"); ap.add_argument("--output","-o",default=None)
    args=ap.parse_args()
    try: BatchLabeledEditor(args.input_folder,args.output).run()
    except KeyboardInterrupt: print("\n⚠  Interrupted"); return 130
    except Exception as e: print(f"\n❌ Fatal: {e}"); import traceback; traceback.print_exc(); return 1
    return 0

if __name__=="__main__": exit(main())
