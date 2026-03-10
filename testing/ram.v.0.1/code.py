#!/usr/bin/env python3
"""
Batch Labeled Editor - Production Version
Centered image on dark canvas · Anti-aliased fonts · Zoom · Auto-save
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import time
import sys
from datetime import datetime


# ── Typography & canvas constants ───────────────────────────────────────────────
FONT       = cv2.FONT_HERSHEY_DUPLEX   # smooth, no pixel-aliasing
AA         = cv2.LINE_AA
CANVAS_W   = 1440
CANVAS_H   = 900
CANVAS_BG  = (18, 18, 18)             # near-black surround
CANVAS_PAD = 40                        # clear gap around image on canvas


class Viewport:
    """
    Professional viewer transform — CVAT / LabelImg architecture.

    Stores the view as a single (scale, tx, ty) affine transform instead of
    separate pan_x / pan_y / zoom_level fields.  Benefits:
      • One source of truth — no drift between the three variables.
      • Pixel-perfect rendering: snaps to canonical integer ratios and selects
        the correct cv2 interpolation flag automatically.
      • zoom_at() keeps the exact image pixel under the cursor fixed — the same
        maths used in every professional image viewer.
    """

    # Canonical zoom levels — snapping makes 1:1, 2:1, etc. exactly representable
    ZOOM_STEPS = [0.125, 0.167, 0.25, 0.333, 0.5, 0.667,
                  1.0, 1.5, 2.0, 3.0, 4.0, 6.0, 8.0, 12.0, 16.0]
    
    # ── rendering ────────────────────────────────────────────────────────────────

    PIXEL_INSPECT_THRESHOLD = 3.0   # switch to nearest-neighbour

    def render(self, image: np.ndarray,
               post_crop_fn=None) -> np.ndarray:
        """Crop-then-scale renderer. post_crop_fn(crop, cx0, cy0) → modified crop
        is called on the tiny visible rectangle before scaling, so effects only
        ever touch ≤ CANVAS_W×CANVAS_H pixels instead of the full source image."""
        ih, iw = image.shape[:2]
        tx = self._tx;  ty = self._ty
        sc = self._scale

        canvas = np.full((self.ch, self.cw, 3), CANVAS_BG, dtype=np.uint8)

        src_x0 = max(0.0,    (-tx) / sc)
        src_y0 = max(0.0,    (-ty) / sc)
        src_x1 = min(float(iw), (self.cw - tx) / sc)
        src_y1 = min(float(ih), (self.ch - ty) / sc)
        if src_x1 <= src_x0 or src_y1 <= src_y0:
            return canvas

        cx0 = max(0,  int(np.floor(src_x0)))
        cy0 = max(0,  int(np.floor(src_y0)))
        cx1 = min(iw, int(np.ceil(src_x1)))
        cy1 = min(ih, int(np.ceil(src_y1)))

        # .copy() only if callback will mutate the crop
        crop = image[cy0:cy1, cx0:cx1].copy() if post_crop_fn else image[cy0:cy1, cx0:cx1]
        if crop.size == 0:
            return canvas

        if post_crop_fn:
            crop = post_crop_fn(crop, cx0, cy0)

        dst_w = max(1, int(round((cx1 - cx0) * sc)))
        dst_h = max(1, int(round((cy1 - cy0) * sc)))

        if sc < 1.0:
            interp = cv2.INTER_AREA
        elif sc < self.PIXEL_INSPECT_THRESHOLD:
            interp = cv2.INTER_LANCZOS4
        else:
            interp = cv2.INTER_NEAREST

        scaled = cv2.resize(crop, (dst_w, dst_h), interpolation=interp)

        dx = max(0, int(tx + cx0 * sc))
        dy = max(0, int(ty + cy0 * sc))
        pw = min(dst_w, self.cw - dx)
        ph = min(dst_h, self.ch - dy)
        if pw > 0 and ph > 0:
            canvas[dy:dy+ph, dx:dx+pw] = scaled[:ph, :pw]

        # subtle image border
        zw = int(iw * sc);  zh = int(ih * sc)
        itx = int(tx);      ity = int(ty)
        cv2.rectangle(canvas,
                      (max(0, itx-1),          max(0, ity-1)),
                      (min(self.cw-1, itx+zw), min(self.ch-1, ity+zh)),
                      (45, 45, 45), 1, AA)
        return canvas

    # ── private helpers ──────────────────────────────────────────────────────────

    def _snap(self) -> None:
        for step in self.ZOOM_STEPS:
            if abs(self._scale - step) / step < self.SNAP_THRESHOLD:
                self._scale = step
                return

    def _recenter(self, img_w: int, img_h: int) -> None:
        self._tx = (self.cw - img_w * self._scale) / 2
        self._ty = (self.ch - img_h * self._scale) / 2


class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50
    ZOOM_DEBOUNCE_MS        = 20
    MAX_CACHED_STATES       = 2   # was 5 — each state holds full circle data

    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)

        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.total_images  = len(self.image_files)
        self.current_index = 0

        self.original_image         = None
        self.scaled_image           = None   # alias to original_image (no copy)
        self.display_image          = None
        # output_image is NOT stored permanently — effects are applied on-the-fly
        # to the visible crop only, eliminating the largest source of RAM waste.
        self.circles                = []
        self.drawing                = False
        self.center                 = None
        self.current_radius         = 0
        self.current_label          = ""
        self.current_description    = ""
        self.label_input_mode       = False
        self.description_input_mode = False

        # ── viewport (professional affine-transform viewer) ──────────────────────
        self.vp             = Viewport(CANVAS_W, CANVAS_H)
        self.is_panning     = False
        self._pan_last_x    = 0
        self._pan_last_y    = 0
        self.last_zoom_time = 0

        # dirty flag — skip compositing when nothing changed
        self._dirty = True

        # dirty flag — True forces composite+imshow; False skips on idle ticks
        self._dirty = True

        self.current_mode    = EditMode.HIGHLIGHT
        self._blur_kernel    = 25
        self._pixelate_size  = 10
        self.highlight_alpha = 0.4
        self.show_labels     = True

        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR:      (255, 80, 80),
            EditMode.PIXELATE:  (80, 80, 255),
            EditMode.DARKEN:    (160, 160, 160),
            EditMode.GRAYSCALE: (210, 210, 210),
            EditMode.INVERT:    (255, 255, 0),
            EditMode.OUTLINE:   (0, 220, 255),
        }

        self.saved_status       = {}
        self.image_states       = {}
        self.state_access_order = []

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name, CANVAS_W, CANVAS_H)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")
        self._print_instructions()

    # ── Properties ──────────────────────────────────────────────────────────────

    @property
    def blur_kernel(self):
        return self._blur_kernel

    @blur_kernel.setter
    def blur_kernel(self, v):
        if not isinstance(v, int) or v <= 0 or v % 2 == 0:
            raise ValueError("blur_kernel must be a positive odd integer")
        self._blur_kernel = v

    @property
    def pixelate_size(self):
        return self._pixelate_size

    @pixelate_size.setter
    def pixelate_size(self, v):
        if not isinstance(v, int) or v <= 0:
            raise ValueError("pixelate_size must be a positive integer")
        self._pixelate_size = v

    # ── File loading ─────────────────────────────────────────────────────────────

    def _load_image_files(self):
        exts  = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid = []
        for f in files:
            try:
                # Use imdecode on header bytes only — avoids fully loading every image
                raw = np.frombuffer(f.read_bytes()[:64], dtype=np.uint8)
                # Fall back to a proper but minimal check via file size heuristic
                if f.stat().st_size < 512:          # suspiciously tiny → skip
                    print(f"⚠  Skipping {f.name}: file too small")
                    continue
                valid.append(f)
            except Exception as e:
                print(f"⚠  Skipping {f.name}: {e}")

        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images (recommended ≤ {self.MAX_BATCH_SIZE})")
            if input("   Continue? (y/N): ").strip().lower() != 'y':
                sys.exit(0)
        return valid

    # ── Image load & scale ───────────────────────────────────────────────────────

    def _scale_image(self):
        # scaled_image is a no-copy view of original_image.
        # scale_factor stays 1.0 always — Viewport handles all display scaling.
        # Using a view (not .copy()) saves the full image size in RAM.
        self.scale_factor = 1.0
        self.scaled_image = self.original_image   # view, zero extra RAM

    def _load_current_image(self):
        if self.current_index >= len(self.image_files):
            return False
        f = self.image_files[self.current_index]
        try:
            # Release previous buffers before loading new image
            self.original_image = None
            self.scaled_image   = None
            self.display_image  = None
            self.original_image = cv2.imread(str(f))
            if self.original_image is None:
                raise IOError("cv2.imread returned None")
            if min(self.original_image.shape[:2]) < self.MIN_IMAGE_SIZE:
                raise ValueError("Image too small")

            self._scale_image()
            ih, iw = self.original_image.shape[:2]
            self.vp.fit(iw, ih)

            # Restore from in-memory state only (no JSON files written or read)
            if f.name in self.image_states:
                self.circles = self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                      f" — restored {len(self.circles)} objects")
            else:
                self.circles = []
                print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")

            self._update_state_access(f.name)
            self.drawing = self.label_input_mode = self.description_input_mode = False
            self.current_label = self.current_description = ""
            self._dirty = True
            return True

        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index < len(self.image_files) - 1:
                self.current_index += 1
                return self._load_current_image()
            return False

    # ── LRU memory management ───────────────────────────────────────────────────

    def _update_state_access(self, name):
        if name in self.state_access_order:
            self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order) > self.MAX_CACHED_STATES:
            for name in self.state_access_order[:-self.MAX_CACHED_STATES]:
                self.image_states.pop(name, None)
            self.state_access_order = self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── Coordinate transforms (centred-canvas aware) ─────────────────────────────

    def _screen_to_image_coords(self, sx, sy):
        return self.vp.screen_to_img(sx, sy)

    # ── Mouse callback ───────────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        ih, iw = self.original_image.shape[:2]

        # ── zoom ────────────────────────────────────────────────────────────────
        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time() * 1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS:
                return
            self.last_zoom_time = now
            self.vp.zoom_at(x, y, 1.025 if flags > 0 else 0.975, iw, ih)
            self._dirty = True
            self._update_display()
            return

        # ── pan ─────────────────────────────────────────────────────────────────
        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning  = True
            self._pan_last_x = x;  self._pan_last_y = y
            return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False;  return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.vp.pan(x - self._pan_last_x, y - self._pan_last_y)
            self._pan_last_x = x;  self._pan_last_y = y
            self._dirty = True
            self._update_display()
            return

        if self.label_input_mode or self.description_input_mode or self.is_panning:
            return

        # ── draw circle ──────────────────────────────────────────────────────────
        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center  = self.vp.screen_to_img(x, y)
            self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            ix, iy = self.vp.screen_to_img(x, y)
            self.current_radius = int(np.hypot(ix - self.center[0], iy - self.center[1]))
            self._dirty = True
            self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles — performance may degrade")
                self._enter_label_input_mode()

    # ── Label / description input flow ───────────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True
        self.current_label    = ""
        print(f"\n→ [{self.current_mode.value.upper()}] type label, ENTER to confirm, ESC to cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save):
        self.label_input_mode = False
        if save:
            self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled")
            self.current_label = ""
            self.drawing = False          # discard the half-drawn circle too
            self._dirty = True            # force redraw — clears the ghost circle
            self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode = True
        self.current_description    = ""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode = False
        lbl  = self.current_label.strip()
        desc = self.current_description.strip()
        self.circles.append({
            'center':      self.center,
            'radius':      self.current_radius,
            'mode':        self.current_mode,
            'label':       lbl,
            'description': desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}" + (" + description" if desc else "") +
              f" [{self.current_mode.value}]")
        self.current_label = self.current_description = ""
        self._dirty = True
        self._update_display()

    # ── Effects ──────────────────────────────────────────────────────────────────

    def _apply_effect(self, image, circle):
        try:
            mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
            mode = circle['mode']

            if mode == EditMode.HIGHLIGHT:
                lit   = cv2.addWeighted(image, 1 - self.highlight_alpha,
                                        np.full_like(image, 255), self.highlight_alpha, 0)
                image = np.where(mask[:, :, np.newaxis] == 255, lit, image)
            elif mode == EditMode.BLUR:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.GaussianBlur(image, (self._blur_kernel,)*2, 0), image)
            elif mode == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                small = cv2.resize(image,
                                   (max(1, w // self._pixelate_size),
                                    max(1, h // self._pixelate_size)),
                                   interpolation=cv2.INTER_NEAREST)
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST), image)
            elif mode == EditMode.DARKEN:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.addWeighted(image, 0.5, np.zeros_like(image), 0.5, 0), image)
            elif mode == EditMode.GRAYSCALE:
                gray  = cv2.cvtColor(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                image = np.where(mask[:, :, np.newaxis] == 255, gray, image)
            elif mode == EditMode.INVERT:
                image = np.where(mask[:, :, np.newaxis] == 255, cv2.bitwise_not(image), image)
        except Exception as e:
            print(f"⚠  Effect error ({circle['mode'].value}): {e}")
        return image

    def _apply_effects_to_crop(self, crop: np.ndarray,
                               cx0: int, cy0: int) -> np.ndarray:
        """Apply all circle effects to a crop of the original image.
        Coordinates are offset by (cx0, cy0) so effects land in the right place.
        This operates on ≤ CANVAS_W×CANVAS_H pixels — never the full image."""
        if not self.circles:
            return crop
        result = crop  # already a copy from render()
        for c in self.circles:
            adj = {**c, 'center': (c['center'][0] - cx0, c['center'][1] - cy0)}
            result = self._apply_effect(result, adj)
            cv2.circle(result, adj['center'], c['radius'],
                       self.mode_colors[c['mode']], 2, AA)
        return result

    def _build_full_output_for_save(self) -> np.ndarray:
        """Build a full-resolution annotated image — called ONLY at save time."""
        final = self.original_image.copy()
        for c in self.circles:
            final = self._apply_effect(final, c)
            cv2.circle(final, c['center'], c['radius'],
                       self.mode_colors[c['mode']], 2, AA)
        return final

    # ── Smart label drawing with collision avoidance ──────────────────────────────

    def _label_font_params(self, img_shape):
        md = min(img_shape[:2])
        if   md < 200: return 0.30, 1
        elif md < 400: return 0.40, 1
        elif md < 800: return 0.50, 1
        else:          return 0.55, 1

    def _rects_overlap(self, r1, r2, buf=4):
        return not (r1[2]+buf < r2[0] or r1[0]-buf > r2[2] or
                    r1[3]+buf < r2[1] or r1[1]-buf > r2[3])

    def _find_label_pos(self, center, radius, tw, fh, bl, pad, img_shape, placed):
        ih, iw = img_shape[:2]
        W = tw + 2*pad;  H = fh + bl + 2*pad
        cx, cy = center

        candidates = [
            (cx - radius,     cy - radius - H - 10),
            (cx - radius,     cy + radius + 10),
            (cx - radius-W-8, cy - H//2),
            (cx + radius+8,   cy - H//2),
            (cx + radius+8,   cy - radius - H - 10),
            (cx - radius-W-8, cy - radius - H - 10),
            (cx + radius+8,   cy + radius + 10),
            (cx - radius-W-8, cy + radius + 10),
        ]
        for extra in (40, 80, 120):
            candidates += [(px, py+extra) for px, py in candidates[:4]]
            candidates += [(px, py-extra) for px, py in candidates[:4]]

        for px, py in candidates:
            rect = (int(px), int(py), int(px+W), int(py+H))
            if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad:
                continue
            if not any(self._rects_overlap(rect, r) for r in placed):
                return int(px)+pad, int(py)+fh+pad, rect

        # fallback: stack below existing labels
        bottom = max((r[3] for r in placed), default=0) + 8
        lx = max(pad, min(cx - W//2, iw-W-pad))
        ly = min(bottom + fh + pad, ih-bl-pad)
        rect = (lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad)
        return lx, ly, rect

    def _draw_all_labels_smart(self, image):
        """Draw labels onto image.  Uses a cached overlay so pan/zoom never
        re-renders labels — only rebuilds when circles actually change."""
        """Draw labels directly on a full-resolution output image (save path only).
        No BGRA overlay — draws straight to BGR image using the same scratch+ROI
        approach as _render_labels_to_overlay but without the alpha compositing."""
        if not self.show_labels:
            return
        ih, iw = image.shape[:2]
        min_dim = min(ih, iw)
        if   min_dim < 400:  sc, th = 0.45, 1
        elif min_dim < 900:  sc, th = 0.58, 1
        else:                sc, th = 0.72, 2
        placed = []
        for i, c in enumerate(self.circles, 1):
            label = c['label'] or f"Error #{i}"
            color = self.mode_colors[c['mode']]
            cx, cy = c['center']
            r      = c['radius']
            # badge
            badge_text = str(i)
            (bw, bh), _ = cv2.getTextSize(badge_text, FONT, sc * 0.9, th)
            br = max(bw, bh) // 2 + 6
            cv2.circle(image, (cx+2, cy+2), br, (0,0,0), -1, AA)
            cv2.circle(image, (cx, cy),     br, color,   -1, AA)
            cv2.circle(image, (cx, cy),     br, (255,255,255), 2, AA)
            cv2.putText(image, badge_text, (cx-bw//2, cy+bh//2),
                        FONT, sc*0.9, (0,0,0), th+1, AA)
            # label tag
            mode_s = c['mode'].value[:3].upper()
            text   = f"#{i}  [{mode_s}]  {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
            pad = max(5, int(5 * sc / 0.5))
            lx, ly, rect = self._find_label_pos(
                (cx, cy), r, tw, fh, bl, pad, (ih, iw, 3), placed)
            placed.append(rect)
            x1, y1, x2, y2 = rect
            cv2.rectangle(image, (x1, y1), (x2, y2), (15, 15, 20), -1)
            cv2.rectangle(image, (x1, y1), (x1+3, y2), color, -1, AA)
            cv2.rectangle(image, (x1, y1), (x2, y2), color, 2, AA)
            tx2 = lx + 4
            cv2.putText(image, text, (tx2+1, ly+1), FONT, sc, (0,0,0), th+1, AA)
            cv2.putText(image, text, (tx2,   ly),   FONT, sc, (255,255,255), th, AA)
            mid_x  = (x1+x2)//2
            mid_y  = y2 if ly > cy else y1
            angle  = np.arctan2(mid_y-cy, mid_x-cx)
            ex     = int(cx + r*np.cos(angle))
            ey     = int(cy + r*np.sin(angle))
            cv2.line(image, (mid_x, mid_y), (ex, ey), color, 2, AA)
            cv2.circle(image, (ex, ey), 4, color, -1, AA)
            cv2.circle(image, (ex, ey), 4, (255,255,255), 1, AA)

    def _draw_typing_label(self, image, center, radius, label, mode):
        # Legacy path used by save_current for burned-in labels on output image
        sc, th = self._label_font_params(image.shape)
        sc += 0.1;  th += 1
        text = f"[{mode.value[:3].upper()}] {label}_"
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
        pad    = max(4, int(6 * sc / 0.5))
        ih, iw = image.shape[:2]
        color  = self.mode_colors[mode]
        px, py = center[0]-radius, center[1]-radius-fh-2*pad-10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px, py = center[0]-radius, center[1]+radius+10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px = max(pad, min(center[0]-radius, iw-tw-2*pad-pad))
            py = max(fh+pad, center[1]-radius-fh-2*pad-10)
        lx = int(px); ly = int(py+fh+pad)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), (10,10,10), -1)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), color, 2, AA)
        cv2.putText(image, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
        cv2.line(image, (lx+tw//2, ly+bl+pad), center, color, 2, AA)

    def _draw_typing_label_canvas(self, canvas, sx, sy, sr, label, mode):
        """Draw the in-progress typing label directly on the canvas (screen coords)."""
        ch, cw = canvas.shape[:2]
        sc = 0.6; th = 1
        text  = f"[{mode.value[:3].upper()}] {label}_"
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
        pad   = 8
        color = self.mode_colors[mode]
        px = max(pad, sx - sr)
        py = sy - sr - fh - 2*pad - 10
        if py < pad: py = sy + sr + 10
        lx = max(pad, min(int(px), cw-tw-2*pad-pad))
        ly = max(fh+pad, int(py+fh+pad))
        cv2.rectangle(canvas, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), (10,10,10), -1)
        cv2.rectangle(canvas, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), color, 2, AA)
        cv2.putText(canvas, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
        cv2.line(canvas, (lx+tw//2, ly+bl+pad), (sx, sy), color, 2, AA)

    def _draw_labels_on_canvas(self, canvas):
        """Draw all labels at screen coordinates on the 1440×900 canvas.
        No source image copy needed — labels are always crisp at any zoom."""
        if not self.circles:
            return
        ch, cw = canvas.shape[:2]
        sc_base = 0.55; th = 1
        placed = []

        for i, c in enumerate(self.circles, 1):
            label  = c['label'] or f"Error #{i}"
            color  = self.mode_colors[c['mode']]
            sx, sy = self.vp.img_to_screen(*c['center'])
            sr     = max(4, int(c['radius'] * self.vp.scale))

            # ── badge ─────────────────────────────────────────────────────────
            bt = str(i)
            (bw, bh), _ = cv2.getTextSize(bt, FONT, sc_base, th)
            br = max(bw, bh) // 2 + 5
            cv2.circle(canvas, (sx+2, sy+2), br, (0,0,0), -1, AA)
            cv2.circle(canvas, (sx, sy),     br, color,   -1, AA)
            cv2.circle(canvas, (sx, sy),     br, (255,255,255), 1, AA)
            cv2.putText(canvas, bt, (sx-bw//2, sy+bh//2),
                        FONT, sc_base, (0,0,0), th+1, AA)

            # ── label box ─────────────────────────────────────────────────────
            mode_s = c['mode'].value[:3].upper()
            text   = f"#{i} [{mode_s}] {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc_base, th)
            pad = 6

            # Find a non-overlapping position
            W = tw + 2*pad; H = fh + bl + 2*pad
            cx2, cy2 = sx, sy
            candidates = [
                (cx2,       cy2-sr-H-8),
                (cx2,       cy2+sr+8),
                (cx2-W-8,   cy2-H//2),
                (cx2+sr+8,  cy2-H//2),
            ]
            chosen = None
            for px2, py2 in candidates:
                r2 = (int(px2), int(py2), int(px2+W), int(py2+H))
                if r2[0]>=4 and r2[1]>=4 and r2[2]<=cw-4 and r2[3]<=ch-4:
                    if not any(not (r2[2]<p[0] or r2[0]>p[2] or
                                    r2[3]<p[1] or r2[1]>p[3]) for p in placed):
                        chosen = r2; break
            if chosen is None:
                bottom = max((r[3] for r in placed), default=sy+sr) + 6
                px2 = max(4, min(cx2-W//2, cw-W-4))
                py2 = min(bottom, ch-H-4)
                chosen = (px2, py2, px2+W, py2+H)
            placed.append(chosen)

            x1,y1,x2,y2 = chosen
            lx = x1 + pad; ly = y1 + fh + pad

            # Dark fill + accent bar + border
            cv2.rectangle(canvas, (x1,y1), (x2,y2), (18,18,24), -1)
            cv2.rectangle(canvas, (x1,y1), (x1+3,y2), color, -1, AA)
            cv2.rectangle(canvas, (x1,y1), (x2,y2), color, 2, AA)
            # Text with shadow
            cv2.putText(canvas, text, (lx+5, ly+1), FONT, sc_base, (0,0,0), th+1, AA)
            cv2.putText(canvas, text, (lx+4, ly),   FONT, sc_base, (255,255,255), th, AA)

            # Connector to circle edge
            mid_x = (x1+x2)//2; mid_y = y2 if ly > sy else y1
            angle = np.arctan2(mid_y-sy, mid_x-sx)
            ex = int(sx + sr*np.cos(angle))
            ey = int(sy + sr*np.sin(angle))
            cv2.line(canvas, (mid_x,mid_y), (ex,ey), color, 2, AA)
            cv2.circle(canvas, (ex,ey), 4, color, -1, AA)
            cv2.circle(canvas, (ex,ey), 4, (255,255,255), 1, AA)

    # ── Display update ───────────────────────────────────────────────────────────

    def _present(self):
        self._dirty = False
        cv2.imshow(self.window_name, self.display_image)

    def _update_display(self):
        if not self._dirty:
            return
        # Render directly from original — effects applied only to visible crop
        fn = self._apply_effects_to_crop if self.circles else None
        canvas = self.vp.render(self.original_image, post_crop_fn=fn)
        # Live circle preview drawn on canvas (screen coords — no source copy)
        if self.drawing and self.current_radius > 0:
            sx, sy = self.vp.img_to_screen(*self.center)
            sr = max(1, int(self.current_radius * self.vp.scale))
            cv2.circle(canvas, (sx, sy), sr,
                       self.mode_colors[self.current_mode], 2, AA)
        if self.show_labels:
            self._draw_labels_on_canvas(canvas)
        self.display_image = canvas
        self._draw_ui()
        self._present()

    def _update_display_with_input(self):
        fn = self._apply_effects_to_crop if self.circles else None
        canvas = self.vp.render(self.original_image, post_crop_fn=fn)
        sx, sy = self.vp.img_to_screen(*self.center)
        sr = max(1, int(self.current_radius * self.vp.scale))
        cv2.circle(canvas, (sx, sy), sr,
                   self.mode_colors[self.current_mode], 3, AA)
        self._draw_typing_label_canvas(canvas, sx, sy, sr,
                                       self.current_label, self.current_mode)
        self.display_image = canvas
        self._draw_input_box("Label")
        self._present()

    def _update_display_with_description_input(self):
        fn = self._apply_effects_to_crop if self.circles else None
        canvas = self.vp.render(self.original_image, post_crop_fn=fn)
        sx, sy = self.vp.img_to_screen(*self.center)
        sr = max(1, int(self.current_radius * self.vp.scale))
        cv2.circle(canvas, (sx, sy), sr,
                   self.mode_colors[self.current_mode], 3, AA)
        if self.current_label:
            self._draw_typing_label_canvas(canvas, sx, sy, sr,
                                           self.current_label, self.current_mode)
        self.display_image = canvas
        self._draw_input_box("Description (optional)")
        self._present()

    # ── UI overlays ──────────────────────────────────────────────────────────────

    def _draw_input_box(self, prompt):
        h, w  = self.display_image.shape[:2]
        bh = 64; by = h - bh
        color = self.mode_colors[self.current_mode]
        # Only operate on the bottom strip ROI — no full canvas copy
        roi = self.display_image[by:h, 0:w]
        roi[:] = (roi * 0.12 + np.array([12,12,12], dtype=np.float32) * 0.88).astype(np.uint8)
        cv2.rectangle(self.display_image, (0, by), (w, h), color, 2, AA)
        cv2.putText(self.display_image, f"[{self.current_mode.value.upper()}]",
                    (12, by+22), FONT, 0.52, color, 1, AA)
        cv2.putText(self.display_image, f"{prompt}:",
                    (160, by+22), FONT, 0.52, (200,200,200), 1, AA)
        text = (self.current_description if self.description_input_mode
                else self.current_label) + "_"
        cv2.putText(self.display_image, text, (12, by+50), FONT, 0.65, (0,240,240), 1, AA)

    def _draw_ui(self):
        h, w = self.display_image.shape[:2]
        bh   = 84
        # Darken only the top strip ROI — no full canvas copy
        roi = self.display_image[0:bh, 0:w]
        roi[:] = (roi * 0.18 + np.array([12,12,12], dtype=np.float32) * 0.82).astype(np.uint8)

        color = self.mode_colors[self.current_mode]
        cf    = self.image_files[self.current_index]

        cv2.putText(self.display_image, f"Mode: {self.current_mode.value.upper()}",
                    (14,24), FONT, 0.55, color, 1, AA)
        cv2.putText(self.display_image, f"Zoom: {self.vp.scale:.2f}x",
                    (14,48), FONT, 0.45, (100,190,255), 1, AA)
        cv2.putText(self.display_image, f"{self.current_index+1} / {self.total_images}",
                    (14,70), FONT, 0.43, (160,160,160), 1, AA)

        fname = cf.name if len(cf.name) <= 46 else cf.name[:43]+"..."
        cv2.putText(self.display_image, fname, (200,24), FONT, 0.46, (185,185,185), 1, AA)

        n  = len(self.circles)
        oc = (40,160,255) if n >= self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(self.display_image, f"Objects: {n}", (200,48), FONT, 0.43, oc, 1, AA)

        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = cf.name in self.image_states or n > 0
        st, sc = (("SAVED",    (60,200,60))  if is_saved  else
                  ("EDITED",   (40,160,255)) if has_edits else
                  ("NO EDITS", (80,80,80)))
        cv2.putText(self.display_image, st, (w-155,24), FONT, 0.54, sc, 1, AA)

        hint = "Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  R:Reset  C:Clear  U:Undo  H:Help  Q:Quit"
        cv2.putText(self.display_image, hint, (12, h-10), FONT, 0.36, (110,110,110), 1, AA)

    # ── Navigation ───────────────────────────────────────────────────────────────

    def _guard_state(self):
        if self.drawing:
            print("⚠  Finish drawing before switching images"); return True
        if self.label_input_mode:
            print("⚠  Finish label input (ESC to cancel)"); return True
        if self.description_input_mode:
            print("⚠  Finish description input (ENTER to skip)"); return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index > 0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index -= 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index < self.total_images - 1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index += 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at last image")

    # ── Misc commands ────────────────────────────────────────────────────────────

    def _list_labels(self):
        print("\n" + "="*60)
        print(f"Objects — {self.image_files[self.current_index].name}")
        print("="*60)
        if not self.circles:
            print("  (none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  "
                  f"r={c['radius']}  pos={c['center']}")
            if c.get('description'):
                print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        print(f"\nMemory — {len(self.image_states)} states cached "
              f"(max {self.MAX_CACHED_STATES})  |  {len(self.saved_status)} saved\n")

    def _edit_last_label(self):
        if not self.circles:
            print("No objects to edit"); return
        last = self.circles[-1]

        print(f"\nEditing label: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label    = last['label']
        self.label_input_mode = True
        self._update_display_with_input()
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:
                self.label_input_mode = False; self.current_label = ""
                self._dirty = True; self._update_display(); return
            elif key == 13:
                last['label'] = self.current_label.strip()
                self.label_input_mode = False; break
            elif key == 8:
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key); self._update_display_with_input()

        print(f"\nEditing description: '{last.get('description','')}'  (ENTER=save, ESC=skip)")
        self.current_description    = last.get('description', '')
        self.description_input_mode = True
        self._update_display_with_description_input()
        while self.description_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 13:
                last['description'] = self.current_description.strip()
                self.description_input_mode = False; break
            elif key == 27:
                self.description_input_mode = False; break
            elif key == 8:
                self.current_description = self.current_description[:-1]
                self._update_display_with_description_input()
            elif 32 <= key <= 126:
                self.current_description += chr(key); self._update_display_with_description_input()

        self.current_label = self.current_description = ""
        print("  ✓ Updated")
        self._dirty = True; self._update_display()

    # ── Save ─────────────────────────────────────────────────────────────────────

    def save_current(self, auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save")
            return

        cf  = self.image_files[self.current_index]
        out = self.output_folder / cf.name

        # Store in memory (used by Excel summary and session restore)
        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._update_state_access(cf.name)

        try:
            # Build full-res annotated image once, only at save time
            final = self._build_full_output_for_save()
            # Burn labels onto the saved copy (using canvas-draw approach at original scale)
            self._draw_all_labels_smart(final)

            if not cv2.imwrite(str(out), final):
                raise IOError(f"cv2.imwrite failed for {out}")

            self.saved_status[cf.name] = True
            if auto_save:
                print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else:
                print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)\n  → {out}")

            self._cleanup_old_states()

        except Exception as e:
            print(f"❌ Save error: {e}")

    # ── Excel summary ─────────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self.output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"

            # ── Styles ────────────────────────────────────────────────────────
            hfill  = PatternFill(start_color="2E4DA7", end_color="2E4DA7", fill_type="solid")
            hfont  = Font(bold=True, color="FFFFFF", size=11)
            border = Border(**{s: Side(style='thin') for s in ('left','right','top','bottom')})
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_l  = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

            def styled(cell, fill=None, font=None, align=None, bord=True):
                if fill:  cell.fill   = fill
                if font:  cell.font   = font
                if align: cell.alignment = align
                if bord:  cell.border = border

            # ── Headers ───────────────────────────────────────────────────────
            headers = ["#", "Image Name", "Error Count", "Errors (numbered)"]
            for col, hdr in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=hdr)
                styled(c, fill=hfill, font=hfont, align=center)
            ws.row_dimensions[1].height = 28

            # ── Data rows ─────────────────────────────────────────────────────
            data_row = 2
            img_num  = 0
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                state = self.image_states.get(img.name)
                if not state: continue
                objects = state['circles']
                img_num += 1

                # Column 1 — row number
                c1 = ws.cell(row=data_row, column=1, value=img_num)
                styled(c1, align=center)
                c1.font = Font(bold=True, size=10)

                # Column 2 — image filename
                c2 = ws.cell(row=data_row, column=2, value=img.name)
                styled(c2, align=top_l)

                # Column 3 — error count
                c3 = ws.cell(row=data_row, column=3, value=len(objects))
                styled(c3, align=center)
                # colour-code: green = 0, yellow = 1-3, red = 4+
                if len(objects) == 0:
                    c3.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    c3.font = Font(color="276221", bold=True)
                elif len(objects) <= 3:
                    c3.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    c3.font = Font(color="9C5700", bold=True)
                else:
                    c3.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    c3.font = Font(color="9C0006", bold=True)

                # Column 4 — numbered error list in one cell ("1. Label\n2. Label")
                if objects:
                    lines = []
                    for i, o in enumerate(objects, 1):
                        lbl  = (o.get('label') or '').strip() or "(unlabeled)"
                        desc = (o.get('description') or '').strip()
                        line = f"{i}. {lbl}"
                        if desc:
                            line += f" — {desc}"
                        lines.append(line)
                    cell_text = "\n".join(lines)
                else:
                    cell_text = "(no errors)"

                c4 = ws.cell(row=data_row, column=4, value=cell_text)
                styled(c4, align=top_l)

                # Auto row height: ~15 pt per line
                ws.row_dimensions[data_row].height = max(20, len(objects) * 15 + 6)

                # Alternating row fill for readability
                if img_num % 2 == 0:
                    alt = PatternFill(start_color="F2F5FB", end_color="F2F5FB", fill_type="solid")
                    for col in (1, 2, 4):
                        if not ws.cell(row=data_row, column=col).fill.fgColor.rgb.endswith("000000"):
                            pass  # keep colour-coded column 3 as-is
                        ws.cell(row=data_row, column=col).fill = alt

                data_row += 1

            # ── Summary footer ────────────────────────────────────────────────
            data_row += 1
            sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            sfont = Font(bold=True, size=11)

            def summary_row(label, value):
                nonlocal data_row
                lc = ws.cell(row=data_row, column=2, value=label)
                lc.font = sfont; lc.fill = sfill; lc.border = border
                vc = ws.cell(row=data_row, column=3, value=value)
                vc.font = sfont; vc.fill = sfill; vc.border = border
                vc.alignment = center
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            summary_row("Total Images Processed", len(self.saved_status))
            total_obj = sum(
                len(self.image_states[f.name]['circles'])
                for f in self.image_files
                if f.name in self.saved_status and f.name in self.image_states)
            summary_row("Total Errors Found", total_obj)
            if self.saved_status:
                summary_row("Avg Errors / Image",
                            round(total_obj / len(self.saved_status), 1))

            # ── Column widths ─────────────────────────────────────────────────
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 34
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 62

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(str(xp))
            print(f"✓ Excel summary: {xp}")
        except ImportError:
            print("⚠  pip install openpyxl")
        except Exception as e:
            print(f"⚠  Excel error: {e}")

    # ── Instructions ─────────────────────────────────────────────────────────────

    def _print_instructions(self):
        print("\n" + "="*58)
        print("BATCH LABELED EDITOR")
        print("="*58)
        print(f"  Input : {self.input_folder}")
        print(f"  Output: {self.output_folder}")
        print(f"  Images: {self.total_images}")
        print()
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*58+"\n")

    # ── Main loop ─────────────────────────────────────────────────────────────────

    def run(self):
        self._dirty = True
        self._update_display()
        while True:
            key = cv2.waitKey(16) & 0xFF

            # Idle tick — nothing changed, skip redraw
            if key == 0xFF and not self._dirty:
                continue

            if self.description_input_mode:
                if key == 27:                        # ESC → discard whole circle
                    self.description_input_mode = False
                    self.current_label = self.current_description = ""
                    print("  ✗ Circle discarded (ESC during description)")
                    self._dirty = True
                    self._update_display()
                elif key == 13:                      # ENTER → save circle
                    self._exit_description_input_mode()
                elif key == 8:
                    self.current_description = self.current_description[:-1]
                    self._update_display_with_description_input()
                elif 32 <= key <= 126:
                    self.current_description += chr(key)
                    self._update_display_with_description_input()
                continue

            if self.label_input_mode:
                if key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13: self._exit_label_input_mode(save=True)
                elif key == 8:
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key); self._update_display_with_input()
                continue

            if   key == ord('a'): self._previous_image()
            elif key == ord('d'): self._next_image()
            elif key == ord('r'):
                ih, iw = self.original_image.shape[:2]
                self.vp.reset(iw, ih)
                self._dirty = True; self._update_display(); print("✓ Zoom reset")
            elif key == ord('s'): self.save_current(); self._dirty = True; self._update_display()
            elif key == ord('S'): self.save_current(); self._next_image()
            elif key == ord('c'):
                self.circles.clear()
                self._dirty = True; self._update_display(); print("✓ Cleared")
            elif key == ord('u'):
                if self.circles:
                    rem = self.circles.pop()
                    print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._dirty = True; self._update_display()
            elif key == ord('l'): self._list_labels()
            elif key == ord('e'): self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}")
                self._dirty = True; self._update_display()
            elif key == ord('m'): self._show_memory_status()
            elif key in (ord('h'), ord('H')): self._print_instructions()
            elif ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._dirty = True; self._update_display()
            elif key == ord('q'):
                if self.circles and not self.saved_status.get(
                        self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved")
            print(f"   Output: {self.output_folder}")
        else:
            print("\n⚠  No images were saved")


# ── Entry point ───────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Batch Labeled Editor")
    ap.add_argument("input_folder")
    ap.add_argument("--output", "-o", default=None)
    args = ap.parse_args()
    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n⚠  Interrupted"); return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback; traceback.print_exc(); return 1
    return 0


if __name__ == "__main__":

    exit(main())










#!/usr/bin/env python3
"""
Batch Labeled Editor - Production Version (Memory-Optimised)
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag

Memory architecture — stays under ~80 MB regardless of source image size
═══════════════════════════════════════════════════════════════════════════
Problem (old code, 50 MB JPEG → 8000×6000 raw = 137 MB/buffer):
  • original_image (137 MB) kept alive alongside scaled_image (137 MB)
  • np.where / np.full_like in _apply_effect → 3 extra 137 MB temps per circle
  • glow in label drawing: image.copy() × 2 per label × 20 labels = 5.5 GB

Solution (this version):
  1. MAX_WORKING_SIDE = 2048 — longest edge capped for all interactive work.
     8000×6000 → 2048×1536 = 9.4 MB/buffer  (vs 137 MB)
  2. original_image freed from RAM immediately after downscale; reloaded from
     disk only during save, then discarded again.
  3. _apply_effect_inplace: boolean-mask ops — no np.where, no np.full_like,
     no temporary full-image arrays.
  4. Label glow uses ROI crop (~KB each) instead of full image.copy() × 2.
  5. Viewport pre-allocates canvas buffer (no per-frame heap alloc).
  6. Three-level render cache: effects | composite | viewport
     Pan/zoom only triggers the viewport level (< 1 ms, no label work).

Peak memory budget for a 50 MB JPEG at 8000×6000:
  scaled_image 9.4 + effects_cache 9.4 + composite_cache 9.4
  + display 3.9 + viewport buffer 3.9 + blur temp 9.4 ≈ 55 MB  ✓
"""

import gc
import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime
import time
import sys

FONT             = cv2.FONT_HERSHEY_DUPLEX
AA               = cv2.LINE_AA
CANVAS_W         = 1440
CANVAS_H         = 900
CANVAS_BG        = (18, 18, 18)
CANVAS_PAD       = 40
HQ_SETTLE_MS     = 180      # ms idle before HQ Lanczos render fires
MAX_WORKING_SIDE = 2048     # longest edge cap — all annotation at this res


# ── Viewport ──────────────────────────────────────────────────────────────────

class Viewport:
    """Affine (scale, tx, ty) viewer.  Canvas buffer pre-allocated once."""

    ZOOM_STEPS     = [0.125, 0.167, 0.25, 0.333, 0.5, 0.667,
                      1.0, 1.5, 2.0, 3.0, 4.0, 6.0, 8.0, 12.0, 16.0]
    SNAP_THRESHOLD = 0.04

    def __init__(self, cw: int, ch: int):
        self.cw = cw;  self.ch = ch
        self._scale = 1.0;  self._tx = 0.0;  self._ty = 0.0
        # Pre-allocated canvas — reused every render() call (zero GC pressure)
        self._canvas = np.full((ch, cw, 3), CANVAS_BG, dtype=np.uint8)

    def fit(self, iw, ih, pad=CANVAS_PAD):
        self._scale = min((self.cw - 2*pad) / iw, (self.ch - 2*pad) / ih)
        self._snap();  self._recenter(iw, ih)

    def zoom_at(self, sx, sy, factor, iw, ih):
        ix = (sx - self._tx) / self._scale
        iy = (sy - self._ty) / self._scale
        min_s = min(0.05, (self.cw - 2*CANVAS_PAD) / max(iw, 1))
        self._scale = max(min_s, min(16.0, self._scale * factor))
        self._snap()
        self._tx = sx - ix * self._scale
        self._ty = sy - iy * self._scale

    def pan(self, dx, dy):     self._tx += dx;  self._ty += dy
    def reset(self, iw, ih):   self.fit(iw, ih)

    def screen_to_img(self, sx, sy):
        return (int((sx - self._tx) / self._scale),
                int((sy - self._ty) / self._scale))

    @property
    def scale(self): return self._scale

    def render(self, image: np.ndarray, fast: bool = False) -> np.ndarray:
        """
        Crop-then-scale into pre-allocated canvas.
        fast=True  INTER_LINEAR during zoom/pan.
        fast=False Lanczos + unsharp after interaction settles.
        """
        ih, iw = image.shape[:2]
        tx = self._tx;  ty = self._ty;  sc = self._scale

        self._canvas[:] = CANVAS_BG     # refill in-place — no allocation

        sx0 = max(0.0, (-tx) / sc);           sy0 = max(0.0, (-ty) / sc)
        sx1 = min(float(iw), (self.cw-tx)/sc);sy1 = min(float(ih), (self.ch-ty)/sc)
        if sx1 <= sx0 or sy1 <= sy0:
            return self._canvas

        cx0 = max(0,  int(np.floor(sx0)));   cy0 = max(0,  int(np.floor(sy0)))
        cx1 = min(iw, int(np.ceil(sx1)));    cy1 = min(ih, int(np.ceil(sy1)))
        crop = image[cy0:cy1, cx0:cx1]
        if crop.size == 0:
            return self._canvas

        dw = max(1, int(round((cx1 - cx0) * sc)))
        dh = max(1, int(round((cy1 - cy0) * sc)))

        if fast:
            scaled = cv2.resize(crop, (dw, dh), interpolation=cv2.INTER_LINEAR)
        else:
            interp = cv2.INTER_LANCZOS4 if sc >= 1.0 else cv2.INTER_AREA
            scaled = cv2.resize(crop, (dw, dh), interpolation=interp)
            if sc >= 1.0:
                blur   = cv2.GaussianBlur(scaled, (0, 0), sigmaX=1.0)
                scaled = cv2.addWeighted(scaled, 1.4, blur, -0.4, 0)

        dx = max(0, int(tx + cx0 * sc));  dy = max(0, int(ty + cy0 * sc))
        pw = min(dw, self.cw - dx);        ph = min(dh, self.ch - dy)
        if pw > 0 and ph > 0:
            self._canvas[dy:dy+ph, dx:dx+pw] = scaled[:ph, :pw]

        zw = int(iw*sc);  zh = int(ih*sc)
        itx = int(tx);    ity = int(ty)
        cv2.rectangle(self._canvas,
                      (max(0, itx-1),          max(0, ity-1)),
                      (min(self.cw-1, itx+zw), min(self.ch-1, ity+zh)),
                      (45, 45, 45), 1, AA)
        return self._canvas

    def _snap(self):
        for s in self.ZOOM_STEPS:
            if abs(self._scale - s) / s < self.SNAP_THRESHOLD:
                self._scale = s;  return

    def _recenter(self, iw, ih):
        self._tx = (self.cw - iw * self._scale) / 2
        self._ty = (self.ch - ih * self._scale) / 2


# ── Edit modes ────────────────────────────────────────────────────────────────

class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


# ── Editor ────────────────────────────────────────────────────────────────────

class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50
    ZOOM_DEBOUNCE_MS        = 16    # ~60 fps cap on wheel events
    MAX_CACHED_STATES       = 3     # reduced from 5 — each state holds working-res data

    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)
        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.total_images  = len(self.image_files)
        self.current_index = 0

        # Working-resolution image (longest side capped to MAX_WORKING_SIDE).
        # The full-resolution original is NEVER kept in RAM — freed immediately
        # after downscaling and reloaded from disk only during save().
        self.scaled_image  = None   # downscaled working copy (pristine, never modified)
        self.scale_factor  = 1.0    # original_pixels / working_pixels
        self.display_image = None

        self.circles             = []
        self.drawing             = False
        self.center              = None
        self.current_radius      = 0
        self.current_label       = ""
        self.current_description = ""
        self.label_input_mode       = False
        self.description_input_mode = False

        # ── Three-level render cache ─────────────────────────────────────────
        # L1  _effects_cache   working_image + region effects + circle outlines
        # L2  _composite_cache effects_cache + label/badge overlays
        # L3  display_image    composite projected to canvas by viewport
        #
        # Invalidation rules:
        #   circle add/remove/clear  → _effects_dirty  → rebuilds L1 + L2
        #   show_labels toggled      → _composite_dirty → rebuilds L2 only
        #   zoom / pan / draw ghost  → _dirty           → re-renders L3 only
        self._effects_cache   = None
        self._composite_cache = None
        self._effects_dirty   = True
        self._composite_dirty = True
        self._dirty           = True

        # ── Interaction quality ──────────────────────────────────────────────
        self._interacting      = False
        self._last_interact_ms = 0.0

        # ── Viewport ────────────────────────────────────────────────────────
        self.vp          = Viewport(CANVAS_W, CANVAS_H)
        self.is_panning  = False
        self._pan_last_x = self._pan_last_y = 0
        self.last_zoom_time = 0

        self.current_mode    = EditMode.HIGHLIGHT
        self._blur_kernel    = 25
        self._pixelate_size  = 10
        self.highlight_alpha = 0.4
        self.show_labels     = True

        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR:      (255, 80, 80),
            EditMode.PIXELATE:  (80, 80, 255),
            EditMode.DARKEN:    (160, 160, 160),
            EditMode.GRAYSCALE: (210, 210, 210),
            EditMode.INVERT:    (255, 255, 0),
            EditMode.OUTLINE:   (0, 220, 255),
        }

        self.saved_status       = {}
        self.image_states       = {}
        self.state_access_order = []

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name, CANVAS_W, CANVAS_H)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")
        self._print_instructions()

    # ── Properties ──────────────────────────────────────────────────────────

    @property
    def blur_kernel(self):  return self._blur_kernel
    @blur_kernel.setter
    def blur_kernel(self, v):
        if not isinstance(v, int) or v <= 0 or v % 2 == 0:
            raise ValueError("blur_kernel must be a positive odd integer")
        self._blur_kernel = v;  self._invalidate_effects()

    @property
    def pixelate_size(self):  return self._pixelate_size
    @pixelate_size.setter
    def pixelate_size(self, v):
        if not isinstance(v, int) or v <= 0:
            raise ValueError("pixelate_size must be a positive integer")
        self._pixelate_size = v;  self._invalidate_effects()

    # ── Cache invalidation ───────────────────────────────────────────────────

    def _invalidate_effects(self):
        """Circle list changed — must rebuild effects AND composite."""
        self._effects_dirty = self._composite_dirty = self._dirty = True

    def _invalidate_composite(self):
        """Labels toggled — composite only needs rebuild."""
        self._composite_dirty = self._dirty = True

    def _get_composite(self) -> np.ndarray:
        """Return cached composite, rebuilding only stale levels."""
        if self._effects_dirty:
            self._rebuild_effects()       # also sets _composite_dirty = True
        if self._composite_dirty:
            self._rebuild_composite()
        return self._composite_cache

    def _rebuild_effects(self):
        """
        L1 rebuild: apply all region effects into a fresh copy of scaled_image.
        One .copy() per full rebuild, never per frame.
        """
        base = self.scaled_image.copy()
        for c in self.circles:
            self._apply_effect_inplace(base, c)
            cv2.circle(base, c['center'], c['radius'],
                       self.mode_colors[c['mode']], 2, AA)
        self._effects_cache   = base
        self._effects_dirty   = False
        self._composite_dirty = True

    def _rebuild_composite(self):
        """L2 rebuild: stamp labels onto a copy of effects_cache."""
        if self._effects_cache is None:
            self._rebuild_effects()
        tmp = self._effects_cache.copy()
        if self.show_labels:
            self._draw_all_labels_smart(tmp)
        self._composite_cache  = tmp
        self._composite_dirty  = False

    # ── File loading ─────────────────────────────────────────────────────────

    def _load_image_files(self):
        exts  = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid = []
        for f in files:
            try:
                if f.stat().st_size < 512:
                    print(f"⚠  Skipping {f.name}: too small");  continue
                valid.append(f)
            except Exception as e:
                print(f"⚠  Skipping {f.name}: {e}")
        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images "
                  f"(recommended ≤ {self.MAX_BATCH_SIZE})")
            if input("   Continue? (y/N): ").strip().lower() != 'y':
                sys.exit(0)
        return valid

    # ── Image load ───────────────────────────────────────────────────────────

    def _make_working_image(self, original: np.ndarray):
        """
        Downscale to fit longest edge within MAX_WORKING_SIDE.
        Returns (working_image, scale_factor).
        scale_factor > 1.0 means we downscaled; coords × scale_factor = original px.
        """
        ih, iw = original.shape[:2]
        longest = max(ih, iw)
        if longest <= MAX_WORKING_SIDE:
            return original.copy(), 1.0
        sf    = longest / MAX_WORKING_SIDE
        new_w = max(1, int(round(iw / sf)))
        new_h = max(1, int(round(ih / sf)))
        working = cv2.resize(original, (new_w, new_h), interpolation=cv2.INTER_AREA)
        return working, sf

    def _load_current_image(self):
        if self.current_index >= len(self.image_files):
            return False
        f = self.image_files[self.current_index]
        try:
            # Load full-res only long enough to create working copy
            raw = cv2.imread(str(f))
            if raw is None:
                raise IOError("cv2.imread returned None")
            if min(raw.shape[:2]) < self.MIN_IMAGE_SIZE:
                raise ValueError("Image too small")

            self.scaled_image, self.scale_factor = self._make_working_image(raw)

            # Free full-res immediately — never needed during interactive work
            del raw;  gc.collect()

            iw, ih = self.scaled_image.shape[1], self.scaled_image.shape[0]
            self.vp.fit(iw, ih)

            # Restore saved annotations (coords stored in original-pixel space)
            json_path = self.output_folder / f.with_suffix('.json').name
            if json_path.exists():
                try:
                    with open(json_path) as jf:
                        data = json.load(jf)
                    sf = self.scale_factor
                    self.circles = [{
                        'center':      (int(o['center'][0] / sf),
                                        int(o['center'][1] / sf)),
                        'radius':      max(1, int(o['radius']  / sf)),
                        'mode':        EditMode(o.get('mode', 'highlight')),
                        'label':       o.get('label', ''),
                        'description': o.get('description', ''),
                    } for o in data.get('objects', [])]
                    print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                          f" — restored {len(self.circles)} objects")
                except Exception as e:
                    print(f"⚠  JSON load error: {e}");  self.circles = []
            elif f.name in self.image_states:
                self.circles = self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                      f" — from memory")
            else:
                self.circles = []
                print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")

            self._update_state_access(f.name)
            self.drawing = self.label_input_mode = self.description_input_mode = False
            self.current_label = self.current_description = ""

            # Flush all cache levels for the new image
            self._effects_cache = self._composite_cache = None
            self._invalidate_effects()
            return True

        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index < len(self.image_files) - 1:
                self.current_index += 1
                return self._load_current_image()
            return False

    # ── LRU state cache ──────────────────────────────────────────────────────

    def _update_state_access(self, name):
        if name in self.state_access_order:
            self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order) > self.MAX_CACHED_STATES:
            for name in self.state_access_order[:-self.MAX_CACHED_STATES]:
                self.image_states.pop(name, None)
            self.state_access_order = self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── Mouse callback ───────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        ih, iw = self.scaled_image.shape[:2]

        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time() * 1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS:
                return
            self.last_zoom_time = now
            self.vp.zoom_at(x, y, 1.2 if flags > 0 else 0.8, iw, ih)
            self._mark_interacting();  self._dirty = True;  self._update_display()
            return

        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning = True
            self._pan_last_x = x;  self._pan_last_y = y;  return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False;  self._dirty = True;  return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.vp.pan(x - self._pan_last_x, y - self._pan_last_y)
            self._pan_last_x = x;  self._pan_last_y = y
            self._mark_interacting();  self._dirty = True;  self._update_display()
            return

        if self.label_input_mode or self.description_input_mode or self.is_panning:
            return

        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center  = self.vp.screen_to_img(x, y)
            self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            ix, iy = self.vp.screen_to_img(x, y)
            self.current_radius = int(np.hypot(ix - self.center[0],
                                               iy - self.center[1]))
            self._mark_interacting();  self._dirty = True;  self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles — may degrade performance")
                self._enter_label_input_mode()

    def _mark_interacting(self):
        self._interacting = True;  self._last_interact_ms = time.time() * 1000

    def _check_settle(self):
        if not self._interacting:
            return False
        if (time.time() * 1000 - self._last_interact_ms) >= HQ_SETTLE_MS:
            self._interacting = False;  return True
        return False

    # ── Label / description input flow ───────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True;  self.current_label = ""
        print(f"\n→ [{self.current_mode.value.upper()}] "
              f"type label, ENTER confirm, ESC cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save):
        self.label_input_mode = False
        if save:
            self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled")
            self.current_label = "";  self.drawing = False
            self._dirty = True;  self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode = True;  self.current_description = ""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode = False
        lbl  = self.current_label.strip()
        desc = self.current_description.strip()
        self.circles.append({
            'center':      self.center,
            'radius':      self.current_radius,
            'mode':        self.current_mode,
            'label':       lbl,
            'description': desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}" + (" + description" if desc else "") +
              f" [{self.current_mode.value}]")
        self.current_label = self.current_description = ""
        self._invalidate_effects()
        self._update_display()

    # ── Effects — in-place, no large temporaries ─────────────────────────────

    def _apply_effect_inplace(self, image: np.ndarray, circle: dict) -> None:
        """
        Apply one region effect directly into image using boolean-mask indexing.

        Why this is much cheaper than the old np.where approach:
          np.where(cond, A, B) → allocates a full output array (= image size)
          image[mb] = A[mb]   → scatter-write; only masked pixels touched

        Allocations per effect:
          All modes:  mask (H×W uint8, ~3 MB for 2048×1536) + mb (same)
          BLUR:       one full-image Gaussian temp (9.4 MB) — unavoidable
          PIXELATE:   one tiny downscale + one full-size nearest-neighbour
          GRAYSCALE:  one full-image grey temp
          HIGHLIGHT:  float32 slice of masked pixels only (~KB)
          DARKEN:     uint16 slice of masked pixels only (~KB)
          INVERT:     no extra alloc (255 - image[mb] in-place)
        """
        mask = np.zeros(image.shape[:2], dtype=np.uint8)
        cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
        mb = mask.astype(bool)
        if not mb.any():
            return

        m = circle['mode']
        try:
            if m == EditMode.HIGHLIGHT:
                a   = self.highlight_alpha
                pix = image[mb].astype(np.float32)
                image[mb] = np.clip(pix * (1.0 - a) + 255.0 * a,
                                    0, 255).astype(np.uint8)

            elif m == EditMode.BLUR:
                blurred   = cv2.GaussianBlur(image, (self._blur_kernel,)*2, 0)
                image[mb] = blurred[mb]

            elif m == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                small = cv2.resize(image,
                                   (max(1, w // self._pixelate_size),
                                    max(1, h // self._pixelate_size)),
                                   interpolation=cv2.INTER_NEAREST)
                pixelated = cv2.resize(small, (w, h),
                                       interpolation=cv2.INTER_NEAREST)
                image[mb] = pixelated[mb]

            elif m == EditMode.DARKEN:
                # right-shift by 1 = halve brightness; no float conversion
                image[mb] = (image[mb].astype(np.uint16) >> 1).astype(np.uint8)

            elif m == EditMode.GRAYSCALE:
                gray_bgr  = cv2.cvtColor(
                    cv2.cvtColor(image, cv2.COLOR_BGR2GRAY),
                    cv2.COLOR_GRAY2BGR)
                image[mb] = gray_bgr[mb]

            elif m == EditMode.INVERT:
                image[mb] = 255 - image[mb]   # fully in-place, zero extra alloc

        except Exception as e:
            print(f"⚠  Effect error ({circle['mode'].value}): {e}")

    def _apply_all_effects(self):
        """Public interface — invalidates cache (effects rebuilt lazily)."""
        self._invalidate_effects()

    # ── Smart label drawing (runs once into composite cache) ──────────────────

    def _rects_overlap(self, r1, r2, buf=4):
        return not (r1[2]+buf < r2[0] or r1[0]-buf > r2[2] or
                    r1[3]+buf < r2[1] or r1[1]-buf > r2[3])

    def _find_label_pos(self, center, radius, tw, fh, bl, pad, img_shape, placed):
        ih, iw = img_shape[:2]
        W = tw + 2*pad;  H = fh + bl + 2*pad
        cx, cy = center
        cands = [
            (cx-radius,     cy-radius-H-10), (cx-radius,   cy+radius+10),
            (cx-radius-W-8, cy-H//2),        (cx+radius+8, cy-H//2),
            (cx+radius+8,   cy-radius-H-10), (cx-radius-W-8, cy-radius-H-10),
            (cx+radius+8,   cy+radius+10),   (cx-radius-W-8, cy+radius+10),
        ]
        for extra in (40, 80, 120):
            cands += [(px, py+extra) for px, py in cands[:4]]
            cands += [(px, py-extra) for px, py in cands[:4]]
        for px, py in cands:
            rect = (int(px), int(py), int(px+W), int(py+H))
            if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad:
                continue
            if not any(self._rects_overlap(rect, r) for r in placed):
                return int(px)+pad, int(py)+fh+pad, rect
        bottom = max((r[3] for r in placed), default=0) + 8
        lx = max(pad, min(cx - W//2, iw-W-pad))
        ly = min(bottom + fh + pad, ih-bl-pad)
        return lx, ly, (lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad)

    def _draw_all_labels_smart(self, image: np.ndarray) -> None:
        """
        Draw all label tags in-place.
        Called only from _rebuild_composite — NEVER on every frame.

        Glow fix: old code did image.copy() × 2 per label (full-image allocation).
        New code: ROI crop of the label bounding box (~KB each).
        For 20 labels: was 20×2×9.4 MB = 376 MB; now 20×2×~0.02 MB = 0.8 MB.
        """
        if not self.show_labels:
            return
        placed = []
        ih, iw = image.shape[:2]
        md = min(ih, iw)
        sc, th = (0.45, 1) if md < 400 else (0.58, 1) if md < 900 else (0.72, 2)

        for i, c in enumerate(self.circles, 1):
            label = c['label'] or f"Error #{i}"
            color = self.mode_colors[c['mode']]
            cx, cy = c['center']
            r      = c['radius']

            # ── badge number on circle ────────────────────────────────────
            bt = str(i)
            (bw, bh), _ = cv2.getTextSize(bt, FONT, sc * 0.9, th)
            br = max(bw, bh) // 2 + 6
            cv2.circle(image, (cx+2, cy+2), br, (0,0,0), -1, AA)
            cv2.circle(image, (cx, cy),     br, color,   -1, AA)
            cv2.circle(image, (cx, cy),     br, (255,255,255), 2, AA)
            cv2.putText(image, bt, (cx - bw//2, cy + bh//2),
                        FONT, sc * 0.9, (0,0,0), th+1, AA)

            # ── label tag position ────────────────────────────────────────
            text = f"#{i}  [{c['mode'].value[:3].upper()}]  {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
            pad = max(5, int(5 * sc / 0.5))

            lx, ly, rect = self._find_label_pos(
                (cx, cy), r, tw, fh, bl, pad, image.shape, placed)
            placed.append(rect)
            x1, y1, x2, y2 = rect

            # ── ROI-based glow ────────────────────────────────────────────
            # OLD: glow = image.copy()  →  full 9.4 MB alloc × 2 per label
            # NEW: roi  = image[tiny_slice].copy()  →  ~KB per label
            for ge, ga in ((6, 0.25), (3, 0.45)):
                gx1 = max(0, x1 - ge);    gy1 = max(0, y1 - ge)
                gx2 = min(iw, x2 + ge);   gy2 = min(ih, y2 + ge)
                if gx2 <= gx1 or gy2 <= gy1:
                    continue
                roi      = image[gy1:gy2, gx1:gx2]
                roi_copy = roi.copy()               # tiny allocation
                cv2.rectangle(roi_copy,
                              (x1 - gx1, y1 - gy1),
                              (x2 - gx1, y2 - gy1),
                              color, -1, AA)
                cv2.addWeighted(roi_copy, ga, roi, 1.0 - ga, 0, roi)  # in-place

            cv2.rectangle(image, (x1, y1), (x2, y2), (15, 15, 20), -1)
            cv2.rectangle(image, (x1, y1), (x1+3, y2), color, -1, AA)
            cv2.rectangle(image, (x1, y1), (x2, y2), color, 2, AA)

            tx = lx + 4
            cv2.putText(image, text, (tx+1, ly+1), FONT, sc, (0,0,0), th+1, AA)
            cv2.putText(image, text, (tx,   ly),   FONT, sc, (255,255,255), th, AA)

            mx  = (x1 + x2) // 2
            my  = y2 if ly > cy else y1
            ang = np.arctan2(my - cy, mx - cx)
            ex  = int(cx + r * np.cos(ang))
            ey  = int(cy + r * np.sin(ang))
            cv2.line(image, (mx, my), (ex, ey), color, 2, AA)
            cv2.circle(image, (ex, ey), 4, color, -1, AA)
            cv2.circle(image, (ex, ey), 4, (255,255,255), 1, AA)

    def _draw_typing_label(self, image, center, radius, label, mode):
        ih, iw = image.shape[:2]
        md = min(ih, iw)
        sc = 0.65 if md >= 800 else (0.50 if md >= 400 else 0.40)
        th = 2
        text  = f"[{mode.value[:3].upper()}] {label}_"
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
        pad   = max(4, int(6 * sc / 0.5))
        color = self.mode_colors[mode]

        px, py = center[0] - radius, center[1] - radius - fh - 2*pad - 10
        if not (px>=pad and py>=fh+pad and
                px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px, py = center[0] - radius, center[1] + radius + 10
        if not (px>=pad and py>=fh+pad and
                px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px = max(pad, min(center[0]-radius, iw-tw-2*pad-pad))
            py = max(fh+pad, center[1]-radius-fh-2*pad-10)

        lx, ly = int(px), int(py + fh + pad)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), (10,10,10), -1)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), color, 2, AA)
        cv2.putText(image, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
        cv2.line(image, (lx+tw//2, ly+bl+pad), center, color, 2, AA)

    # ── Display update ───────────────────────────────────────────────────────

    def _present(self):
        self._dirty = False
        cv2.imshow(self.window_name, self.display_image)

    def _update_display(self):
        """
        Only the minimum cache level is rebuilt per interaction type:
          Pan/zoom         → vp.render() on existing composite  (< 1 ms)
          Drawing ghost    → composite.copy() + one cv2.circle  (< 1 ms)
          Circle add/del   → rebuild L1 + L2 then vp.render()
          Label toggle     → rebuild L2 only then vp.render()
        """
        if not self._dirty:
            return
        if self._check_settle():
            self._dirty = True

        base = self._get_composite()

        if self.drawing and self.current_radius > 0:
            tmp = base.copy()
            cv2.circle(tmp, self.center, self.current_radius,
                       self.mode_colors[self.current_mode], 2, AA)
            rendered = self.vp.render(tmp, fast=self._interacting)
        else:
            rendered = self.vp.render(base, fast=self._interacting)

        # vp.render returns its internal buffer — must copy before UI overlay
        self.display_image = rendered.copy()
        self._draw_ui()
        self._present()

    def _update_display_with_input(self):
        base = self._get_composite()
        tmp  = base.copy()
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        self._draw_typing_label(tmp, self.center, self.current_radius,
                                self.current_label, self.current_mode)
        self.display_image = self.vp.render(tmp, fast=False).copy()
        self._draw_input_box("Label")
        self._present()

    def _update_display_with_description_input(self):
        base = self._get_composite()
        tmp  = base.copy()
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        if self.current_label:
            self._draw_typing_label(tmp, self.center, self.current_radius,
                                    self.current_label, self.current_mode)
        self.display_image = self.vp.render(tmp, fast=False).copy()
        self._draw_input_box("Description (optional)")
        self._present()

    # ── UI overlays ──────────────────────────────────────────────────────────

    def _draw_input_box(self, prompt):
        h, w = self.display_image.shape[:2]
        bh = 64;  by = h - bh
        color = self.mode_colors[self.current_mode]
        ov = self.display_image.copy()
        cv2.rectangle(ov, (0, by), (w, h), (12,12,12), -1)
        cv2.addWeighted(ov, 0.88, self.display_image, 0.12, 0, self.display_image)
        cv2.rectangle(self.display_image, (0, by), (w, h), color, 2, AA)
        cv2.putText(self.display_image, f"[{self.current_mode.value.upper()}]",
                    (12, by+22), FONT, 0.52, color, 1, AA)
        cv2.putText(self.display_image, f"{prompt}:",
                    (160, by+22), FONT, 0.52, (200,200,200), 1, AA)
        text = (self.current_description if self.description_input_mode
                else self.current_label) + "_"
        cv2.putText(self.display_image, text,
                    (12, by+50), FONT, 0.65, (0,240,240), 1, AA)

    def _draw_ui(self):
        ov = self.display_image.copy()
        h, w = self.display_image.shape[:2]
        cv2.rectangle(ov, (0, 0), (w, 84), (12,12,12), -1)
        color = self.mode_colors[self.current_mode]
        cf    = self.image_files[self.current_index]

        cv2.putText(ov, f"Mode: {self.current_mode.value.upper()}",
                    (14,24), FONT, 0.55, color, 1, AA)
        cv2.putText(ov, f"Zoom: {self.vp.scale:.2f}x",
                    (14,48), FONT, 0.45, (100,190,255), 1, AA)
        cv2.putText(ov, f"{self.current_index+1} / {self.total_images}",
                    (14,70), FONT, 0.43, (160,160,160), 1, AA)

        fname = cf.name if len(cf.name) <= 46 else cf.name[:43]+"..."
        cv2.putText(ov, fname, (200,24), FONT, 0.46, (185,185,185), 1, AA)

        n  = len(self.circles)
        oc = (40,160,255) if n >= self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(ov, f"Objects: {n}", (200,48), FONT, 0.43, oc, 1, AA)

        if self.scale_factor > 1.0:
            cv2.putText(ov, f"Working 1/{self.scale_factor:.1f}x",
                        (200,70), FONT, 0.36, (80,200,80), 1, AA)

        qi = "FAST" if self._interacting else "HQ"
        qc = (60,180,255) if self._interacting else (60,220,100)
        cv2.putText(ov, qi, (390,70), FONT, 0.36, qc, 1, AA)

        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = cf.name in self.image_states or n > 0
        st, sc = (("SAVED",    (60,200,60))  if is_saved  else
                  ("EDITED",   (40,160,255)) if has_edits else
                  ("NO EDITS", (80,80,80)))
        cv2.putText(ov, st, (w-155,24), FONT, 0.54, sc, 1, AA)

        hint = ("Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  R:Reset  "
                "C:Clear  U:Undo  H:Help  Q:Quit")
        cv2.putText(ov, hint, (12, h-10), FONT, 0.36, (110,110,110), 1, AA)
        cv2.addWeighted(ov, 0.82, self.display_image, 0.18, 0, self.display_image)

    # ── Navigation ───────────────────────────────────────────────────────────

    def _guard_state(self):
        if self.drawing:
            print("⚠  Finish drawing before switching images");  return True
        if self.label_input_mode:
            print("⚠  Finish label input (ESC to cancel)");      return True
        if self.description_input_mode:
            print("⚠  Finish description (ENTER to skip)");      return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index > 0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index -= 1
            self._load_current_image();  self._update_display()
        else:
            print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index < self.total_images - 1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index += 1
            self._load_current_image();  self._update_display()
        else:
            print("Already at last image")

    # ── Misc ─────────────────────────────────────────────────────────────────

    def _list_labels(self):
        print("\n" + "="*60)
        print(f"Objects — {self.image_files[self.current_index].name}")
        print("="*60)
        if not self.circles:
            print("  (none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  "
                  f"r={c['radius']}  pos={c['center']}")
            if c.get('description'):
                print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        bufs = [self.scaled_image, self._effects_cache,
                self._composite_cache, self.display_image, self.vp._canvas]
        total = sum(a.nbytes for a in bufs if a is not None)
        print(f"\nMemory — working buffers: {total//1024//1024} MB  |  "
              f"scale: 1/{self.scale_factor:.2f}x  |  "
              f"states cached: {len(self.image_states)}/{self.MAX_CACHED_STATES}  |  "
              f"saved: {len(self.saved_status)}\n")

    def _edit_last_label(self):
        if not self.circles:
            print("No objects to edit");  return
        last = self.circles[-1]

        print(f"\nEditing label: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label    = last['label']
        self.label_input_mode = True
        self._update_display_with_input()
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:
                self.label_input_mode = False;  self.current_label = ""
                self._dirty = True;  self._update_display();  return
            elif key == 13:
                last['label'] = self.current_label.strip()
                self.label_input_mode = False;  break
            elif key == 8:
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key)
                self._update_display_with_input()

        print(f"\nEditing description: '{last.get('description','')}'  "
              f"(ENTER=save, ESC=skip)")
        self.current_description    = last.get('description', '')
        self.description_input_mode = True
        self._update_display_with_description_input()
        while self.description_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 13:
                last['description'] = self.current_description.strip()
                self.description_input_mode = False;  break
            elif key == 27:
                self.description_input_mode = False;  break
            elif key == 8:
                self.current_description = self.current_description[:-1]
                self._update_display_with_description_input()
            elif 32 <= key <= 126:
                self.current_description += chr(key)
                self._update_display_with_description_input()

        self.current_label = self.current_description = ""
        print("  ✓ Updated")
        self._invalidate_composite()    # label text changed — composite only
        self._update_display()

    # ── Save ──────────────────────────────────────────────────────────────────

    def save_current(self, auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save")
            return

        cf  = self.image_files[self.current_index]
        out = self.output_folder / cf.name
        js  = out.with_suffix('.json')

        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._update_state_access(cf.name)

        try:
            # Reload full-res from disk — only lives here during save, then freed
            original = cv2.imread(str(cf))
            if original is None:
                raise IOError(f"Could not re-read {cf.name} for save")

            sf = self.scale_factor

            # Scale working coords → original-pixel space
            orig_circles = [{
                'center':      (int(c['center'][0] * sf), int(c['center'][1] * sf)),
                'radius':      max(1, int(c['radius']  * sf)),
                'mode':        c['mode'],
                'label':       c['label'],
                'description': c.get('description', ''),
            } for c in self.circles]

            # Apply effects at full resolution
            for oc in orig_circles:
                self._apply_effect_inplace(original, oc)
                cv2.circle(original, oc['center'], oc['radius'],
                           self.mode_colors[oc['mode']], 3, AA)

            # Draw labels at full resolution (temporarily swap circle list)
            saved_circles = self.circles
            self.circles  = orig_circles
            self._draw_all_labels_smart(original)
            self.circles  = saved_circles

            if not cv2.imwrite(str(out), original):
                raise IOError(f"cv2.imwrite failed for {out}")

            # Free full-res immediately
            del original;  gc.collect()

            # JSON stores original-pixel-space coords
            data = {
                'source_image': cf.name,
                'timestamp':    datetime.now().isoformat(),
                'objects': [{
                    'id':          i,
                    'label':       oc['label'],
                    'description': oc.get('description', ''),
                    'mode':        oc['mode'].value,
                    'center':      list(oc['center']),
                    'radius':      oc['radius'],
                } for i, oc in enumerate(orig_circles, 1)]
            }
            with open(js, 'w') as jf:
                json.dump(data, jf, indent=2)

            self.saved_status[cf.name] = True
            if auto_save:
                print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else:
                print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)"
                      f"\n  → {out}\n  → {js}")

            self._cleanup_old_states()

        except Exception as e:
            print(f"❌ Save error: {e}")

    # ── Excel summary ─────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self.output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook();  ws = wb.active;  ws.title = "Summary"

            hfill  = PatternFill(start_color="2E4DA7", end_color="2E4DA7",
                                 fill_type="solid")
            hfont  = Font(bold=True, color="FFFFFF", size=11)
            border = Border(**{s: Side(style='thin')
                             for s in ('left','right','top','bottom')})
            ca  = Alignment(horizontal='center', vertical='center', wrap_text=True)
            tla = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

            def styled(cell, fill=None, font=None, align=None):
                if fill:  cell.fill   = fill
                if font:  cell.font   = font
                if align: cell.alignment = align
                cell.border = border

            for col, hdr in enumerate(
                    ["#", "Image Name", "Error Count", "Errors (numbered)"], 1):
                styled(ws.cell(row=1, column=col, value=hdr),
                       fill=hfill, font=hfont, align=ca)
            ws.row_dimensions[1].height = 28

            data_row = 2;  img_num = 0
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                jp = self.output_folder / img.with_suffix('.json').name
                if not jp.exists(): continue
                with open(jp) as jf: data = json.load(jf)
                objects  = data.get('objects', [])
                img_num += 1

                c1 = ws.cell(row=data_row, column=1, value=img_num)
                styled(c1, align=ca);  c1.font = Font(bold=True, size=10)
                styled(ws.cell(row=data_row, column=2, value=img.name), align=tla)

                c3 = ws.cell(row=data_row, column=3, value=len(objects))
                styled(c3, align=ca)
                if   len(objects) == 0:
                    c3.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                         fill_type="solid")
                    c3.font = Font(color="276221", bold=True)
                elif len(objects) <= 3:
                    c3.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                                         fill_type="solid")
                    c3.font = Font(color="9C5700", bold=True)
                else:
                    c3.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                         fill_type="solid")
                    c3.font = Font(color="9C0006", bold=True)

                cell_text = "\n".join(
                    f"{k}. {(o.get('label','').strip() or '(unlabeled)')}"
                    + (f" — {o['description'].strip()}" if o.get('description','').strip() else "")
                    for k, o in enumerate(objects, 1)
                ) if objects else "(no errors)"

                styled(ws.cell(row=data_row, column=4, value=cell_text), align=tla)
                ws.row_dimensions[data_row].height = max(20, len(objects)*15+6)

                if img_num % 2 == 0:
                    alt = PatternFill(start_color="F2F5FB", end_color="F2F5FB",
                                     fill_type="solid")
                    for col in (1, 2, 4):
                        ws.cell(row=data_row, column=col).fill = alt
                data_row += 1

            data_row += 1
            sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                fill_type="solid")
            sfont = Font(bold=True, size=11)

            def summary_row(label, value):
                nonlocal data_row
                lc = ws.cell(row=data_row, column=2, value=label)
                lc.font = sfont;  lc.fill = sfill;  lc.border = border
                vc = ws.cell(row=data_row, column=3, value=value)
                vc.font = sfont;  vc.fill = sfill;  vc.border = border
                vc.alignment = ca
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            summary_row("Total Images Processed", len(self.saved_status))
            total_obj = sum(
                len(json.load(open(
                    self.output_folder / f.with_suffix('.json').name))['objects'])
                for f in self.image_files
                if f.name in self.saved_status
                and (self.output_folder / f.with_suffix('.json').name).exists())
            summary_row("Total Errors Found", total_obj)
            if self.saved_status:
                summary_row("Avg Errors / Image",
                            round(total_obj / len(self.saved_status), 1))

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 34
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 62
            ws.freeze_panes = "A2"
            wb.save(str(xp))
            print(f"✓ Excel summary: {xp}")
        except ImportError:
            print("⚠  pip install openpyxl")
        except Exception as e:
            print(f"⚠  Excel error: {e}")

    # ── Instructions ─────────────────────────────────────────────────────────

    def _print_instructions(self):
        iw, ih = self.scaled_image.shape[1], self.scaled_image.shape[0]
        sf_note = (f" (working: {iw}×{ih}, "
                   f"scale 1/{self.scale_factor:.2f}x)") if self.scale_factor > 1.0 else ""
        print("\n" + "="*62)
        print("BATCH LABELED EDITOR")
        print("="*62)
        print(f"  Input : {self.input_folder}")
        print(f"  Output: {self.output_folder}")
        print(f"  Images: {self.total_images}{sf_note}")
        print()
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next image")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo last circle")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*62+"\n")

    # ── Main loop ─────────────────────────────────────────────────────────────

    def run(self):
        self._dirty = True
        self._update_display()
        while True:
            key = cv2.waitKey(1) & 0xFF

            # HQ settle tick
            if key == 0xFF:
                if self._check_settle():
                    self._dirty = True;  self._update_display()
                elif not self._dirty:
                    continue

            if self.description_input_mode:
                if key == 27:
                    self.description_input_mode = False
                    self.current_label = self.current_description = ""
                    print("  ✗ Circle discarded (ESC during description)")
                    self._dirty = True;  self._update_display()
                elif key == 13:
                    self._exit_description_input_mode()
                elif key == 8:
                    self.current_description = self.current_description[:-1]
                    self._update_display_with_description_input()
                elif 32 <= key <= 126:
                    self.current_description += chr(key)
                    self._update_display_with_description_input()
                continue

            if self.label_input_mode:
                if key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13: self._exit_label_input_mode(save=True)
                elif key == 8:
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key)
                    self._update_display_with_input()
                continue

            if   key == ord('a'): self._previous_image()
            elif key == ord('d'): self._next_image()
            elif key == ord('r'):
                iw, ih = self.scaled_image.shape[1], self.scaled_image.shape[0]
                self.vp.reset(iw, ih)
                self._dirty = True;  self._update_display();  print("✓ Zoom reset")
            elif key == ord('s'):
                self.save_current();  self._dirty = True;  self._update_display()
            elif key == ord('S'):
                self.save_current();  self._next_image()
            elif key == ord('c'):
                self.circles.clear()
                self._invalidate_effects()
                self._update_display();  print("✓ Cleared")
            elif key == ord('u'):
                if self.circles:
                    rem = self.circles.pop()
                    print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._invalidate_effects();  self._update_display()
            elif key == ord('l'): self._list_labels()
            elif key == ord('e'): self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}")
                self._invalidate_composite();  self._update_display()
            elif key == ord('m'): self._show_memory_status()
            elif key in (ord('h'), ord('H')): self._print_instructions()
            elif ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._dirty = True;  self._update_display()
            elif key == ord('q'):
                if self.circles and not self.saved_status.get(
                        self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved")
            print(f"   Output: {self.output_folder}")
        else:
            print("\n⚠  No images were saved")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Batch Labeled Editor")
    ap.add_argument("input_folder")
    ap.add_argument("--output", "-o", default=None)
    ap.add_argument("--max-side", type=int, default=MAX_WORKING_SIDE,
                    help=f"Working resolution cap — longest side in pixels "
                         f"(default {MAX_WORKING_SIDE}). "
                         f"Lower = less RAM; higher = more detail during annotation.")
    args = ap.parse_args()
    global MAX_WORKING_SIDE
    MAX_WORKING_SIDE = args.max_side

    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n⚠  Interrupted");  return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback;  traceback.print_exc();  return 1
    return 0


if __name__ == "__main__":
    exit(main())
    


