#!/usr/bin/env python3
"""
Batch Labeled Editor — Qt Edition
PySide6 + OpenCV  ·  Full-Res Save  ·  Low-Memory Display  ·  Background Loading

Navigate: A / D    Zoom: Mouse Wheel    Pan: Right-click drag    Draw: Left-click drag

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ARCHITECTURE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ImageLoader (QThread)
    └─ Loads full-res from disk, downscales to MAX_DISPLAY_PX,
       frees full-res, emits (display_img_numpy, scale, w, h).
       UI never blocks during disk I/O.

  ImageViewer (QGraphicsView)
    └─ QGraphicsScene with ONE QGraphicsPixmapItem (the image).
    └─ GhostCircleItem overlay drawn during left-drag (circle preview).
    └─ Zoom: mouse-wheel → scale transform anchored at cursor.
    └─ Pan:  right-drag  → translate transform.
    └─ All transforms are GPU-accelerated Qt operations — zero numpy work.

  MainWindow (QMainWindow)
    └─ Owns the annotation state (circles in ORIGINAL pixel coords).
    └─ Render cache: two numpy levels (effects_img) → one QPixmap
       (composite_pixmap).  Rebuilt only when annotations change.
       Pan / zoom never re-render: only the Qt viewport transform updates.
    └─ Save: reloads full-res from disk, applies effects at orig coords,
       writes, frees immediately.  Peak save RAM ≈ 2× raw image size
       for < 1 second, then back to display-scale baseline.

MEMORY BUDGET (example 8000×5333 source, no GPU, full-res display)
  display_img numpy  (8000×5333):  ~128 MB
  effects_img numpy  (same):       ~128 MB
  QPixmap in Qt/CPU:               ~128 MB
  Python + Qt + OpenCV overhead:    ~40 MB
  ─────────────────────────────────────────
  Total steady-state:              ~424 MB  (full-res; one image at a time)
  QGraphicsView handles zoom/pan natively — no extra decode per interaction.
  Old image buffers are freed before the next image loads.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import gc
import sys
import json
import argparse
from datetime import datetime
from enum import Enum
from pathlib import Path

import cv2
import numpy as np

from PySide6.QtCore import (
    Qt, QThread, Signal, QPointF, QRectF, QTimer,
)
from PySide6.QtGui import (
    QImage, QPixmap, QPainter, QPen, QColor, QKeySequence,
    QWheelEvent, QMouseEvent, QBrush, QAction, QPalette,
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QGraphicsItem, QDialog, QLineEdit,
    QDialogButtonBox, QFormLayout, QVBoxLayout, QMessageBox, QToolBar,
    QLabel, QStatusBar, QWidget, QHBoxLayout, QSizePolicy,
)

# ── Constants ─────────────────────────────────────────────────────────────────

MAX_DISPLAY_PX          = 1800   # longest-edge cap for in-RAM working copy
MAX_BATCH_SIZE          = 200
MIN_IMAGE_SIZE          = 50
MAX_RECOMMENDED_CIRCLES = 30

# OpenCV drawing constants (used in numpy-based rendering)
_FONT = cv2.FONT_HERSHEY_DUPLEX
_AA   = cv2.LINE_AA


# ── Edit modes ────────────────────────────────────────────────────────────────

class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


# BGR tuples used for OpenCV drawing / saving
_MODE_BGR: dict[EditMode, tuple] = {
    EditMode.HIGHLIGHT: (0,   255, 0),
    EditMode.BLUR:      (255, 80,  80),
    EditMode.PIXELATE:  (80,  80,  255),
    EditMode.DARKEN:    (160, 160, 160),
    EditMode.GRAYSCALE: (210, 210, 210),
    EditMode.INVERT:    (255, 255, 0),
    EditMode.OUTLINE:   (0,   220, 255),
}

# QColor equivalents used in Qt overlay drawing
_MODE_QT: dict[EditMode, QColor] = {
    EditMode.HIGHLIGHT: QColor(0,   255, 0),
    EditMode.BLUR:      QColor(255, 80,  80),
    EditMode.PIXELATE:  QColor(80,  80,  255),
    EditMode.DARKEN:    QColor(160, 160, 160),
    EditMode.GRAYSCALE: QColor(210, 210, 210),
    EditMode.INVERT:    QColor(255, 255, 0),
    EditMode.OUTLINE:   QColor(0,   220, 255),
}


# ── Utilities ─────────────────────────────────────────────────────────────────

def bgr_to_pixmap(img: np.ndarray) -> QPixmap:
    """
    Convert an OpenCV BGR uint8 array to a QPixmap.
    Uses QImage.Format_RGB888 after in-place channel swap.
    The .copy() call detaches the QImage from the numpy buffer so
    the numpy array can be freed independently.
    """
    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    h, w, ch = rgb.shape
    qi = QImage(rgb.data, w, h, w * ch, QImage.Format.Format_RGB888)
    return QPixmap.fromImage(qi.copy())


# ── Background image loader ───────────────────────────────────────────────────

class ImageLoader(QThread):
    """
    Loads one image in a worker thread.

    Signals
    -------
    loaded(display_img, display_scale, orig_w, orig_h)
        Emitted when the image has been loaded AND downscaled.
        display_img is a BGR numpy array at most MAX_DISPLAY_PX on the
        longest edge.  The full-resolution array is freed before the
        signal fires — no peak-RAM bloat from keeping both in memory.
    error(message)
        Emitted if the file cannot be read.
    """
    loaded = Signal(object, float, int, int)  # np.ndarray, scale, w, h
    error  = Signal(str)

    def __init__(self, path: Path, max_px: int = MAX_DISPLAY_PX):
        super().__init__()
        self._path   = path
        self._max_px = max_px
        self._abort  = False

    def abort(self):
        self._abort = True

    def run(self):
        try:
            full = cv2.imread(str(self._path))
            if full is None:
                self.error.emit(f"cv2.imread returned None for {self._path.name}")
                return
            if self._abort:
                return

            orig_h, orig_w = full.shape[:2]
            if min(orig_h, orig_w) < MIN_IMAGE_SIZE:
                self.error.emit(f"{self._path.name}: image too small ({orig_w}×{orig_h})")
                return

            # ── Full-resolution display: no downscaling ────────────────────
            # QGraphicsView handles zoom and pan natively on the QPixmap, so
            # we never need a pre-scaled working copy.  scale=1.0 means all
            # annotation coordinates are already in original pixel space.
            if not self._abort:
                self.loaded.emit(full, 1.0, orig_w, orig_h)

        except Exception as exc:
            if not self._abort:
                self.error.emit(str(exc))


# ── Ghost circle overlay (live drawing preview) ───────────────────────────────

class GhostCircleItem(QGraphicsItem):
    """
    Lightweight QGraphicsItem that draws a single dashed circle preview
    while the user is dragging.  Sits at Z=10, above the image pixmap.
    All coordinates are in scene (display-image) space.
    """

    def __init__(self):
        super().__init__()
        self._cx    = 0.0
        self._cy    = 0.0
        self._r     = 0.0
        self._color = QColor(0, 255, 0)
        self.setVisible(False)
        self.setZValue(10)

    def update_circle(self, cx: float, cy: float, r: float, color: QColor):
        self.prepareGeometryChange()
        self._cx    = cx
        self._cy    = cy
        self._r     = r
        self._color = color
        self.update()

    # QGraphicsItem interface ──────────────────────────────────────────────────

    def boundingRect(self) -> QRectF:
        pad = 6.0
        d   = self._r + pad
        return QRectF(self._cx - d, self._cy - d, 2 * d, 2 * d)

    def paint(self, painter: QPainter, option, widget=None):
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Dashed circle outline
        pen = QPen(self._color, 2.0, Qt.PenStyle.DashLine)
        pen.setCosmetic(True)   # constant screen-pixel width regardless of zoom
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawEllipse(QPointF(self._cx, self._cy), self._r, self._r)

        # Small crosshair at centre
        s   = 7.0
        pen2 = QPen(self._color, 1.0, Qt.PenStyle.SolidLine)
        pen2.setCosmetic(True)
        painter.setPen(pen2)
        painter.drawLine(QPointF(self._cx - s, self._cy),
                         QPointF(self._cx + s, self._cy))
        painter.drawLine(QPointF(self._cx, self._cy - s),
                         QPointF(self._cx, self._cy + s))


# ── Image viewer ──────────────────────────────────────────────────────────────

class ImageViewer(QGraphicsView):
    """
    QGraphicsView for smooth zoom / pan / circle-drawing.

    Zoom  — mouse wheel, anchored at cursor position.
    Pan   — right-mouse drag.
    Draw  — left-mouse drag; emits circle_drawn(center_orig, radius_orig)
            on release when draw_enabled is True.

    The scene holds exactly two items:
      • _pixmap_item   — the rendered composite image
      • _ghost         — the live-preview circle (hidden when not drawing)

    Pan and zoom are pure Qt transform operations — zero numpy work,
    zero cache rebuilds.
    """

    # Emitted when user finishes drawing a circle.
    # Carries (cx_orig, cy_orig) and radius in ORIGINAL image pixel coords.
    circle_drawn = Signal(tuple, int)

    # Zoom change signal so the main window can refresh the status bar
    zoom_changed = Signal(float)

    ZOOM_FACTOR = 1.20
    MIN_SCALE   = 0.03
    MAX_SCALE   = 16.0

    def __init__(self, parent=None):
        super().__init__(parent)

        # Scene setup
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)

        # Image item
        self._pixmap_item = QGraphicsPixmapItem()
        self._pixmap_item.setTransformationMode(
            Qt.TransformationMode.SmoothTransformation)
        self._pixmap_item.setZValue(0)
        self._scene.addItem(self._pixmap_item)

        # Ghost overlay
        self._ghost = GhostCircleItem()
        self._scene.addItem(self._ghost)

        # Rendering settings
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.setOptimizationFlag(
            QGraphicsView.OptimizationFlag.DontAdjustForAntialiasing, True)
        self.setViewportUpdateMode(
            QGraphicsView.ViewportUpdateMode.SmartViewportUpdate)
        self.setDragMode(QGraphicsView.DragMode.NoDrag)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.NoAnchor)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.NoAnchor)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setBackgroundBrush(QBrush(QColor(18, 18, 18)))

        # Interaction state
        self._draw_enabled   = True
        self._draw_mode      = EditMode.HIGHLIGHT
        self._display_scale  = 1.0  # display_img / original pixel ratio
        self._panning        = False
        self._pan_origin     = None
        self._drawing        = False
        self._draw_start_scene = None   # scene-space start of current circle

    # ── Public API ────────────────────────────────────────────────────────────

    def set_pixmap(self, pixmap: QPixmap):
        """Replace the displayed image.  Scene rect is updated automatically."""
        self._pixmap_item.setPixmap(pixmap)
        rect = QRectF(pixmap.rect())
        self._scene.setSceneRect(rect)

    def fit_view(self):
        """Scale-to-fit with a small margin, keeping aspect ratio."""
        rect = self._scene.sceneRect()
        if rect.isEmpty():
            return
        self.fitInView(rect.adjusted(-20, -20, 20, 20),
                       Qt.AspectRatioMode.KeepAspectRatio)

    def set_draw_enabled(self, enabled: bool):
        self._draw_enabled = enabled
        if not enabled:
            self._ghost.setVisible(False)
            self._drawing = False

    def set_draw_mode(self, mode: EditMode):
        self._draw_mode = mode

    def set_display_scale(self, scale: float):
        self._display_scale = scale

    @property
    def current_zoom(self) -> float:
        """Current horizontal scale factor of the view transform."""
        return self.transform().m11()

    # ── Event handlers ────────────────────────────────────────────────────────

    def wheelEvent(self, event: QWheelEvent):
        """Zoom in/out anchored at the cursor position."""
        delta  = event.angleDelta().y()
        factor = self.ZOOM_FACTOR if delta > 0 else (1.0 / self.ZOOM_FACTOR)

        # Clamp to allowed range
        new_scale = self.current_zoom * factor
        if not (self.MIN_SCALE <= new_scale <= self.MAX_SCALE):
            return

        # Zoom toward the cursor (translate so the scene point under the
        # cursor remains under the cursor after the scale)
        anchor    = event.position().toPoint()
        old_scene = self.mapToScene(anchor)
        self.scale(factor, factor)
        new_scene = self.mapToScene(anchor)
        diff      = new_scene - old_scene
        self.translate(diff.x(), diff.y())

        self.zoom_changed.emit(self.current_zoom)

    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.RightButton:
            self._panning    = True
            self._pan_origin = event.position()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
            return

        if event.button() == Qt.MouseButton.LeftButton and self._draw_enabled:
            self._drawing          = True
            self._draw_start_scene = self.mapToScene(event.position().toPoint())
            return

        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent):
        if self._panning and self._pan_origin is not None:
            delta = event.position() - self._pan_origin
            self._pan_origin = event.position()
            # Divide by scale so translate is in scene units, not screen pixels
            s = self.current_zoom
            self.translate(delta.x() / s, delta.y() / s)
            return

        if self._drawing and self._draw_start_scene is not None:
            cur = self.mapToScene(event.position().toPoint())
            r   = float(np.hypot(cur.x() - self._draw_start_scene.x(),
                                 cur.y() - self._draw_start_scene.y()))
            if r > 3:
                self._ghost.update_circle(
                    self._draw_start_scene.x(),
                    self._draw_start_scene.y(),
                    r,
                    _MODE_QT[self._draw_mode])
                self._ghost.setVisible(True)
            return

        super().mouseMoveEvent(event)

    def keyPressEvent(self, event):
        """
        Forward ALL key events to the parent MainWindow.
        Without this override, QGraphicsView consumes arrow keys and letter
        keys for its built-in scrolling — silently blocking A/D/S/E/etc.
        """
        if self.parent() is not None:
            QApplication.sendEvent(self.parent(), event)
        else:
            super().keyPressEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.RightButton and self._panning:
            self._panning = False
            self.setCursor(Qt.CursorShape.ArrowCursor)
            return

        if event.button() == Qt.MouseButton.LeftButton and self._drawing:
            self._drawing = False
            self._ghost.setVisible(False)

            if self._draw_start_scene is not None:
                cur     = self.mapToScene(event.position().toPoint())
                r_disp  = float(np.hypot(cur.x() - self._draw_start_scene.x(),
                                         cur.y() - self._draw_start_scene.y()))
                if r_disp > 5:
                    ds      = self._display_scale
                    cx_orig = int(self._draw_start_scene.x() / ds)
                    cy_orig = int(self._draw_start_scene.y() / ds)
                    r_orig  = max(1, int(r_disp / ds))
                    self.circle_drawn.emit((cx_orig, cy_orig), r_orig)

            self._draw_start_scene = None
            return

        super().mouseReleaseEvent(event)


# ── Annotation input dialog ───────────────────────────────────────────────────

class AnnotationDialog(QDialog):
    """
    Compact, semi-transparent input dialog that appears near the drawn circle.

    Design goals:
    • 88% window opacity so the underlying image stays visible
    • Frameless — no title bar chrome, blends into the canvas
    • Positioned automatically beside the circle on screen
    • Accent colour matches the current edit mode
    • Light, airy layout — minimal padding, soft border only
    """

    def __init__(self, mode: EditMode, parent=None, screen_pos=None):
        super().__init__(parent)
        self._screen_pos = screen_pos

        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, False)
        self.setWindowOpacity(0.88)
        self.setMinimumWidth(320)
        self.setMaximumWidth(420)

        # Accent from mode — softer, desaturated version for backgrounds
        qc     = _MODE_QT[mode]
        accent = qc.name()                      # full-sat for borders/buttons
        # Muted accent for subtle header strip: reduce saturation 40%
        h, s, v, _ = qc.getHsvF()
        soft_qc    = QColor.fromHsvF(h, s * 0.55, min(1.0, v * 1.1))
        soft       = soft_qc.name()

        self.setStyleSheet(f"""
            QDialog {{
                background: rgba(20, 20, 26, 230);
                border: 1px solid {accent};
                border-radius: 6px;
                color: #d8d8d8;
            }}
            QLabel {{
                color: #909090;
                font-size: 11px;
                padding: 0;
            }}
            QLineEdit {{
                background: rgba(38, 38, 48, 200);
                color: #e8e8e8;
                border: 1px solid rgba(255,255,255,40);
                border-bottom: 1px solid {accent};
                border-radius: 3px;
                padding: 5px 7px;
                font-size: 12px;
                selection-background-color: {soft};
            }}
            QLineEdit:focus {{
                border-bottom: 2px solid {accent};
                background: rgba(45, 45, 58, 210);
            }}
            QDialogButtonBox QPushButton {{
                background: rgba(30, 30, 40, 180);
                color: {accent};
                border: 1px solid {accent};
                border-radius: 3px;
                padding: 4px 16px;
                font-size: 11px;
                font-weight: 600;
                min-width: 58px;
            }}
            QDialogButtonBox QPushButton:hover {{
                background: rgba(255,255,255,18);
            }}
            QDialogButtonBox QPushButton:default {{
                background: {accent};
                color: #111;
            }}
        """)

        # ── Mode pill header ──────────────────────────────────────────────
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 10)
        outer.setSpacing(0)

        pill = QLabel(f"  ● {mode.value.upper()}")
        pill.setStyleSheet(f"""
            QLabel {{
                background: {soft};
                color: rgba(0,0,0,200);
                font-size: 10px;
                font-weight: 700;
                letter-spacing: 1px;
                padding: 4px 10px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }}
        """)
        outer.addWidget(pill)

        form = QFormLayout()
        form.setContentsMargins(14, 10, 14, 4)
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        outer.addLayout(form)

        self._lbl  = QLineEdit()
        self._desc = QLineEdit()
        self._lbl.setPlaceholderText("scratch, dent, misalignment …")
        self._desc.setPlaceholderText("optional detail")

        form.addRow("Label", self._lbl)
        form.addRow("Detail", self._desc)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        # Make OK the default (Enter confirms)
        ok_btn = btns.button(QDialogButtonBox.StandardButton.Ok)
        if ok_btn:
            ok_btn.setDefault(True)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(14, 0, 14, 0)
        btn_row.addStretch()
        btn_row.addWidget(btns)
        outer.addLayout(btn_row)

        self._lbl.setFocus()

    def showEvent(self, event):
        """Position the dialog beside the circle once Qt knows its size."""
        super().showEvent(event)
        if self._screen_pos is not None:
            self.adjustSize()
            screen = QApplication.primaryScreen().availableGeometry()
            x = self._screen_pos.x() + 28
            y = self._screen_pos.y() - self.height() // 2
            # Keep within screen bounds
            x = max(screen.left(), min(x, screen.right()  - self.width()))
            y = max(screen.top(),  min(y, screen.bottom() - self.height()))
            self.move(x, y)

    @property
    def label(self) -> str:
        return self._lbl.text().strip()

    @property
    def description(self) -> str:
        return self._desc.text().strip()


# ── Main window ───────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    """
    Top-level application window.

    Responsibilities
    ----------------
    • File discovery and navigation (A / D keys)
    • Owns the annotation list (circles in ORIGINAL pixel coords)
    • Manages the three-level render cache:
        display_img  (numpy)  →  effects_img (numpy)  →  composite_pixmap (QPixmap)
      Cache is invalidated only when annotations change.
      Pan / zoom never touch the cache — only the QGraphicsView transform.
    • Coordinates the background ImageLoader thread
    • Save (full-res, from disk) and Excel summary generation
    """

    def __init__(
        self,
        input_folder: str,
        output_folder: str | None = None,
        max_display_px: int = MAX_DISPLAY_PX,
    ):
        super().__init__()
        self.setWindowTitle("Batch Labeled Editor  ·  Qt Edition")
        self.resize(1440, 900)

        # ── Folders ───────────────────────────────────────────────────────────
        self._input_folder = Path(input_folder)
        self._max_display  = max_display_px
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._output_folder = (
            Path(output_folder) if output_folder
            else self._input_folder.parent / f"labeled_output_{ts}"
        )
        self._output_folder.mkdir(parents=True, exist_ok=True)

        # ── Image file list ───────────────────────────────────────────────────
        self._image_files: list[Path] = self._discover_images()
        self._total    = len(self._image_files)
        self._idx      = 0

        # ── Current-image state ───────────────────────────────────────────────
        self._display_img:   np.ndarray | None = None
        self._display_scale: float             = 1.0
        self._orig_w:        int               = 0
        self._orig_h:        int               = 0
        self._loader:        ImageLoader | None = None

        # ── Annotation state ──────────────────────────────────────────────────
        # Circles are always stored in ORIGINAL (full-res) pixel coordinates.
        # _circle_disp() converts them to display-image coords for rendering.
        self._circles:      list[dict] = []
        self._saved_status: dict[str, bool] = {}
        self._img_circles:  dict[str, list] = {}   # in-memory navigation cache

        # ── Render cache ──────────────────────────────────────────────────────
        self._effects_img:      np.ndarray | None = None
        self._composite_pixmap: QPixmap | None    = None
        self._effects_dirty    = True
        self._composite_dirty  = True
        self._show_labels      = True

        # ── Edit mode & effect parameters ─────────────────────────────────────
        self._mode        = EditMode.HIGHLIGHT
        self._blur_kernel = 25    # must be odd and > 0
        self._pix_size    = 10
        self._hl_alpha    = 0.4

        # ── UI ────────────────────────────────────────────────────────────────
        self._mode_actions: dict[EditMode, QAction] = {}
        self._build_ui()
        self._build_toolbar()
        self._build_statusbar()
        self._setup_shortcuts()

        # ── Deferred zoom-label refresh (batches rapid wheel events) ──────────
        self._zoom_timer = QTimer(self)
        self._zoom_timer.setSingleShot(True)
        self._zoom_timer.setInterval(60)   # ms
        self._zoom_timer.timeout.connect(self._update_status_bar)

        if not self._image_files:
            QMessageBox.critical(
                self, "No images found",
                f"No supported image files found in:\n{self._input_folder}")
            sys.exit(1)

        self._load_image(0)
        self._print_instructions()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        self._viewer = ImageViewer(self)
        self._viewer.circle_drawn.connect(self._on_circle_drawn)
        self._viewer.zoom_changed.connect(lambda _: self._zoom_timer.start())
        self.setCentralWidget(self._viewer)

    def _build_toolbar(self):
        tb = QToolBar("Controls", self)
        tb.setMovable(False)
        tb.setStyleSheet("""
            QToolBar { background: #232323; border: none; spacing: 4px; padding: 2px 6px; }
            QToolButton {
                color: #ccc; background: #2e2e2e;
                border: 1px solid #444; border-radius: 3px;
                padding: 3px 8px; font-size: 11px;
            }
            QToolButton:checked { border-color: #5af; color: #5af; }
            QToolButton:hover   { background: #3a3a3a; }
        """)
        self.addToolBar(tb)

        mode_keys = ["1", "2", "3", "4", "5", "6", "7"]
        mode_abbr = ["HLT", "BLR", "PIX", "DRK", "GRY", "INV", "OUT"]
        for key, abbr, mode in zip(mode_keys, mode_abbr, EditMode):
            act = QAction(f"{key}:{abbr}", self)
            act.setCheckable(True)
            act.setChecked(mode == EditMode.HIGHLIGHT)
            act.setToolTip(f"Mode: {mode.value}  (shortcut: {key})")
            act.triggered.connect(lambda _, m=mode: self._set_mode(m))
            tb.addAction(act)
            self._mode_actions[mode] = act

        tb.addSeparator()

        for label, tooltip, slot in [
            ("Save",    "Save current image (S)",                    self.save_current),
            ("Undo",    "Undo last annotation (U)",                  self._undo),
            ("Edit",    "Edit last annotation label/desc (E)",       self._edit_last_label),
            ("Clear",   "Clear all annotations (C)",                 self._clear),
            ("Labels",  "Toggle label visibility (T)",               self._toggle_labels),
            ("Fit",     "Fit image to window (R)",                   self._reset_zoom),
            ("List",    "List annotations in console (L)",           self._list_objects),
            ("Help",    "Show keyboard shortcuts (H)",               self._show_help),
        ]:
            act = QAction(label, self)
            act.setToolTip(tooltip)
            act.triggered.connect(slot)
            tb.addAction(act)

    def _build_statusbar(self):
        sb = self.statusBar()
        sb.setStyleSheet("QStatusBar { background: #1a1a1a; color: #999; }")

        def _lbl(text="–", bold=False):
            w = QLabel(text)
            w.setStyleSheet(
                f"color: {'#eee' if bold else '#999'}; padding: 0 8px;"
                + (" font-weight: bold;" if bold else ""))
            return w

        self._lbl_file  = _lbl()
        self._lbl_nav   = _lbl(bold=True)
        self._lbl_mode  = _lbl()
        self._lbl_obj   = _lbl()
        self._lbl_res   = _lbl()
        self._lbl_zoom  = _lbl()
        self._lbl_saved = _lbl(bold=True)

        # Left side: filename
        sb.addWidget(self._lbl_file)
        # Right side permanent widgets
        for w in (self._lbl_res, self._lbl_zoom, self._lbl_obj,
                  self._lbl_mode, self._lbl_nav, self._lbl_saved):
            sb.addPermanentWidget(w)

    def _setup_shortcuts(self):
        """
        Build a dispatch table used by keyPressEvent().

        WHY NOT QAction shortcuts:
          QGraphicsView grabs keyboard focus and consumes key events for its
          own scrolling before they can bubble up to QMainWindow QActions.
          The only reliable fix is to override keyPressEvent directly on the
          MainWindow, which always receives events after ImageViewer forwards
          them via its own keyPressEvent override.
        """
        modes = list(EditMode)
        self._key_dispatch = {
            Qt.Key.Key_A:      self._prev_image,
            Qt.Key.Key_D:      self._next_image,
            Qt.Key.Key_S:      self._handle_s_key,   # S vs Shift+S handled below
            Qt.Key.Key_C:      self._clear,
            Qt.Key.Key_U:      self._undo,
            Qt.Key.Key_E:      self._edit_last_label,
            Qt.Key.Key_T:      self._toggle_labels,
            Qt.Key.Key_R:      self._reset_zoom,
            Qt.Key.Key_L:      self._list_objects,
            Qt.Key.Key_H:      self._show_help,
            Qt.Key.Key_Q:      self.close,
            Qt.Key.Key_1:      lambda: self._set_mode(modes[0]),
            Qt.Key.Key_2:      lambda: self._set_mode(modes[1]),
            Qt.Key.Key_3:      lambda: self._set_mode(modes[2]),
            Qt.Key.Key_4:      lambda: self._set_mode(modes[3]),
            Qt.Key.Key_5:      lambda: self._set_mode(modes[4]),
            Qt.Key.Key_6:      lambda: self._set_mode(modes[5]),
            Qt.Key.Key_7:      lambda: self._set_mode(modes[6]),
        }

    def _handle_s_key(self, shift: bool = False):
        if shift:
            self._save_and_next()
        else:
            self.save_current()

    def keyPressEvent(self, event):
        """
        Central keyboard dispatcher.  Handles all single-key shortcuts so
        they work regardless of which child widget currently has focus.
        """
        key  = event.key()
        mods = event.modifiers()

        # Shift+S → save and advance
        if key == Qt.Key.Key_S and (mods & Qt.KeyboardModifier.ShiftModifier):
            self._save_and_next()
            return

        fn = self._key_dispatch.get(key)
        if fn is not None:
            fn()
            return

        super().keyPressEvent(event)

    # ── File discovery ────────────────────────────────────────────────────────

    def _discover_images(self) -> list[Path]:
        exts  = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}
        files = sorted(f for e in exts for f in self._input_folder.glob(f"*{e}"))
        valid = []
        for f in files:
            try:
                if f.stat().st_size >= 512:
                    valid.append(f)
                else:
                    print(f"  Skipping {f.name}: too small")
            except Exception as exc:
                print(f"  Skipping {f.name}: {exc}")
        if len(valid) > MAX_BATCH_SIZE:
            print(f"\nLarge batch: {len(valid)} images (recommended ≤ {MAX_BATCH_SIZE})")
        print(f"Found {len(valid)} image(s) in {self._input_folder}")
        return valid

    # ── Image loading ─────────────────────────────────────────────────────────

    def _free_current_image(self):
        """
        Explicitly release all pixel buffers for the current image.
        Called BEFORE starting to load the next one so the two images
        never coexist in RAM.
        """
        self._display_img      = None
        self._effects_img      = None
        self._composite_pixmap = None
        gc.collect()

    def _load_image(self, idx: int):
        """Begin loading image at index `idx` in the background thread."""
        if not (0 <= idx < self._total):
            return
        self._idx = idx
        f = self._image_files[idx]

        # ── Cancel any in-flight load ──────────────────────────────────────
        if self._loader and self._loader.isRunning():
            self._loader.abort()
            self._loader.quit()
            self._loader.wait(200)

        # ── Restore circle list from JSON or navigation cache ──────────────
        json_path = self._output_folder / f.with_suffix(".json").name
        if json_path.exists():
            try:
                with open(json_path) as jf:
                    data = json.load(jf)
                self._circles = [
                    {
                        "center":      tuple(o["center"]),
                        "radius":      o["radius"],
                        "mode":        EditMode(o.get("mode", "highlight")),
                        "label":       o.get("label", ""),
                        "description": o.get("description", ""),
                    }
                    for o in data.get("objects", [])
                ]
                print(f"  {f.name}  — restored {len(self._circles)} object(s) from JSON")
            except Exception as exc:
                print(f"  JSON load error: {exc}")
                self._circles = []
        elif f.name in self._img_circles:
            self._circles = [c.copy() for c in self._img_circles[f.name]]
            print(f"  {f.name}  — restored {len(self._circles)} object(s) from memory cache")
        else:
            self._circles = []

        # ── Free old buffers, then start loader ───────────────────────────
        self._free_current_image()
        self._viewer.set_draw_enabled(False)

        self._loader = ImageLoader(f, self._max_display)
        self._loader.loaded.connect(self._on_image_loaded, Qt.ConnectionType.QueuedConnection)
        self._loader.error.connect(self._on_image_error,   Qt.ConnectionType.QueuedConnection)
        self._loader.start()

        self.setWindowTitle(f"Batch Labeled Editor  —  {f.name}  (loading…)")
        self._update_status_bar()

    def _on_image_loaded(
        self,
        display_img: np.ndarray,
        scale: float,
        orig_w: int,
        orig_h: int,
    ):
        self._display_img   = display_img
        self._display_scale = scale
        self._orig_w        = orig_w
        self._orig_h        = orig_h

        self._viewer.set_display_scale(scale)
        self._effects_dirty   = True
        self._composite_dirty = True

        self._refresh_display()
        self._viewer.set_draw_enabled(True)
        self._viewer.fit_view()
        self._update_status_bar()

        f = self._image_files[self._idx]
        dh, dw = display_img.shape[:2]
        self.setWindowTitle(f"Batch Labeled Editor  —  {f.name}")
        print(f"  {f.name}  [{orig_w}×{orig_h}  disp:{dw}×{dh}"
              f"  scale:{scale:.2f}]  circles:{len(self._circles)}")

    def _on_image_error(self, msg: str):
        print(f"  Load error: {msg}")
        self.statusBar().showMessage(f"Load failed: {msg}", 4000)
        self._viewer.set_draw_enabled(True)
        # Skip to the next image if possible
        if self._idx < self._total - 1:
            self._load_image(self._idx + 1)

    # ── Render cache ──────────────────────────────────────────────────────────

    def _get_composite_pixmap(self) -> QPixmap | None:
        """
        Return the cached composite QPixmap, rebuilding as needed.
        Rebuilds effects layer if annotations changed,
        rebuilds composite (label overlay) if labels changed.
        """
        if self._display_img is None:
            return None
        if self._effects_dirty:
            self._rebuild_effects()
        if self._composite_dirty:
            self._rebuild_composite()
        return self._composite_pixmap

    def _rebuild_effects(self):
        """Apply region effects + circle outlines to a fresh display copy."""
        base = self._display_img.copy()
        for c in self._circles:
            dc   = self._circle_disp(c)
            base = self._apply_effect(base, dc)
            cv2.circle(base, dc["center"], dc["radius"],
                       _MODE_BGR[c["mode"]], 2, _AA)
        self._effects_img     = base
        self._effects_dirty   = False
        self._composite_dirty = True

    def _rebuild_composite(self):
        """
        Stamp label badges onto the effects layer.
        The O(n²) collision-detection logic runs only here — never per-frame.
        """
        if self._effects_img is None:
            self._rebuild_effects()
        tmp = self._effects_img.copy()
        if self._show_labels and self._circles:
            disp_circles = [self._circle_disp(c) for c in self._circles]
            self._draw_all_labels(tmp, disp_circles)
        self._composite_pixmap = bgr_to_pixmap(tmp)
        self._composite_dirty  = False

    def _refresh_display(self):
        """Push the current composite pixmap into the viewer."""
        pxm = self._get_composite_pixmap()
        if pxm is not None:
            self._viewer.set_pixmap(pxm)
        self._update_status_bar()

    # ── Coordinate helpers ────────────────────────────────────────────────────

    def _circle_disp(self, c: dict) -> dict:
        """Return a circle dict with coords scaled to display-image space."""
        ds = self._display_scale
        cx = int(c["center"][0] * ds)
        cy = int(c["center"][1] * ds)
        r  = max(1, int(c["radius"] * ds))
        return {**c, "center": (cx, cy), "radius": r}

    # ── OpenCV effects ────────────────────────────────────────────────────────

    def _apply_effect(self, image: np.ndarray, dc: dict) -> np.ndarray:
        """
        Apply the edit effect for `dc` in-place on `image`.
        Uses a boolean mask scatter-write to avoid allocating a full
        output array on every call (saves ~8 MB per invocation on a
        2000-px display image).
        `dc` must be in display-image pixel coords.
        """
        try:
            mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.circle(mask, dc["center"], dc["radius"], 255, -1)
            mb   = mask == 255
            mode = dc["mode"]

            if mode == EditMode.HIGHLIGHT:
                lit = cv2.addWeighted(
                    image, 1 - self._hl_alpha,
                    np.full_like(image, 255), self._hl_alpha, 0)
                image[mb] = lit[mb]

            elif mode == EditMode.BLUR:
                k = self._blur_kernel
                blurred = cv2.GaussianBlur(image, (k, k), 0)
                image[mb] = blurred[mb]

            elif mode == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                sz    = self._pix_size
                small = cv2.resize(image,
                                   (max(1, w // sz), max(1, h // sz)),
                                   interpolation=cv2.INTER_NEAREST)
                pix   = cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST)
                image[mb] = pix[mb]

            elif mode == EditMode.DARKEN:
                image[mb] = (image[mb] * 0.5).astype(np.uint8)

            elif mode == EditMode.GRAYSCALE:
                gray = cv2.cvtColor(
                    cv2.cvtColor(image, cv2.COLOR_BGR2GRAY),
                    cv2.COLOR_GRAY2BGR)
                image[mb] = gray[mb]

            elif mode == EditMode.INVERT:
                image[mb] = 255 - image[mb]

            # EditMode.OUTLINE: no fill effect; circle outline drawn separately

        except Exception as exc:
            print(f"  Effect error ({dc.get('mode', '?')}): {exc}")

        return image

    # ── Label drawing (OpenCV) ────────────────────────────────────────────────

    @staticmethod
    def _rects_overlap(r1: tuple, r2: tuple, buf: int = 4) -> bool:
        return not (
            r1[2] + buf < r2[0] or r1[0] - buf > r2[2] or
            r1[3] + buf < r2[1] or r1[1] - buf > r2[3]
        )

    def _find_label_pos(
        self,
        cx: int, cy: int, r: int,
        tw: int, fh: int, bl: int,
        pad: int,
        img_shape: tuple,
        placed: list,
    ) -> tuple:
        """
        Find a non-overlapping position for a label tag around a circle.
        Returns (text_x, text_y, bounding_rect).
        """
        ih, iw = img_shape[:2]
        W = tw + 2 * pad
        H = fh + bl + 2 * pad

        candidates = [
            (cx - r,     cy - r - H - 10),
            (cx - r,     cy + r + 10),
            (cx - r - W - 8, cy - H // 2),
            (cx + r + 8, cy - H // 2),
            (cx + r + 8, cy - r - H - 10),
            (cx - r - W - 8, cy - r - H - 10),
            (cx + r + 8, cy + r + 10),
            (cx - r - W - 8, cy + r + 10),
        ]
        for extra in (40, 80, 120):
            candidates += [(px, py + extra) for px, py in candidates[:4]]
            candidates += [(px, py - extra) for px, py in candidates[:4]]

        for px, py in candidates:
            rect = (int(px), int(py), int(px + W), int(py + H))
            if rect[0] < pad or rect[1] < pad or rect[2] > iw - pad or rect[3] > ih - pad:
                continue
            if not any(self._rects_overlap(rect, r2) for r2 in placed):
                return int(px) + pad, int(py) + fh + pad, rect

        # Fallback: stack below previous labels
        bottom = max((r2[3] for r2 in placed), default=0) + 8
        lx = max(pad, min(cx - W // 2, iw - W - pad))
        ly = min(bottom + fh + pad, ih - bl - pad)
        return lx, ly, (lx - pad, ly - fh - pad, lx + tw + pad, ly + bl + pad)

    def _draw_all_labels(self, image: np.ndarray, disp_circles: list):
        """
        Render badge numbers + semi-transparent label tags onto `image`.
        Label backgrounds are blended at LABEL_BG_ALPHA so the underlying
        image shows through.  Text is kept fully opaque for readability.
        Called ONLY from _rebuild_composite — never per-frame.
        """
        LABEL_BG_ALPHA  = 0.55   # how opaque the label box background is
        BADGE_ALPHA     = 0.72   # badge circle fill opacity
        GLOW_ALPHAS     = ((6, 0.12), (3, 0.22))   # softer glow passes

        placed: list[tuple] = []
        ih, iw = image.shape[:2]
        min_dim = min(ih, iw)
        sc = 0.68 if min_dim >= 900 else (0.54 if min_dim >= 400 else 0.42)
        th = 1    # single stroke throughout — less dense

        for i, (c, dc) in enumerate(zip(self._circles, disp_circles), 1):
            label  = c["label"] or f"Object #{i}"
            color  = _MODE_BGR[c["mode"]]
            cx, cy = dc["center"]
            r      = dc["radius"]

            # ── Semi-transparent numbered badge ───────────────────────────
            badge       = str(i)
            (bw, bh), _ = cv2.getTextSize(badge, _FONT, sc * 0.85, th)
            br          = max(bw, bh) // 2 + 5

            # Badge fill: blend colour into image at BADGE_ALPHA
            badge_mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.circle(badge_mask, (cx, cy), br, 255, -1)
            bm = badge_mask == 255
            colored = np.full_like(image, color)
            image[bm] = cv2.addWeighted(
                colored, BADGE_ALPHA,
                image,   1 - BADGE_ALPHA, 0)[bm]
            # Thin white border — cosmetic only, stays solid
            cv2.circle(image, (cx, cy), br, (220, 220, 220), 1, _AA)
            # Number text — solid black for legibility
            cv2.putText(image, badge, (cx - bw // 2, cy + bh // 2),
                        _FONT, sc * 0.85, (0, 0, 0), th + 1, _AA)

            # ── Label tag ─────────────────────────────────────────────────
            tag          = f"#{i}  {label}"   # shorter — mode not repeated
            (tw, fh), bl = cv2.getTextSize(tag, _FONT, sc, th)
            pad          = max(4, int(4 * sc / 0.5))
            lx, ly, rect = self._find_label_pos(
                cx, cy, r, tw, fh, bl, pad, image.shape, placed)
            placed.append(rect)
            x1, y1, x2, y2 = rect

            # ── Soft glow (very subtle) ────────────────────────────────────
            for exp, alpha in GLOW_ALPHAS:
                ry1 = max(0, y1 - exp);  ry2 = min(ih, y2 + exp)
                rx1 = max(0, x1 - exp);  rx2 = min(iw, x2 + exp)
                if ry2 <= ry1 or rx2 <= rx1:
                    continue
                roi = image[ry1:ry2, rx1:rx2].copy()
                cv2.rectangle(roi, (0, 0),
                              (roi.shape[1] - 1, roi.shape[0] - 1), color, -1, _AA)
                image[ry1:ry2, rx1:rx2] = cv2.addWeighted(
                    roi, alpha, image[ry1:ry2, rx1:rx2], 1 - alpha, 0)

            # ── Semi-transparent label background ──────────────────────────
            if y2 > y1 and x2 > x1:
                roi = image[y1:y2, x1:x2].copy()
                bg  = np.full_like(roi, (12, 12, 18))
                image[y1:y2, x1:x2] = cv2.addWeighted(
                    bg, LABEL_BG_ALPHA, roi, 1 - LABEL_BG_ALPHA, 0)

            # Left accent bar (solid, 2 px) and thin border
            cv2.rectangle(image, (x1, y1), (x1 + 2, y2), color, -1, _AA)
            cv2.rectangle(image, (x1, y1), (x2, y2),      color,  1, _AA)

            # Text: subtle shadow then white
            cv2.putText(image, tag, (lx + 5, ly),     _FONT, sc, (0, 0, 0),       th + 1, _AA)
            cv2.putText(image, tag, (lx + 4, ly - 1), _FONT, sc, (235, 235, 235), th,     _AA)

            # ── Thin connector line ────────────────────────────────────────
            mid_x = (x1 + x2) // 2
            mid_y = y2 if ly > cy else y1
            ang   = np.arctan2(mid_y - cy, mid_x - cx)
            ex    = int(cx + r * np.cos(ang))
            ey    = int(cy + r * np.sin(ang))
            # Blend a thin line instead of drawing it solid
            line_mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.line(line_mask, (mid_x, mid_y), (ex, ey), 200, 1)
            lm = line_mask > 0
            colored_line = np.full_like(image, color)
            image[lm] = cv2.addWeighted(
                colored_line, 0.65, image, 0.35, 0)[lm]
            # Small dot at circle edge
            cv2.circle(image, (ex, ey), 3, color,           -1, _AA)
            cv2.circle(image, (ex, ey), 3, (200, 200, 200),  1, _AA)

    # ── Circle annotation flow ────────────────────────────────────────────────

    def _on_circle_drawn(self, center_orig: tuple, radius_orig: int):
        """Called when the user finishes dragging a circle."""
        if len(self._circles) >= MAX_RECOMMENDED_CIRCLES:
            print(f"  Warning: {len(self._circles)} annotations — "
                  f"rendering may slow down")

        # Convert circle centre from scene (original-pixel) coords to
        # screen coords so the dialog can appear right beside the circle.
        scene_pt  = QPointF(center_orig[0] * self._display_scale,
                            center_orig[1] * self._display_scale)
        view_pt   = self._viewer.mapFromScene(scene_pt)
        screen_pt = self._viewer.mapToGlobal(view_pt)

        dlg = AnnotationDialog(self._mode, self, screen_pos=screen_pt)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return     # user cancelled — discard the drawn circle

        lbl  = dlg.label
        desc = dlg.description
        self._circles.append({
            "center":      center_orig,
            "radius":      radius_orig,
            "mode":        self._mode,
            "label":       lbl,
            "description": desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  Added {tag}" +
              (f"  + description" if desc else "") +
              f"  [{self._mode.value}]  r={radius_orig}  pos={center_orig}")

        self._effects_dirty   = True
        self._composite_dirty = True
        self._refresh_display()

    # ── Commands ──────────────────────────────────────────────────────────────

    def _set_mode(self, mode: EditMode):
        self._mode = mode
        self._viewer.set_draw_mode(mode)
        for m, act in self._mode_actions.items():
            act.setChecked(m == mode)
        self._update_status_bar()
        print(f"  Mode → {mode.value.upper()}")

    def _undo(self):
        if not self._circles:
            self.statusBar().showMessage("Nothing to undo", 1500)
            return
        rem = self._circles.pop()
        print(f"  Removed: {rem['label'] or '(unlabeled)'}")
        self._effects_dirty   = True
        self._composite_dirty = True
        self._refresh_display()

    def _clear(self):
        if not self._circles:
            return
        self._circles.clear()
        self._effects_dirty   = True
        self._composite_dirty = True
        self._refresh_display()
        print("  Cleared all annotations")

    def _toggle_labels(self):
        self._show_labels     = not self._show_labels
        self._composite_dirty = True
        self._refresh_display()
        print(f"  Labels {'ON' if self._show_labels else 'OFF'}")

    def _reset_zoom(self):
        self._viewer.fit_view()
        self._update_status_bar()

    def _list_objects(self):
        f = self._image_files[self._idx]
        print("\n" + "=" * 64)
        print(f"Annotations — {f.name}  [{self._orig_w}×{self._orig_h}]")
        print("=" * 64)
        if not self._circles:
            print("  (none)")
        for i, c in enumerate(self._circles, 1):
            print(f"  #{i}  {c['label'] or '(no label)':30s}"
                  f"  [{c['mode'].value}]"
                  f"  r={c['radius']}  pos={c['center']}")
            if c.get("description"):
                print(f"       {c['description']}")
        print("=" * 64 + "\n")

    def _edit_last_label(self):
        """
        Re-open the annotation dialog pre-filled with the last circle's
        label + description so the user can correct a typo without
        having to undo and redraw.
        """
        if not self._circles:
            self.statusBar().showMessage("No annotations to edit", 1500)
            return

        last = self._circles[-1]

        # Position dialog near the last circle
        scene_pt  = QPointF(last["center"][0] * self._display_scale,
                            last["center"][1] * self._display_scale)
        view_pt   = self._viewer.mapFromScene(scene_pt)
        screen_pt = self._viewer.mapToGlobal(view_pt)

        dlg = AnnotationDialog(last["mode"], self, screen_pos=screen_pt)
        dlg.setWindowTitle(
            f"Edit Annotation #{len(self._circles)}  ·  {last['mode'].value.upper()}")
        dlg._lbl.setText(last["label"])
        dlg._desc.setText(last.get("description", ""))
        dlg._lbl.selectAll()
        dlg._lbl.setFocus()

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        old_label = last["label"]
        last["label"]       = dlg.label
        last["description"] = dlg.description

        print(f"  Edited #{len(self._circles)}: "
              f"'{old_label}' → '{last['label']}'")
        self._composite_dirty = True
        self._refresh_display()

    def _show_help(self):
        QMessageBox.information(self, "Keyboard Shortcuts", """\
Navigation
  A / D       Previous / Next image
  S           Save current image + JSON
  Shift+S     Save and go to next
  Q           Quit (auto-saves unsaved)

Annotations
  Left drag   Draw circle
  E           Edit last annotation (label & description)
  U           Undo last annotation
  C           Clear all annotations
  T           Toggle label visibility
  1 – 7       Switch edit mode
                1 Highlight   2 Blur    3 Pixelate
                4 Darken      5 Grayscale
                6 Invert      7 Outline

View
  Mouse wheel Zoom (anchored at cursor)
  Right drag  Pan
  R           Fit image to window
  L           List annotations in console
  H           This help dialog
""")

    # ── Navigation ────────────────────────────────────────────────────────────

    def _cache_circles(self):
        """Save current circle list to the in-memory navigation cache."""
        f = self._image_files[self._idx]
        self._img_circles[f.name] = [c.copy() for c in self._circles]

    def _navigate(self, delta: int):
        # Block navigation while loading
        if self._loader and self._loader.isRunning():
            return
        new_idx = self._idx + delta
        if not (0 <= new_idx < self._total):
            edge = "first" if delta < 0 else "last"
            self.statusBar().showMessage(f"Already at the {edge} image", 1500)
            return
        # Auto-save unsaved annotations before leaving
        if self._circles:
            self._cache_circles()
            f = self._image_files[self._idx]
            if not self._saved_status.get(f.name):
                self.save_current(auto_save=True)
        self._load_image(new_idx)

    def _prev_image(self):    self._navigate(-1)
    def _next_image(self):    self._navigate(+1)
    def _save_and_next(self): self.save_current(); self._next_image()

    # ── Save (full-resolution) ────────────────────────────────────────────────

    def save_current(self, auto_save: bool = False):
        """
        Reload the source file at FULL resolution, apply all effects at
        original pixel coords, stamp labels, write the output image + JSON,
        then immediately free the full-res buffer.

        Peak RAM during save ≈ 2× raw image size (original + working copy),
        which lasts < 1 second before dropping back to display-scale baseline.
        """
        if not self._circles:
            if not auto_save:
                self.statusBar().showMessage("No annotations to save", 2000)
            return

        f   = self._image_files[self._idx]
        out = self._output_folder / f.name
        js  = out.with_suffix(".json")
        self._cache_circles()

        try:
            # Step 1: reload full-res (only for the duration of the save)
            full = cv2.imread(str(f))
            if full is None:
                raise IOError(f"Cannot reload {f.name} for full-res save")

            # Step 2: apply effects at original-pixel coords
            #         (circles are already stored in original coords)
            for c in self._circles:
                full = self._apply_effect(full, c)
                cv2.circle(full, c["center"], c["radius"],
                           _MODE_BGR[c["mode"]], 3, _AA)

            # Step 3: draw labels at full scale
            #         Temporarily set display_scale=1.0 so _circle_disp()
            #         returns original coords unchanged.
            saved_scale = self._display_scale
            self._display_scale = 1.0
            self._draw_all_labels(full, self._circles)
            self._display_scale = saved_scale

            # Step 4: write image
            if not cv2.imwrite(str(out), full):
                raise IOError("cv2.imwrite failed")

            # Step 5: free full-res immediately — back to display-scale baseline
            del full
            gc.collect()

            # Step 6: write JSON (original pixel coords)
            data = {
                "source_image":  f.name,
                "original_size": [self._orig_w, self._orig_h],
                "timestamp":     datetime.now().isoformat(),
                "objects": [
                    {
                        "id":          i,
                        "label":       c["label"],
                        "description": c.get("description", ""),
                        "mode":        c["mode"].value,
                        "center":      list(c["center"]),
                        "radius":      c["radius"],
                    }
                    for i, c in enumerate(self._circles, 1)
                ],
            }
            with open(js, "w") as jf:
                json.dump(data, jf, indent=2)

            self._saved_status[f.name] = True
            msg = (
                f"Auto-saved {len(self._circles)} annotation(s)  [{f.name}]"
                if auto_save else
                f"Saved  {f.name}  ({len(self._circles)} annotation(s))"
            )
            print(f"  {msg}\n  → {out}\n  → {js}")
            self.statusBar().showMessage(msg, 3000)
            self._update_status_bar()

        except Exception as exc:
            print(f"  Save error: {exc}")
            self.statusBar().showMessage(f"Save failed: {exc}", 5000)

    # ── Excel summary ─────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self._output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"

            hfill   = PatternFill(start_color="2E4DA7", end_color="2E4DA7",
                                  fill_type="solid")
            hfont   = Font(bold=True, color="FFFFFF", size=11)
            border  = Border(
                **{s: Side(style="thin")
                   for s in ("left", "right", "top", "bottom")})
            center_a = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
            top_l    = Alignment(horizontal="left",   vertical="top",
                                 wrap_text=True)

            def styled(cell, fill=None, font=None, align=None):
                if fill:  cell.fill      = fill
                if font:  cell.font      = font
                if align: cell.alignment = align
                cell.border = border

            for col, hdr in enumerate(
                    ["#", "Image Name", "Error Count", "Errors (numbered)"], 1):
                c = ws.cell(row=1, column=col, value=hdr)
                styled(c, fill=hfill, font=hfont, align=center_a)
            ws.row_dimensions[1].height = 28

            data_row = 2
            img_num  = 0
            for img in sorted(self._image_files):
                if img.name not in self._saved_status:
                    continue
                jp = self._output_folder / img.with_suffix(".json").name
                if not jp.exists():
                    continue
                with open(jp) as jf:
                    data = json.load(jf)
                objects  = data.get("objects", [])
                img_num += 1

                c1 = ws.cell(row=data_row, column=1, value=img_num)
                styled(c1, align=center_a); c1.font = Font(bold=True, size=10)

                c2 = ws.cell(row=data_row, column=2, value=img.name)
                styled(c2, align=top_l)

                c3 = ws.cell(row=data_row, column=3, value=len(objects))
                styled(c3, align=center_a)
                if len(objects) == 0:
                    c3.fill = PatternFill(start_color="C6EFCE",
                                          end_color="C6EFCE", fill_type="solid")
                    c3.font = Font(color="276221", bold=True)
                elif len(objects) <= 3:
                    c3.fill = PatternFill(start_color="FFEB9C",
                                          end_color="FFEB9C", fill_type="solid")
                    c3.font = Font(color="9C5700", bold=True)
                else:
                    c3.fill = PatternFill(start_color="FFC7CE",
                                          end_color="FFC7CE", fill_type="solid")
                    c3.font = Font(color="9C0006", bold=True)

                lines = [
                    f"{i}. {o.get('label', '').strip() or '(unlabeled)'}"
                    + (f" — {o['description']}"
                       if o.get("description", "").strip() else "")
                    for i, o in enumerate(objects, 1)
                ] or ["(no errors)"]
                c4 = ws.cell(row=data_row, column=4, value="\n".join(lines))
                styled(c4, align=top_l)
                ws.row_dimensions[data_row].height = max(20, len(objects) * 15 + 6)

                if img_num % 2 == 0:
                    alt = PatternFill(start_color="F2F5FB",
                                      end_color="F2F5FB", fill_type="solid")
                    for col in (1, 2, 4):
                        ws.cell(row=data_row, column=col).fill = alt
                data_row += 1

            data_row += 1
            sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                fill_type="solid")
            sfont = Font(bold=True, size=11)

            def summary_row(label: str, value):
                nonlocal data_row
                for col, val in ((2, label), (3, value)):
                    c = ws.cell(row=data_row, column=col, value=val)
                    c.font   = sfont
                    c.fill   = sfill
                    c.border = border
                    if col == 3:
                        c.alignment = center_a
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            total_obj = sum(
                len(json.load(
                    open(self._output_folder / f_.with_suffix(".json").name)
                )["objects"])
                for f_ in self._image_files
                if f_.name in self._saved_status
                and (self._output_folder / f_.with_suffix(".json").name).exists()
            )
            summary_row("Total Images Processed", len(self._saved_status))
            summary_row("Total Errors Found",     total_obj)
            if self._saved_status:
                summary_row("Avg Errors / Image",
                            round(total_obj / len(self._saved_status), 1))

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 34
            ws.column_dimensions["C"].width = 13
            ws.column_dimensions["D"].width = 62
            ws.freeze_panes = "A2"
            wb.save(str(xp))
            print(f"  Excel summary → {xp}")

        except ImportError:
            print("  pip install openpyxl  (skipping Excel summary)")
        except Exception as exc:
            print(f"  Excel error: {exc}")

    # ── Status bar ────────────────────────────────────────────────────────────

    def _update_status_bar(self):
        if not self._image_files:
            return
        f          = self._image_files[self._idx]
        is_saved   = self._saved_status.get(f.name, False)
        has_edits  = len(self._circles) > 0 or f.name in self._img_circles
        save_state = ("SAVED" if is_saved else ("EDITED" if has_edits else "NO EDITS"))

        self._lbl_file.setText(f.name[:55])
        self._lbl_nav.setText(f"{self._idx + 1} / {self._total}")
        self._lbl_mode.setText(f"Mode: {self._mode.value.upper()}")
        self._lbl_obj.setText(f"Objects: {len(self._circles)}")
        self._lbl_res.setText(f"{self._orig_w}×{self._orig_h}")
        self._lbl_zoom.setText(f"Zoom: {self._viewer.current_zoom:.2f}×")
        self._lbl_saved.setText(save_state)

        colors = {"SAVED": "#5f5", "EDITED": "#5af", "NO EDITS": "#666"}
        self._lbl_saved.setStyleSheet(
            f"color: {colors[save_state]}; font-weight: bold; padding: 0 8px;")

    # ── Close ─────────────────────────────────────────────────────────────────

    def closeEvent(self, event):
        # Auto-save unsaved annotations for the current image
        if self._circles:
            f = self._image_files[self._idx]
            if not self._saved_status.get(f.name):
                self.save_current(auto_save=True)

        # Stop any in-flight loader
        if self._loader and self._loader.isRunning():
            self._loader.abort()
            self._loader.quit()
            self._loader.wait(500)

        # Generate Excel summary
        if self._saved_status:
            self.generate_summary()
            print(f"\n  Done — {len(self._saved_status)}/{self._total} image(s) saved"
                  f"\n  Output: {self._output_folder}")
        else:
            print("\n  No images were saved")

        event.accept()

    # ── Console instructions ──────────────────────────────────────────────────

    @staticmethod
    def _print_instructions():
        print("\n" + "=" * 66)
        print("  BATCH LABELED EDITOR  ·  Qt Edition")
        print("  Full-Res Save  ·  Background Loading  ·  Low-Memory Display")
        print("=" * 66)
        print("  Left drag    Draw annotation circle")
        print("  Right drag   Pan                    R     Fit to window")
        print("  Mouse wheel  Zoom                   A/D   Prev / Next image")
        print("  S            Save                   U     Undo last circle")
        print("  Shift+S      Save + Next            C     Clear all")
        print("  E            Edit last annotation   L     List annotations")
        print("  T            Toggle labels          H     Help dialog")
        print("  1–7          Edit mode              Q     Quit (auto-saves)")
        print("=" * 66 + "\n")


# ── Entry point ───────────────────────────────────────────────────────────────

def _dark_palette() -> QPalette:
    """Return a dark Fusion palette for the entire application."""
    p = QPalette()
    p.setColor(QPalette.ColorRole.Window,          QColor(30,  30,  30))
    p.setColor(QPalette.ColorRole.WindowText,      QColor(220, 220, 220))
    p.setColor(QPalette.ColorRole.Base,            QColor(18,  18,  18))
    p.setColor(QPalette.ColorRole.AlternateBase,   QColor(25,  25,  25))
    p.setColor(QPalette.ColorRole.ToolTipBase,     QColor(40,  40,  40))
    p.setColor(QPalette.ColorRole.ToolTipText,     QColor(220, 220, 220))
    p.setColor(QPalette.ColorRole.Text,            QColor(220, 220, 220))
    p.setColor(QPalette.ColorRole.Button,          QColor(45,  45,  45))
    p.setColor(QPalette.ColorRole.ButtonText,      QColor(220, 220, 220))
    p.setColor(QPalette.ColorRole.BrightText,      QColor(255, 80,  80))
    p.setColor(QPalette.ColorRole.Highlight,       QColor(60,  120, 200))
    p.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
    return p


def main() -> int:
    ap = argparse.ArgumentParser(description="Batch Labeled Editor — Qt Edition")
    ap.add_argument("input_folder",
                    help="Folder containing source images")
    ap.add_argument("--output", "-o", default=None,
                    help="Output folder (default: auto-named beside input)")
    ap.add_argument("--display-px", type=int, default=MAX_DISPLAY_PX,
                    help=f"Longest-edge cap for display copy "
                         f"(default {MAX_DISPLAY_PX})")
    args = ap.parse_args()

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setPalette(_dark_palette())
    app.setStyleSheet("""
        QToolTip {
            background: #2a2a2a; color: #ddd;
            border: 1px solid #555; padding: 4px;
        }
        QScrollBar { width: 0px; height: 0px; }
    """)

    try:
        win = MainWindow(args.input_folder, args.output, args.display_px)
        win.show()
        return app.exec()
    except KeyboardInterrupt:
        print("\n  Interrupted")
        return 130
    except Exception as exc:
        print(f"\n  Fatal: {exc}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
  
