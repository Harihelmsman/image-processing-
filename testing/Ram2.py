#!/usr/bin/env python3
"""
Batch Labeled Editor — Full Quality, Low Memory
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag

═══════════════════════════════════════════════════════════════
MEMORY MODEL
═══════════════════════════════════════════════════════════════
Steady-state = raw_image_size + ~12 MB overhead.
For a 50 MB JPEG (typical decode: 20–60 MB raw) → ~30–70 MB total.

Original image is kept at FULL RESOLUTION in RAM.
No quality reduction, no working-resolution downscale.

How memory stays low:
  1. VIEWPORT-FIRST RENDERING
     Every frame, only the VISIBLE CROP is processed:
       zoom 1:1 on 8 K image → crop ≈ 1440×900 pixels (3.9 MB)
       zoom-out (full image)  → crop = full image, but freed immediately
     The full-res buffer is NEVER copied for display purposes.

  2. IN-PLACE EFFECTS (boolean-mask indexing)
     Old:  np.where(mask, A, B)  → allocates a full output array every call
     New:  image[mb] = A[mb]     → scatter-write, masked pixels only (~KB)

  3. ROI-BASED LABEL GLOW
     Old:  image.copy() × 2 per label → 20 labels × 2 × 60 MB = 2.4 GB (!!)
     New:  roi.copy()  × 2 per label → 20 × 2 × ~0.02 MB = 0.8 MB

  4. SCREEN-SPACE LABELS
     Labels are drawn on the canvas-sized frame (3.9 MB), not the full
     image — so label overlay never touches the original buffer.

  5. SINGLE DISPLAY BUFFER
     display_image (3.9 MB) is the only persistent non-original array.
     It is reused every frame; _dirty=False skips re-render entirely.

  6. SAVE PATH
     Full-res effects applied in-place to original_image.copy() only
     during cv2.imwrite(), then freed. Peak during save: 2 × raw size.
═══════════════════════════════════════════════════════════════
"""

import gc
import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime
import time
import sys

FONT          = cv2.FONT_HERSHEY_DUPLEX
AA            = cv2.LINE_AA
CANVAS_W      = 1440
CANVAS_H      = 900
CANVAS_BG     = (18, 18, 18)
CANVAS_PAD    = 40
HQ_SETTLE_MS  = 180    # ms after last event before Lanczos quality fires


# ── Viewport ──────────────────────────────────────────────────────────────────

class Viewport:
    """Affine (scale, tx, ty) viewer with pre-allocated canvas buffer."""

    ZOOM_STEPS     = [0.125, 0.167, 0.25, 0.333, 0.5, 0.667,
                      1.0, 1.5, 2.0, 3.0, 4.0, 6.0, 8.0, 12.0, 16.0]
    SNAP_THRESHOLD = 0.04

    def __init__(self, cw: int, ch: int):
        self.cw = cw;  self.ch = ch
        self._scale = 1.0;  self._tx = 0.0;  self._ty = 0.0

    def fit(self, iw, ih, pad=CANVAS_PAD):
        self._scale = min((self.cw - 2*pad) / iw, (self.ch - 2*pad) / ih)
        self._snap();  self._recenter(iw, ih)

    def zoom_at(self, sx, sy, factor, iw, ih):
        ix = (sx - self._tx) / self._scale
        iy = (sy - self._ty) / self._scale
        min_s = min(0.05, (self.cw - 2*CANVAS_PAD) / max(iw, 1))
        self._scale = max(min_s, min(16.0, self._scale * factor))
        self._snap()
        self._tx = sx - ix * self._scale
        self._ty = sy - iy * self._scale

    def pan(self, dx, dy):  self._tx += dx;  self._ty += dy
    def reset(self, iw, ih): self.fit(iw, ih)

    def img_to_screen(self, ix, iy):
        return int(ix * self._scale + self._tx), int(iy * self._scale + self._ty)

    def screen_to_img(self, sx, sy):
        return (int((sx - self._tx) / self._scale),
                int((sy - self._ty) / self._scale))

    @property
    def scale(self): return self._scale

    def _snap(self):
        for s in self.ZOOM_STEPS:
            if abs(self._scale - s) / s < self.SNAP_THRESHOLD:
                self._scale = s;  return

    def _recenter(self, iw, ih):
        self._tx = (self.cw - iw * self._scale) / 2
        self._ty = (self.ch - ih * self._scale) / 2


# ── Edit modes ────────────────────────────────────────────────────────────────

class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


# ── Editor ────────────────────────────────────────────────────────────────────

class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50
    ZOOM_DEBOUNCE_MS        = 16
    MAX_CACHED_STATES       = 3

    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)
        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.total_images  = len(self.image_files)
        self.current_index = 0

        # Full-resolution original — read-only reference, never modified
        # scale_factor = 1.0 always (no downscaling)
        self.original_image = None
        self.display_image  = None     # canvas-sized rendered frame (3.9 MB)

        self.circles             = []
        self.drawing             = False
        self.center              = None
        self.current_radius      = 0
        self.current_label       = ""
        self.current_description = ""
        self.label_input_mode       = False
        self.description_input_mode = False

        # _dirty = True → re-render display_image next tick
        # _dirty = False → show existing display_image (idle, no work done)
        self._dirty          = True
        self._interacting    = False
        self._last_interact  = 0.0

        self.vp          = Viewport(CANVAS_W, CANVAS_H)
        self.is_panning  = False
        self._pan_last_x = self._pan_last_y = 0
        self.last_zoom_time = 0

        self.current_mode    = EditMode.HIGHLIGHT
        self._blur_kernel    = 25
        self._pixelate_size  = 10
        self.highlight_alpha = 0.4
        self.show_labels     = True

        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR:      (255, 80, 80),
            EditMode.PIXELATE:  (80, 80, 255),
            EditMode.DARKEN:    (160, 160, 160),
            EditMode.GRAYSCALE: (210, 210, 210),
            EditMode.INVERT:    (255, 255, 0),
            EditMode.OUTLINE:   (0, 220, 255),
        }

        self.saved_status       = {}
        self.image_states       = {}
        self.state_access_order = []

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name, CANVAS_W, CANVAS_H)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")
        self._print_instructions()

    # ── Properties ──────────────────────────────────────────────────────────

    @property
    def blur_kernel(self):  return self._blur_kernel
    @blur_kernel.setter
    def blur_kernel(self, v):
        if not isinstance(v, int) or v <= 0 or v % 2 == 0:
            raise ValueError("blur_kernel must be positive odd int")
        self._blur_kernel = v;  self._dirty = True

    @property
    def pixelate_size(self):  return self._pixelate_size
    @pixelate_size.setter
    def pixelate_size(self, v):
        if not isinstance(v, int) or v <= 0:
            raise ValueError("pixelate_size must be positive int")
        self._pixelate_size = v;  self._dirty = True

    # ── File loading ─────────────────────────────────────────────────────────

    def _load_image_files(self):
        exts  = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid = []
        for f in files:
            try:
                if f.stat().st_size < 512:
                    print(f"⚠  Skipping {f.name}: too small");  continue
                valid.append(f)
            except Exception as e:
                print(f"⚠  Skipping {f.name}: {e}")
        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images "
                  f"(recommended ≤ {self.MAX_BATCH_SIZE})")
            if input("   Continue? (y/N): ").strip().lower() != 'y':
                sys.exit(0)
        return valid

    # ── Image load — full resolution, no downscale ───────────────────────────

    def _load_current_image(self):
        if self.current_index >= len(self.image_files):
            return False
        f = self.image_files[self.current_index]
        try:
            img = cv2.imread(str(f))
            if img is None:
                raise IOError("cv2.imread returned None")
            if min(img.shape[:2]) < self.MIN_IMAGE_SIZE:
                raise ValueError("Image too small")

            # Free previous image before storing new one
            self.original_image = None;  gc.collect()
            self.original_image = img   # full-resolution, read-only reference

            ih, iw = self.original_image.shape[:2]
            self.vp.fit(iw, ih)

            # Restore saved annotations (coords in original-pixel space)
            json_path = self.output_folder / f.with_suffix('.json').name
            if json_path.exists():
                try:
                    with open(json_path) as jf:
                        data = json.load(jf)
                    self.circles = [{
                        'center':      tuple(o['center']),
                        'radius':      o['radius'],
                        'mode':        EditMode(o.get('mode', 'highlight')),
                        'label':       o.get('label', ''),
                        'description': o.get('description', ''),
                    } for o in data.get('objects', [])]
                    print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                          f" — restored {len(self.circles)} objects")
                except Exception as e:
                    print(f"⚠  JSON load: {e}");  self.circles = []
            elif f.name in self.image_states:
                self.circles = self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                      f" — from memory")
            else:
                self.circles = []
                print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")

            self._update_state_access(f.name)
            self.drawing = self.label_input_mode = self.description_input_mode = False
            self.current_label = self.current_description = ""
            self._dirty = True
            return True

        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index < len(self.image_files) - 1:
                self.current_index += 1
                return self._load_current_image()
            return False

    # ── LRU state cache ──────────────────────────────────────────────────────

    def _update_state_access(self, name):
        if name in self.state_access_order:
            self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order) > self.MAX_CACHED_STATES:
            for name in self.state_access_order[:-self.MAX_CACHED_STATES]:
                self.image_states.pop(name, None)
            self.state_access_order = self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── Mouse callback ───────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        ih, iw = self.original_image.shape[:2]

        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time() * 1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS:
                return
            self.last_zoom_time = now
            self.vp.zoom_at(x, y, 1.2 if flags > 0 else 0.8, iw, ih)
            self._mark_interacting();  self._dirty = True;  self._update_display()
            return

        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning = True
            self._pan_last_x = x;  self._pan_last_y = y;  return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False;  self._dirty = True;  return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.vp.pan(x - self._pan_last_x, y - self._pan_last_y)
            self._pan_last_x = x;  self._pan_last_y = y
            self._mark_interacting();  self._dirty = True;  self._update_display()
            return

        if self.label_input_mode or self.description_input_mode or self.is_panning:
            return

        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center  = self.vp.screen_to_img(x, y)
            self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            ix, iy = self.vp.screen_to_img(x, y)
            self.current_radius = int(np.hypot(ix - self.center[0],
                                               iy - self.center[1]))
            self._mark_interacting();  self._dirty = True;  self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles")
                self._enter_label_input_mode()

    def _mark_interacting(self):
        self._interacting = True;  self._last_interact = time.time() * 1000

    def _check_settle(self):
        """Returns True once after interaction stops (triggers HQ re-render)."""
        if not self._interacting:
            return False
        if (time.time() * 1000 - self._last_interact) >= HQ_SETTLE_MS:
            self._interacting = False;  return True
        return False

    # ── Label / description input ────────────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True;  self.current_label = ""
        print(f"\n→ [{self.current_mode.value.upper()}] "
              f"type label, ENTER confirm, ESC cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save):
        self.label_input_mode = False
        if save:
            self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled")
            self.current_label = "";  self.drawing = False
            self._dirty = True;  self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode = True;  self.current_description = ""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode = False
        lbl  = self.current_label.strip()
        desc = self.current_description.strip()
        self.circles.append({
            'center':      self.center,
            'radius':      self.current_radius,
            'mode':        self.current_mode,
            'label':       lbl,
            'description': desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}" + (" + description" if desc else "") +
              f" [{self.current_mode.value}]")
        self.current_label = self.current_description = ""
        self._dirty = True;  self._update_display()

    # ── Core render pipeline ─────────────────────────────────────────────────
    #
    # _render_frame() is the ONLY place that allocates temporary large arrays.
    # It works entirely in SCREEN SPACE:
    #
    #   1. Compute visible crop rectangle in image space
    #   2. Copy crop (ONLY when effects exist; otherwise use a zero-copy view)
    #   3. Apply effects in crop-pixel space — in-place, no temporaries
    #   4. Scale crop to canvas footprint (at most CANVAS_W × CANVAS_H)
    #   5. Blit scaled crop onto pre-allocated canvas
    #   6. Free crop copy immediately  ← memory spike lasts < 1 frame
    #   7. Draw circle outlines, labels, UI all in screen space (no large allocs)
    #
    # Peak RAM during step 2-6:
    #   original_image  (e.g. 60 MB)
    #   + crop copy      (max = original if fully zoomed out, freed after step 6)
    #   + blur temp      (one crop-sized temp for BLUR mode, freed after blur)
    #   + canvas frame   (3.9 MB, kept as display_image)
    #   ─────────────────────────────────────────────────────────────────────
    #   worst case = 2 × original + 3.9 MB  (only during zoomed-out blur)
    #   typical    = original + crop_portion + 3.9 MB  ≈ original + ~5 MB

    def _render_frame(self, fast: bool = False,
                      ghost: dict = None) -> np.ndarray:
        """
        Render the full display frame from scratch.
        Allocations bounded to: one visible crop copy + one canvas.
        """
        ih, iw = self.original_image.shape[:2]
        sc = self.vp._scale;  tx = self.vp._tx;  ty = self.vp._ty

        # ── Allocate canvas ──────────────────────────────────────────────────
        canvas = np.full((CANVAS_H, CANVAS_W, 3), CANVAS_BG, dtype=np.uint8)

        # ── Visible source rectangle in image space ──────────────────────────
        sx0 = max(0.0, -tx / sc);               sy0 = max(0.0, -ty / sc)
        sx1 = min(float(iw), (CANVAS_W - tx) / sc)
        sy1 = min(float(ih), (CANVAS_H - ty) / sc)
        if sx1 <= sx0 or sy1 <= sy0:
            return canvas

        cx0 = max(0,  int(np.floor(sx0)));  cy0 = max(0,  int(np.floor(sy0)))
        cx1 = min(iw, int(np.ceil(sx1)));   cy1 = min(ih, int(np.ceil(sy1)))
        dst_w = max(1, int(round((cx1 - cx0) * sc)))
        dst_h = max(1, int(round((cy1 - cy0) * sc)))

        # ── Find circles that overlap the visible crop ───────────────────────
        vis = [c for c in self.circles
               if (c['center'][0] + c['radius'] > cx0 and
                   c['center'][0] - c['radius'] < cx1 and
                   c['center'][1] + c['radius'] > cy0 and
                   c['center'][1] - c['radius'] < cy1)]

        # ── Crop + effects ───────────────────────────────────────────────────
        if vis:
            # Copy only the visible crop — at most CANVAS_W × CANVAS_H at 1:1
            crop = self.original_image[cy0:cy1, cx0:cx1].copy()
            for c in vis:
                # Translate circle coords to crop-local space
                self._apply_effect_inplace(crop, {
                    'center': (c['center'][0] - cx0, c['center'][1] - cy0),
                    'radius': c['radius'],
                    'mode':   c['mode'],
                    'label':  c['label'],
                    'description': c.get('description', ''),
                })
            interp = (cv2.INTER_LINEAR if fast else
                      (cv2.INTER_LANCZOS4 if sc >= 1.0 else cv2.INTER_AREA))
            scaled = cv2.resize(crop, (dst_w, dst_h), interpolation=interp)
            del crop   # ← free immediately; only canvas-sized 'scaled' remains
        else:
            view   = self.original_image[cy0:cy1, cx0:cx1]   # zero-copy NumPy view
            interp = (cv2.INTER_LINEAR if fast else
                      (cv2.INTER_LANCZOS4 if sc >= 1.0 else cv2.INTER_AREA))
            scaled = cv2.resize(view, (dst_w, dst_h), interpolation=interp)

        # Unsharp mask — sharpens after upscale (zoom-in only, not during drag)
        if not fast and sc >= 1.0:
            blur   = cv2.GaussianBlur(scaled, (0, 0), sigmaX=1.0)
            scaled = cv2.addWeighted(scaled, 1.4, blur, -0.4, 0)

        # ── Blit onto canvas ─────────────────────────────────────────────────
        dx = max(0, int(tx + cx0 * sc));  dy = max(0, int(ty + cy0 * sc))
        pw = min(dst_w, CANVAS_W - dx);   ph = min(dst_h, CANVAS_H - dy)
        if pw > 0 and ph > 0:
            canvas[dy:dy+ph, dx:dx+pw] = scaled[:ph, :pw]
        del scaled   # ← free canvas-sized temp

        # Image border (1 px on dark surround)
        zw = int(iw * sc);  zh = int(ih * sc)
        itx = int(tx);      ity = int(ty)
        cv2.rectangle(canvas,
                      (max(0, itx-1),           max(0, ity-1)),
                      (min(CANVAS_W-1, itx+zw), min(CANVAS_H-1, ity+zh)),
                      (45, 45, 45), 1, AA)

        # ── Circle outlines in screen space (2 px regardless of zoom) ────────
        for c in self.circles:
            sx, sy = self.vp.img_to_screen(c['center'][0], c['center'][1])
            sr     = max(1, int(round(c['radius'] * sc)))
            if -sr-2 <= sx <= CANVAS_W+sr and -sr-2 <= sy <= CANVAS_H+sr:
                cv2.circle(canvas, (sx, sy), sr,
                           self.mode_colors[c['mode']], 2, AA)

        # ── Ghost circle (while drawing) ─────────────────────────────────────
        if ghost:
            gsx, gsy = self.vp.img_to_screen(ghost['center'][0], ghost['center'][1])
            gsr = max(1, int(round(ghost['radius'] * sc)))
            cv2.circle(canvas, (gsx, gsy), gsr,
                       self.mode_colors[self.current_mode], 2, AA)

        # ── Labels in screen space (no full-image allocation) ────────────────
        if self.show_labels and self.circles:
            self._draw_labels_screen(canvas)

        return canvas

    # ── In-place effects (no np.where, no np.full_like) ──────────────────────

    def _apply_effect_inplace(self, image: np.ndarray, circle: dict) -> None:
        """
        Mutate image pixels within the circle using boolean-mask scatter-writes.
        Largest allocation: one boolean mask (H×W uint8) + one mode-specific temp.
        """
        mask = np.zeros(image.shape[:2], dtype=np.uint8)
        cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
        mb = mask.astype(bool)
        if not mb.any():
            return
        m = circle['mode']
        try:
            if m == EditMode.HIGHLIGHT:
                a = self.highlight_alpha
                # Float math only on the masked pixel subset (~KB, not full image)
                pix = image[mb].astype(np.float32)
                image[mb] = np.clip(pix * (1.0 - a) + 255.0 * a,
                                    0, 255).astype(np.uint8)

            elif m == EditMode.BLUR:
                blurred   = cv2.GaussianBlur(image, (self._blur_kernel,)*2, 0)
                image[mb] = blurred[mb]   # blurred is one full-crop temp, freed after

            elif m == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                small = cv2.resize(image,
                                   (max(1, w // self._pixelate_size),
                                    max(1, h // self._pixelate_size)),
                                   interpolation=cv2.INTER_NEAREST)
                px    = cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST)
                image[mb] = px[mb]

            elif m == EditMode.DARKEN:
                # Right-shift = halve brightness; uint16 prevents overflow
                image[mb] = (image[mb].astype(np.uint16) >> 1).astype(np.uint8)

            elif m == EditMode.GRAYSCALE:
                gray = cv2.cvtColor(
                    cv2.cvtColor(image, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                image[mb] = gray[mb]

            elif m == EditMode.INVERT:
                image[mb] = 255 - image[mb]   # fully in-place, zero extra alloc

        except Exception as e:
            print(f"⚠  Effect error ({circle['mode'].value}): {e}")

    # ── Screen-space label drawing ────────────────────────────────────────────
    #
    # All coordinates are in canvas pixels (≤ 1440×900).
    # Glow uses tiny ROI crops — never touches the original image buffer.

    def _rects_overlap(self, r1, r2, buf=4):
        return not (r1[2]+buf < r2[0] or r1[0]-buf > r2[2] or
                    r1[3]+buf < r2[1] or r1[1]-buf > r2[3])

    def _find_label_pos_screen(self, cx, cy, sr, tw, fh, bl, pad, placed):
        """Find a non-overlapping label position in canvas coordinates."""
        W = tw + 2*pad;  H = fh + bl + 2*pad
        cands = [
            (cx-sr,     cy-sr-H-10),   (cx-sr,   cy+sr+10),
            (cx-sr-W-8, cy-H//2),      (cx+sr+8, cy-H//2),
            (cx+sr+8,   cy-sr-H-10),   (cx-sr-W-8, cy-sr-H-10),
            (cx+sr+8,   cy+sr+10),     (cx-sr-W-8, cy+sr+10),
        ]
        for extra in (40, 80, 120):
            cands += [(px, py+extra) for px, py in cands[:4]]
            cands += [(px, py-extra) for px, py in cands[:4]]
        for px, py in cands:
            rect = (int(px), int(py), int(px+W), int(py+H))
            if (rect[0] < pad or rect[1] < pad or
                    rect[2] > CANVAS_W-pad or rect[3] > CANVAS_H-pad):
                continue
            if not any(self._rects_overlap(rect, r) for r in placed):
                return int(px)+pad, int(py)+fh+pad, rect
        bottom = max((r[3] for r in placed), default=0) + 8
        lx = max(pad, min(cx - W//2, CANVAS_W - W - pad))
        ly = min(bottom + fh + pad, CANVAS_H - bl - pad)
        return lx, ly, (lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad)

    def _draw_labels_screen(self, canvas: np.ndarray) -> None:
        """
        Draw all badge numbers + label tags onto canvas in screen coordinates.
        Font size is fixed — labels are always readable regardless of zoom.
        ROI glow: allocates only the label bounding box (~KB per label).
        """
        placed = []
        sc_f = 0.55;  th_f = 1   # fixed screen-space font scale

        for i, c in enumerate(self.circles, 1):
            label = c['label'] or f"Error #{i}"
            color = self.mode_colors[c['mode']]

            # Screen-space circle position
            cx, cy = self.vp.img_to_screen(c['center'][0], c['center'][1])
            sr     = max(1, int(round(c['radius'] * self.vp.scale)))

            # Skip completely off-screen circles (with generous margin for labels)
            if cx + sr + 220 < 0 or cx - sr - 220 > CANVAS_W:
                continue
            if cy + sr + 80  < 0 or cy - sr - 80  > CANVAS_H:
                continue

            # ── Badge number centred on circle ────────────────────────────
            bt = str(i)
            (bw, bh), _ = cv2.getTextSize(bt, FONT, sc_f * 0.9, th_f)
            br = max(bw, bh) // 2 + 6
            # Clamp badge centre to canvas so it stays visible even if circle
            # extends off-screen
            bcx = max(br+1, min(CANVAS_W-br-1, cx))
            bcy = max(br+1, min(CANVAS_H-br-1, cy))
            cv2.circle(canvas, (bcx+2, bcy+2), br, (0,0,0),        -1, AA)
            cv2.circle(canvas, (bcx,   bcy),   br, color,           -1, AA)
            cv2.circle(canvas, (bcx,   bcy),   br, (255,255,255),    2, AA)
            cv2.putText(canvas, bt, (bcx - bw//2, bcy + bh//2),
                        FONT, sc_f * 0.9, (0,0,0), th_f+1, AA)

            # ── Label tag ─────────────────────────────────────────────────
            text = f"#{i}  [{c['mode'].value[:3].upper()}]  {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc_f, th_f)
            pad = 6

            lx, ly, rect = self._find_label_pos_screen(
                cx, cy, sr, tw, fh, bl, pad, placed)
            placed.append(rect)
            x1, y1, x2, y2 = rect

            # ROI glow — only allocates the label's bounding box (~KB, not MB)
            for ge, ga in ((6, 0.25), (3, 0.45)):
                gx1 = max(0, x1-ge);    gy1 = max(0, y1-ge)
                gx2 = min(CANVAS_W, x2+ge); gy2 = min(CANVAS_H, y2+ge)
                if gx2 <= gx1 or gy2 <= gy1:
                    continue
                roi      = canvas[gy1:gy2, gx1:gx2]
                roi_copy = roi.copy()                   # tiny (~KB) allocation
                cv2.rectangle(roi_copy,
                              (x1-gx1, y1-gy1), (x2-gx1, y2-gy1),
                              color, -1, AA)
                cv2.addWeighted(roi_copy, ga, roi, 1.0-ga, 0, roi)

            cv2.rectangle(canvas, (x1, y1), (x2, y2), (15,15,20), -1)
            cv2.rectangle(canvas, (x1, y1), (x1+3, y2), color, -1, AA)
            cv2.rectangle(canvas, (x1, y1), (x2, y2), color, 2, AA)

            tx_pos = lx + 4
            cv2.putText(canvas, text, (tx_pos+1, ly+1), FONT, sc_f,
                        (0,0,0), th_f+1, AA)
            cv2.putText(canvas, text, (tx_pos, ly), FONT, sc_f,
                        (255,255,255), th_f, AA)

            # Connector line from label to circle edge
            mx  = (x1 + x2) // 2
            my  = y2 if ly > cy else y1
            ang = np.arctan2(my - cy, mx - cx)
            ex  = int(cx + sr * np.cos(ang))
            ey  = int(cy + sr * np.sin(ang))
            cv2.line(canvas, (mx, my), (ex, ey), color, 2, AA)
            cv2.circle(canvas, (ex, ey), 4, color, -1, AA)
            cv2.circle(canvas, (ex, ey), 4, (255,255,255), 1, AA)

    # ── Typing label overlay (during label/description input) ────────────────

    def _draw_typing_label_screen(self, canvas, label, mode):
        """Draw the live-typing indicator near the circle being annotated."""
        if self.center is None:
            return
        sx, sy = self.vp.img_to_screen(self.center[0], self.center[1])
        sr     = max(1, int(round(self.current_radius * self.vp.scale)))

        color = self.mode_colors[mode]
        text  = f"[{mode.value[:3].upper()}] {label}_"
        sc_f  = 0.62;  th_f = 2
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc_f, th_f)
        pad = 8

        # Prefer above, fall back to below
        candidates = [
            (sx - sr,      sy - sr - fh - 2*pad - 10),
            (sx - sr,      sy + sr + 10),
            (max(pad, sx - tw//2), max(fh+pad, sy - sr - fh - 2*pad - 30)),
        ]
        lx = ly = None
        for px, py in candidates:
            x1 = int(px) - pad;  y1 = int(py) - fh - pad
            x2 = int(px) + tw + pad;  y2 = int(py) + bl + pad
            if x1 >= 0 and y1 >= 0 and x2 <= CANVAS_W and y2 <= CANVAS_H:
                lx, ly = int(px), int(py)
                break
        if lx is None:
            lx = max(pad, min(sx, CANVAS_W - tw - 2*pad))
            ly = max(fh + pad, sy - sr - fh - 2*pad - 10)

        x1, y1 = lx - pad, ly - fh - pad
        x2, y2 = lx + tw + pad, ly + bl + pad
        cv2.rectangle(canvas, (x1, y1), (x2, y2), (10,10,10), -1)
        cv2.rectangle(canvas, (x1, y1), (x2, y2), color, 2, AA)
        cv2.putText(canvas, text, (lx, ly), FONT, sc_f, (255,255,255), th_f, AA)
        cv2.line(canvas, (lx + tw//2, y2), (sx, sy), color, 2, AA)

    # ── Display update ───────────────────────────────────────────────────────

    def _present(self):
        self._dirty = False
        cv2.imshow(self.window_name, self.display_image)

    def _update_display(self):
        """
        Re-render only when dirty. Idle ticks (nothing changed) skip entirely.
        fast=True during interaction (INTER_LINEAR); Lanczos fires after settle.
        """
        if not self._dirty:
            return
        if self._check_settle():
            self._dirty = True

        ghost = None
        if self.drawing and self.current_radius > 0:
            ghost = {'center': self.center, 'radius': self.current_radius}

        self.display_image = self._render_frame(
            fast=self._interacting, ghost=ghost)
        self._draw_ui()
        self._present()

    def _update_display_with_input(self):
        ghost = {'center': self.center, 'radius': self.current_radius}
        self.display_image = self._render_frame(fast=False, ghost=ghost)
        self._draw_typing_label_screen(
            self.display_image, self.current_label, self.current_mode)
        self._draw_input_box("Label")
        self._present()

    def _update_display_with_description_input(self):
        ghost = {'center': self.center, 'radius': self.current_radius}
        self.display_image = self._render_frame(fast=False, ghost=ghost)
        if self.current_label:
            self._draw_typing_label_screen(
                self.display_image, self.current_label, self.current_mode)
        self._draw_input_box("Description (optional)")
        self._present()

    # ── UI overlays ──────────────────────────────────────────────────────────

    def _draw_input_box(self, prompt):
        h, w  = self.display_image.shape[:2]
        bh = 64;  by = h - bh
        color = self.mode_colors[self.current_mode]
        ov = self.display_image.copy()
        cv2.rectangle(ov, (0, by), (w, h), (12,12,12), -1)
        cv2.addWeighted(ov, 0.88, self.display_image, 0.12, 0, self.display_image)
        cv2.rectangle(self.display_image, (0, by), (w, h), color, 2, AA)
        cv2.putText(self.display_image,
                    f"[{self.current_mode.value.upper()}]",
                    (12, by+22), FONT, 0.52, color, 1, AA)
        cv2.putText(self.display_image, f"{prompt}:",
                    (160, by+22), FONT, 0.52, (200,200,200), 1, AA)
        text = (self.current_description if self.description_input_mode
                else self.current_label) + "_"
        cv2.putText(self.display_image, text,
                    (12, by+50), FONT, 0.65, (0,240,240), 1, AA)

    def _draw_ui(self):
        ov = self.display_image.copy()
        h, w = self.display_image.shape[:2]
        cv2.rectangle(ov, (0, 0), (w, 84), (12,12,12), -1)
        color = self.mode_colors[self.current_mode]
        cf    = self.image_files[self.current_index]
        ih, iw = self.original_image.shape[:2]

        cv2.putText(ov, f"Mode: {self.current_mode.value.upper()}",
                    (14,24), FONT, 0.55, color, 1, AA)
        cv2.putText(ov, f"Zoom: {self.vp.scale:.2f}x",
                    (14,48), FONT, 0.45, (100,190,255), 1, AA)
        cv2.putText(ov, f"{self.current_index+1} / {self.total_images}",
                    (14,70), FONT, 0.43, (160,160,160), 1, AA)

        fname = cf.name if len(cf.name) <= 42 else cf.name[:39]+"..."
        cv2.putText(ov, fname, (200,24), FONT, 0.46, (185,185,185), 1, AA)

        # Image resolution (full-quality indicator)
        cv2.putText(ov, f"{iw}×{ih}  FULL RES",
                    (200,48), FONT, 0.38, (80,200,80), 1, AA)

        n  = len(self.circles)
        oc = (40,160,255) if n >= self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(ov, f"Objects: {n}", (200,70), FONT, 0.43, oc, 1, AA)

        qi = "FAST" if self._interacting else "HQ"
        qc = (60,180,255) if self._interacting else (60,220,100)
        cv2.putText(ov, qi, (390,70), FONT, 0.38, qc, 1, AA)

        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = cf.name in self.image_states or n > 0
        st, sc = (("SAVED",    (60,200,60))  if is_saved  else
                  ("EDITED",   (40,160,255)) if has_edits else
                  ("NO EDITS", (80,80,80)))
        cv2.putText(ov, st, (w-155,24), FONT, 0.54, sc, 1, AA)

        hint = ("Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  "
                "R:Reset  C:Clear  U:Undo  H:Help  Q:Quit")
        cv2.putText(ov, hint, (12, h-10), FONT, 0.36, (110,110,110), 1, AA)
        cv2.addWeighted(ov, 0.82, self.display_image, 0.18, 0, self.display_image)

    # ── Navigation ───────────────────────────────────────────────────────────

    def _guard_state(self):
        if self.drawing:
            print("⚠  Finish drawing first");             return True
        if self.label_input_mode:
            print("⚠  Finish label input (ESC to cancel)"); return True
        if self.description_input_mode:
            print("⚠  Finish description (ENTER to skip)"); return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index > 0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index -= 1
            self._load_current_image();  self._dirty = True;  self._update_display()
        else:
            print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index < self.total_images - 1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index += 1
            self._load_current_image();  self._dirty = True;  self._update_display()
        else:
            print("Already at last image")

    # ── Misc ─────────────────────────────────────────────────────────────────

    def _list_labels(self):
        print("\n" + "="*60)
        print(f"Objects — {self.image_files[self.current_index].name}")
        print("="*60)
        if not self.circles:
            print("  (none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  "
                  f"r={c['radius']}  pos={c['center']}")
            if c.get('description'):
                print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        raw_mb = (self.original_image.nbytes // 1024 // 1024
                  if self.original_image is not None else 0)
        disp_mb = (self.display_image.nbytes // 1024 // 1024
                   if self.display_image is not None else 0)
        ih, iw = self.original_image.shape[:2]
        print(f"\nMemory — original: {raw_mb} MB ({iw}×{ih} full-res)  |  "
              f"display: {disp_mb} MB  |  "
              f"states cached: {len(self.image_states)}/{self.MAX_CACHED_STATES}  |  "
              f"saved: {len(self.saved_status)}\n")

    def _edit_last_label(self):
        if not self.circles:
            print("No objects to edit");  return
        last = self.circles[-1]

        # Temporarily set center/radius so typing overlay renders correctly
        self.center         = last['center']
        self.current_radius = last['radius']

        print(f"\nEditing label: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label    = last['label']
        self.label_input_mode = True
        self._update_display_with_input()

        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:
                self.label_input_mode = False;  self.current_label = ""
                self._dirty = True;  self._update_display();  return
            elif key == 13:
                last['label'] = self.current_label.strip()
                self.label_input_mode = False;  break
            elif key == 8:
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key);  self._update_display_with_input()

        print(f"\nEditing description: '{last.get('description','')}'  "
              f"(ENTER=save, ESC=skip)")
        self.current_description    = last.get('description', '')
        self.description_input_mode = True
        self._update_display_with_description_input()

        while self.description_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 13:
                last['description'] = self.current_description.strip()
                self.description_input_mode = False;  break
            elif key == 27:
                self.description_input_mode = False;  break
            elif key == 8:
                self.current_description = self.current_description[:-1]
                self._update_display_with_description_input()
            elif 32 <= key <= 126:
                self.current_description += chr(key)
                self._update_display_with_description_input()

        self.current_label = self.current_description = ""
        self.center = self.current_radius = None
        print("  ✓ Updated")
        self._dirty = True;  self._update_display()

    # ── Save ──────────────────────────────────────────────────────────────────

    def _draw_all_labels_full_res(self, image: np.ndarray) -> None:
        """
        Draw labels on the full-res output image using image-space coords.
        Called only during save. ROI glow — no full image.copy().
        """
        ih, iw = image.shape[:2]
        md = min(ih, iw)
        sc, th = (0.45, 1) if md < 400 else (0.58, 1) if md < 900 else (0.72, 2)
        placed = []

        for i, c in enumerate(self.circles, 1):
            label = c['label'] or f"Error #{i}"
            color = self.mode_colors[c['mode']]
            cx, cy = c['center'];  r = c['radius']

            # Badge
            bt = str(i)
            (bw, bh), _ = cv2.getTextSize(bt, FONT, sc * 0.9, th)
            br = max(bw, bh) // 2 + 6
            cv2.circle(image, (cx+2, cy+2), br, (0,0,0), -1, AA)
            cv2.circle(image, (cx, cy),     br, color,   -1, AA)
            cv2.circle(image, (cx, cy),     br, (255,255,255), 2, AA)
            cv2.putText(image, bt, (cx - bw//2, cy + bh//2),
                        FONT, sc * 0.9, (0,0,0), th+1, AA)

            # Label tag
            text = f"#{i}  [{c['mode'].value[:3].upper()}]  {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
            pad = max(5, int(5 * sc / 0.5))

            W = tw + 2*pad;  H = fh + bl + 2*pad
            cands = [
                (cx-r, cy-r-H-10), (cx-r, cy+r+10),
                (cx-r-W-8, cy-H//2), (cx+r+8, cy-H//2),
            ]
            for extra in (40, 80, 120):
                cands += [(px, py+extra) for px, py in cands[:4]]
                cands += [(px, py-extra) for px, py in cands[:4]]
            lx = ly = None
            for px, py in cands:
                rect = (int(px), int(py), int(px+W), int(py+H))
                if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad:
                    continue
                if not any(self._rects_overlap(rect, rr) for rr in placed):
                    lx, ly = int(px)+pad, int(py)+fh+pad
                    placed.append(rect);  break
            if lx is None:
                bottom = max((rr[3] for rr in placed), default=0) + 8
                lx = max(pad, min(cx - W//2, iw-W-pad))
                ly = min(bottom + fh + pad, ih-bl-pad)
                placed.append((lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad))

            x1, y1 = lx-pad, ly-fh-pad
            x2, y2 = lx+tw+pad, ly+bl+pad

            # ROI glow
            for ge, ga in ((6, 0.25), (3, 0.45)):
                gx1 = max(0, x1-ge);   gy1 = max(0, y1-ge)
                gx2 = min(iw, x2+ge);  gy2 = min(ih, y2+ge)
                if gx2 <= gx1 or gy2 <= gy1: continue
                roi = image[gy1:gy2, gx1:gx2]
                roi_copy = roi.copy()
                cv2.rectangle(roi_copy, (x1-gx1, y1-gy1), (x2-gx1, y2-gy1), color, -1, AA)
                cv2.addWeighted(roi_copy, ga, roi, 1.0-ga, 0, roi)

            cv2.rectangle(image, (x1, y1), (x2, y2), (15,15,20), -1)
            cv2.rectangle(image, (x1, y1), (x1+3, y2), color, -1, AA)
            cv2.rectangle(image, (x1, y1), (x2, y2), color, 2, AA)
            tx_pos = lx + 4
            cv2.putText(image, text, (tx_pos+1, ly+1), FONT, sc, (0,0,0), th+1, AA)
            cv2.putText(image, text, (tx_pos, ly), FONT, sc, (255,255,255), th, AA)

            mx  = (x1+x2)//2;  my = y2 if ly > cy else y1
            ang = np.arctan2(my-cy, mx-cx)
            ex  = int(cx + r*np.cos(ang));  ey = int(cy + r*np.sin(ang))
            cv2.line(image, (mx, my), (ex, ey), color, 2, AA)
            cv2.circle(image, (ex, ey), 4, color, -1, AA)
            cv2.circle(image, (ex, ey), 4, (255,255,255), 1, AA)

    def save_current(self, auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save")
            return

        cf  = self.image_files[self.current_index]
        out = self.output_folder / cf.name
        js  = out.with_suffix('.json')

        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._update_state_access(cf.name)

        try:
            # Make one full-res copy for output (freed after imwrite)
            final = self.original_image.copy()

            # Apply effects at full resolution (in-place, no extra allocations)
            for c in self.circles:
                self._apply_effect_inplace(final, c)
                cv2.circle(final, c['center'], c['radius'],
                           self.mode_colors[c['mode']], 3, AA)

            # Draw labels at full resolution
            self._draw_all_labels_full_res(final)

            if not cv2.imwrite(str(out), final):
                raise IOError(f"cv2.imwrite failed for {out}")

            del final;  gc.collect()   # ← free full-res copy immediately after save

            # JSON: coords are already in original-pixel space
            data = {
                'source_image': cf.name,
                'timestamp':    datetime.now().isoformat(),
                'objects': [{
                    'id':          i,
                    'label':       c['label'],
                    'description': c.get('description', ''),
                    'mode':        c['mode'].value,
                    'center':      list(c['center']),
                    'radius':      c['radius'],
                } for i, c in enumerate(self.circles, 1)]
            }
            with open(js, 'w') as jf:
                json.dump(data, jf, indent=2)

            self.saved_status[cf.name] = True
            if auto_save:
                print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else:
                print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)"
                      f"\n  → {out}\n  → {js}")
            self._cleanup_old_states()

        except Exception as e:
            print(f"❌ Save error: {e}")

    # ── Excel summary ─────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self.output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook();  ws = wb.active;  ws.title = "Summary"
            hfill  = PatternFill(start_color="2E4DA7", end_color="2E4DA7",
                                 fill_type="solid")
            hfont  = Font(bold=True, color="FFFFFF", size=11)
            border = Border(**{s: Side(style='thin')
                             for s in ('left','right','top','bottom')})
            ca  = Alignment(horizontal='center', vertical='center', wrap_text=True)
            tla = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

            def styled(cell, fill=None, font=None, align=None):
                if fill:  cell.fill   = fill
                if font:  cell.font   = font
                if align: cell.alignment = align
                cell.border = border

            for col, hdr in enumerate(
                    ["#", "Image Name", "Error Count", "Errors (numbered)"], 1):
                styled(ws.cell(row=1, column=col, value=hdr),
                       fill=hfill, font=hfont, align=ca)
            ws.row_dimensions[1].height = 28

            data_row = 2;  img_num = 0
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                jp = self.output_folder / img.with_suffix('.json').name
                if not jp.exists(): continue
                with open(jp) as jf: data = json.load(jf)
                objects  = data.get('objects', [])
                img_num += 1

                c1 = ws.cell(row=data_row, column=1, value=img_num)
                styled(c1, align=ca);  c1.font = Font(bold=True, size=10)
                styled(ws.cell(row=data_row, column=2, value=img.name), align=tla)

                c3 = ws.cell(row=data_row, column=3, value=len(objects))
                styled(c3, align=ca)
                if   len(objects) == 0:
                    c3.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                         fill_type="solid")
                    c3.font = Font(color="276221", bold=True)
                elif len(objects) <= 3:
                    c3.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                                         fill_type="solid")
                    c3.font = Font(color="9C5700", bold=True)
                else:
                    c3.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                         fill_type="solid")
                    c3.font = Font(color="9C0006", bold=True)

                cell_text = "\n".join(
                    f"{k}. {o.get('label','').strip() or '(unlabeled)'}"
                    + (f" — {o['description'].strip()}"
                       if o.get('description','').strip() else "")
                    for k, o in enumerate(objects, 1)
                ) if objects else "(no errors)"
                styled(ws.cell(row=data_row, column=4, value=cell_text), align=tla)
                ws.row_dimensions[data_row].height = max(20, len(objects)*15+6)

                if img_num % 2 == 0:
                    alt = PatternFill(start_color="F2F5FB", end_color="F2F5FB",
                                     fill_type="solid")
                    for col in (1, 2, 4):
                        ws.cell(row=data_row, column=col).fill = alt
                data_row += 1

            data_row += 1
            sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                fill_type="solid")
            sfont = Font(bold=True, size=11)

            def summary_row(label, value):
                nonlocal data_row
                lc = ws.cell(row=data_row, column=2, value=label)
                lc.font = sfont;  lc.fill = sfill;  lc.border = border
                vc = ws.cell(row=data_row, column=3, value=value)
                vc.font = sfont;  vc.fill = sfill;  vc.border = border
                vc.alignment = ca
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            summary_row("Total Images Processed", len(self.saved_status))
            total_obj = sum(
                len(json.load(open(
                    self.output_folder / f.with_suffix('.json').name))['objects'])
                for f in self.image_files
                if f.name in self.saved_status
                and (self.output_folder / f.with_suffix('.json').name).exists())
            summary_row("Total Errors Found", total_obj)
            if self.saved_status:
                summary_row("Avg Errors / Image",
                            round(total_obj / len(self.saved_status), 1))

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 34
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 62
            ws.freeze_panes = "A2"
            wb.save(str(xp))
            print(f"✓ Excel summary: {xp}")
        except ImportError:
            print("⚠  pip install openpyxl")
        except Exception as e:
            print(f"⚠  Excel error: {e}")

    # ── Instructions ─────────────────────────────────────────────────────────

    def _print_instructions(self):
        ih, iw = self.original_image.shape[:2]
        raw_mb = self.original_image.nbytes // 1024 // 1024
        print("\n" + "="*62)
        print("BATCH LABELED EDITOR  —  Full Quality Mode")
        print("="*62)
        print(f"  Input : {self.input_folder}")
        print(f"  Output: {self.output_folder}")
        print(f"  Images: {self.total_images}"
              f"  |  {iw}×{ih}  ({raw_mb} MB raw)")
        print()
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next image")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo last circle")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*62+"\n")

    # ── Main loop ─────────────────────────────────────────────────────────────

    def run(self):
        self._dirty = True
        self._update_display()
        while True:
            key = cv2.waitKey(1) & 0xFF

            # HQ settle tick
            if key == 0xFF:
                if self._check_settle():
                    self._dirty = True;  self._update_display()
                elif not self._dirty:
                    continue

            if self.description_input_mode:
                if key == 27:
                    self.description_input_mode = False
                    self.current_label = self.current_description = ""
                    print("  ✗ Circle discarded")
                    self._dirty = True;  self._update_display()
                elif key == 13:
                    self._exit_description_input_mode()
                elif key == 8:
                    self.current_description = self.current_description[:-1]
                    self._update_display_with_description_input()
                elif 32 <= key <= 126:
                    self.current_description += chr(key)
                    self._update_display_with_description_input()
                continue

            if self.label_input_mode:
                if key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13: self._exit_label_input_mode(save=True)
                elif key == 8:
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key);  self._update_display_with_input()
                continue

            if   key == ord('a'): self._previous_image()
            elif key == ord('d'): self._next_image()
            elif key == ord('r'):
                ih, iw = self.original_image.shape[:2]
                self.vp.reset(iw, ih)
                self._dirty = True;  self._update_display();  print("✓ Zoom reset")
            elif key == ord('s'):
                self.save_current();  self._dirty = True;  self._update_display()
            elif key == ord('S'):
                self.save_current();  self._next_image()
            elif key == ord('c'):
                self.circles.clear()
                self._dirty = True;  self._update_display();  print("✓ Cleared")
            elif key == ord('u'):
                if self.circles:
                    rem = self.circles.pop()
                    print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._dirty = True;  self._update_display()
            elif key == ord('l'): self._list_labels()
            elif key == ord('e'): self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}")
                self._dirty = True;  self._update_display()
            elif key == ord('m'): self._show_memory_status()
            elif key in (ord('h'), ord('H')): self._print_instructions()
            elif ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._dirty = True;  self._update_display()
            elif key == ord('q'):
                if self.circles and not self.saved_status.get(
                        self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved")
            print(f"   Output: {self.output_folder}")
        else:
            print("\n⚠  No images were saved")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Batch Labeled Editor — Full Quality")
    ap.add_argument("input_folder")
    ap.add_argument("--output", "-o", default=None)
    args = ap.parse_args()
    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n⚠  Interrupted");  return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback;  traceback.print_exc();  return 1
    return 0


if __name__ == "__main__":
    exit(main())
