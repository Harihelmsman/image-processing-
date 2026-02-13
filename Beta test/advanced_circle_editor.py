#!/usr/bin/env python3
"""
Batch Advanced Labeled Editor - Process entire folders of images
Navigate with A (previous) and D (next) keys
Zoom with mouse wheel
Saves output to separate folder with labels visible
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime


class EditMode(Enum):
    """Available editing modes"""
    HIGHLIGHT = "highlight"
    BLUR = "blur"
    PIXELATE = "pixelate"
    DARKEN = "darken"
    GRAYSCALE = "grayscale"
    INVERT = "invert"
    OUTLINE = "outline"


class BatchLabeledEditor:
    """Batch editor for processing multiple images in a folder"""
    
    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)
        
        # Setup output folder
        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{timestamp}"
        
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # Load all image files
        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No images found in {input_folder}")
        
        self.current_index = 0
        self.total_images = len(self.image_files)
        
        # Editor state
        self.original_image = None
        self.scaled_image = None
        self.display_image = None
        self.output_image = None
        self.circles = []
        self.drawing = False
        self.center = None
        self.current_radius = 0
        self.current_label = ""
        self.label_input_mode = False
        
        # Zoom and pan state
        self.zoom_level = 1.0
        self.min_zoom = 0.7
        self.max_zoom = 7.0
        self.pan_x = 0  # Integer pan offset
        self.pan_y = 0  # Integer pan offset
        self.is_panning = False
        self.pan_start_x = 0
        self.pan_start_y = 0
        
        # Current editing mode
        self.current_mode = EditMode.HIGHLIGHT
        
        # Effect parameters
        self.blur_kernel = 25
        self.pixelate_size = 10
        self.highlight_alpha = 0.4
        
        # Label settings
        self.label_font = cv2.FONT_HERSHEY_SIMPLEX
        self.label_scale = 1.0
        self.label_thickness = 2
        self.show_labels = True
        
        # Mode colors
        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR: (255, 0, 0),
            EditMode.PIXELATE: (0, 0, 255),
            EditMode.DARKEN: (128, 128, 128),
            EditMode.GRAYSCALE: (200, 200, 200),
            EditMode.INVERT: (255, 255, 0),
            EditMode.OUTLINE: (0, 255, 255),
        }
        
        # Track saved images and their states
        self.saved_status = {}
        self.image_states = {}
        
        # Window setup
        self.window_name = "Batch Labeled Editor - Zoom with Mouse Wheel"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)
        
        # Load first image
        self._load_current_image()
        self._print_instructions()
    
    def _load_image_files(self):
        """Load all image files from folder"""
        extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = []
        
        for ext in extensions:
            files.extend(self.input_folder.glob(f'*{ext}'))
            #files.extend(self.input_folder.glob(f'*{ext.upper()}'))
        
        return sorted(files)
    
    def _load_current_image(self):
        """Load current image and restore its state from JSON if exists"""
        current_file = self.image_files[self.current_index]
        
        self.original_image = cv2.imread(str(current_file))
        if self.original_image is None:
            print(f"Error loading: {current_file}")
            return False
        
        self._scale_image()
        
        # Reset zoom and pan
        self.zoom_level = 1.0
        self.pan_x = 0
        self.pan_y = 0
        
        # Try to load from JSON file first (persistent storage)
        json_path = self.output_folder / current_file.with_suffix('.json').name
        
        if json_path.exists():
            # Load from saved JSON file
            try:
                with open(json_path, 'r') as f:
                    data = json.load(f)
                
                self.circles = []
                for obj in data.get('objects', []):
                    # Restore circle with mode
                    mode_value = obj.get('mode', 'highlight')
                    mode = EditMode(mode_value)
                    
                    self.circles.append({
                        'center': tuple(obj['center']),
                        'radius': obj['radius'],
                        'mode': mode,
                        'label': obj.get('label', '')
                    })
                
                print(f"\n✓ Loaded: {current_file.name} ({self.current_index + 1}/{self.total_images}) - RESTORED FROM JSON ({len(self.circles)} objects)")
            except Exception as e:
                print(f"Error loading JSON: {e}")
                self.circles = []
        elif current_file.name in self.image_states:
            # Load from memory state (legacy)
            self.circles = self.image_states[current_file.name]['circles'].copy()
            print(f"\n✓ Loaded: {current_file.name} ({self.current_index + 1}/{self.total_images}) - RESTORED FROM MEMORY")
        else:
            # Fresh image
            self.circles = []
            print(f"\nLoaded: {current_file.name} ({self.current_index + 1}/{self.total_images})")
        
        # Reset input state
        self.drawing = False
        self.label_input_mode = False
        self.current_label = ""
        
        # Apply effects
        self.output_image = self.scaled_image.copy()
        if self.circles:
            self._apply_all_effects()
        
        self.display_image = self.output_image.copy()
        
        return True
    
    def _print_instructions(self):
        """Print usage instructions"""
        print("\n" + "="*70)
        print("BATCH LABELED EDITOR - FOLDER PROCESSING WITH ZOOM")
        print("="*70)
        print(f"\nInput Folder:  {self.input_folder}")
        print(f"Output Folder: {self.output_folder}")
        print(f"Total Images:  {self.total_images}")
        print("\n🔍 Zoom & Pan Controls:")
        print("  Mouse Wheel  - Zoom in/out (0.5x to 5x)")
        print("  Right Click  - Pan/move image while zoomed")
        print("  R            - Reset zoom to 100%")
        print("\n🔄 Auto-Save Feature:")
        print("  - Work auto-saved when navigating between images")
        print("  - Press 'S' to save to disk permanently")
        print("\n⌨️  Navigation Controls:")
        print("  A         - Previous image")
        print("  D         - Next image")
        print("  S         - Save current image to disk")
        print("  SHIFT+S   - Save and go to next")
        print("\n✏️  Editing Controls:")
        print("  1-7       - Switch editing mode")
        print("  C         - Clear all circles")
        print("  U         - Undo last circle")
        print("  L         - List all objects")
        print("  E         - Edit last label")
        print("  T         - Toggle labels")
        print("\n💾 Label Input:")
        print("  Type      - Enter label (appears above circle)")
        print("  ENTER     - Confirm")
        print("  ESC       - Skip")
        print("\n🚪 Other:")
        print("  Q         - Quit and save")
        print("="*70 + "\n")
    
    def _scale_image(self):
        """Scale image for display if needed"""
        max_height = 900
        max_width = 1400
        
        height, width = self.original_image.shape[:2]
        self.scale_factor = 1.0
        
        if height > max_height or width > max_width:
            scale_h = max_height / height
            scale_w = max_width / width
            self.scale_factor = min(scale_h, scale_w)
            
            new_width = int(width * self.scale_factor)
            new_height = int(height * self.scale_factor)
            
            self.scaled_image = cv2.resize(
                self.original_image,
                (new_width, new_height),
                interpolation=cv2.INTER_NEAREST
            )
        else:
            self.scaled_image = self.original_image.copy()
    
    def _screen_to_image_coords(self, screen_x, screen_y):
        """Convert screen coordinates to image coordinates considering zoom and pan"""
        # Account for zoom and pan
        img_x = int((screen_x - self.pan_x) / self.zoom_level)
        img_y = int((screen_y - self.pan_y) / self.zoom_level)
        return img_x, img_y
    
    def _image_to_screen_coords(self, img_x, img_y):
        """Convert image coordinates to screen coordinates considering zoom and pan"""
        screen_x = int(img_x * self.zoom_level + self.pan_x)
        screen_y = int(img_y * self.zoom_level + self.pan_y)
        return screen_x, screen_y
    
    def _mouse_callback(self, event, x, y, flags, param):
        """Handle mouse events including zoom and pan"""
        
        # Handle mouse wheel for zooming
        if event == cv2.EVENT_MOUSEWHEEL:
            # Get zoom direction
            if flags > 0:  # Scroll up - zoom in
                zoom_factor = 1.1
            else:  # Scroll down - zoom out
                zoom_factor = 0.9
            
            # Calculate new zoom level
            new_zoom = self.zoom_level * zoom_factor
            new_zoom = max(self.min_zoom, min(self.max_zoom, new_zoom))
            
            # Zoom towards mouse cursor
            # Calculate the position in the image before zoom
            img_x, img_y = self._screen_to_image_coords(x, y)
            
            # Update zoom
            self.zoom_level = new_zoom
            
            # Calculate new pan to keep the same point under cursor
            self.pan_x = int(x - img_x * self.zoom_level)
            self.pan_y = int(y - img_y * self.zoom_level)
            
            self._update_display()
            print(f"Zoom: {self.zoom_level:.2f}x")
            return
        
        # Handle right-click pan
        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning = True
            self.pan_start_x = x - self.pan_x
            self.pan_start_y = y - self.pan_y
            return
        
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False
            return
        
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.pan_x = int(x - self.pan_start_x)
            self.pan_y = int(y - self.pan_start_y)
            self._update_display()
            return
        
        # Skip drawing during label input or panning
        if self.label_input_mode or self.is_panning:
            return
        
        # Handle drawing
        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            # Convert screen coords to image coords
            self.center = self._screen_to_image_coords(x, y)
            self.current_radius = 0
        
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            img_x, img_y = self._screen_to_image_coords(x, y)
            self.current_radius = int(np.sqrt((img_x - self.center[0])**2 + 
                                             (img_y - self.center[1])**2))
            self._update_display()
        
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                self._enter_label_input_mode()
    
    def _enter_label_input_mode(self):
        """Enter label input mode"""
        self.label_input_mode = True
        self.current_label = ""
        print(f"\n→ Circle drawn in {self.current_mode.value.upper()} mode")
        print("  Enter label (or ESC to skip):")
        self._update_display_with_input()
    
    def _exit_label_input_mode(self, save=True):
        """Exit label input mode"""
        self.label_input_mode = False
        
        if save:
            label = self.current_label.strip()
            self.circles.append({
                'center': self.center,
                'radius': self.current_radius,
                'mode': self.current_mode,
                'label': label
            })
            
            if label:
                print(f"  ✓ Added: '{label}' [{self.current_mode.value}]")
            else:
                print(f"  ✓ Added unlabeled [{self.current_mode.value}]")
            
            self._apply_all_effects()
        else:
            print("  ✗ Cancelled")
        
        self.current_label = ""
        self._update_display()
    
    def _apply_effect(self, image, circle):
        """Apply specific effect to circular region"""
        mask = np.zeros(image.shape[:2], dtype=np.uint8)
        cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
        
        mode = circle['mode']
        
        if mode == EditMode.HIGHLIGHT:
            highlighted = cv2.addWeighted(
                image, 1 - self.highlight_alpha,
                np.full_like(image, 255), self.highlight_alpha, 0
            )
            image = np.where(mask[:, :, np.newaxis] == 255, highlighted, image)
        
        elif mode == EditMode.BLUR:
            blurred = cv2.GaussianBlur(image, (self.blur_kernel, self.blur_kernel), 0)
            image = np.where(mask[:, :, np.newaxis] == 255, blurred, image)
        
        elif mode == EditMode.PIXELATE:
            h, w = image.shape[:2]
            temp = cv2.resize(
                image,
                (w // self.pixelate_size, h // self.pixelate_size),
                interpolation=cv2.INTER_NEAREST
            )
            pixelated = cv2.resize(temp, (w, h), interpolation=cv2.INTER_NEAREST)
            image = np.where(mask[:, :, np.newaxis] == 255, pixelated, image)
        
        elif mode == EditMode.DARKEN:
            darkened = cv2.addWeighted(image, 0.5, np.zeros_like(image), 0.5, 0)
            image = np.where(mask[:, :, np.newaxis] == 255, darkened, image)
        
        elif mode == EditMode.GRAYSCALE:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            gray_bgr = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
            image = np.where(mask[:, :, np.newaxis] == 255, gray_bgr, image)
        
        elif mode == EditMode.INVERT:
            inverted = cv2.bitwise_not(image)
            image = np.where(mask[:, :, np.newaxis] == 255, inverted, image)
        
        return image
    
    def _apply_all_effects(self):
        """Apply all effects"""
        self.output_image = self.scaled_image.copy()
        
        for circle in self.circles:
            self.output_image = self._apply_effect(self.output_image, circle)
            
            # Draw circle border
            color = self.mode_colors[circle['mode']]
            cv2.circle(self.output_image, circle['center'],
                      circle['radius'], color, 2)
    
    def _draw_label(self, image, circle, number):
        """Draw label for a circle with smart positioning to keep within bounds"""
        if not self.show_labels or not circle['label']:
            return
        
        center = circle['center']
        radius = circle['radius']
        label = circle['label']
        
        # Format label
        mode_short = circle['mode'].value[:3].upper()
        full_label = f"#{number} [{mode_short}] {label}"
        
        # Get text size
        (text_w, text_h), baseline = cv2.getTextSize(
            full_label, self.label_font, self.label_scale, self.label_thickness
        )
        
        padding = 4
        total_width = text_w + 2 * padding
        total_height = text_h + baseline + 2 * padding
        
        # Get image dimensions
        img_h, img_w = image.shape[:2]
        
        # Try different positions in order of preference
        positions = [
            # Above circle
            (center[0] - radius, center[1] - radius - total_height - 10),
            # Below circle
            (center[0] - radius, center[1] + radius + 20),
            # Left of circle
            (center[0] - radius - total_width - 10, center[1] - total_height // 2),
            # Right of circle
            (center[0] + radius + 10, center[1] - total_height // 2),
            # Top-left of circle
            (center[0] - radius - total_width - 10, center[1] - radius - total_height - 10),
            # Top-right of circle
            (center[0] + radius + 10, center[1] - radius - total_height - 10),
            # Bottom-left of circle
            (center[0] - radius - total_width - 10, center[1] + radius + 20),
            # Bottom-right of circle
            (center[0] + radius + 10, center[1] + radius + 20),
        ]
        
        # Find first position that fits completely within image
        label_x, label_y = None, None
        for pos_x, pos_y in positions:
            # Check if label fits within image bounds
            if (pos_x >= padding and 
                pos_y >= text_h + padding and 
                pos_x + total_width <= img_w - padding and 
                pos_y + baseline + padding <= img_h - padding):
                label_x = pos_x
                label_y = pos_y + text_h  # Adjust to text baseline
                break
        
        # If no position fits, force it to fit by clamping
        if label_x is None:
            label_x = max(padding, min(center[0] - radius, img_w - total_width - padding))
            label_y = max(text_h + padding, min(center[1] - radius - 10, img_h - baseline - padding))
            label_y = label_y if label_y > text_h else center[1] + radius + text_h + 10
        
        # Ensure final position is within bounds
        label_x = int(max(padding, min(label_x, img_w - total_width - padding)))
        label_y = int(max(text_h + padding, min(label_y, img_h - baseline - padding)))
        
        # Draw background
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     (0, 0, 0), -1)
        
        # Draw border
        color = self.mode_colors[circle['mode']]
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     color, 1)
        
        # Draw text
        cv2.putText(image, full_label, (label_x, label_y),
                   self.label_font, self.label_scale,
                   (255, 255, 255), self.label_thickness)
        
        # Draw connector line
        line_start = (label_x + text_w // 2, label_y + baseline + padding)
        cv2.line(image, line_start, center, color, 1)
    
    def _draw_typing_label(self, image, center, radius, label, mode):
        """Draw label text above circle in real-time while typing with smart positioning"""
        # Format label
        mode_short = mode.value[:3].upper()
        display_label = f"[{mode_short}] {label}_"
        
        # Get text size with larger font
        (text_w, text_h), baseline = cv2.getTextSize(
            display_label, self.label_font, self.label_scale + 0.2, self.label_thickness + 1
        )
        
        padding = 6
        total_width = text_w + 2 * padding
        total_height = text_h + baseline + 2 * padding
        
        # Get image dimensions
        img_h, img_w = image.shape[:2]
        
        # Try positions (prefer above, then below, then sides)
        positions = [
            (center[0] - radius, center[1] - radius - total_height - 10),  # Above
            (center[0] - radius, center[1] + radius + 20),  # Below
            (center[0] - radius - total_width - 10, center[1] - total_height // 2),  # Left
            (center[0] + radius + 10, center[1] - total_height // 2),  # Right
        ]
        
        # Find best position
        label_x, label_y = None, None
        for pos_x, pos_y in positions:
            if (pos_x >= padding and 
                pos_y >= text_h + padding and 
                pos_x + total_width <= img_w - padding and 
                pos_y + baseline + padding <= img_h - padding):
                label_x = pos_x
                label_y = pos_y + text_h
                break
        
        # Force fit if no good position
        if label_x is None:
            label_x = max(padding, min(center[0] - radius, img_w - total_width - padding))
            label_y = max(text_h + padding, min(center[1] - radius - 10, img_h - baseline - padding))
            if label_y <= text_h + padding:
                label_y = center[1] + radius + text_h + 20
        
        # Ensure within bounds
        label_x = int(max(padding, min(label_x, img_w - total_width - padding)))
        label_y = int(max(text_h + padding, min(label_y, img_h - baseline - padding)))
        
        color = self.mode_colors[mode]
        
        # Dark background
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     (0, 0, 0), -1)
        
        # Bright colored border
        cv2.rectangle(image,
                     (label_x - padding, label_y - text_h - padding),
                     (label_x + text_w + padding, label_y + baseline + padding),
                     color, 2)
        
        # Draw text
        cv2.putText(image, display_label, (label_x, label_y),
                   self.label_font, self.label_scale + 0.2,
                   (255, 255, 255), self.label_thickness + 1)
        
        # Draw connector line
        line_start = (label_x + text_w // 2, label_y + baseline + padding)
        cv2.line(image, line_start, center, color, 2)
    
    def _get_zoomed_view(self, image):
        """Get zoomed and panned view of image"""
        h, w = image.shape[:2]
        
        # Create zoomed image
        new_w = int(w * self.zoom_level)
        new_h = int(h * self.zoom_level)
        
        if self.zoom_level != 1.0:
            zoomed = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_NEAREST)
        else:
            zoomed = image.copy()
        
        # Create canvas for display
        canvas = np.zeros((h, w, 3), dtype=np.uint8)
        
        # Calculate visible region - ensure integers
        start_x = int(max(0, -self.pan_x))
        start_y = int(max(0, -self.pan_y))
        end_x = int(min(new_w, w - self.pan_x))
        end_y = int(min(new_h, h - self.pan_y))
        
        # Calculate target position on canvas - ensure integers
        target_x = int(max(0, self.pan_x))
        target_y = int(max(0, self.pan_y))
        target_end_x = int(target_x + (end_x - start_x))
        target_end_y = int(target_y + (end_y - start_y))
        
        # Ensure all values are valid
        if start_x < end_x and start_y < end_y and target_x < w and target_y < h:
            # Clip to valid ranges
            target_end_x = min(target_end_x, w)
            target_end_y = min(target_end_y, h)
            end_x = min(end_x, new_w)
            end_y = min(end_y, new_h)
            
            # Adjust if sizes don't match
            copy_w = min(end_x - start_x, target_end_x - target_x)
            copy_h = min(end_y - start_y, target_end_y - target_y)
            
            if copy_w > 0 and copy_h > 0:
                canvas[target_y:target_y+copy_h, target_x:target_x+copy_w] = \
                    zoomed[start_y:start_y+copy_h, start_x:start_x+copy_w]
        
        return canvas
    
    def _update_display(self):
        """Update display with zoom"""
        # Start with output image
        temp_image = self.output_image.copy()
        
        # Draw labels
        for idx, circle in enumerate(self.circles):
            self._draw_label(temp_image, circle, idx + 1)
        
        # Draw current circle being drawn
        if self.drawing and self.current_radius > 0:
            color = self.mode_colors[self.current_mode]
            cv2.circle(temp_image, self.center, self.current_radius, color, 2)
        
        # Apply zoom and pan
        self.display_image = self._get_zoomed_view(temp_image)
        
        # Draw UI
        self._draw_ui()
        cv2.imshow(self.window_name, self.display_image)
    
    def _update_display_with_input(self):
        """Update display during label input"""
        temp_image = self.output_image.copy()
        
        # Draw existing labels
        for idx, circle in enumerate(self.circles):
            self._draw_label(temp_image, circle, idx + 1)
        
        # Draw current circle
        color = self.mode_colors[self.current_mode]
        cv2.circle(temp_image, self.center, self.current_radius, color, 3)
        
        # Draw typing label
        if self.current_label:
            self._draw_typing_label(temp_image, self.center,
                                   self.current_radius, self.current_label,
                                   self.current_mode)
        
        # Apply zoom
        self.display_image = self._get_zoomed_view(temp_image)
        
        # Draw input box
        self._draw_input_box()
        cv2.imshow(self.window_name, self.display_image)
    
    def _draw_input_box(self):
        """Draw label input box"""
        h, w = self.display_image.shape[:2]
        box_height = 70
        box_y = h - box_height
        
        overlay = self.display_image.copy()
        cv2.rectangle(overlay, (0, box_y), (w, h), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.8, self.display_image, 0.2, 0, self.display_image)
        
        color = self.mode_colors[self.current_mode]
        cv2.rectangle(self.display_image, (0, box_y), (w, h), color, 2)
        
        mode_text = f"[{self.current_mode.value.upper()}]"
        cv2.putText(self.display_image, mode_text, (10, box_y + 25),
                   self.label_font, 0.6, color, 2)
        
        prompt = "Label:"
        cv2.putText(self.display_image, prompt, (150, box_y + 25),
                   self.label_font, 0.6, (255, 255, 255), 1)
        
        input_text = self.current_label + "_"
        cv2.putText(self.display_image, input_text, (10, box_y + 55),
                   self.label_font, 0.7, (0, 255, 255), 2)
    
    def _draw_ui(self):
        """Draw UI overlay"""
        overlay = self.display_image.copy()
        h, w = self.display_image.shape[:2]
        
        bar_height = 100
        cv2.rectangle(overlay, (0, 0), (w, bar_height), (0, 0, 0), -1)
        
        # Mode
        mode_text = f"Mode: {self.current_mode.value.upper()}"
        cv2.putText(overlay, mode_text, (15, 25),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, self.mode_colors[self.current_mode], 2)
        
        # Zoom level
        zoom_text = f"Zoom: {self.zoom_level:.2f}x"
        cv2.putText(overlay, zoom_text, (15, 50),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (100, 200, 255), 1)
        
        # Image info
        current_file = self.image_files[self.current_index]
        nav_text = f"Image: {self.current_index + 1}/{self.total_images}"
        cv2.putText(overlay, nav_text, (15, 75),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Filename
        filename = current_file.name
        if len(filename) > 40:
            filename = filename[:37] + "..."
        cv2.putText(overlay, filename, (250, 25),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Status
        has_edits = current_file.name in self.image_states or len(self.circles) > 0
        is_saved = self.saved_status.get(current_file.name, False)
        
        if is_saved:
            status_text = "SAVED"
            status_color = (0, 255, 0)
        elif has_edits:
            status_text = "EDITED"
            status_color = (0, 165, 255)
        else:
            status_text = "NO EDITS"
            status_color = (100, 100, 100)
        
        cv2.putText(overlay, status_text, (w - 180, 25),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, status_color, 2)
        
        # Objects count
        obj_text = f"Objects: {len(self.circles)}"
        cv2.putText(overlay, obj_text, (250, 50),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        
        # Bottom hints
        nav_hint = "Wheel:Zoom | Right-Click:Pan | A/D:Nav | S:Save | R:Reset | Q:Quit"
        cv2.putText(overlay, nav_hint, (15, h - 15),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200, 200, 200), 1)
        
        cv2.addWeighted(overlay, 0.7, self.display_image, 0.3, 0, self.display_image)
    
    def _previous_image(self):
        """Go to previous image with auto-save"""
        if self.current_index > 0:
            # Auto-save current work before leaving
            if self.circles:
                print(f"  → Auto-saving current work...")
                self.save_current(auto_save=True)
            
            self.current_index -= 1
            self._load_current_image()
            self._update_display()
        else:
            print("Already at first image")
    
    def _next_image(self):
        """Go to next image with auto-save"""
        if self.current_index < self.total_images - 1:
            # Auto-save current work before leaving
            if self.circles:
                print(f"  → Auto-saving current work...")
                self.save_current(auto_save=True)
            
            self.current_index += 1
            self._load_current_image()
            self._update_display()
        else:
            print("Already at last image")
    
    def _list_labels(self):
        """List all labels"""
        print("\n" + "="*70)
        print(f"LABELED OBJECTS - {self.image_files[self.current_index].name}")
        print("="*70)
        if not self.circles:
            print("No objects marked yet.")
        else:
            for idx, circle in enumerate(self.circles, 1):
                label = circle['label'] if circle['label'] else "(no label)"
                mode = circle['mode'].value
                print(f"#{idx}: {label}")
                print(f"     Mode: {mode}, Position: {circle['center']}, "
                      f"Radius: {circle['radius']}px")
        print("="*70 + "\n")
    
    def _edit_last_label(self):
        """Edit last circle's label"""
        if not self.circles:
            print("No objects to edit!")
            return
        
        last_circle = self.circles[-1]
        current_label = last_circle['label']
        
        print(f"\nCurrent label: '{current_label}'")
        print("Enter new label (ESC to cancel):")
        
        self.current_label = current_label
        self.label_input_mode = True
        self._update_display_with_input()
        
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            
            if key == 27:  # ESC
                print("  ✗ Edit cancelled")
                self.label_input_mode = False
                self.current_label = ""
            elif key == 13:  # ENTER
                last_circle['label'] = self.current_label.strip()
                print(f"  ✓ Updated to: '{self.current_label.strip()}'")
                self.label_input_mode = False
                self.current_label = ""
                self._apply_all_effects()
            elif key == 8:  # BACKSPACE
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key)
                self._update_display_with_input()
        
        self._update_display()
    
    def save_current(self, auto_save=False):
        """Save current image and labels with labels visible on image
        
        Args:
            auto_save: If True, saves silently without verbose output
        """
        if not self.circles:
            if not auto_save:
                print("No objects to save")
            return
        
        current_file = self.image_files[self.current_index]
        
        # Store state in memory (backup)
        self.image_states[current_file.name] = {
            'circles': [circle.copy() for circle in self.circles]
        }
        
        # Prepare output paths
        output_image_path = self.output_folder / current_file.name
        output_json_path = output_image_path.with_suffix('.json')
        
        # Create final output image with labels at ORIGINAL resolution
        if self.scale_factor != 1.0:
            # Scale back to original size
            h, w = self.original_image.shape[:2]
            
            # Apply effects at original resolution
            final_image = self.original_image.copy()
            
            for circle in self.circles:
                # Scale circle coordinates back to original
                orig_center = (
                    int(circle['center'][0] / self.scale_factor),
                    int(circle['center'][1] / self.scale_factor)
                )
                orig_radius = int(circle['radius'] / self.scale_factor)
                
                # Create scaled circle
                scaled_circle = {
                    'center': orig_center,
                    'radius': orig_radius,
                    'mode': circle['mode'],
                    'label': circle['label']
                }
                
                # Apply effect
                final_image = self._apply_effect(final_image, scaled_circle)
                
                # Draw circle border
                color = self.mode_colors[circle['mode']]
                cv2.circle(final_image, orig_center, orig_radius, color, 3)
            
            # Draw labels on final image
            for idx, circle in enumerate(self.circles, 1):
                orig_center = (
                    int(circle['center'][0] / self.scale_factor),
                    int(circle['center'][1] / self.scale_factor)
                )
                orig_radius = int(circle['radius'] / self.scale_factor)
                
                scaled_circle = {
                    'center': orig_center,
                    'radius': orig_radius,
                    'mode': circle['mode'],
                    'label': circle['label']
                }
                
                self._draw_label(final_image, scaled_circle, idx)
        else:
            # Already at original resolution
            final_image = self.output_image.copy()
            
            # Draw labels
            for idx, circle in enumerate(self.circles, 1):
                self._draw_label(final_image, circle, idx)
        
        # Save image
        cv2.imwrite(str(output_image_path), final_image)
        
        # Save JSON with proper structure
        labels_data = {
            'source_image': current_file.name,
            'timestamp': datetime.now().isoformat(),
            'objects': []
        }
        
        for idx, circle in enumerate(self.circles, 1):
            labels_data['objects'].append({
                'id': idx,
                'label': circle['label'],
                'mode': circle['mode'].value,
                'center': list(circle['center']),  # Convert tuple to list for JSON
                'radius': circle['radius']
            })
        
        with open(output_json_path, 'w') as f:
            json.dump(labels_data, f, indent=2)
        
        # Mark as saved
        self.saved_status[current_file.name] = True
        
        if not auto_save:
            print(f"\n✓ Saved: {output_image_path.name}")
            print(f"  - Image: {output_image_path} (with labels)")
            print(f"  - Labels: {output_json_path}")
            print(f"  - {len(self.circles)} objects saved")
        else:
            print(f"    ✓ Auto-saved {len(self.circles)} objects")
    
    def generate_summary(self):
        """Generate Excel summary"""
        excel_path = self.output_folder / "processing_summary.xlsx"
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Processing Summary"
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ["Image Name", "Number of Mistakes", "Error Names"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row = 2
            for img_file in sorted(self.image_files):
                if img_file.name not in self.saved_status:
                    continue
                
                json_path = self.output_folder / img_file.with_suffix('.json').name
                
                if json_path.exists():
                    with open(json_path, 'r') as f:
                        data = json.load(f)
                    
                    ws.cell(row=row, column=1).value = img_file.name
                    ws.cell(row=row, column=1).border = border
                    
                    num_labels = len(data['objects'])
                    ws.cell(row=row, column=2).value = num_labels
                    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2).border = border
                    
                    labels = [obj['label'] for obj in data['objects'] if obj['label']]
                    label_names = ", ".join(labels) if labels else "(no labels)"
                    ws.cell(row=row, column=3).value = label_names
                    ws.cell(row=row, column=3).border = border
                    
                    row += 1
            
            # Summary
            row += 1
            summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            summary_font = Font(bold=True, size=11)
            
            ws.cell(row=row, column=1).value = "SUMMARY"
            ws.cell(row=row, column=1).font = summary_font
            ws.cell(row=row, column=1).fill = summary_fill
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws.cell(row=row, column=1).value = "Total Images Processed"
            ws.cell(row=row, column=2).value = len(self.saved_status)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
            ws.cell(row=row, column=1).value = "Total Objects Labeled"
            total_objects = 0
            for img_file in self.image_files:
                if img_file.name in self.saved_status:
                    json_path = self.output_folder / img_file.with_suffix('.json').name
                    if json_path.exists():
                        with open(json_path, 'r') as f:
                            data = json.load(f)
                        total_objects += len(data['objects'])
            ws.cell(row=row, column=2).value = total_objects
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 60
            
            wb.save(str(excel_path))
            print(f"✓ Excel summary saved: {excel_path}")
            
        except ImportError:
            print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            print(f"⚠️  Could not create Excel: {e}")
    
    def run(self):
        """Main loop"""
        self._update_display()
        
        while True:
            key = cv2.waitKey(1) & 0xFF
            
            if self.label_input_mode:
                if key == 27:  # ESC
                    self._exit_label_input_mode(save=True)
                elif key == 13:  # ENTER
                    self._exit_label_input_mode(save=True)
                elif key == 8:  # BACKSPACE
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key)
                    self._update_display_with_input()
                continue
            
            # Navigation
            if key == ord('a'):
                self._previous_image()
            elif key == ord('d'):
                self._next_image()
            
            # Reset zoom
            elif key == ord('r'):
                self.zoom_level = 1.0
                self.pan_x = 0
                self.pan_y = 0
                self._update_display()
                print("Zoom reset to 100%")
            
            # Save
            elif key == ord('s'):
                self.save_current(auto_save=False)
                self._update_display()
            elif key == ord('S'):
                self.save_current(auto_save=False)
                self._next_image()
            
            # Editing
            elif key == ord('c'):
                self.circles.clear()
                self.output_image = self.scaled_image.copy()
                self._update_display()
                print("✓ Cleared all")
            elif key == ord('u'):
                if self.circles:
                    removed = self.circles.pop()
                    label = removed['label'] if removed['label'] else "(unlabeled)"
                    print(f"✓ Removed: {label}")
                    self._apply_all_effects()
                    self._update_display()
            elif key == ord('l'):
                self._list_labels()
            elif key == ord('e'):
                self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels: {'ON' if self.show_labels else 'OFF'}")
                self._update_display()
            
            # Mode switching
            elif ord('1') <= key <= ord('7'):
                modes = list(EditMode)
                self.current_mode = modes[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._update_display()
            
            # Quit
            elif key == ord('q'):
                break
        
        cv2.destroyAllWindows()
        
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Processed {len(self.saved_status)}/{self.total_images} images")
        else:
            print("\n⚠️  No images saved")


def main():
    parser = argparse.ArgumentParser(
        description="Batch Advanced Labeled Editor with Zoom"
    )
    parser.add_argument("input_folder", type=str,
                       help="Input folder containing images")
    parser.add_argument("--output", "-o", type=str, default=None,
                       help="Output folder (default: labeled_output_TIMESTAMP)")
    
    args = parser.parse_args()
    
    try:
        editor = BatchLabeledEditor(args.input_folder, args.output)
        editor.run()
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
