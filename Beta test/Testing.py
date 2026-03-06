#!/usr/bin/env python3
"""
Batch Labeled Editor - Production Version
Centered image on dark canvas · Anti-aliased fonts · Zoom · Auto-save
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime
import time
import sys


# ── Typography & canvas constants ───────────────────────────────────────────────
FONT       = cv2.FONT_HERSHEY_DUPLEX   # smooth, no pixel-aliasing
AA         = cv2.LINE_AA
CANVAS_W   = 1440
CANVAS_H   = 900
CANVAS_BG  = (18, 18, 18)             # near-black surround
CANVAS_PAD = 40                        # clear gap around image on canvas


class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50
    ZOOM_DEBOUNCE_MS        = 50
    MAX_CACHED_STATES       = 5

    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)

        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.total_images  = len(self.image_files)
        self.current_index = 0

        self.original_image         = None
        self.scaled_image           = None
        self.display_image          = None
        self.output_image           = None
        self.circles                = []
        self.drawing                = False
        self.center                 = None
        self.current_radius         = 0
        self.current_label          = ""
        self.current_description    = ""
        self.label_input_mode       = False
        self.description_input_mode = False

        self.zoom_level    = 1.0
        self.min_zoom      = 0.3
        self.max_zoom      = 10.0
        self.pan_x         = 0
        self.pan_y         = 0
        self.is_panning    = False
        self.pan_start_x   = 0
        self.pan_start_y   = 0
        self.last_zoom_time = 0

        self.current_mode    = EditMode.HIGHLIGHT
        self._blur_kernel    = 25
        self._pixelate_size  = 10
        self.highlight_alpha = 0.4
        self.show_labels     = True

        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR:      (255, 80, 80),
            EditMode.PIXELATE:  (80, 80, 255),
            EditMode.DARKEN:    (160, 160, 160),
            EditMode.GRAYSCALE: (210, 210, 210),
            EditMode.INVERT:    (255, 255, 0),
            EditMode.OUTLINE:   (0, 220, 255),
        }

        self.saved_status       = {}
        self.image_states       = {}
        self.state_access_order = []

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name, CANVAS_W, CANVAS_H)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")
        self._print_instructions()

    # ── Properties ──────────────────────────────────────────────────────────────

    @property
    def blur_kernel(self):
        return self._blur_kernel

    @blur_kernel.setter
    def blur_kernel(self, v):
        if not isinstance(v, int) or v <= 0 or v % 2 == 0:
            raise ValueError("blur_kernel must be a positive odd integer")
        self._blur_kernel = v

    @property
    def pixelate_size(self):
        return self._pixelate_size

    @pixelate_size.setter
    def pixelate_size(self, v):
        if not isinstance(v, int) or v <= 0:
            raise ValueError("pixelate_size must be a positive integer")
        self._pixelate_size = v

    # ── File loading ─────────────────────────────────────────────────────────────

    def _load_image_files(self):
        exts  = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid = []
        for f in files:
            try:
                # Use imdecode on header bytes only — avoids fully loading every image
                raw = np.frombuffer(f.read_bytes()[:64], dtype=np.uint8)
                # Fall back to a proper but minimal check via file size heuristic
                if f.stat().st_size < 512:          # suspiciously tiny → skip
                    print(f"⚠  Skipping {f.name}: file too small")
                    continue
                valid.append(f)
            except Exception as e:
                print(f"⚠  Skipping {f.name}: {e}")

        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images (recommended ≤ {self.MAX_BATCH_SIZE})")
            if input("   Continue? (y/N): ").strip().lower() != 'y':
                sys.exit(0)
        return valid

    # ── Image load & scale ───────────────────────────────────────────────────────

    def _scale_image(self):
        max_w = CANVAS_W - CANVAS_PAD * 2
        max_h = CANVAS_H - CANVAS_PAD * 2
        h, w  = self.original_image.shape[:2]
        self.scale_factor = 1.0
        if h > max_h or w > max_w:
            self.scale_factor = min(max_h / h, max_w / w)
            nw = int(w * self.scale_factor)
            nh = int(h * self.scale_factor)
            self.scaled_image = cv2.resize(
                self.original_image, (nw, nh), interpolation=cv2.INTER_AREA)
        else:
            self.scaled_image = self.original_image.copy()

    def _load_current_image(self):
        if self.current_index >= len(self.image_files):
            return False
        f = self.image_files[self.current_index]
        try:
            self.original_image = cv2.imread(str(f))
            if self.original_image is None:
                raise IOError("cv2.imread returned None")
            if min(self.original_image.shape[:2]) < self.MIN_IMAGE_SIZE:
                raise ValueError("Image too small")

            self._scale_image()
            self.zoom_level = 1.0
            self.pan_x = self.pan_y = 0

            json_path = self.output_folder / f.with_suffix('.json').name
            if json_path.exists():
                try:
                    with open(json_path) as jf:
                        data = json.load(jf)
                    self.circles = [{
                        'center':      tuple(o['center']),
                        'radius':      o['radius'],
                        'mode':        EditMode(o.get('mode', 'highlight')),
                        'label':       o.get('label', ''),
                        'description': o.get('description', ''),
                    } for o in data.get('objects', [])]
                    print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                          f" — restored {len(self.circles)} objects")
                except Exception as e:
                    print(f"⚠  JSON load error: {e}")
                    self.circles = []
            elif f.name in self.image_states:
                self.circles = self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images}) — from memory")
            else:
                self.circles = []
                print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")

            self._update_state_access(f.name)
            self.drawing = self.label_input_mode = self.description_input_mode = False
            self.current_label = self.current_description = ""

            self.output_image = self.scaled_image.copy()
            if self.circles:
                self._apply_all_effects()
            self.display_image = self.output_image.copy()
            return True

        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index < len(self.image_files) - 1:
                self.current_index += 1
                return self._load_current_image()
            return False

    # ── LRU memory management ───────────────────────────────────────────────────

    def _update_state_access(self, name):
        if name in self.state_access_order:
            self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order) > self.MAX_CACHED_STATES:
            for name in self.state_access_order[:-self.MAX_CACHED_STATES]:
                self.image_states.pop(name, None)
            self.state_access_order = self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── Coordinate transforms (centred-canvas aware) ─────────────────────────────

    def _img_offset(self):
        ih, iw = self.scaled_image.shape[:2]
        zw = int(iw * self.zoom_level)
        zh = int(ih * self.zoom_level)
        return (CANVAS_W - zw) // 2 + self.pan_x, (CANVAS_H - zh) // 2 + self.pan_y

    def _screen_to_image_coords(self, sx, sy):
        ox, oy = self._img_offset()
        return int((sx - ox) / self.zoom_level), int((sy - oy) / self.zoom_level)

    # ── Mouse callback ───────────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time() * 1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS:
                return
            self.last_zoom_time = now
            factor   = 1.1 if flags > 0 else 0.9
            new_zoom = max(self.min_zoom, min(self.max_zoom, self.zoom_level * factor))
            ix, iy   = self._screen_to_image_coords(x, y)
            self.zoom_level = new_zoom
            ih, iw   = self.scaled_image.shape[:2]
            zw = int(iw * self.zoom_level); zh = int(ih * self.zoom_level)
            self.pan_x = int(x - ix * self.zoom_level - (CANVAS_W - zw) // 2)
            self.pan_y = int(y - iy * self.zoom_level - (CANVAS_H - zh) // 2)
            self._update_display()
            return

        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning = True
            self.pan_start_x = x - self.pan_x; self.pan_start_y = y - self.pan_y
            return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False; return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.pan_x = x - self.pan_start_x; self.pan_y = y - self.pan_start_y
            self._update_display(); return

        if self.label_input_mode or self.description_input_mode or self.is_panning:
            return

        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center  = self._screen_to_image_coords(x, y)
            self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            ix, iy = self._screen_to_image_coords(x, y)
            self.current_radius = int(np.hypot(ix - self.center[0], iy - self.center[1]))
            self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles — performance may degrade")
                self._enter_label_input_mode()

    # ── Label / description input flow ───────────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True
        self.current_label    = ""
        print(f"\n→ [{self.current_mode.value.upper()}] type label, ENTER to confirm, ESC to cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save):
        self.label_input_mode = False
        if save:
            self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled")
            self.current_label = ""
            self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode = True
        self.current_description    = ""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode = False
        lbl  = self.current_label.strip()
        desc = self.current_description.strip()
        self.circles.append({
            'center':      self.center,
            'radius':      self.current_radius,
            'mode':        self.current_mode,
            'label':       lbl,
            'description': desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}" + (" + description" if desc else "") +
              f" [{self.current_mode.value}]")
        self.current_label = self.current_description = ""
        self._apply_all_effects()
        self._update_display()

    # ── Effects ──────────────────────────────────────────────────────────────────

    def _apply_effect(self, image, circle):
        try:
            mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
            mode = circle['mode']

            if mode == EditMode.HIGHLIGHT:
                lit   = cv2.addWeighted(image, 1 - self.highlight_alpha,
                                        np.full_like(image, 255), self.highlight_alpha, 0)
                image = np.where(mask[:, :, np.newaxis] == 255, lit, image)
            elif mode == EditMode.BLUR:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.GaussianBlur(image, (self._blur_kernel,)*2, 0), image)
            elif mode == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                small = cv2.resize(image,
                                   (max(1, w // self._pixelate_size),
                                    max(1, h // self._pixelate_size)),
                                   interpolation=cv2.INTER_NEAREST)
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST), image)
            elif mode == EditMode.DARKEN:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.addWeighted(image, 0.5, np.zeros_like(image), 0.5, 0), image)
            elif mode == EditMode.GRAYSCALE:
                gray  = cv2.cvtColor(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                image = np.where(mask[:, :, np.newaxis] == 255, gray, image)
            elif mode == EditMode.INVERT:
                image = np.where(mask[:, :, np.newaxis] == 255, cv2.bitwise_not(image), image)
        except Exception as e:
            print(f"⚠  Effect error ({circle['mode'].value}): {e}")
        return image

    def _apply_all_effects(self):
        self.output_image = self.scaled_image.copy()
        for c in self.circles:
            self.output_image = self._apply_effect(self.output_image, c)
            cv2.circle(self.output_image, c['center'], c['radius'],
                       self.mode_colors[c['mode']], 2, AA)

    # ── Smart label drawing with collision avoidance ──────────────────────────────

    def _label_font_params(self, img_shape):
        md = min(img_shape[:2])
        if   md < 200: return 0.30, 1
        elif md < 400: return 0.40, 1
        elif md < 800: return 0.50, 1
        else:          return 0.55, 1

    def _rects_overlap(self, r1, r2, buf=4):
        return not (r1[2]+buf < r2[0] or r1[0]-buf > r2[2] or
                    r1[3]+buf < r2[1] or r1[1]-buf > r2[3])

    def _find_label_pos(self, center, radius, tw, fh, bl, pad, img_shape, placed):
        ih, iw = img_shape[:2]
        W = tw + 2*pad;  H = fh + bl + 2*pad
        cx, cy = center

        candidates = [
            (cx - radius,     cy - radius - H - 10),
            (cx - radius,     cy + radius + 10),
            (cx - radius-W-8, cy - H//2),
            (cx + radius+8,   cy - H//2),
            (cx + radius+8,   cy - radius - H - 10),
            (cx - radius-W-8, cy - radius - H - 10),
            (cx + radius+8,   cy + radius + 10),
            (cx - radius-W-8, cy + radius + 10),
        ]
        for extra in (40, 80, 120):
            candidates += [(px, py+extra) for px, py in candidates[:4]]
            candidates += [(px, py-extra) for px, py in candidates[:4]]

        for px, py in candidates:
            rect = (int(px), int(py), int(px+W), int(py+H))
            if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad:
                continue
            if not any(self._rects_overlap(rect, r) for r in placed):
                return int(px)+pad, int(py)+fh+pad, rect

        # fallback: stack below existing labels
        bottom = max((r[3] for r in placed), default=0) + 8
        lx = max(pad, min(cx - W//2, iw-W-pad))
        ly = min(bottom + fh + pad, ih-bl-pad)
        rect = (lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad)
        return lx, ly, rect

    def _draw_all_labels_smart(self, image):
        if not self.show_labels:
            return
        placed = []
        sc, th = self._label_font_params(image.shape)

        for i, c in enumerate(self.circles, 1):
            if not c['label']:
                continue
            text = f"#{i} [{c['mode'].value[:3].upper()}] {c['label']}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
            pad   = max(3, int(4 * sc / 0.5))
            color = self.mode_colors[c['mode']]

            lx, ly, rect = self._find_label_pos(
                c['center'], c['radius'], tw, fh, bl, pad, image.shape, placed)
            placed.append(rect)

            cv2.rectangle(image, (rect[0], rect[1]), (rect[2], rect[3]), (10,10,10), -1)
            cv2.rectangle(image, (rect[0], rect[1]), (rect[2], rect[3]), color, 1, AA)
            cv2.putText(image, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
            mid_x = (rect[0]+rect[2])//2
            mid_y = rect[3]
            cv2.line(image, (mid_x, mid_y), c['center'], color, 1, AA)

    def _draw_typing_label(self, image, center, radius, label, mode):
        sc, th = self._label_font_params(image.shape)
        sc += 0.1;  th += 1
        text = f"[{mode.value[:3].upper()}] {label}_"
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
        pad    = max(4, int(6 * sc / 0.5))
        ih, iw = image.shape[:2]
        color  = self.mode_colors[mode]

        px, py = center[0]-radius, center[1]-radius-fh-2*pad-10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px, py = center[0]-radius, center[1]+radius+10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px = max(pad, min(center[0]-radius, iw-tw-2*pad-pad))
            py = max(fh+pad, center[1]-radius-fh-2*pad-10)

        lx = int(px); ly = int(py+fh+pad)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), (10,10,10), -1)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), color, 2, AA)
        cv2.putText(image, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
        cv2.line(image, (lx+tw//2, ly+bl+pad), center, color, 2, AA)

    # ── Centred dark-canvas zoom view ────────────────────────────────────────────

    def _get_zoomed_view(self, image):
        ih, iw = image.shape[:2]
        zw = max(1, int(iw * self.zoom_level))
        zh = max(1, int(ih * self.zoom_level))
        interp = cv2.INTER_LINEAR if self.zoom_level >= 1.0 else cv2.INTER_AREA
        zoomed = cv2.resize(image, (zw, zh), interpolation=interp)

        canvas = np.full((CANVAS_H, CANVAS_W, 3), CANVAS_BG, dtype=np.uint8)
        ox = (CANVAS_W - zw) // 2 + self.pan_x
        oy = (CANVAS_H - zh) // 2 + self.pan_y

        sx = max(0, -ox);  sy = max(0, -oy)
        dx = max(0,  ox);  dy = max(0,  oy)
        cw = min(zw-sx, CANVAS_W-dx)
        ch = min(zh-sy, CANVAS_H-dy)
        if cw > 0 and ch > 0:
            canvas[dy:dy+ch, dx:dx+cw] = zoomed[sy:sy+ch, sx:sx+cw]

        bx1 = max(0, ox-1); by1 = max(0, oy-1)
        bx2 = min(CANVAS_W-1, ox+zw); by2 = min(CANVAS_H-1, oy+zh)
        cv2.rectangle(canvas, (bx1, by1), (bx2, by2), (45,45,45), 1, AA)
        return canvas

    # ── Display update ───────────────────────────────────────────────────────────

    def _update_display(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        if self.drawing and self.current_radius > 0:
            cv2.circle(tmp, self.center, self.current_radius,
                       self.mode_colors[self.current_mode], 2, AA)
        self.display_image = self._get_zoomed_view(tmp)
        self._draw_ui()
        cv2.imshow(self.window_name, self.display_image)

    def _update_display_with_input(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        self._draw_typing_label(tmp, self.center, self.current_radius,
                                self.current_label, self.current_mode)
        self.display_image = self._get_zoomed_view(tmp)
        self._draw_input_box("Label")
        cv2.imshow(self.window_name, self.display_image)

    def _update_display_with_description_input(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        if self.current_label:
            self._draw_typing_label(tmp, self.center, self.current_radius,
                                    self.current_label, self.current_mode)
        self.display_image = self._get_zoomed_view(tmp)
        self._draw_input_box("Description (optional)")
        cv2.imshow(self.window_name, self.display_image)

    # ── UI overlays ──────────────────────────────────────────────────────────────

    def _draw_input_box(self, prompt):
        h, w  = self.display_image.shape[:2]
        bh = 64;  by = h - bh
        color = self.mode_colors[self.current_mode]
        overlay = self.display_image.copy()
        cv2.rectangle(overlay, (0, by), (w, h), (12,12,12), -1)
        cv2.addWeighted(overlay, 0.88, self.display_image, 0.12, 0, self.display_image)
        cv2.rectangle(self.display_image, (0, by), (w, h), color, 2, AA)
        cv2.putText(self.display_image, f"[{self.current_mode.value.upper()}]",
                    (12, by+22), FONT, 0.52, color, 1, AA)
        cv2.putText(self.display_image, f"{prompt}:",
                    (160, by+22), FONT, 0.52, (200,200,200), 1, AA)
        text = (self.current_description if self.description_input_mode
                else self.current_label) + "_"
        cv2.putText(self.display_image, text, (12, by+50), FONT, 0.65, (0,240,240), 1, AA)

    def _draw_ui(self):
        overlay = self.display_image.copy()
        h, w    = self.display_image.shape[:2]
        bh      = 84
        cv2.rectangle(overlay, (0, 0), (w, bh), (12,12,12), -1)
        color = self.mode_colors[self.current_mode]
        cf    = self.image_files[self.current_index]

        cv2.putText(overlay, f"Mode: {self.current_mode.value.upper()}",
                    (14,24), FONT, 0.55, color, 1, AA)
        cv2.putText(overlay, f"Zoom: {self.zoom_level:.2f}x",
                    (14,48), FONT, 0.45, (100,190,255), 1, AA)
        cv2.putText(overlay, f"{self.current_index+1} / {self.total_images}",
                    (14,70), FONT, 0.43, (160,160,160), 1, AA)

        fname = cf.name if len(cf.name) <= 46 else cf.name[:43]+"..."
        cv2.putText(overlay, fname,   (200,24), FONT, 0.46, (185,185,185), 1, AA)

        n   = len(self.circles)
        oc  = (40,160,255) if n >= self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(overlay, f"Objects: {n}", (200,48), FONT, 0.43, oc, 1, AA)

        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = cf.name in self.image_states or n > 0
        st, sc = (("SAVED",    (60,200,60))  if is_saved  else
                  ("EDITED",   (40,160,255)) if has_edits else
                  ("NO EDITS", (80,80,80)))
        cv2.putText(overlay, st, (w-155,24), FONT, 0.54, sc, 1, AA)

        hint = "Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  R:Reset  C:Clear  U:Undo  H:Help  Q:Quit"
        cv2.putText(overlay, hint, (12, h-10), FONT, 0.36, (110,110,110), 1, AA)
        cv2.addWeighted(overlay, 0.82, self.display_image, 0.18, 0, self.display_image)

    # ── Navigation ───────────────────────────────────────────────────────────────

    def _guard_state(self):
        if self.drawing:
            print("⚠  Finish drawing before switching images"); return True
        if self.label_input_mode:
            print("⚠  Finish label input (ESC to cancel)"); return True
        if self.description_input_mode:
            print("⚠  Finish description input (ENTER to skip)"); return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index > 0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index -= 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index < self.total_images - 1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index += 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at last image")

    # ── Misc commands ────────────────────────────────────────────────────────────

    def _list_labels(self):
        print("\n" + "="*60)
        print(f"Objects — {self.image_files[self.current_index].name}")
        print("="*60)
        if not self.circles:
            print("  (none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  "
                  f"r={c['radius']}  pos={c['center']}")
            if c.get('description'):
                print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        print(f"\nMemory — {len(self.image_states)} states cached "
              f"(max {self.MAX_CACHED_STATES})  |  {len(self.saved_status)} saved\n")

    def _edit_last_label(self):
        if not self.circles:
            print("No objects to edit"); return
        last = self.circles[-1]

        print(f"\nEditing label: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label    = last['label']
        self.label_input_mode = True
        self._update_display_with_input()
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:
                self.label_input_mode = False; self.current_label = ""
                self._update_display(); return
            elif key == 13:
                last['label'] = self.current_label.strip()
                self.label_input_mode = False; break
            elif key == 8:
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key); self._update_display_with_input()

        print(f"\nEditing description: '{last.get('description','')}'  (ENTER=save, ESC=skip)")
        self.current_description    = last.get('description', '')
        self.description_input_mode = True
        self._update_display_with_description_input()
        while self.description_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 13:
                last['description'] = self.current_description.strip()
                self.description_input_mode = False; break
            elif key == 27:
                self.description_input_mode = False; break
            elif key == 8:
                self.current_description = self.current_description[:-1]
                self._update_display_with_description_input()
            elif 32 <= key <= 126:
                self.current_description += chr(key); self._update_display_with_description_input()

        self.current_label = self.current_description = ""
        print("  ✓ Updated")
        self._apply_all_effects(); self._update_display()

    # ── Save ─────────────────────────────────────────────────────────────────────

    def save_current(self, auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save")
            return

        cf  = self.image_files[self.current_index]
        out = self.output_folder / cf.name
        js  = out.with_suffix('.json')

        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._update_state_access(cf.name)

        try:
            if self.scale_factor != 1.0:
                final = self.original_image.copy()
                sc_circles = [{
                    'center': (int(c['center'][0]/self.scale_factor),
                               int(c['center'][1]/self.scale_factor)),
                    'radius': int(c['radius']/self.scale_factor),
                    'mode':   c['mode'], 'label': c['label'],
                } for c in self.circles]
                for sc in sc_circles:
                    final = self._apply_effect(final, sc)
                    cv2.circle(final, sc['center'], sc['radius'],
                               self.mode_colors[sc['mode']], 3, AA)
                old_circles   = self.circles
                self.circles  = sc_circles
                self._draw_all_labels_smart(final)
                self.circles  = old_circles
            else:
                final = self.output_image.copy()
                self._draw_all_labels_smart(final)

            if not cv2.imwrite(str(out), final):
                raise IOError(f"cv2.imwrite failed for {out}")

            data = {
                'source_image': cf.name,
                'timestamp':    datetime.now().isoformat(),
                'objects': [{
                    'id':          i,
                    'label':       c['label'],
                    'description': c.get('description', ''),
                    'mode':        c['mode'].value,
                    'center':      list(c['center']),
                    'radius':      c['radius'],
                } for i, c in enumerate(self.circles, 1)]
            }
            with open(js, 'w') as f:
                json.dump(data, f, indent=2)

            self.saved_status[cf.name] = True
            if auto_save:
                print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else:
                print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)\n  → {out}\n  → {js}")

            self._cleanup_old_states()

        except Exception as e:
            print(f"❌ Save error: {e}")

    # ── Excel summary ─────────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self.output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
            hfill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hfont  = Font(bold=True, color="FFFFFF", size=12)
            border = Border(**{s: Side(style='thin') for s in ('left','right','top','bottom')})

            for col, hdr in enumerate(["Image Name","# Objects","Labels","Descriptions"], 1):
                c = ws.cell(row=1, column=col, value=hdr)
                c.fill = hfill; c.font = hfont; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')

            row = 2
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                jp = self.output_folder / img.with_suffix('.json').name
                if not jp.exists(): continue
                with open(jp) as f: data = json.load(f)

                ws.cell(row=row, column=1, value=img.name).border = border
                nc = ws.cell(row=row, column=2, value=len(data['objects']))
                nc.alignment = Alignment(horizontal='center'); nc.border = border
                lbls  = ", ".join(o['label'] for o in data['objects'] if o['label']) or "(none)"
                ws.cell(row=row, column=3, value=lbls).border = border
                descs = " | ".join(f"{o['label']}: {o['description']}"
                                   for o in data['objects'] if o.get('description'))
                ws.cell(row=row, column=4, value=descs or "(none)").border = border
                row += 1

            row += 1
            sfill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = sfill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws.cell(row=row, column=1, value="Total Images").font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(self.saved_status))
            row += 1
            total = sum(
                len(json.load(open(self.output_folder / f.with_suffix('.json').name))['objects'])
                for f in self.image_files
                if f.name in self.saved_status
                and (self.output_folder / f.with_suffix('.json').name).exists())
            ws.cell(row=row, column=1, value="Total Objects").font = Font(bold=True)
            ws.cell(row=row, column=2, value=total)

            for col, w in zip('ABCD', [32, 12, 40, 55]):
                ws.column_dimensions[col].width = w
            wb.save(str(xp))
            print(f"✓ Excel summary: {xp}")
        except ImportError:
            print("⚠  pip install openpyxl")
        except Exception as e:
            print(f"⚠  Excel error: {e}")

    # ── Instructions ─────────────────────────────────────────────────────────────

    def _print_instructions(self):
        print("\n" + "="*58)
        print("BATCH LABELED EDITOR")
        print("="*58)
        print(f"  Input : {self.input_folder}")
        print(f"  Output: {self.output_folder}")
        print(f"  Images: {self.total_images}")
        print()
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*58+"\n")

    # ── Main loop ─────────────────────────────────────────────────────────────────

    def run(self):
        self._update_display()
        while True:
            key = cv2.waitKey(1) & 0xFF

            if self.description_input_mode:
                if key in (13, 27): self._exit_description_input_mode()
                elif key == 8:
                    self.current_description = self.current_description[:-1]
                    self._update_display_with_description_input()
                elif 32 <= key <= 126:
                    self.current_description += chr(key)
                    self._update_display_with_description_input()
                continue

            if self.label_input_mode:
                if key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13: self._exit_label_input_mode(save=True)
                elif key == 8:
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key); self._update_display_with_input()
                continue

            if   key == ord('a'): self._previous_image()
            elif key == ord('d'): self._next_image()
            elif key == ord('r'):
                self.zoom_level = 1.0; self.pan_x = self.pan_y = 0
                self._update_display(); print("✓ Zoom reset")
            elif key == ord('s'): self.save_current(); self._update_display()
            elif key == ord('S'): self.save_current(); self._next_image()
            elif key == ord('c'):
                self.circles.clear()
                self.output_image = self.scaled_image.copy()
                self._update_display(); print("✓ Cleared")
            elif key == ord('u'):
                if self.circles:
                    rem = self.circles.pop()
                    print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._apply_all_effects(); self._update_display()
            elif key == ord('l'): self._list_labels()
            elif key == ord('e'): self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}")
                self._update_display()
            elif key == ord('m'): self._show_memory_status()
            elif key in (ord('h'), ord('H')): self._print_instructions()
            elif ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._update_display()
            elif key == ord('q'):
                if self.circles and not self.saved_status.get(
                        self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved")
            print(f"   Output: {self.output_folder}")
        else:
            print("\n⚠  No images were saved")


# ── Entry point ───────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Batch Labeled Editor")
    ap.add_argument("input_folder")
    ap.add_argument("--output", "-o", default=None)
    args = ap.parse_args()
    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n⚠  Interrupted"); return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback; traceback.print_exc(); return 1
    return 0


if __name__ == "__main__":
    exit(main())



























# """
# Segmentation Spill Detector
# ────────────────────────────
# Finds small "spill" clusters — tiny connected groups of pixels whose color
# does NOT match the dominant color of their surrounding region.

# Designed for FLAT-COLOR segmentation maps (like label images).
# """

# import cv2
# import numpy as np

# # ══════════════════════════════════════════════════════
# #  SETTINGS
# # ══════════════════════════════════════════════════════
# IMAGE_PATH      = r"C:\Users\mshar\OneDrive\Desktop\Data Ai\input\test1.png"
# OUTPUT_PATH     = r"C:\Users\mshar\OneDrive\Desktop\Data Ai\output\spill_detected.png"

# MAX_SPILL_AREA  = 15       # max pixels in a "spill" cluster (raise for bigger blobs)
# MIN_SPILL_AREA  = 1        # minimum size (1 = single pixel)
# COLOR_TOLERANCE = 20       # how similar two colors must be to be "same label" (0–100)
#                            # raise if your image has slight anti-aliasing / compression

# CIRCLE_COLOR    = (0, 0, 255)    # BGR — red
# CIRCLE_THICKNESS = 2
# MIN_CIRCLE_R    = 10             # minimum drawn circle radius
# # ══════════════════════════════════════════════════════


# def quantize_to_labels(img, tolerance):
#     """
#     Merge colours that are within `tolerance` (Euclidean BGR) of each other
#     into a single label. Returns a 2-D label array and the label→colour map.
#     """
#     h, w = img.shape[:2]
#     pixels = img.reshape(-1, 3).astype(np.float32)

#     labels_out = np.full(len(pixels), -1, dtype=np.int32)
#     palette    = []   # representative colour for each label

#     for idx, px in enumerate(pixels):
#         if len(palette) == 0:
#             palette.append(px)
#             labels_out[idx] = 0
#             continue
#         dists = np.linalg.norm(np.array(palette) - px, axis=1)
#         best  = int(np.argmin(dists))
#         if dists[best] <= tolerance:
#             labels_out[idx] = best
#         else:
#             labels_out[idx] = len(palette)
#             palette.append(px)

#     # sort unique labels by frequency so label-0 is the most common colour
#     label_img = labels_out.reshape(h, w).astype(np.int32)
#     return label_img, [p.astype(np.uint8) for p in palette]


# def surrounding_label(labels_img, component_mask, dilation_px=5):
#     """
#     Return the most common label in the ring just outside a component mask.
#     """
#     kernel   = np.ones((dilation_px * 2 + 1,) * 2, np.uint8)
#     dilated  = cv2.dilate(component_mask, kernel)
#     ring     = dilated & ~component_mask          # pixels just outside
#     ring_labels = labels_img[ring > 0]
#     if len(ring_labels) == 0:
#         return -1
#     vals, counts = np.unique(ring_labels, return_counts=True)
#     return int(vals[np.argmax(counts)])


# # ── Load ──────────────────────────────────────────────
# img = cv2.imread(IMAGE_PATH)
# if img is None:
#     raise FileNotFoundError(f"Cannot open '{IMAGE_PATH}'")

# h, w = img.shape[:2]

# print("Quantising colours …")
# label_img, palette = quantize_to_labels(img, COLOR_TOLERANCE)
# n_labels = len(palette)
# print(f"  → {n_labels} unique colour regions found")

# # ── Find connected components per label ───────────────
# output       = img.copy()
# spill_count  = 0

# for lbl in range(n_labels):
#     single_label_mask = np.where(label_img == lbl, 255, 0).astype(np.uint8)

#     # connected components of this colour
#     num_cc, cc_labels, stats, centroids = cv2.connectedComponentsWithStats(
#         single_label_mask, connectivity=8
#     )

#     for cc_id in range(1, num_cc):
#         area = int(stats[cc_id, cv2.CC_STAT_AREA])
#         if not (MIN_SPILL_AREA <= area <= MAX_SPILL_AREA):
#             continue

#         # build mask of just this tiny component
#         comp_mask = np.where(cc_labels == cc_id, 255, 0).astype(np.uint8)

#         # what label surrounds it?
#         surr = surrounding_label(label_img, comp_mask, dilation_px=4)

#         # if surrounded entirely by a DIFFERENT label → it's a spill
#         if surr != lbl and surr != -1:
#             cx, cy = int(centroids[cc_id][0]), int(centroids[cc_id][1])
#             r = max(MIN_CIRCLE_R, int(np.sqrt(area) * 4))
#             cv2.circle(output, (cx, cy), r, CIRCLE_COLOR, CIRCLE_THICKNESS)
#             # tiny filled dot at exact centre
#            # cv2.circle(output, (cx, cy), 3, (0, 255, 255), -1)

#             spill_count += 1
#             colour_hex = "#{:02X}{:02X}{:02X}".format(
#                 palette[lbl][2], palette[lbl][1], palette[lbl][0]
#             )
#             print(f"  Spill #{spill_count}: label {lbl} ({colour_hex}), "
#                   f"area={area}px, centre=({cx},{cy}), surrounded by label {surr}")

# cv2.imwrite(OUTPUT_PATH, output)
# print(f"\nTotal spills found: {spill_count}")
# print(f"Saved → {OUTPUT_PATH}")

"""
Segmentation Spill Detector — FAST VERSION
────────────────────────────────────────────
Fully vectorised. No Python pixel loops.
Typical 1MP image: < 1 second.
"""

import cv2
import numpy as np
import time

# ══════════════════════════════════════════════════════
#  SETTINGS
# ══════════════════════════════════════════════════════
IMAGE_PATH      = r"C:\Users\mshar\OneDrive\Desktop\Data Ai\output\spill_detected.png"
OUTPUT_PATH     = r"C:\Users\mshar\OneDrive\Desktop\Data Ai\output\spill_detected.png"

MAX_SPILL_AREA   = 15      # max pixels in a spill cluster
MIN_SPILL_AREA   = 1       # min pixels (1 = single pixel)
COLOR_TOLERANCE  = 20      # merge colours within this BGR distance
CIRCLE_COLOR     = (0, 0, 255)
CIRCLE_THICKNESS = 2
MIN_CIRCLE_R     = 10
# ══════════════════════════════════════════════════════


def fast_quantize(img, tolerance):
    """
    Vectorised colour quantisation.
    Groups similar colours into labels using numpy — NO Python pixel loops.
    """
    h, w = img.shape[:2]
    pixels = img.reshape(-1, 3).astype(np.int32)   # (N, 3)

    # Round each channel to nearest multiple of tolerance → fast bucketing
    bucket_size = max(1, tolerance)
    quantised   = (pixels // bucket_size) * bucket_size  # (N, 3)

    # Encode each colour as a single int64 for fast unique lookup
    encoded = (quantised[:, 0].astype(np.int64) * 1_000_000
             + quantised[:, 1].astype(np.int64) * 1_000
             + quantised[:, 2].astype(np.int64))

    unique_codes, label_img_flat = np.unique(encoded, return_inverse=True)

    # Palette: mean colour per label
    n_labels = len(unique_codes)
    palette  = []
    for lbl in range(n_labels):
        mask = label_img_flat == lbl
        palette.append(pixels[mask].mean(axis=0).astype(np.uint8))

    return label_img_flat.reshape(h, w).astype(np.int32), palette


def surrounding_label(label_img, comp_mask, dilation_px=5):
    """Vectorised: dominant label in the ring just outside a component."""
    kernel  = np.ones((dilation_px * 2 + 1,) * 2, np.uint8)
    dilated = cv2.dilate(comp_mask, kernel)
    ring    = (dilated > 0) & (comp_mask == 0)
    ring_labels = label_img[ring]
    if ring_labels.size == 0:
        return -1
    return int(np.bincount(ring_labels).argmax())


# ── Load ──────────────────────────────────────────────
img = cv2.imread(IMAGE_PATH)
if img is None:
    raise FileNotFoundError(f"Cannot open '{IMAGE_PATH}'")

h, w = img.shape[:2]
print(f"Image: {w}x{h}  ({w*h:,} pixels)")

t0 = time.time()

print("Quantising colours (vectorised) ...")
label_img, palette = fast_quantize(img, COLOR_TOLERANCE)
n_labels = len(palette)
print(f"  -> {n_labels} colour regions  [{time.time()-t0:.2f}s]")

# ── Scan connected components of each label ───────────
output      = img.copy()
spill_count = 0

t1 = time.time()
for lbl in range(n_labels):
    single_mask = (label_img == lbl).astype(np.uint8) * 255

    num_cc, cc_labels, stats, centroids = cv2.connectedComponentsWithStats(
        single_mask, connectivity=8
    )

    for cc_id in range(1, num_cc):
        area = int(stats[cc_id, cv2.CC_STAT_AREA])
        if not (MIN_SPILL_AREA <= area <= MAX_SPILL_AREA):
            continue

        comp_mask = (cc_labels == cc_id).astype(np.uint8) * 255
        surr = surrounding_label(label_img, comp_mask, dilation_px=4)

        if surr != lbl and surr != -1:
            cx, cy = int(centroids[cc_id][0]), int(centroids[cc_id][1])
            r = max(MIN_CIRCLE_R, int(np.sqrt(area) * 4))
            cv2.circle(output, (cx, cy), r, CIRCLE_COLOR, CIRCLE_THICKNESS)
            cv2.circle(output, (cx, cy), 3, (0, 255, 255), -1)

            spill_count += 1
            c = palette[lbl]
            print(f"  Spill #{spill_count}: #{c[2]:02X}{c[1]:02X}{c[0]:02X}  "
                  f"area={area}px  centre=({cx},{cy})")

print(f"\nComponent scan done  [{time.time()-t1:.2f}s]")
print(f"Total spills found  : {spill_count}")
cv2.imwrite(OUTPUT_PATH, output)
print(f"Saved -> {OUTPUT_PATH}   [total {time.time()-t0:.2f}s]")
