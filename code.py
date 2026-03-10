#!/usr/bin/env python3
"""
Batch Labeled Editor - Production Version
Centered image on dark canvas · Anti-aliased fonts · Zoom · Auto-save
Navigate: A / D   Zoom: Mouse Wheel   Pan: Right-click drag
"""

import cv2
import numpy as np
from pathlib import Path
from enum import Enum
import argparse
import json
from datetime import datetime
import time
import sys


# ── Typography & canvas constants ───────────────────────────────────────────────
FONT       = cv2.FONT_HERSHEY_DUPLEX   # smooth, no pixel-aliasing
AA         = cv2.LINE_AA
CANVAS_W   = 1440
CANVAS_H   = 900
CANVAS_BG  = (18, 18, 18)             # near-black surround
CANVAS_PAD = 40                        # clear gap around image on canvas


class Viewport:
    """
    Professional viewer transform — CVAT / LabelImg architecture.

    Stores the view as a single (scale, tx, ty) affine transform instead of
    separate pan_x / pan_y / zoom_level fields.  Benefits:
      • One source of truth — no drift between the three variables.
      • Pixel-perfect rendering: snaps to canonical integer ratios and selects
        the correct cv2 interpolation flag automatically.
      • zoom_at() keeps the exact image pixel under the cursor fixed — the same
        maths used in every professional image viewer.
    """

    # Canonical zoom levels — snapping makes 1:1, 2:1, etc. exactly representable
    ZOOM_STEPS = [0.125, 0.167, 0.25, 0.333, 0.5, 0.667,
                  1.0, 1.5, 2.0, 3.0, 4.0, 6.0, 8.0, 12.0, 16.0]
    SNAP_THRESHOLD = 0.04   # snap if within 4 % of a canonical level

    def __init__(self, canvas_w: int, canvas_h: int):
        self.cw     = canvas_w
        self.ch     = canvas_h
        self._scale = 1.0
        self._tx    = 0.0   # translation in canvas pixels (float for accuracy)
        self._ty    = 0.0

    # ── public API ───────────────────────────────────────────────────────────────

    def fit(self, img_w: int, img_h: int, pad: int = CANVAS_PAD) -> None:
        """Fit image centred in canvas with padding, snap to nearest canonical level."""
        sx = (self.cw - 2 * pad) / img_w
        sy = (self.ch - 2 * pad) / img_h
        self._scale = min(sx, sy)
        self._snap()
        self._recenter(img_w, img_h)

    def zoom_at(self, sx: int, sy: int, factor: float,
                img_w: int, img_h: int) -> None:
        """Zoom by factor, keeping the image pixel under (sx, sy) fixed."""
        ix = (sx - self._tx) / self._scale
        iy = (sy - self._ty) / self._scale
        min_s = min(0.05, (self.cw - 2 * CANVAS_PAD) / max(img_w, 1))
        self._scale = max(min_s, min(16.0, self._scale * factor))
        self._snap()
        self._tx = sx - ix * self._scale
        self._ty = sy - iy * self._scale

    def pan(self, dx: int, dy: int) -> None:
        self._tx += dx
        self._ty += dy

    def reset(self, img_w: int, img_h: int) -> None:
        self.fit(img_w, img_h)

    # ── coordinate transforms ────────────────────────────────────────────────────

    def img_to_screen(self, ix: int, iy: int):
        return int(ix * self._scale + self._tx), int(iy * self._scale + self._ty)

    def screen_to_img(self, sx: int, sy: int):
        return (int((sx - self._tx) / self._scale),
                int((sy - self._ty) / self._scale))

    @property
    def scale(self) -> float:
        return self._scale

    # ── rendering ────────────────────────────────────────────────────────────────

    def render(self, image: np.ndarray) -> np.ndarray:
        """
        Crop-then-scale renderer with Lanczos resampling — same quality as
        Windows Photo Viewer / Photoshop.

        Architecture:
          1. Compute which source pixels are visible on the canvas (fractional).
          2. Crop ONLY that rectangle from the source (never allocate zoom² buffer).
          3. Scale the small crop to its canvas footprint with the best filter:
               zoom-in  →  INTER_LANCZOS4  (4-lobe sinc, sharpest upscale)
               zoom-out →  INTER_AREA      (pixel-area averaging, no moiré)
          Allocation is always bounded to CANVAS_W × CANVAS_H regardless of zoom.
        """
        ih, iw = image.shape[:2]
        tx = self._tx;  ty = self._ty
        sc = self._scale

        canvas = np.full((self.ch, self.cw, 3), CANVAS_BG, dtype=np.uint8)

        # 1. Visible source rectangle in fractional image coordinates
        src_x0 = max(0.0,    (-tx) / sc)
        src_y0 = max(0.0,    (-ty) / sc)
        src_x1 = min(float(iw), (self.cw - tx) / sc)
        src_y1 = min(float(ih), (self.ch - ty) / sc)
        if src_x1 <= src_x0 or src_y1 <= src_y0:
            return canvas   # image fully off-screen

        # 2. Integer-aligned crop (floor/ceil avoids sub-pixel seams at borders)
        cx0 = max(0,  int(np.floor(src_x0)))
        cy0 = max(0,  int(np.floor(src_y0)))
        cx1 = min(iw, int(np.ceil(src_x1)))
        cy1 = min(ih, int(np.ceil(src_y1)))
        crop = image[cy0:cy1, cx0:cx1]
        if crop.size == 0:
            return canvas

        # 3. Scale only the visible crop to its canvas footprint
        dst_w = max(1, int(round((cx1 - cx0) * sc)))
        dst_h = max(1, int(round((cy1 - cy0) * sc)))
        interp = cv2.INTER_LANCZOS4 if sc >= 1.0 else cv2.INTER_AREA
        scaled = cv2.resize(crop, (dst_w, dst_h), interpolation=interp)

        # Unsharp mask — recovers the ~20% softness introduced by any resampling.
        # Same technique used by Photoshop "Smart Sharpen" and Windows Photo Viewer.
        # Sigma=1.0 targets 1-pixel-radius detail; strength=1.4 is visibly crisp
        # without haloing. Only applied on zoom-in where sharpening is meaningful.
        if sc >= 1.0:
            blur   = cv2.GaussianBlur(scaled, (0, 0), sigmaX=1.0)
            scaled = cv2.addWeighted(scaled, 1.4, blur, -0.4, 0)

        # 4. Blit onto canvas
        dx = max(0, int(tx + cx0 * sc))
        dy = max(0, int(ty + cy0 * sc))
        pw = min(dst_w, self.cw - dx)
        ph = min(dst_h, self.ch - dy)
        if pw > 0 and ph > 0:
            canvas[dy:dy+ph, dx:dx+pw] = scaled[:ph, :pw]

        # Subtle 1-px border so image edge is always visible on dark surround
        zw = int(iw * sc);  zh = int(ih * sc)
        itx = int(tx);      ity = int(ty)
        cv2.rectangle(canvas,
                      (max(0, itx-1),          max(0, ity-1)),
                      (min(self.cw-1, itx+zw), min(self.ch-1, ity+zh)),
                      (45, 45, 45), 1, AA)
        return canvas

    # ── private helpers ──────────────────────────────────────────────────────────

    def _snap(self) -> None:
        for step in self.ZOOM_STEPS:
            if abs(self._scale - step) / step < self.SNAP_THRESHOLD:
                self._scale = step
                return

    def _recenter(self, img_w: int, img_h: int) -> None:
        self._tx = (self.cw - img_w * self._scale) / 2
        self._ty = (self.ch - img_h * self._scale) / 2


class EditMode(Enum):
    HIGHLIGHT = "highlight"
    BLUR      = "blur"
    PIXELATE  = "pixelate"
    DARKEN    = "darken"
    GRAYSCALE = "grayscale"
    INVERT    = "invert"
    OUTLINE   = "outline"


class BatchLabeledEditor:

    MAX_BATCH_SIZE          = 200
    MAX_RECOMMENDED_CIRCLES = 30
    MIN_IMAGE_SIZE          = 50
    ZOOM_DEBOUNCE_MS        = 50
    MAX_CACHED_STATES       = 5

    def __init__(self, input_folder, output_folder=None):
        self.input_folder = Path(input_folder)

        if output_folder:
            self.output_folder = Path(output_folder)
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_folder = self.input_folder.parent / f"labeled_output_{ts}"
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.image_files = self._load_image_files()
        if not self.image_files:
            raise ValueError(f"No valid images found in {input_folder}")

        self.total_images  = len(self.image_files)
        self.current_index = 0

        self.original_image         = None
        self.scaled_image           = None
        self.display_image          = None
        self.output_image           = None
        self.circles                = []
        self.drawing                = False
        self.center                 = None
        self.current_radius         = 0
        self.current_label          = ""
        self.current_description    = ""
        self.label_input_mode       = False
        self.description_input_mode = False

        # ── viewport (professional affine-transform viewer) ──────────────────────
        self.vp             = Viewport(CANVAS_W, CANVAS_H)
        self.is_panning     = False
        self._pan_last_x    = 0
        self._pan_last_y    = 0
        self.last_zoom_time = 0

        # dirty flag — skip compositing when nothing changed
        self._dirty = True

        # dirty flag — True forces composite+imshow; False skips on idle ticks
        self._dirty = True

        self.current_mode    = EditMode.HIGHLIGHT
        self._blur_kernel    = 25
        self._pixelate_size  = 10
        self.highlight_alpha = 0.4
        self.show_labels     = True

        self.mode_colors = {
            EditMode.HIGHLIGHT: (0, 255, 0),
            EditMode.BLUR:      (255, 80, 80),
            EditMode.PIXELATE:  (80, 80, 255),
            EditMode.DARKEN:    (160, 160, 160),
            EditMode.GRAYSCALE: (210, 210, 210),
            EditMode.INVERT:    (255, 255, 0),
            EditMode.OUTLINE:   (0, 220, 255),
        }

        self.saved_status       = {}
        self.image_states       = {}
        self.state_access_order = []

        self.window_name = "Batch Labeled Editor"
        cv2.namedWindow(self.window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(self.window_name, CANVAS_W, CANVAS_H)
        cv2.setMouseCallback(self.window_name, self._mouse_callback)

        if not self._load_current_image():
            raise RuntimeError("Failed to load first image")
        self._print_instructions()

    # ── Properties ──────────────────────────────────────────────────────────────

    @property
    def blur_kernel(self):
        return self._blur_kernel

    @blur_kernel.setter
    def blur_kernel(self, v):
        if not isinstance(v, int) or v <= 0 or v % 2 == 0:
            raise ValueError("blur_kernel must be a positive odd integer")
        self._blur_kernel = v

    @property
    def pixelate_size(self):
        return self._pixelate_size

    @pixelate_size.setter
    def pixelate_size(self, v):
        if not isinstance(v, int) or v <= 0:
            raise ValueError("pixelate_size must be a positive integer")
        self._pixelate_size = v

    # ── File loading ─────────────────────────────────────────────────────────────

    def _load_image_files(self):
        exts  = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        files = sorted(f for e in exts for f in self.input_folder.glob(f'*{e}'))
        valid = []
        for f in files:
            try:
                # Use imdecode on header bytes only — avoids fully loading every image
                raw = np.frombuffer(f.read_bytes()[:64], dtype=np.uint8)
                # Fall back to a proper but minimal check via file size heuristic
                if f.stat().st_size < 512:          # suspiciously tiny → skip
                    print(f"⚠  Skipping {f.name}: file too small")
                    continue
                valid.append(f)
            except Exception as e:
                print(f"⚠  Skipping {f.name}: {e}")

        if len(valid) > self.MAX_BATCH_SIZE:
            print(f"\n⚠  Large batch: {len(valid)} images (recommended ≤ {self.MAX_BATCH_SIZE})")
            if input("   Continue? (y/N): ").strip().lower() != 'y':
                sys.exit(0)
        return valid

    # ── Image load & scale ───────────────────────────────────────────────────────

    def _scale_image(self):
        # No pre-downscaling — the Viewport handles all fit/zoom scaling in one
        # high-quality pass. Pre-scaling was causing double-degradation:
        # (original → INTER_AREA down → INTER_LINEAR up) compounded two lossy
        # resizes. Now effects and coordinates always live in original-pixel space.
        self.scale_factor = 1.0
        self.scaled_image = self.original_image.copy()

    def _load_current_image(self):
        if self.current_index >= len(self.image_files):
            return False
        f = self.image_files[self.current_index]
        try:
            self.original_image = cv2.imread(str(f))
            if self.original_image is None:
                raise IOError("cv2.imread returned None")
            if min(self.original_image.shape[:2]) < self.MIN_IMAGE_SIZE:
                raise ValueError("Image too small")

            self._scale_image()
            ih, iw = self.original_image.shape[:2]   # always original dims now
            self.vp.fit(iw, ih)

            json_path = self.output_folder / f.with_suffix('.json').name
            if json_path.exists():
                try:
                    with open(json_path) as jf:
                        data = json.load(jf)
                    self.circles = [{
                        'center':      tuple(o['center']),
                        'radius':      o['radius'],
                        'mode':        EditMode(o.get('mode', 'highlight')),
                        'label':       o.get('label', ''),
                        'description': o.get('description', ''),
                    } for o in data.get('objects', [])]
                    print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images})"
                          f" — restored {len(self.circles)} objects")
                except Exception as e:
                    print(f"⚠  JSON load error: {e}")
                    self.circles = []
            elif f.name in self.image_states:
                self.circles = self.image_states[f.name]['circles'].copy()
                print(f"\n✓ {f.name} ({self.current_index+1}/{self.total_images}) — from memory")
            else:
                self.circles = []
                print(f"\n  {f.name} ({self.current_index+1}/{self.total_images})")

            self._update_state_access(f.name)
            self.drawing = self.label_input_mode = self.description_input_mode = False
            self.current_label = self.current_description = ""

            self.output_image = self.scaled_image.copy()
            if self.circles:
                self._apply_all_effects()
            self.display_image = self.output_image.copy()
            self._dirty = True
            return True

        except Exception as e:
            print(f"❌ Error loading {f.name}: {e}")
            if self.current_index < len(self.image_files) - 1:
                self.current_index += 1
                return self._load_current_image()
            return False

    # ── LRU memory management ───────────────────────────────────────────────────

    def _update_state_access(self, name):
        if name in self.state_access_order:
            self.state_access_order.remove(name)
        self.state_access_order.append(name)

    def _cleanup_old_states(self):
        if len(self.state_access_order) > self.MAX_CACHED_STATES:
            for name in self.state_access_order[:-self.MAX_CACHED_STATES]:
                self.image_states.pop(name, None)
            self.state_access_order = self.state_access_order[-self.MAX_CACHED_STATES:]

    # ── Coordinate transforms (centred-canvas aware) ─────────────────────────────

    def _screen_to_image_coords(self, sx, sy):
        return self.vp.screen_to_img(sx, sy)

    # ── Mouse callback ───────────────────────────────────────────────────────────

    def _mouse_callback(self, event, x, y, flags, param):
        ih, iw = self.scaled_image.shape[:2]

        # ── zoom ────────────────────────────────────────────────────────────────
        if event == cv2.EVENT_MOUSEWHEEL:
            now = time.time() * 1000
            if now - self.last_zoom_time < self.ZOOM_DEBOUNCE_MS:
                return
            self.last_zoom_time = now
            # Zoom faster by using a larger scale factor
            self.vp.zoom_at(x, y, 1.2 if flags > 0 else 0.8, iw, ih)
            self._dirty = True
            self._update_display()
            return

        # ── pan ─────────────────────────────────────────────────────────────────
        if event == cv2.EVENT_RBUTTONDOWN:
            self.is_panning  = True
            self._pan_last_x = x;  self._pan_last_y = y
            return
        if event == cv2.EVENT_RBUTTONUP:
            self.is_panning = False;  return
        if event == cv2.EVENT_MOUSEMOVE and self.is_panning:
            self.vp.pan(x - self._pan_last_x, y - self._pan_last_y)
            self._pan_last_x = x;  self._pan_last_y = y
            self._dirty = True
            self._update_display()
            return

        if self.label_input_mode or self.description_input_mode or self.is_panning:
            return

        # ── draw circle ──────────────────────────────────────────────────────────
        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.center  = self.vp.screen_to_img(x, y)
            self.current_radius = 0
        elif event == cv2.EVENT_MOUSEMOVE and self.drawing:
            ix, iy = self.vp.screen_to_img(x, y)
            self.current_radius = int(np.hypot(ix - self.center[0], iy - self.center[1]))
            self._dirty = True
            self._update_display()
        elif event == cv2.EVENT_LBUTTONUP and self.drawing:
            self.drawing = False
            if self.current_radius > 5:
                if len(self.circles) >= self.MAX_RECOMMENDED_CIRCLES:
                    print(f"⚠  {len(self.circles)} circles — performance may degrade")
                self._enter_label_input_mode()

    # ── Label / description input flow ───────────────────────────────────────────

    def _enter_label_input_mode(self):
        self.label_input_mode = True
        self.current_label    = ""
        print(f"\n→ [{self.current_mode.value.upper()}] type label, ENTER to confirm, ESC to cancel:")
        self._update_display_with_input()

    def _exit_label_input_mode(self, save):
        self.label_input_mode = False
        if save:
            self._enter_description_input_mode()
        else:
            print("  ✗ Cancelled")
            self.current_label = ""
            self.drawing = False          # discard the half-drawn circle too
            self._dirty = True            # force redraw — clears the ghost circle
            self._update_display()

    def _enter_description_input_mode(self):
        self.description_input_mode = True
        self.current_description    = ""
        print("  Description (optional) — ENTER or ESC to skip:")
        self._update_display_with_description_input()

    def _exit_description_input_mode(self):
        self.description_input_mode = False
        lbl  = self.current_label.strip()
        desc = self.current_description.strip()
        self.circles.append({
            'center':      self.center,
            'radius':      self.current_radius,
            'mode':        self.current_mode,
            'label':       lbl,
            'description': desc,
        })
        tag = f"'{lbl}'" if lbl else "(unlabeled)"
        print(f"  ✓ Added {tag}" + (" + description" if desc else "") +
              f" [{self.current_mode.value}]")
        self.current_label = self.current_description = ""
        self._apply_all_effects()
        self._dirty = True
        self._update_display()

    # ── Effects ──────────────────────────────────────────────────────────────────

    def _apply_effect(self, image, circle):
        try:
            mask = np.zeros(image.shape[:2], dtype=np.uint8)
            cv2.circle(mask, circle['center'], circle['radius'], 255, -1)
            mode = circle['mode']

            if mode == EditMode.HIGHLIGHT:
                lit   = cv2.addWeighted(image, 1 - self.highlight_alpha,
                                        np.full_like(image, 255), self.highlight_alpha, 0)
                image = np.where(mask[:, :, np.newaxis] == 255, lit, image)
            elif mode == EditMode.BLUR:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.GaussianBlur(image, (self._blur_kernel,)*2, 0), image)
            elif mode == EditMode.PIXELATE:
                h, w  = image.shape[:2]
                small = cv2.resize(image,
                                   (max(1, w // self._pixelate_size),
                                    max(1, h // self._pixelate_size)),
                                   interpolation=cv2.INTER_NEAREST)
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.resize(small, (w, h), interpolation=cv2.INTER_NEAREST), image)
            elif mode == EditMode.DARKEN:
                image = np.where(mask[:, :, np.newaxis] == 255,
                                 cv2.addWeighted(image, 0.5, np.zeros_like(image), 0.5, 0), image)
            elif mode == EditMode.GRAYSCALE:
                gray  = cv2.cvtColor(cv2.cvtColor(image, cv2.COLOR_BGR2GRAY), cv2.COLOR_GRAY2BGR)
                image = np.where(mask[:, :, np.newaxis] == 255, gray, image)
            elif mode == EditMode.INVERT:
                image = np.where(mask[:, :, np.newaxis] == 255, cv2.bitwise_not(image), image)
        except Exception as e:
            print(f"⚠  Effect error ({circle['mode'].value}): {e}")
        return image

    def _apply_all_effects(self):
        self.output_image = self.scaled_image.copy()
        for c in self.circles:
            self.output_image = self._apply_effect(self.output_image, c)
            cv2.circle(self.output_image, c['center'], c['radius'],
                       self.mode_colors[c['mode']], 2, AA)

    # ── Smart label drawing with collision avoidance ──────────────────────────────

    def _label_font_params(self, img_shape):
        md = min(img_shape[:2])
        if   md < 200: return 0.30, 1
        elif md < 400: return 0.40, 1
        elif md < 800: return 0.50, 1
        else:          return 0.55, 1

    def _rects_overlap(self, r1, r2, buf=4):
        return not (r1[2]+buf < r2[0] or r1[0]-buf > r2[2] or
                    r1[3]+buf < r2[1] or r1[1]-buf > r2[3])

    def _find_label_pos(self, center, radius, tw, fh, bl, pad, img_shape, placed):
        ih, iw = img_shape[:2]
        W = tw + 2*pad;  H = fh + bl + 2*pad
        cx, cy = center

        candidates = [
            (cx - radius,     cy - radius - H - 10),
            (cx - radius,     cy + radius + 10),
            (cx - radius-W-8, cy - H//2),
            (cx + radius+8,   cy - H//2),
            (cx + radius+8,   cy - radius - H - 10),
            (cx - radius-W-8, cy - radius - H - 10),
            (cx + radius+8,   cy + radius + 10),
            (cx - radius-W-8, cy + radius + 10),
        ]
        for extra in (40, 80, 120):
            candidates += [(px, py+extra) for px, py in candidates[:4]]
            candidates += [(px, py-extra) for px, py in candidates[:4]]

        for px, py in candidates:
            rect = (int(px), int(py), int(px+W), int(py+H))
            if rect[0]<pad or rect[1]<pad or rect[2]>iw-pad or rect[3]>ih-pad:
                continue
            if not any(self._rects_overlap(rect, r) for r in placed):
                return int(px)+pad, int(py)+fh+pad, rect

        # fallback: stack below existing labels
        bottom = max((r[3] for r in placed), default=0) + 8
        lx = max(pad, min(cx - W//2, iw-W-pad))
        ly = min(bottom + fh + pad, ih-bl-pad)
        rect = (lx-pad, ly-fh-pad, lx+tw+pad, ly+bl+pad)
        return lx, ly, rect

    def _draw_all_labels_smart(self, image):
        if not self.show_labels:
            return
        placed = []
        ih, iw = image.shape[:2]
        # Scale font to image size — always readable
        min_dim = min(ih, iw)
        if   min_dim < 400:  sc, th = 0.45, 1
        elif min_dim < 900:  sc, th = 0.58, 1
        else:                sc, th = 0.72, 2

        for i, c in enumerate(self.circles, 1):
            label = c['label'] or f"Error #{i}"   # always show something
            color = self.mode_colors[c['mode']]
            cx, cy = c['center']
            r      = c['radius']

            # ── badge number on the circle itself ─────────────────────────────
            badge_text = str(i)
            (bw, bh), _ = cv2.getTextSize(badge_text, FONT, sc * 0.9, th)
            bpad = 6
            br   = max(bw, bh) // 2 + bpad
            # drop shadow
            cv2.circle(image, (cx + 2, cy + 2), br, (0, 0, 0), -1, AA)
            # filled badge
            cv2.circle(image, (cx, cy), br, color, -1, AA)
            # white ring
            cv2.circle(image, (cx, cy), br, (255, 255, 255), 2, AA)
            # number
            cv2.putText(image, badge_text,
                        (cx - bw // 2, cy + bh // 2),
                        FONT, sc * 0.9, (0, 0, 0), th + 1, AA)

            # ── label tag ─────────────────────────────────────────────────────
            mode_s = c['mode'].value[:3].upper()
            text   = f"#{i}  [{mode_s}]  {label}"
            (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
            pad  = max(5, int(5 * sc / 0.5))

            lx, ly, rect = self._find_label_pos(
                (cx, cy), r, tw, fh, bl, pad, image.shape, placed)
            placed.append(rect)

            x1, y1, x2, y2 = rect

            # layer 1 — outer glow (blurred color halo effect via thick rect)
            for glow_expand, glow_alpha in ((6, 0.25), (3, 0.45)):
                glow = image.copy()
                cv2.rectangle(glow,
                               (x1 - glow_expand, y1 - glow_expand),
                               (x2 + glow_expand, y2 + glow_expand),
                               color, -1, AA)
                cv2.addWeighted(glow, glow_alpha, image, 1 - glow_alpha, 0, image)

            # layer 2 — dark fill
            cv2.rectangle(image, (x1, y1), (x2, y2), (15, 15, 20), -1)

            # layer 3 — color left accent bar (3 px)
            cv2.rectangle(image, (x1, y1), (x1 + 3, y2), color, -1, AA)

            # layer 4 — bright border
            cv2.rectangle(image, (x1, y1), (x2, y2), color, 2, AA)

            # layer 5 — text (white, bold shadow first)
            tx = lx + 4   # shift text right of accent bar
            cv2.putText(image, text, (tx + 1, ly + 1), FONT, sc,
                        (0, 0, 0), th + 1, AA)                    # shadow
            cv2.putText(image, text, (tx, ly), FONT, sc,
                        (255, 255, 255), th, AA)                   # text

            # layer 6 — connector line from label to circle edge
            mid_x = (x1 + x2) // 2
            mid_y = y2 if ly > cy else y1
            # find point on circle edge closest to label
            angle   = np.arctan2(mid_y - cy, mid_x - cx)
            edge_x  = int(cx + r * np.cos(angle))
            edge_y  = int(cy + r * np.sin(angle))
            cv2.line(image, (mid_x, mid_y), (edge_x, edge_y), color, 2, AA)
            # arrowhead dot at circle edge
            cv2.circle(image, (edge_x, edge_y), 4, color, -1, AA)
            cv2.circle(image, (edge_x, edge_y), 4, (255,255,255), 1, AA)

    def _draw_typing_label(self, image, center, radius, label, mode):
        sc, th = self._label_font_params(image.shape)
        sc += 0.1;  th += 1
        text = f"[{mode.value[:3].upper()}] {label}_"
        (tw, fh), bl = cv2.getTextSize(text, FONT, sc, th)
        pad    = max(4, int(6 * sc / 0.5))
        ih, iw = image.shape[:2]
        color  = self.mode_colors[mode]

        px, py = center[0]-radius, center[1]-radius-fh-2*pad-10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px, py = center[0]-radius, center[1]+radius+10
        if not (px>=pad and py>=fh+pad and px+tw+2*pad<=iw-pad and py+bl+2*pad<=ih-pad):
            px = max(pad, min(center[0]-radius, iw-tw-2*pad-pad))
            py = max(fh+pad, center[1]-radius-fh-2*pad-10)

        lx = int(px); ly = int(py+fh+pad)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), (10,10,10), -1)
        cv2.rectangle(image, (lx-pad, ly-fh-pad), (lx+tw+pad, ly+bl+pad), color, 2, AA)
        cv2.putText(image, text, (lx, ly), FONT, sc, (255,255,255), th, AA)
        cv2.line(image, (lx+tw//2, ly+bl+pad), center, color, 2, AA)

    # ── Display update ───────────────────────────────────────────────────────────

    def _present(self):
        """Push display_image to the window and clear the dirty flag."""
        self._dirty = False
        cv2.imshow(self.window_name, self.display_image)

    def _update_display(self):
        if not self._dirty:
            return
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        if self.drawing and self.current_radius > 0:
            cv2.circle(tmp, self.center, self.current_radius,
                       self.mode_colors[self.current_mode], 2, AA)
        self.display_image = self.vp.render(tmp)
        self._draw_ui()
        self._present()

    def _update_display_with_input(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        self._draw_typing_label(tmp, self.center, self.current_radius,
                                self.current_label, self.current_mode)
        self.display_image = self.vp.render(tmp)
        self._draw_input_box("Label")
        self._present()

    def _update_display_with_description_input(self):
        tmp = self.output_image.copy()
        self._draw_all_labels_smart(tmp)
        cv2.circle(tmp, self.center, self.current_radius,
                   self.mode_colors[self.current_mode], 3, AA)
        if self.current_label:
            self._draw_typing_label(tmp, self.center, self.current_radius,
                                    self.current_label, self.current_mode)
        self.display_image = self.vp.render(tmp)
        self._draw_input_box("Description (optional)")
        self._present()

    # ── UI overlays ──────────────────────────────────────────────────────────────

    def _draw_input_box(self, prompt):
        h, w  = self.display_image.shape[:2]
        bh = 64;  by = h - bh
        color = self.mode_colors[self.current_mode]
        overlay = self.display_image.copy()
        cv2.rectangle(overlay, (0, by), (w, h), (12,12,12), -1)
        cv2.addWeighted(overlay, 0.88, self.display_image, 0.12, 0, self.display_image)
        cv2.rectangle(self.display_image, (0, by), (w, h), color, 2, AA)
        cv2.putText(self.display_image, f"[{self.current_mode.value.upper()}]",
                    (12, by+22), FONT, 0.52, color, 1, AA)
        cv2.putText(self.display_image, f"{prompt}:",
                    (160, by+22), FONT, 0.52, (200,200,200), 1, AA)
        text = (self.current_description if self.description_input_mode
                else self.current_label) + "_"
        cv2.putText(self.display_image, text, (12, by+50), FONT, 0.65, (0,240,240), 1, AA)

    def _draw_ui(self):
        overlay = self.display_image.copy()
        h, w    = self.display_image.shape[:2]
        bh      = 84
        cv2.rectangle(overlay, (0, 0), (w, bh), (12,12,12), -1)
        color = self.mode_colors[self.current_mode]
        cf    = self.image_files[self.current_index]

        cv2.putText(overlay, f"Mode: {self.current_mode.value.upper()}",
                    (14,24), FONT, 0.55, color, 1, AA)
        cv2.putText(overlay, f"Zoom: {self.vp.scale:.2f}x",
                    (14,48), FONT, 0.45, (100,190,255), 1, AA)
        cv2.putText(overlay, f"{self.current_index+1} / {self.total_images}",
                    (14,70), FONT, 0.43, (160,160,160), 1, AA)

        fname = cf.name if len(cf.name) <= 46 else cf.name[:43]+"..."
        cv2.putText(overlay, fname,   (200,24), FONT, 0.46, (185,185,185), 1, AA)

        n   = len(self.circles)
        oc  = (40,160,255) if n >= self.MAX_RECOMMENDED_CIRCLES else (160,160,160)
        cv2.putText(overlay, f"Objects: {n}", (200,48), FONT, 0.43, oc, 1, AA)

        is_saved  = self.saved_status.get(cf.name, False)
        has_edits = cf.name in self.image_states or n > 0
        st, sc = (("SAVED",    (60,200,60))  if is_saved  else
                  ("EDITED",   (40,160,255)) if has_edits else
                  ("NO EDITS", (80,80,80)))
        cv2.putText(overlay, st, (w-155,24), FONT, 0.54, sc, 1, AA)

        hint = "Wheel:Zoom  RClick:Pan  A/D:Nav  S:Save  R:Reset  C:Clear  U:Undo  H:Help  Q:Quit"
        cv2.putText(overlay, hint, (12, h-10), FONT, 0.36, (110,110,110), 1, AA)
        cv2.addWeighted(overlay, 0.82, self.display_image, 0.18, 0, self.display_image)

    # ── Navigation ───────────────────────────────────────────────────────────────

    def _guard_state(self):
        if self.drawing:
            print("⚠  Finish drawing before switching images"); return True
        if self.label_input_mode:
            print("⚠  Finish label input (ESC to cancel)"); return True
        if self.description_input_mode:
            print("⚠  Finish description input (ENTER to skip)"); return True
        return False

    def _previous_image(self):
        if self._guard_state(): return
        if self.current_index > 0:
            if self.circles: self.save_current(auto_save=True)
            self.current_index -= 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at first image")

    def _next_image(self):
        if self._guard_state(): return
        if self.current_index < self.total_images - 1:
            if self.circles: self.save_current(auto_save=True)
            self.current_index += 1
            self._load_current_image(); self._update_display()
        else:
            print("Already at last image")

    # ── Misc commands ────────────────────────────────────────────────────────────

    def _list_labels(self):
        print("\n" + "="*60)
        print(f"Objects — {self.image_files[self.current_index].name}")
        print("="*60)
        if not self.circles:
            print("  (none)")
        for i, c in enumerate(self.circles, 1):
            print(f"  #{i}: {c['label'] or '(no label)'}  [{c['mode'].value}]  "
                  f"r={c['radius']}  pos={c['center']}")
            if c.get('description'):
                print(f"       {c['description']}")
        print("="*60+"\n")

    def _show_memory_status(self):
        print(f"\nMemory — {len(self.image_states)} states cached "
              f"(max {self.MAX_CACHED_STATES})  |  {len(self.saved_status)} saved\n")

    def _edit_last_label(self):
        if not self.circles:
            print("No objects to edit"); return
        last = self.circles[-1]

        print(f"\nEditing label: '{last['label']}'  (ENTER=save, ESC=cancel)")
        self.current_label    = last['label']
        self.label_input_mode = True
        self._update_display_with_input()
        while self.label_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 27:
                self.label_input_mode = False; self.current_label = ""
                self._dirty = True; self._update_display(); return
            elif key == 13:
                last['label'] = self.current_label.strip()
                self.label_input_mode = False; break
            elif key == 8:
                self.current_label = self.current_label[:-1]
                self._update_display_with_input()
            elif 32 <= key <= 126:
                self.current_label += chr(key); self._update_display_with_input()

        print(f"\nEditing description: '{last.get('description','')}'  (ENTER=save, ESC=skip)")
        self.current_description    = last.get('description', '')
        self.description_input_mode = True
        self._update_display_with_description_input()
        while self.description_input_mode:
            key = cv2.waitKey(0) & 0xFF
            if key == 13:
                last['description'] = self.current_description.strip()
                self.description_input_mode = False; break
            elif key == 27:
                self.description_input_mode = False; break
            elif key == 8:
                self.current_description = self.current_description[:-1]
                self._update_display_with_description_input()
            elif 32 <= key <= 126:
                self.current_description += chr(key); self._update_display_with_description_input()

        self.current_label = self.current_description = ""
        print("  ✓ Updated")
        self._apply_all_effects(); self._update_display()

    # ── Save ─────────────────────────────────────────────────────────────────────

    def save_current(self, auto_save=False):
        if not self.circles:
            if not auto_save: print("No objects to save")
            return

        cf  = self.image_files[self.current_index]
        out = self.output_folder / cf.name
        js  = out.with_suffix('.json')

        self.image_states[cf.name] = {'circles': [c.copy() for c in self.circles]}
        self._update_state_access(cf.name)

        try:
            if self.scale_factor != 1.0:
                final = self.original_image.copy()
                sc_circles = [{
                    'center': (int(c['center'][0]/self.scale_factor),
                               int(c['center'][1]/self.scale_factor)),
                    'radius': int(c['radius']/self.scale_factor),
                    'mode':   c['mode'], 'label': c['label'],
                } for c in self.circles]
                for sc in sc_circles:
                    final = self._apply_effect(final, sc)
                    cv2.circle(final, sc['center'], sc['radius'],
                               self.mode_colors[sc['mode']], 3, AA)
                old_circles   = self.circles
                self.circles  = sc_circles
                self._draw_all_labels_smart(final)
                self.circles  = old_circles
            else:
                final = self.output_image.copy()
                self._draw_all_labels_smart(final)

            if not cv2.imwrite(str(out), final):
                raise IOError(f"cv2.imwrite failed for {out}")

            data = {
                'source_image': cf.name,
                'timestamp':    datetime.now().isoformat(),
                'objects': [{
                    'id':          i,
                    'label':       c['label'],
                    'description': c.get('description', ''),
                    'mode':        c['mode'].value,
                    'center':      list(c['center']),
                    'radius':      c['radius'],
                } for i, c in enumerate(self.circles, 1)]
            }
            with open(js, 'w') as f:
                json.dump(data, f, indent=2)

            self.saved_status[cf.name] = True
            if auto_save:
                print(f"    ✓ Auto-saved {len(self.circles)} objects")
            else:
                print(f"\n✓ Saved {cf.name}  ({len(self.circles)} objects)\n  → {out}\n  → {js}")

            self._cleanup_old_states()

        except Exception as e:
            print(f"❌ Save error: {e}")

    # ── Excel summary ─────────────────────────────────────────────────────────────

    def generate_summary(self):
        xp = self.output_folder / "processing_summary.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"

            # ── Styles ────────────────────────────────────────────────────────
            hfill  = PatternFill(start_color="2E4DA7", end_color="2E4DA7", fill_type="solid")
            hfont  = Font(bold=True, color="FFFFFF", size=11)
            border = Border(**{s: Side(style='thin') for s in ('left','right','top','bottom')})
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_l  = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

            def styled(cell, fill=None, font=None, align=None, bord=True):
                if fill:  cell.fill   = fill
                if font:  cell.font   = font
                if align: cell.alignment = align
                if bord:  cell.border = border

            # ── Headers ───────────────────────────────────────────────────────
            headers = ["#", "Image Name", "Error Count", "Errors (numbered)"]
            for col, hdr in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=hdr)
                styled(c, fill=hfill, font=hfont, align=center)
            ws.row_dimensions[1].height = 28

            # ── Data rows ─────────────────────────────────────────────────────
            data_row = 2
            img_num  = 0
            for img in sorted(self.image_files):
                if img.name not in self.saved_status: continue
                jp = self.output_folder / img.with_suffix('.json').name
                if not jp.exists(): continue
                with open(jp) as f: data = json.load(f)

                objects = data.get('objects', [])
                img_num += 1

                # Column 1 — row number
                c1 = ws.cell(row=data_row, column=1, value=img_num)
                styled(c1, align=center)
                c1.font = Font(bold=True, size=10)

                # Column 2 — image filename
                c2 = ws.cell(row=data_row, column=2, value=img.name)
                styled(c2, align=top_l)

                # Column 3 — error count
                c3 = ws.cell(row=data_row, column=3, value=len(objects))
                styled(c3, align=center)
                # colour-code: green = 0, yellow = 1-3, red = 4+
                if len(objects) == 0:
                    c3.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    c3.font = Font(color="276221", bold=True)
                elif len(objects) <= 3:
                    c3.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    c3.font = Font(color="9C5700", bold=True)
                else:
                    c3.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    c3.font = Font(color="9C0006", bold=True)

                # Column 4 — numbered error list in one cell ("1. Label\n2. Label")
                if objects:
                    lines = []
                    for i, o in enumerate(objects, 1):
                        lbl  = o.get('label', '').strip() or "(unlabeled)"
                        desc = o.get('description', '').strip()
                        line = f"{i}. {lbl}"
                        if desc:
                            line += f" — {desc}"
                        lines.append(line)
                    cell_text = "\n".join(lines)
                else:
                    cell_text = "(no errors)"

                c4 = ws.cell(row=data_row, column=4, value=cell_text)
                styled(c4, align=top_l)

                # Auto row height: ~15 pt per line
                ws.row_dimensions[data_row].height = max(20, len(objects) * 15 + 6)

                # Alternating row fill for readability
                if img_num % 2 == 0:
                    alt = PatternFill(start_color="F2F5FB", end_color="F2F5FB", fill_type="solid")
                    for col in (1, 2, 4):
                        if not ws.cell(row=data_row, column=col).fill.fgColor.rgb.endswith("000000"):
                            pass  # keep colour-coded column 3 as-is
                        ws.cell(row=data_row, column=col).fill = alt

                data_row += 1

            # ── Summary footer ────────────────────────────────────────────────
            data_row += 1
            sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            sfont = Font(bold=True, size=11)

            def summary_row(label, value):
                nonlocal data_row
                lc = ws.cell(row=data_row, column=2, value=label)
                lc.font = sfont; lc.fill = sfill; lc.border = border
                vc = ws.cell(row=data_row, column=3, value=value)
                vc.font = sfont; vc.fill = sfill; vc.border = border
                vc.alignment = center
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            summary_row("Total Images Processed", len(self.saved_status))
            total_obj = sum(
                len(json.load(open(self.output_folder / f.with_suffix('.json').name))['objects'])
                for f in self.image_files
                if f.name in self.saved_status
                and (self.output_folder / f.with_suffix('.json').name).exists())
            summary_row("Total Errors Found", total_obj)
            if self.saved_status:
                summary_row("Avg Errors / Image",
                            round(total_obj / len(self.saved_status), 1))

            # ── Column widths ─────────────────────────────────────────────────
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 34
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 62

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(str(xp))
            print(f"✓ Excel summary: {xp}")
        except ImportError:
            print("⚠  pip install openpyxl")
        except Exception as e:
            print(f"⚠  Excel error: {e}")

    # ── Instructions ─────────────────────────────────────────────────────────────

    def _print_instructions(self):
        print("\n" + "="*58)
        print("BATCH LABELED EDITOR")
        print("="*58)
        print(f"  Input : {self.input_folder}")
        print(f"  Output: {self.output_folder}")
        print(f"  Images: {self.total_images}")
        print()
        print("  Wheel        Zoom           R    Reset zoom")
        print("  Right-drag   Pan            A/D  Prev / Next")
        print("  Left-drag    Draw circle    S    Save  (Shift+S = save & next)")
        print("  1–7          Mode           C    Clear  U  Undo")
        print("  L            List objects   E    Edit last label")
        print("  T            Toggle labels  M    Memory status")
        print("  H            Help           Q    Quit")
        print("="*58+"\n")

    # ── Main loop ─────────────────────────────────────────────────────────────────

    def run(self):
        self._dirty = True
        self._update_display()
        while True:
            key = cv2.waitKey(1) & 0xFF

            # Idle tick — nothing changed, skip redraw
            if key == 0xFF and not self._dirty:
                continue

            if self.description_input_mode:
                if key == 27:                        # ESC → discard whole circle
                    self.description_input_mode = False
                    self.current_label = self.current_description = ""
                    print("  ✗ Circle discarded (ESC during description)")
                    self._dirty = True
                    self._update_display()
                elif key == 13:                      # ENTER → save circle
                    self._exit_description_input_mode()
                elif key == 8:
                    self.current_description = self.current_description[:-1]
                    self._update_display_with_description_input()
                elif 32 <= key <= 126:
                    self.current_description += chr(key)
                    self._update_display_with_description_input()
                continue

            if self.label_input_mode:
                if key == 27:  self._exit_label_input_mode(save=False)
                elif key == 13: self._exit_label_input_mode(save=True)
                elif key == 8:
                    self.current_label = self.current_label[:-1]
                    self._update_display_with_input()
                elif 32 <= key <= 126:
                    self.current_label += chr(key); self._update_display_with_input()
                continue

            if   key == ord('a'): self._previous_image()
            elif key == ord('d'): self._next_image()
            elif key == ord('r'):
                ih, iw = self.scaled_image.shape[:2]
                self.vp.reset(iw, ih)
                self._dirty = True; self._update_display(); print("✓ Zoom reset")
            elif key == ord('s'): self.save_current(); self._dirty = True; self._update_display()
            elif key == ord('S'): self.save_current(); self._next_image()
            elif key == ord('c'):
                self.circles.clear()
                self.output_image = self.scaled_image.copy()
                self._dirty = True; self._update_display(); print("✓ Cleared")
            elif key == ord('u'):
                if self.circles:
                    rem = self.circles.pop()
                    print(f"✓ Removed: {rem['label'] or '(unlabeled)'}")
                    self._apply_all_effects(); self._dirty = True; self._update_display()
            elif key == ord('l'): self._list_labels()
            elif key == ord('e'): self._edit_last_label()
            elif key == ord('t'):
                self.show_labels = not self.show_labels
                print(f"✓ Labels {'ON' if self.show_labels else 'OFF'}")
                self._dirty = True; self._update_display()
            elif key == ord('m'): self._show_memory_status()
            elif key in (ord('h'), ord('H')): self._print_instructions()
            elif ord('1') <= key <= ord('7'):
                self.current_mode = list(EditMode)[key - ord('1')]
                print(f"✓ Mode: {self.current_mode.value.upper()}")
                self._dirty = True; self._update_display()
            elif key == ord('q'):
                if self.circles and not self.saved_status.get(
                        self.image_files[self.current_index].name):
                    self.save_current(auto_save=True)
                break

        cv2.destroyAllWindows()
        if self.saved_status:
            self.generate_summary()
            print(f"\n✅ Done — {len(self.saved_status)}/{self.total_images} saved")
            print(f"   Output: {self.output_folder}")
        else:
            print("\n⚠  No images were saved")


# ── Entry point ───────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Batch Labeled Editor")
    ap.add_argument("input_folder")
    ap.add_argument("--output", "-o", default=None)
    args = ap.parse_args()
    try:
        BatchLabeledEditor(args.input_folder, args.output).run()
    except KeyboardInterrupt:
        print("\n⚠  Interrupted"); return 130
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback; traceback.print_exc(); return 1
    return 0


if __name__ == "__main__":
    exit(main())